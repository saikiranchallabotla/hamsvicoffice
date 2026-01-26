#!/usr/bin/env python
"""
Test script to verify workslip upload logic.
Simulates parsing an estimate file and building preview rows.
"""

import os
import sys
import django
from io import BytesIO
from openpyxl import Workbook

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
sys.path.insert(0, os.path.dirname(__file__))
django.setup()

from core.views import (
    find_estimate_sheet_and_header_row,
    parse_estimate_items,
    to_number,
    is_merged_cell
)


def create_test_estimate():
    """Create a simple test estimate workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    
    # Add header row at row 3
    ws['A3'] = 'Sl.No'
    ws['B3'] = 'Quantity'
    ws['C3'] = 'Unit'
    ws['D3'] = 'Description of work'
    ws['E3'] = 'Rate'
    ws['F3'] = 'Per'
    ws['G3'] = 'Unit'
    ws['H3'] = 'Amount'
    
    # Add test items
    ws['A4'] = 1
    ws['B4'] = 100
    ws['C4'] = 'm'
    ws['D4'] = 'Excavation'
    ws['E4'] = 50
    ws['H4'] = 5000
    
    ws['A5'] = 2
    ws['B5'] = 50
    ws['C5'] = 'cum'
    ws['D5'] = 'Concrete'
    ws['E5'] = 200
    ws['H5'] = 10000
    
    ws['A7'] = 'Sub Total'
    ws['H7'] = 15000
    
    return wb


def test_upload_estimate():
    """Test parsing an estimate file."""
    print("=" * 60)
    print("TEST: Workslip Upload Estimate")
    print("=" * 60)
    
    # Create test workbook
    wb = create_test_estimate()
    
    # Find header row
    ws, header_row = find_estimate_sheet_and_header_row(wb)
    print(f"✓ Header row found at row {header_row}")
    
    # Parse items
    items = parse_estimate_items(ws, header_row)
    print(f"✓ Parsed {len(items)} items from estimate")
    
    # Build preview rows (as in workslip view)
    preview_rows = []
    ws_estimate_rows = []
    for idx, item_data in enumerate(items, start=1):
        ws_estimate_rows.append({
            "sl": idx,
            "desc": item_data.get("desc", ""),
            "unit": item_data.get("unit", ""),
            "qty": item_data.get("qty", 0),
            "rate": item_data.get("rate", 0),
        })
        
        preview_rows.append({
            "row_type": "base",
            "key": f"base:row{idx}",
            "sl": idx,
            "desc": item_data.get("desc", ""),
            "unit": item_data.get("unit", ""),
            "qty_est": item_data.get("qty", 0),
            "qty_exec": 0,
            "rate": item_data.get("rate", 0),
        })
    
    # Verify items
    assert len(items) == 2, f"Expected 2 items, got {len(items)}"
    
    item1 = items[0]
    assert item1['desc'] == 'Excavation', f"Item 1 desc mismatch: {item1['desc']}"
    assert item1['qty'] == 100, f"Item 1 qty mismatch: {item1['qty']}"
    assert item1['rate'] == 50, f"Item 1 rate mismatch: {item1['rate']}"
    
    item2 = items[1]
    assert item2['desc'] == 'Concrete', f"Item 2 desc mismatch: {item2['desc']}"
    assert item2['qty'] == 50, f"Item 2 qty mismatch: {item2['qty']}"
    assert item2['rate'] == 200, f"Item 2 rate mismatch: {item2['rate']}"
    
    # Verify preview rows
    assert len(preview_rows) == 2, f"Expected 2 preview rows, got {len(preview_rows)}"
    
    preview1 = preview_rows[0]
    assert preview1['qty_est'] == 100, f"Preview 1 qty_est mismatch: {preview1['qty_est']}"
    assert preview1['qty_exec'] == 0, f"Preview 1 qty_exec should be 0: {preview1['qty_exec']}"
    
    print("✓ Preview rows created successfully")
    print()
    print("Preview Rows:")
    for row in preview_rows:
        print(f"  {row['sl']}. {row['desc']} | {row['unit']} | Qty: {row['qty_est']} | Rate: {row['rate']}")
    
    print()
    print("=" * 60)
    print("✓ ALL TESTS PASSED")
    print("=" * 60)


if __name__ == "__main__":
    test_upload_estimate()
