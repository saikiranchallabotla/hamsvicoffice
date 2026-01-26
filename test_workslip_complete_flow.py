#!/usr/bin/env python
"""
Full integration test: Upload estimate, then verify data persists in session.
"""

import os
import sys
import django

# Setup Django FIRST before any other imports
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
sys.path.insert(0, os.path.dirname(__file__))
django.setup()

from io import BytesIO
from openpyxl import Workbook
from django.test import RequestFactory
from django.contrib.sessions.middleware import SessionMiddleware
from django.contrib.sessions.backends.db import SessionStore

from core.views import workslip
from django.core.files.uploadedfile import InMemoryUploadedFile


def add_session_to_request(request):
    """Add session support to a request object."""
    middleware = SessionMiddleware(lambda x: None)
    middleware.process_request(request)
    request.session.save()


def create_test_estimate_file():
    """Create a test estimate file in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    
    ws['A3'] = 'Sl.No'
    ws['B3'] = 'Quantity'
    ws['C3'] = 'Unit'
    ws['D3'] = 'Description of work'
    ws['E3'] = 'Rate'
    ws['F3'] = 'Per'
    ws['G3'] = 'Unit'
    ws['H3'] = 'Amount'
    
    ws['A4'] = 1
    ws['B4'] = 100
    ws['C4'] = 'm'
    ws['D4'] = 'Excavation'
    ws['E4'] = 50
    ws['H4'] = 5000
    
    ws['A5'] = 2
    ws['B5'] = 50
    ws['C5'] = 'cum'
    ws['D5'] = 'Concrete'
    ws['E5'] = 200
    ws['H5'] = 10000
    
    file_bytes = BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    return file_bytes


def test_complete_workslip_flow():
    """Test complete workslip flow: upload → save → reload."""
    print("=" * 70)
    print("TEST: Complete Workslip Flow - Upload and Session Persistence")
    print("=" * 70)
    print()
    
    factory = RequestFactory()
    
    # STEP 1: Upload estimate
    print("STEP 1: Upload Estimate")
    print("-" * 70)
    
    file_bytes = create_test_estimate_file()
    uploaded_file = InMemoryUploadedFile(
        file_bytes,
        field_name='estimate_file',
        name='test_estimate.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        size=file_bytes.getbuffer().nbytes,
        charset=None
    )
    
    request1 = factory.post('/workslip/', {'action': 'upload_estimate'})
    request1.FILES['estimate_file'] = uploaded_file
    add_session_to_request(request1)
    session_key = request1.session.session_key
    
    response1 = workslip(request1)
    assert response1.status_code == 200
    print(f"✓ Upload completed, session key: {session_key}")
    
    preview_rows_after_upload = request1.session.get('ws_preview_rows', [])
    assert len(preview_rows_after_upload) == 2
    print(f"✓ {len(preview_rows_after_upload)} items parsed and saved to session")
    
    print()
    
    # STEP 2: Simulate user editing quantities
    print("STEP 2: Simulate User Editing Quantities (Client-side JSON)")
    print("-" * 70)
    
    # Simulate JavaScript serializing the exec_map
    exec_map_json = '{"base:row1": "80", "base:row2": "45"}'
    tp_percent = 5.5
    tp_type = "Excess"
    
    print(f"✓ User entered exec quantities:")
    print(f"    Row 1: 80 units (was 100)")
    print(f"    Row 2: 45 units (was 50)")
    print(f"✓ TP set to: {tp_percent}% {tp_type}")
    
    print()
    
    # STEP 3: POST to download_workslip (which should persist session data)
    print("STEP 3: POST Download Workslip with Exec Data")
    print("-" * 70)
    
    # Create new request but with same session
    request2 = factory.post('/workslip/', {
        'action': 'download_workslip',
        'exec_map': exec_map_json,
        'tp_percent': str(tp_percent),
        'tp_type': tp_type,
    })
    
    # Restore session to this request
    session_store = SessionStore(session_key=session_key)
    request2.session = session_store
    request2.session['ws_preview_rows'] = preview_rows_after_upload
    request2.session['ws_estimate_rows'] = request1.session.get('ws_estimate_rows', [])
    request2.session['ws_tp_percent'] = tp_percent
    request2.session['ws_tp_type'] = tp_type
    request2.session.save()
    
    response2 = workslip(request2)
    assert response2.status_code == 200
    print(f"✓ Download request processed")
    
    # Verify session data is still accessible
    assert request2.session.get('ws_preview_rows'), "Session lost preview_rows"
    print(f"✓ Session data preserved: {len(request2.session.get('ws_preview_rows', []))} items")
    
    print()
    
    # STEP 4: Verify session cleanup via clear_all
    print("STEP 4: Clear All Data")
    print("-" * 70)
    
    request3 = factory.post('/workslip/', {'action': 'clear_all'})
    request3.session = session_store
    request3.session.save()
    
    response3 = workslip(request3)
    assert response3.status_code == 200
    print(f"✓ Clear all processed")
    
    # Verify session was cleared
    assert request3.session.get('ws_preview_rows') == []
    assert request3.session.get('ws_estimate_rows') == []
    assert request3.session.get('ws_tp_percent') == 0
    print(f"✓ Session data cleared")
    
    print()
    print("=" * 70)
    print("✓ ALL TESTS PASSED - Complete Workslip Flow Works Correctly")
    print("=" * 70)


if __name__ == "__main__":
    test_complete_workslip_flow()
