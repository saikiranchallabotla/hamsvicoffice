import os
from openpyxl import load_workbook
import json

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
path = os.path.join(BASE, "core", "data", "temp_electrical.xlsx")
print("Using file:", path)
wb_vals = load_workbook(path, data_only=True)
wb_for = load_workbook(path, data_only=False)
ws_vals = wb_vals["Master Datas"]
ws_for = wb_for["Master Datas"]
ws_groups = wb_for["Groups"]

# read groups
groups_map = {}
for r in range(2, ws_groups.max_row + 1):
    item = ws_groups.cell(row=r, column=1).value
    group = ws_groups.cell(row=r, column=2).value
    if item and group:
        groups_map.setdefault(str(group).strip(), []).append(str(item).strip())

# detect items (using yellow fill detection approx)
SCAN_COL_START = 1
SCAN_COL_END = 10

def _is_yellow(cell):
    try:
        fill = cell.fill
        if not fill or not getattr(fill, "patternType", None):
            return False
        rgb = getattr(fill.fgColor, "rgb", None)
        if rgb and str(rgb).upper().endswith("FFFF00"):
            return True
    except Exception:
        return False
    return False

items_list = []
max_row = ws_for.max_row
r = 1
while r <= max_row:
    heading_name = None
    for c in range(SCAN_COL_START, SCAN_COL_END + 1):
        cell = ws_for.cell(row=r, column=c)
        if _is_yellow(cell) and str(cell.value or "").strip():
            heading_name = str(cell.value).strip()
            break
    if heading_name:
        start_row = r
        end_row = max_row
        rr = r + 1
        while rr <= max_row:
            found_next = False
            for c in range(SCAN_COL_START, SCAN_COL_END + 1):
                cell = ws_for.cell(row=rr, column=c)
                if _is_yellow(cell) and str(cell.value or "").strip():
                    found_next = True
                    break
            if found_next:
                end_row = rr - 1
                break
            rr += 1
        items_list.append({"name": heading_name, "start_row": start_row, "end_row": end_row})
        r = end_row + 1
    else:
        r += 1

print(f"Detected {len(items_list)} headings")

# build day_rates

def _safe_float(x):
    try:
        if x is None or x == "":
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).replace(",", "").strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def build_temp_day_rates_local(path, items_list):
    wb_vals = load_workbook(path, data_only=True)
    ws_vals = wb_vals["Master Datas"]
    day_rates = {}
    for it in items_list:
        name = (it.get("name") or "").strip()
        sr = int(it.get("start_row") or 0)
        er = int(it.get("end_row") or 0)
        if not name or sr <= 0 or er <= 0:
            continue
        per_item = {}
        for r in range(sr, er + 1):
            day_cell = ws_vals.cell(row=r, column=3).value
            if day_cell in (None, ""):
                continue
            try:
                day_no = int(float(day_cell))
            except Exception:
                continue
            if day_no <= 0:
                continue
            rate_cached = _safe_float(ws_vals.cell(row=r, column=10).value)
            if rate_cached is not None and rate_cached > 0:
                per_item[str(day_no)] = rate_cached
        if per_item:
            day_rates[name] = per_item
    return day_rates


day_rates = build_temp_day_rates_local(path, items_list)

# UI normalization (JS normalizeName: collapse whitespace)
def norm_js(s):
    return " ".join(str(s or "").split()).strip()

# Now for each group, show mapping coverage
for grp, items in groups_map.items():
    print('\nGroup:', grp)
    for it in items:
        key = norm_js(it)
        found = None
        for dk in day_rates.keys():
            if norm_js(dk) == key:
                found = dk
                break
        print(f"  Item: {it!r} -> matched day_rates key: {found!r}")

# show day_rates keys normalized
print('\nDay rates normalized keys:')
for k in day_rates.keys():
    print(' -', norm_js(k))

print('\nSample day_rates for first key:')
if day_rates:
    k = list(day_rates.keys())[0]
    print(k, day_rates[k])
else:
    print('no day_rates')
