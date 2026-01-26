#!/usr/bin/env python
"""Create a test Excel file with multiple bill sheets for debugging multi-sheet Nth bill generation."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

# Create first bill sheet
ws1 = wb.create_sheet("Bill-1", 0)
ws1['A1'] = "Project: Test"
ws1['A2'] = "Bill No: 1"
ws1['A8'] = "Sl.No"
ws1['B8'] = "Quantity"
ws1['C8'] = "Unit"
ws1['D8'] = "Description of work"
ws1['E8'] = "Rate"
ws1['F8'] = "Per"
ws1['G8'] = "Unit"
ws1['H8'] = "Amount"

ws1['A9'] = 1
ws1['B9'] = 100
ws1['C9'] = "m"
ws1['D9'] = "Excavation"
ws1['E9'] = 50
ws1['F9'] = 1
ws1['G9'] = "m"
ws1['H9'] = 5000

ws1['A10'] = 2
ws1['B10'] = 50
ws1['C10'] = "cum"
ws1['D10'] = "Concrete"
ws1['E10'] = 200
ws1['F10'] = 1
ws1['G10'] = "cum"
ws1['H10'] = 10000

ws1['A12'] = "Sub Total"
ws1['H12'] = 15000

# Create second bill sheet
ws2 = wb.create_sheet("Bill-2", 1)
ws2['A1'] = "Project: Test"
ws2['A2'] = "Bill No: 2"
ws2['A8'] = "Sl.No"
ws2['B8'] = "Quantity"
ws2['C8'] = "Unit"
ws2['D8'] = "Description of work"
ws2['E8'] = "Rate"
ws2['F8'] = "Per"
ws2['G8'] = "Unit"
ws2['H8'] = "Amount"

ws2['A9'] = 1
ws2['B9'] = 50
ws2['C9'] = "m"
ws2['D9'] = "Excavation"
ws2['E9'] = 50
ws2['F9'] = 1
ws2['G9'] = "m"
ws2['H9'] = 2500

ws2['A10'] = 2
ws2['B10'] = 25
ws2['C10'] = "cum"
ws2['D10'] = "Concrete"
ws2['E10'] = 200
ws2['F10'] = 1
ws2['G10'] = "cum"
ws2['H10'] = 5000

ws2['A12'] = "Sub Total"
ws2['H12'] = 7500

# Create third bill sheet
ws3 = wb.create_sheet("Bill-3", 2)
ws3['A1'] = "Project: Test"
ws3['A2'] = "Bill No: 3"
ws3['A8'] = "Sl.No"
ws3['B8'] = "Quantity"
ws3['C8'] = "Unit"
ws3['D8'] = "Description of work"
ws3['E8'] = "Rate"
ws3['F8'] = "Per"
ws3['G8'] = "Unit"
ws3['H8'] = "Amount"

ws3['A9'] = 1
ws3['B9'] = 30
ws3['C9'] = "m"
ws3['D9'] = "Excavation"
ws3['E9'] = 50
ws3['F9'] = 1
ws3['G9'] = "m"
ws3['H9'] = 1500

ws3['A10'] = 2
ws3['B10'] = 15
ws3['C10'] = "cum"
ws3['D10'] = "Concrete"
ws3['E10'] = 200
ws3['F10'] = 1
ws3['G10'] = "cum"
ws3['H10'] = 3000

ws3['A12'] = "Sub Total"
ws3['H12'] = 4500

wb.save('test_multi_bill.xlsx')
print("Created test_multi_bill.xlsx with 3 sheets (Bill-1, Bill-2, Bill-3)")
