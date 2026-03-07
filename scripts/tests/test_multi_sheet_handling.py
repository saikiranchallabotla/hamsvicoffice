#!/usr/bin/env python
"""Test the multi-sheet handling of bill generation."""

import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
sys.path.insert(0, r'c:\Users\HP\Documents\Windows x 1')
django.setup()

from openpyxl import load_workbook, Workbook
from core.views import parse_first_bill_for_nth, _extract_header_data_from_sheet, _populate_nth_bill_sheet

# Load the test file we created
wb_in = load_workbook('test_multi_bill.xlsx', data_only=True)
print(f"Loaded workbook with sheets: {[ws.title for ws in wb_in.worksheets]}")
print(f"Total sheets: {len(wb_in.worksheets)}")

# Test the detection logic
bill_sheets = [ws for ws in wb_in.worksheets if ws.title.startswith("Bill")]
print(f"\nDetection by 'Bill' prefix: Found {len(bill_sheets)} sheets")
print(f"Sheet titles: {[ws.title for ws in bill_sheets]}")

# Test parsing each sheet
print("\nTesting parsing of each sheet:")
for ws in bill_sheets:
    print(f"\nSheet: '{ws.title}'")
    
    # Find header row
    header_row = None
    for r in range(1, 30):
        a = str(ws.cell(row=r, column=1).value or "").strip().lower()
        b = str(ws.cell(row=r, column=2).value or "").strip().lower()
        d = str(ws.cell(row=r, column=4).value or "").strip().lower()
        if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
            header_row = r
            print(f"  Found header row: {r}")
            break
    
    if header_row is None:
        header_row = 10
        print(f"  No header found, using default row 10")
    
    # Parse items
    items = parse_first_bill_for_nth(ws, header_row)
    print(f"  Parsed {len(items)} items:")
    for item in items:
        print(f"    - {item['desc']} (qty: {item.get('qty', '?')}, rate: {item.get('rate', '?')}, amount: {item.get('prev_amount', '?')})")

print("\n\nTest complete!")
