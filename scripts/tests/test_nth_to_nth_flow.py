#!/usr/bin/env python
"""Test Nth→Nth bill generation with multiple sheets."""

import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
sys.path.insert(0, r'c:\Users\HP\Documents\Windows x 1')
django.setup()

from openpyxl import load_workbook, Workbook
from core.views import (
    parse_nth_bill_for_next, 
    _extract_header_data_from_sheet, 
    _populate_nth_bill_sheet,
    read_tp_from_sheet,
    ordinal_word
)

# First, create a multi-sheet Nth bill format file (like output from First→Nth conversion)
print("=" * 60)
print("CREATING TEST NTH BILL FILE (3 sheets)")
print("=" * 60)

wb_nth_test = Workbook()
wb_nth_test.remove(wb_nth_test.active)

# Create 3 Nth bill sheets
for sheet_idx in range(1, 4):
    ws = wb_nth_test.create_sheet(f"Bill-{sheet_idx}", sheet_idx - 1)
    
    # Headers for Nth bill (10 columns)
    ws['A1'] = "Project: Test"
    ws['A2'] = f"2nd Bill - Part {sheet_idx}"
    
    ws['A9'] = "Sl.No"
    ws['B9'] = "Item"
    ws['C9'] = "Qty Till Date"
    ws['D9'] = "Rate"
    ws['E9'] = "Total Value Till Date"
    ws['F9'] = "Deduct Previous"
    ws['G9'] = "Since Last"
    ws['H9'] = "Qty"
    ws['I9'] = "Rate"
    ws['J9'] = "Amount"
    
    # Add data rows (Nth bill has different amounts for "till date" vs "this period")
    row = 10
    ws[f'A{row}'] = 1
    ws[f'B{row}'] = "Excavation"
    ws[f'C{row}'] = 100 - (sheet_idx - 1) * 20  # Qty till date
    ws[f'D{row}'] = 50
    ws[f'E{row}'] = 5000 - (sheet_idx - 1) * 2500  # Total till date
    ws[f'F{row}'] = 2500 - (sheet_idx - 1) * 1000  # Deduct previous
    ws[f'G{row}'] = 2500 - (sheet_idx - 1) * 1000  # Since last
    ws[f'H{row}'] = 50 - (sheet_idx - 1) * 10  # Qty this period
    ws[f'I{row}'] = 50
    ws[f'J{row}'] = 2500 - (sheet_idx - 1) * 1000
    
    row = 11
    ws[f'A{row}'] = 2
    ws[f'B{row}'] = "Concrete"
    ws[f'C{row}'] = 50 - (sheet_idx - 1) * 10  # Qty till date
    ws[f'D{row}'] = 200
    ws[f'E{row}'] = 10000 - (sheet_idx - 1) * 5000  # Total till date
    ws[f'F{row}'] = 5000 - (sheet_idx - 1) * 2500  # Deduct previous
    ws[f'G{row}'] = 5000 - (sheet_idx - 1) * 2500  # Since last
    ws[f'H{row}'] = 25 - (sheet_idx - 1) * 5  # Qty this period
    ws[f'I{row}'] = 200
    ws[f'J{row}'] = 5000 - (sheet_idx - 1) * 2500
    
    # Add subtotal row
    ws['A14'] = "Sub Total"
    ws['J14'] = 7500 - (sheet_idx - 1) * 3750

print(f"Created test Nth bill file with 3 sheets")
wb_nth_test.save('test_nth_bill_input.xlsx')
print(f"Saved to: test_nth_bill_input.xlsx\n")

# Now test the Nth→Nth handler
print("=" * 60)
    print("TESTING NTH TO NTH HANDLER (3RD BILL FROM 2ND BILL)")
# Load the file we just created
wb_in = load_workbook('test_nth_bill_input.xlsx', data_only=True)

# Simulate the handler logic
action = "nth_nth_part"
bill_sheets = [ws for ws in wb_in.worksheets if ws.title.startswith("Bill")]
print(f"\nDetected {len(bill_sheets)} bill sheets: {[ws.title for ws in bill_sheets]}")

if not bill_sheets:
    print("Fallback: Looking for 'Quantity Till Date' header...")
    bill_sheets = []
    for ws in wb_in.worksheets:
        for r in range(1, 30):
            c = str(ws.cell(row=r, column=3).value or "").strip().lower()
            if "quantity till date" in c:
                bill_sheets.append(ws)
                print(f"Found: {ws.title}")
                break

print(f"Total sheets to process: {len(bill_sheets)}\n")

# Extract Nth number
nth_val = 3
ord_word = ordinal_word(nth_val)
title_text = f"CC {ord_word} & Part Bill"

# Build output workbook
wb_out = Workbook()
created = 0

print("Processing sheets...")
for idx, ws_nth in enumerate(bill_sheets, start=1):
    print(f"\n[{idx}] Processing sheet: '{ws_nth.title}'")
    
    # Find header row
    header_row = None
    for r in range(1, 30):
        a = str(ws_nth.cell(row=r, column=1).value or "").strip().lower()
        c = str(ws_nth.cell(row=r, column=3).value or "").strip().lower()
        if "sl" in a and "quantity till date" in c:
            header_row = r
            print(f"    Found header row: {r}")
            break
    
    if header_row is None:
        header_row = 10
        print(f"    Using default header row: 10")
    
    # Extract data
    header_data = _extract_header_data_from_sheet(ws_nth)
    items = parse_nth_bill_for_next(ws_nth, header_row)
    print(f"    Parsed {len(items)} items")
    
    if not items:
        print(f"    No items found, skipping")
        continue
    
    created += 1
    print(f"    Creating output sheet (created={created})")
    
    tp_percent, tp_type = read_tp_from_sheet(ws_nth)
    if tp_percent is None:
        tp_percent = 0.0
    if tp_type is None:
        tp_type = "Excess"
    
    # Determine sheet name
    if len(bill_sheets) > 1:
        sheet_name = f"Bill-{created}"
    else:
        sheet_name = "Bill"
    
    # Determine sheet title
    if len(bill_sheets) > 1:
        sheet_title = f"{title_text} ({created})"
    else:
        sheet_title = title_text
    
    # Create output sheet
    if created == 1:
        ws_out = wb_out.active
        ws_out.title = sheet_name
        print(f"    Using active sheet, named: '{sheet_name}'")
    else:
        ws_out = wb_out.create_sheet(title=sheet_name)
        print(f"    Created new sheet: '{sheet_name}'")
    
    # Populate sheet
    print(f"    Populating sheet with {len(items)} items...")
    _populate_nth_bill_sheet(
        ws_out,
        items=items,
        header_data=header_data,
        title_text=sheet_title,
        tp_percent=tp_percent,
        tp_type=tp_type,
        mb_measure_no="",
        mb_measure_p_from="",
        mb_measure_p_to="",
        mb_abs_no="",
        mb_abs_p_from="",
        mb_abs_p_to="",
        doi="",
        doc="",
        domr="",
        dobr="",
    )
    print(f"    Sheet populated successfully")

print(f"\n\nFinal result: Created {created} output sheets")
print(f"Output workbook sheets: {[ws.title for ws in wb_out.worksheets]}")

# Save output
output_file = "test_nth_to_nth_output.xlsx"
wb_out.save(output_file)
print(f"\nSaved to: {output_file}")
print("\n" + "=" * 60)
print("NTH→NTH TEST COMPLETE")
print("=" * 60)
