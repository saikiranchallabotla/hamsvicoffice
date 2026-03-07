#!/usr/bin/env python
"""
Integration test for workslip view: upload_estimate action.
Tests the full flow of uploading an estimate and populating preview rows.
"""

import os
import sys
import django
from io import BytesIO
from openpyxl import Workbook
from django.test import TestCase, RequestFactory
from django.contrib.sessions.middleware import SessionMiddleware

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
sys.path.insert(0, os.path.dirname(__file__))
django.setup()

from core.views import workslip


def add_session_to_request(request):
    """Add session support to a request object."""
    middleware = SessionMiddleware(lambda x: None)
    middleware.process_request(request)
    request.session.save()


def create_test_estimate_file():
    """Create a test estimate file in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    
    # Add header row at row 3
    ws['A3'] = 'Sl.No'
    ws['B3'] = 'Quantity'
    ws['C3'] = 'Unit'
    ws['D3'] = 'Description of work'
    ws['E3'] = 'Rate'
    ws['F3'] = 'Per'
    ws['G3'] = 'Unit'
    ws['H3'] = 'Amount'
    
    # Add test items
    ws['A4'] = 1
    ws['B4'] = 100
    ws['C4'] = 'm'
    ws['D4'] = 'Excavation'
    ws['E4'] = 50
    ws['H4'] = 5000
    
    ws['A5'] = 2
    ws['B5'] = 50
    ws['C5'] = 'cum'
    ws['D5'] = 'Concrete'
    ws['E5'] = 200
    ws['H5'] = 10000
    
    ws['A7'] = 'Sub Total'
    ws['H7'] = 15000
    
    # Save to BytesIO
    file_bytes = BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    return file_bytes


def test_upload_estimate_action():
    """Test the upload_estimate POST action."""
    print("=" * 60)
    print("TEST: Workslip - Upload Estimate Action")
    print("=" * 60)
    
    factory = RequestFactory()
    
    # Create POST request with file upload
    file_bytes = create_test_estimate_file()
    from django.core.files.uploadedfile import InMemoryUploadedFile
    uploaded_file = InMemoryUploadedFile(
        file_bytes,
        field_name='estimate_file',
        name='test_estimate.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        size=file_bytes.getbuffer().nbytes,
        charset=None
    )
    
    request = factory.post('/workslip/', {
        'action': 'upload_estimate',
    }, HTTP_HOST='testserver')
    request.FILES['estimate_file'] = uploaded_file
    
    # Add session support
    add_session_to_request(request)
    
    # Call workslip view (it returns HttpResponse from render)
    from core.views import workslip
    response = workslip(request)
    
    # For template responses, we need to access context differently
    # response.context is a list of Context objects in render() output
    if hasattr(response, 'context_data'):
        context = response.context_data
    else:
        # Try to get from context attribute
        context = {}
        if hasattr(response, 'context'):
            for ctx in response.context:
                context.update(ctx)
    
    # Alternatively, verify via session data (more reliable)
    preview_rows = request.session.get('ws_preview_rows', [])
    
    # Verify response
    assert response.status_code == 200, f"Expected 200, got {response.status_code}"
    print(f"✓ Response status: {response.status_code}")
    
    # Verify items were parsed via session
    assert len(preview_rows) == 2, f"Expected 2 preview rows, got {len(preview_rows)}"
    print(f"✓ Exactly 2 items parsed from estimate")
    
    # Verify first item
    row1 = preview_rows[0]
    assert row1['desc'] == 'Excavation', f"Item 1 desc mismatch: {row1['desc']}"
    assert row1['qty_est'] == 100, f"Item 1 qty_est mismatch: {row1['qty_est']}"
    assert row1['rate'] == 50, f"Item 1 rate mismatch: {row1['rate']}"
    assert row1['qty_exec'] == 0, f"Item 1 qty_exec should be 0"
    print(f"✓ Item 1 (Excavation): qty=100, rate=50")
    
    # Verify second item
    row2 = preview_rows[1]
    assert row2['desc'] == 'Concrete', f"Item 2 desc mismatch: {row2['desc']}"
    assert row2['qty_est'] == 50, f"Item 2 qty_est mismatch: {row2['qty_est']}"
    assert row2['rate'] == 200, f"Item 2 rate mismatch: {row2['rate']}"
    print(f"✓ Item 2 (Concrete): qty=50, rate=200")
    
    # Verify session data was saved
    assert 'ws_preview_rows' in request.session, "Session 'ws_preview_rows' not saved"
    assert 'ws_estimate_rows' in request.session, "Session 'ws_estimate_rows' not saved"
    print(f"✓ Session data saved correctly")
    
    print()
    print("Preview Rows from Session:")
    for row in preview_rows:
        print(f"  {row['sl']}. {row['desc']} | {row['unit']} | Est.Qty: {row['qty_est']} | Rate: {row['rate']}")
    
    print()
    print("=" * 60)
    print("✓ ALL TESTS PASSED")
    print("=" * 60)


if __name__ == "__main__":
    test_upload_estimate_action()
