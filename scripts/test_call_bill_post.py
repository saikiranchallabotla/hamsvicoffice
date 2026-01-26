import os
import sys
sys.path.insert(0, r"C:\Users\HP\Documents\Windows x 1")
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
import django
django.setup()
from django.test import RequestFactory
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from io import BytesIO
from core import views

# build a tiny workbook with a simple estimate sheet
wb = Workbook()
ws = wb.active
ws.title = 'Estimate1'
# header row
ws['A1'] = 'Sl.'
ws['B1'] = 'Quantity'
ws['C1'] = 'Unit'
ws['D1'] = 'Item'
ws['E1'] = 'Rate'
# one item
ws['A2'] = 1
ws['B2'] = 10
ws['C2'] = 'No'
ws['D2'] = 'Test item'
ws['E2'] = 100
buf = BytesIO()
wb.save(buf)
buf.seek(0)
file_data = buf.read()
uploaded = SimpleUploadedFile('test.xlsx', file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

factory = RequestFactory()
req = factory.post('/bill/', {'action': 'estimate_first_part'}, files={'bill_file': uploaded})
res = views.bill(req)
print('TYPE:', type(res))
print('VALUE:', res)
if res is None:
    print('Returned None')
else:
    print('Status:', getattr(res, 'status_code', None))
