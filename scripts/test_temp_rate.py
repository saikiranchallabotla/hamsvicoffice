import os
import sys
from openpyxl import load_workbook

# Ensure project root is on sys.path so `import core` works when running this script
BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, BASE)

path = os.path.join(BASE, "core", "data", "temp_electrical.xlsx")
print("Using file:", path)
wb = load_workbook(path, data_only=False)
ws = wb["Master Datas"]
# We'll implement a small local build_temp_day_rates to avoid importing from core.utils_excel

# Minimal local re-implementation of detection & rate extraction to avoid importing Django models
SCAN_COL_START = 1
SCAN_COL_END = 10

def _is_yellow_and_red(cell):
    # For this quick check we only look at cell fill rgb if available
    try:
        fill = cell.fill
        if not fill or not getattr(fill, "patternType", None):
            return False
        rgb = getattr(fill.fgColor, "rgb", None)
        if rgb and str(rgb).upper().endswith("FFFF00"):
            # ignore red-text check for this smoke test
            return True
    except Exception:
        return False
    return False

def detect_items(ws):
    items = []
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        heading_name = None
        for c in range(SCAN_COL_START, SCAN_COL_END + 1):
            cell = ws.cell(row=r, column=c)
            if _is_yellow_and_red(cell) and str(cell.value or "").strip():
                heading_name = str(cell.value).strip()
                break
        if heading_name:
            start_row = r
            end_row = max_row
            rr = r + 1
            while rr <= max_row:
                found_next = False
                for c in range(SCAN_COL_START, SCAN_COL_END + 1):
                    cell = ws.cell(row=rr, column=c)
                    if _is_yellow_and_red(cell) and str(cell.value or "").strip():
                        found_next = True
                        break
                if found_next:
                    end_row = rr - 1
                    break
                rr += 1
            items.append({"name": heading_name, "start_row": start_row, "end_row": end_row})
            r = end_row + 1
        else:
            r += 1
    return items

def get_item_description_and_rate(ws_data, item_info):
    desc = ws_data.cell(item_info["start_row"] + 2, 4).value
    desc = str(desc).strip() if desc else ""
    rate = ""
    for r in range(item_info["end_row"], item_info["start_row"] - 1, -1):
        val = ws_data.cell(r, 10).value
        if val not in (None, ""):
            rate = val
            break
    return desc, rate
items = detect_items(ws)

print(f"Detected {len(items)} items")
for it in items[:20]:
    desc, rate = get_item_description_and_rate(ws, it)
    print(f"Item: {it['name']} (rows {it['start_row']}-{it['end_row']}) -> desc=\"{desc}\" rate=\"{rate}\"")
def _safe_float(x):
    try:
        if x is None or x == "":
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).replace(",", "").strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None

def build_temp_day_rates_local(path, items_list):
    wb_vals = load_workbook(path, data_only=True)
    ws_vals = wb_vals["Master Datas"]
    day_rates = {}
    for it in items_list:
        name = (it.get("name") or "").strip()
        sr = int(it.get("start_row") or 0)
        er = int(it.get("end_row") or 0)
        if not name or sr <= 0 or er <= 0:
            continue
        per_item = {}
        for r in range(sr, er + 1):
            day_cell = ws_vals.cell(row=r, column=3).value
            if day_cell in (None, ""):
                continue
            try:
                day_no = int(float(day_cell))
            except Exception:
                continue
            if day_no <= 0:
                continue
            rate_cached = _safe_float(ws_vals.cell(row=r, column=10).value)
            if rate_cached is not None and rate_cached > 0:
                per_item[str(day_no)] = rate_cached
        if per_item:
            day_rates[name] = per_item
    return day_rates

day_rates = build_temp_day_rates_local(path, items)
print("Day rates keys:", list(day_rates.keys())[:10])
for k, v in list(day_rates.items())[:5]:
    print(k, v)
