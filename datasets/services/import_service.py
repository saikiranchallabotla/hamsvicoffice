# datasets/services/import_service.py
"""
Dataset Import Service.

Handles Excel/CSV master data uploads with:
- Dry-run validation
- Row parsing with error collection
- Version creation and publishing
- Rollback to previous versions

Usage:
    from datasets.services.import_service import DatasetImportService
    
    # Start import (dry run)
    result = DatasetImportService.start_import(dataset_id, file, dry_run=True, user=admin)
    
    # If valid, publish
    result = DatasetImportService.publish_version(dataset_id, import_job_id)
    
    # Rollback if needed
    result = DatasetImportService.rollback(dataset_id, version_id)
"""

import csv
import io
import json
import logging
from typing import Optional, List, Dict, Any, Iterator, Tuple
from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.utils import timezone
from django.db import transaction
from django.core.files.uploadedfile import UploadedFile

logger = logging.getLogger(__name__)


class DatasetImportService:
    """
    Dataset import with validation, versioning, and rollback.
    """
    
    # Supported file types
    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    MAX_ERRORS_DISPLAYED = 100
    
    # ==========================================================================
    # MAIN IMPORT WORKFLOW
    # ==========================================================================
    
    @classmethod
    def start_import(
        cls,
        dataset_id: int,
        uploaded_file: UploadedFile,
        dry_run: bool = True,
        requested_by=None
    ) -> Dict[str, Any]:
        """
        Start dataset import process.
        
        Args:
            dataset_id: Target dataset ID
            uploaded_file: Django UploadedFile (Excel/CSV)
            dry_run: If True, validate only without publishing
            requested_by: Admin user who initiated import
        
        Returns:
            {ok: bool, reason: str, data: {import_job_id, errors, row_count, ...}}
        """
        from datasets.models import Dataset, DatasetImportJob
        
        # Validate dataset exists
        try:
            dataset = Dataset.objects.get(id=dataset_id)
        except Dataset.DoesNotExist:
            return cls._fail("Dataset not found.", code="DATASET_NOT_FOUND")
        
        # Validate file
        file_errors = cls.validate(uploaded_file)
        if file_errors:
            return cls._fail(
                "File validation failed.",
                code="VALIDATION_ERROR",
                data={"errors": file_errors}
            )
        
        # Create import job
        import_job = DatasetImportJob.objects.create(
            dataset=dataset,
            file_name=uploaded_file.name,
            file_size=uploaded_file.size,
            status='processing',
            initiated_by=requested_by,
            is_dry_run=dry_run,
        )
        
        try:
            # Parse and validate rows
            rows, row_errors, row_count = cls._process_file(uploaded_file, dataset)
            
            # Update job with results
            import_job.total_rows = row_count
            import_job.error_count = len(row_errors)
            import_job.errors_json = row_errors[:cls.MAX_ERRORS_DISPLAYED]
            
            if row_errors:
                import_job.status = 'failed'
                import_job.error_message = f"{len(row_errors)} validation error(s) found."
                import_job.save()
                
                return cls._fail(
                    f"Validation failed with {len(row_errors)} error(s).",
                    code="ROW_ERRORS",
                    data={
                        "import_job_id": import_job.id,
                        "total_rows": row_count,
                        "error_count": len(row_errors),
                        "errors": row_errors[:cls.MAX_ERRORS_DISPLAYED],
                        "dry_run": dry_run,
                    }
                )
            
            # Store parsed data for publishing
            import_job.parsed_data = rows
            import_job.valid_rows = len(rows)
            
            if dry_run:
                import_job.status = 'validated'
                import_job.save()
                
                return cls._success(
                    f"Validation passed. {len(rows)} rows ready to publish.",
                    data={
                        "import_job_id": import_job.id,
                        "total_rows": row_count,
                        "valid_rows": len(rows),
                        "dry_run": True,
                        "ready_to_publish": True,
                    }
                )
            else:
                # Auto-publish if not dry run
                return cls.publish_version(dataset_id, import_job.id)
                
        except Exception as e:
            logger.error(f"Import failed for dataset {dataset_id}: {e}")
            import_job.status = 'failed'
            import_job.error_message = str(e)
            import_job.save()
            
            return cls._fail(f"Import failed: {str(e)}", code="IMPORT_ERROR")
    
    # ==========================================================================
    # VALIDATION
    # ==========================================================================
    
    @classmethod
    def validate(cls, uploaded_file: UploadedFile) -> List[str]:
        """
        Validate uploaded file (type, size, structure).
        
        Returns:
            List of error messages (empty if valid)
        """
        errors = []
        
        # Check file exists
        if not uploaded_file:
            errors.append("No file provided.")
            return errors
        
        # Check file size
        if uploaded_file.size > cls.MAX_FILE_SIZE:
            errors.append(f"File too large. Maximum size is {cls.MAX_FILE_SIZE // (1024*1024)}MB.")
        
        # Check extension
        ext = cls._get_extension(uploaded_file.name)
        if ext not in cls.ALLOWED_EXTENSIONS:
            errors.append(f"Invalid file type. Allowed: {', '.join(cls.ALLOWED_EXTENSIONS)}")
        
        # Try to read file
        try:
            uploaded_file.seek(0)
            if ext == 'csv':
                # Check CSV is readable
                content = uploaded_file.read().decode('utf-8', errors='replace')
                reader = csv.reader(io.StringIO(content))
                header = next(reader, None)
                if not header:
                    errors.append("CSV file is empty or has no header row.")
            else:
                # Check Excel is readable
                try:
                    import openpyxl
                    uploaded_file.seek(0)
                    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                    ws = wb.active
                    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    if not any(header):
                        errors.append("Excel file is empty or has no header row.")
                    wb.close()
                except ImportError:
                    errors.append("Excel support requires openpyxl. Install with: pip install openpyxl")
                except Exception as e:
                    errors.append(f"Cannot read Excel file: {str(e)}")
            
            uploaded_file.seek(0)
            
        except Exception as e:
            errors.append(f"Cannot read file: {str(e)}")
        
        return errors
    
    # ==========================================================================
    # PARSING
    # ==========================================================================
    
    @classmethod
    def parse_rows(cls, uploaded_file: UploadedFile) -> Iterator[Tuple[int, Dict[str, Any]]]:
        """
        Parse file and yield (row_number, row_dict) tuples.
        
        Yields:
            (row_number: int, row_data: dict)
        """
        ext = cls._get_extension(uploaded_file.name)
        uploaded_file.seek(0)
        
        if ext == 'csv':
            yield from cls._parse_csv(uploaded_file)
        else:
            yield from cls._parse_excel(uploaded_file)
    
    @classmethod
    def _parse_csv(cls, uploaded_file: UploadedFile) -> Iterator[Tuple[int, Dict[str, Any]]]:
        """Parse CSV file."""
        content = uploaded_file.read().decode('utf-8', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        
        for row_num, row in enumerate(reader, start=2):  # Row 1 is header
            # Clean keys and values
            cleaned = {
                cls._clean_key(k): cls._clean_value(v)
                for k, v in row.items()
                if k is not None
            }
            yield row_num, cleaned
    
    @classmethod
    def _parse_excel(cls, uploaded_file: UploadedFile) -> Iterator[Tuple[int, Dict[str, Any]]]:
        """Parse Excel file."""
        import openpyxl
        
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        
        if not header_row:
            wb.close()
            return
        
        # Clean header names
        headers = [cls._clean_key(h) if h else f"column_{i}" for i, h in enumerate(header_row)]
        
        for row_num, row in enumerate(rows_iter, start=2):
            if not any(row):  # Skip empty rows
                continue
            
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = cls._clean_value(value)
            
            yield row_num, row_dict
        
        wb.close()
    
    @classmethod
    def _process_file(
        cls,
        uploaded_file: UploadedFile,
        dataset
    ) -> Tuple[List[Dict], List[Dict], int]:
        """
        Process file: parse all rows, validate each, collect errors.
        
        Returns:
            (valid_rows, errors, total_count)
        """
        valid_rows = []
        errors = []
        total_count = 0
        
        # Get expected schema from dataset
        schema = dataset.schema or {}
        required_fields = schema.get('required', [])
        field_types = schema.get('types', {})
        
        for row_num, row_data in cls.parse_rows(uploaded_file):
            total_count += 1
            row_errors = []
            
            # Validate required fields
            for field in required_fields:
                if field not in row_data or row_data[field] in (None, ''):
                    row_errors.append(f"Missing required field: {field}")
            
            # Validate field types
            for field, expected_type in field_types.items():
                if field in row_data and row_data[field] not in (None, ''):
                    value = row_data[field]
                    if not cls._validate_type(value, expected_type):
                        row_errors.append(f"Invalid type for '{field}': expected {expected_type}")
            
            if row_errors:
                errors.append({
                    "row": row_num,
                    "errors": row_errors,
                    "data": {k: str(v)[:100] for k, v in list(row_data.items())[:5]}
                })
            else:
                row_data['_row_num'] = row_num
                valid_rows.append(row_data)
        
        return valid_rows, errors, total_count
    
    # ==========================================================================
    # PUBLISHING
    # ==========================================================================
    
    @classmethod
    @transaction.atomic
    def publish_version(cls, dataset_id: int, import_job_id: int) -> Dict[str, Any]:
        """
        Create a new DatasetVersion from validated import job.
        
        Args:
            dataset_id: Dataset ID
            import_job_id: Validated import job ID
        
        Returns:
            {ok: bool, reason: str, data: {version_id, version_number}}
        """
        from datasets.models import Dataset, DatasetImportJob, DatasetVersion
        
        try:
            dataset = Dataset.objects.get(id=dataset_id)
            import_job = DatasetImportJob.objects.get(id=import_job_id, dataset=dataset)
        except (Dataset.DoesNotExist, DatasetImportJob.DoesNotExist):
            return cls._fail("Dataset or import job not found.", code="NOT_FOUND")
        
        if import_job.status not in ('validated', 'processing'):
            return cls._fail(
                f"Import job cannot be published (status: {import_job.status}).",
                code="INVALID_STATUS"
            )
        
        if not import_job.parsed_data:
            return cls._fail("No parsed data available.", code="NO_DATA")
        
        # Get next version number
        last_version = DatasetVersion.objects.filter(dataset=dataset).order_by('-version').first()
        next_version = (last_version.version + 1) if last_version else 1
        
        # Unpublish current version
        DatasetVersion.objects.filter(dataset=dataset, is_published=True).update(is_published=False)
        
        # Create new version
        version = DatasetVersion.objects.create(
            dataset=dataset,
            version=next_version,
            data=import_job.parsed_data,
            row_count=len(import_job.parsed_data),
            is_published=True,
            published_at=timezone.now(),
            published_by=import_job.initiated_by,
            import_job=import_job,
            changelog=f"Imported {len(import_job.parsed_data)} rows from {import_job.file_name}",
        )
        
        # Update dataset
        dataset.current_version = version
        dataset.row_count = version.row_count
        dataset.last_updated_at = timezone.now()
        dataset.save(update_fields=['current_version', 'row_count', 'last_updated_at'])
        
        # Update import job
        import_job.status = 'completed'
        import_job.completed_at = timezone.now()
        import_job.save()
        
        logger.info(f"Published version {next_version} for dataset {dataset.slug}")
        
        return cls._success(
            f"Version {next_version} published successfully.",
            data={
                "version_id": version.id,
                "version_number": next_version,
                "row_count": version.row_count,
                "published_at": version.published_at.isoformat(),
            }
        )
    
    # ==========================================================================
    # ROLLBACK
    # ==========================================================================
    
    @classmethod
    @transaction.atomic
    def rollback(cls, dataset_id: int, version_id: int) -> Dict[str, Any]:
        """
        Rollback dataset to a previous version.
        
        Args:
            dataset_id: Dataset ID
            version_id: Version ID to rollback to
        
        Returns:
            {ok: bool, reason: str, data: {restored_version}}
        """
        from datasets.models import Dataset, DatasetVersion
        
        try:
            dataset = Dataset.objects.get(id=dataset_id)
            target_version = DatasetVersion.objects.get(id=version_id, dataset=dataset)
        except (Dataset.DoesNotExist, DatasetVersion.DoesNotExist):
            return cls._fail("Dataset or version not found.", code="NOT_FOUND")
        
        if target_version.is_published:
            return cls._fail("This version is already the current version.", code="ALREADY_CURRENT")
        
        # Get current version for logging
        current_version = dataset.current_version
        current_version_num = current_version.version if current_version else 0
        
        # Unpublish all versions
        DatasetVersion.objects.filter(dataset=dataset, is_published=True).update(is_published=False)
        
        # Publish target version
        target_version.is_published = True
        target_version.save(update_fields=['is_published'])
        
        # Update dataset
        dataset.current_version = target_version
        dataset.row_count = target_version.row_count
        dataset.last_updated_at = timezone.now()
        dataset.save(update_fields=['current_version', 'row_count', 'last_updated_at'])
        
        logger.info(f"Rolled back dataset {dataset.slug} from v{current_version_num} to v{target_version.version}")
        
        return cls._success(
            f"Rolled back to version {target_version.version}.",
            data={
                "restored_version": target_version.version,
                "previous_version": current_version_num,
                "row_count": target_version.row_count,
            }
        )
    
    # ==========================================================================
    # HELPERS
    # ==========================================================================
    
    @classmethod
    def _get_extension(cls, filename: str) -> str:
        """Get lowercase file extension."""
        return filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    
    @classmethod
    def _clean_key(cls, key) -> str:
        """Clean column header to valid key."""
        if key is None:
            return ''
        return str(key).strip().lower().replace(' ', '_').replace('-', '_')
    
    @classmethod
    def _clean_value(cls, value) -> Any:
        """Clean cell value."""
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value if value else None
        if isinstance(value, datetime):
            return value.isoformat()
        return value
    
    @classmethod
    def _validate_type(cls, value, expected_type: str) -> bool:
        """Validate value against expected type."""
        try:
            if expected_type == 'string':
                return isinstance(value, str)
            elif expected_type == 'number':
                if isinstance(value, (int, float, Decimal)):
                    return True
                Decimal(str(value))
                return True
            elif expected_type == 'integer':
                int(value)
                return True
            elif expected_type == 'boolean':
                return str(value).lower() in ('true', 'false', '1', '0', 'yes', 'no')
            elif expected_type == 'date':
                # Basic date validation
                if isinstance(value, datetime):
                    return True
                datetime.fromisoformat(str(value).replace('/', '-'))
                return True
            return True  # Unknown type, accept
        except (ValueError, InvalidOperation, TypeError):
            return False
    
    @classmethod
    def _success(cls, reason: str, data: Optional[dict] = None) -> dict:
        return {"ok": True, "reason": reason, "data": data or {}}
    
    @classmethod
    def _fail(cls, reason: str, code: str = "ERROR", data: Optional[dict] = None) -> dict:
        return {"ok": False, "reason": reason, "code": code, "data": data or {}}
