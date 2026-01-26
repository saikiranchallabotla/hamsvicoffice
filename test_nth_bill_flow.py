#!/usr/bin/env python
"""Simulate the complete multi-sheet Nth bill generation flow."""

import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'estimate_site.settings')
sys.path.insert(0, r'c:\Users\HP\Documents\Windows x 1')
django.setup()

from openpyxl import load_workbook, Workbook
from core.views import (
    parse_first_bill_for_nth, 
    _extract_header_data_from_sheet, 
    _populate_nth_bill_sheet,
    read_tp_from_sheet,
    ordinal_word
)

# Load the test file
wb_in = load_workbook('test_multi_bill.xlsx', data_only=True)

# Simulate the handler logic
action = "firstpart_nth_part"
bill_sheets = [ws for ws in wb_in.worksheets if ws.title.startswith("Bill")]
print(f"Detected {len(bill_sheets)} bill sheets: {[ws.title for ws in bill_sheets]}")

# Extract Nth number
nth_val = 2
ord_word = ordinal_word(nth_val)
title_text = f"CC {ord_word} & Part Bill"

# Build output workbook
wb_out = Workbook()
created = 0

print(f"\nStarting to process {len(bill_sheets)} sheets...")

for idx, ws_first in enumerate(bill_sheets, start=1):
    print(f"\n[{idx}] Processing sheet: '{ws_first.title}'")
    
    # Find header row
    header_row = None
    for r in range(1, 30):
        a = str(ws_first.cell(row=r, column=1).value or "").strip().lower()
        b = str(ws_first.cell(row=r, column=2).value or "").strip().lower()
        d = str(ws_first.cell(row=r, column=4).value or "").strip().lower()
        if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
            header_row = r
            print(f"    Found header row: {r}")
            break
    
    if header_row is None:
        header_row = 10
        print(f"    Using default header row: 10")
    
    # Extract data
    header_data = _extract_header_data_from_sheet(ws_first)
    items = parse_first_bill_for_nth(ws_first, header_row)
    print(f"    Parsed {len(items)} items")
    
    if not items:
        print(f"    No items found, skipping")
        continue
    
    created += 1
    print(f"    Creating output sheet (created={created})")
    
    tp_percent, tp_type = read_tp_from_sheet(ws_first)
    if tp_percent is None:
        tp_percent = 0.0
    if tp_type is None:
        tp_type = "Excess"
    
    # Determine sheet name
    if len(bill_sheets) > 1:
        sheet_name = f"Bill-{created}"
    else:
        sheet_name = "Bill"
    
    # Determine sheet title
    if len(bill_sheets) > 1:
        sheet_title = f"{title_text} ({created})"
    else:
        sheet_title = title_text
    
    # Create output sheet
    if created == 1:
        ws_nth = wb_out.active
        ws_nth.title = sheet_name
        print(f"    Using active sheet, named: '{sheet_name}'")
    else:
        ws_nth = wb_out.create_sheet(title=sheet_name)
        print(f"    Created new sheet: '{sheet_name}'")
    
    # Populate sheet
    print(f"    Populating sheet with {len(items)} items...")
    _populate_nth_bill_sheet(
        ws_nth,
        items=items,
        header_data=header_data,
        title_text=sheet_title,
        tp_percent=tp_percent,
        tp_type=tp_type,
        mb_measure_no="",
        mb_measure_p_from="",
        mb_measure_p_to="",
        mb_abs_no="",
        mb_abs_p_from="",
        mb_abs_p_to="",
        doi="",
        doc="",
        domr="",
        dobr="",
    )
    print(f"    Sheet populated successfully")

print(f"\n\nFinal result: Created {created} output sheets")
print(f"Output workbook sheets: {[ws.title for ws in wb_out.worksheets]}")

# Save output
output_file = "test_nth_bill_output.xlsx"
wb_out.save(output_file)
print(f"\nSaved to: {output_file}")
