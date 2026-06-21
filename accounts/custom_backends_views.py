# accounts/custom_backends_views.py
"""
Per-user custom item uploads.
Items are detected from ALL sheets of the uploaded Excel (yellow-fill +
red-text headings). The Group name and per-item Units are entered in the UI.
"""

import os

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import FileResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.views.decorators.http import require_POST

from openpyxl import load_workbook

from .models import UserCustomBackend


def _scan_all_sheets_for_items(file_or_path):
    """
    Open the workbook and scan EVERY sheet for item blocks.
    Returns list of (sheet_name, item_name) tuples in order found.
    """
    wb = load_workbook(file_or_path, data_only=False, read_only=False)
    from core.utils_excel import _extract_items_from_sheet

    found = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        names, _blocks = _extract_items_from_sheet(ws)
        for n in names:
            found.append((sheet_name, n))
    return found, wb


@login_required
def custom_backend_list_view(request):
    backends = UserCustomBackend.objects.filter(user=request.user).order_by('-uploaded_at')
    return render(request, 'accounts/custom_backends/list.html', {
        'backends': backends,
    })


@login_required
def custom_backend_upload_view(request):
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        group_name = (request.POST.get('group_name') or '').strip()
        category = (request.POST.get('category') or '').strip().lower()
        uploaded = request.FILES.get('file')

        applies_estimate = request.POST.get('applies_estimate') == 'on'
        applies_tempworks = request.POST.get('applies_tempworks') == 'on'
        applies_amc = request.POST.get('applies_amc') == 'on'

        if not name:
            messages.error(request, "Please enter a name for this upload.")
            return redirect('custom_backend_upload')
        if not group_name:
            messages.error(request, "Please enter a Group name.")
            return redirect('custom_backend_upload')
        if category not in ('electrical', 'civil'):
            messages.error(request, "Please select a valid category.")
            return redirect('custom_backend_upload')
        if not uploaded:
            messages.error(request, "Please choose an Excel (.xlsx) file.")
            return redirect('custom_backend_upload')
        if not (applies_estimate or applies_tempworks or applies_amc):
            messages.error(request, "Select at least one module to apply this to.")
            return redirect('custom_backend_upload')
        if not uploaded.name.lower().endswith('.xlsx'):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect('custom_backend_upload')

        # Validate by scanning all sheets for at least one item block
        try:
            found, _wb = _scan_all_sheets_for_items(uploaded)
        except Exception as e:
            messages.error(request, f"Could not read Excel file: {e}")
            return redirect('custom_backend_upload')
        if not found:
            messages.error(
                request,
                "No item blocks detected. Mark item headings with yellow fill + red text in any sheet."
            )
            return redirect('custom_backend_upload')
        try:
            uploaded.seek(0)
        except Exception:
            pass

        # Build initial units map using existing heuristic
        from core.utils_excel import _determine_unit_from_heading
        units = {}
        for _sheet, item_name in found:
            units[item_name] = _determine_unit_from_heading(item_name)

        backend = UserCustomBackend.objects.create(
            user=request.user,
            name=name,
            group_name=group_name,
            file=uploaded,
            category=category,
            applies_estimate=applies_estimate,
            applies_tempworks=applies_tempworks,
            applies_amc=applies_amc,
            units_override=units,
        )
        messages.success(request, f"Uploaded '{name}'. Detected {len(found)} item(s). Set the units below.")
        return redirect('custom_backend_edit_units', backend_id=backend.pk)

    # Pre-fill from query params (e.g., ?module=estimate&category=civil)
    pre_module = (request.GET.get('module') or '').strip().lower()
    pre_category = (request.GET.get('category') or '').strip().lower()
    if pre_category.startswith('temp_'):
        pre_category = pre_category.replace('temp_', '')
    elif pre_category.startswith('amc_'):
        pre_category = pre_category.replace('amc_', '')
    return render(request, 'accounts/custom_backends/upload.html', {
        'pre_module': pre_module,
        'pre_category': pre_category if pre_category in ('electrical', 'civil') else '',
    })


@login_required
def custom_backend_edit_units_view(request, backend_id):
    backend = get_object_or_404(UserCustomBackend, pk=backend_id, user=request.user)

    # Re-scan items so the list reflects the actual file (sheet + item)
    items = []
    try:
        cb_path = backend.file.path
    except (NotImplementedError, Exception):
        cb_path = None
    if cb_path is None:
        try:
            data = backend.file.read()
            backend.file.seek(0)
            import tempfile
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            tmp.write(data)
            tmp.close()
            cb_path = tmp.name
        except Exception:
            cb_path = None

    if cb_path:
        try:
            found, _wb = _scan_all_sheets_for_items(cb_path)
            import urllib.parse as _up
            rp = backend.repair_prefixes or {}
            uo = backend.units_override or {}
            items = [{
                'sheet': s,
                'name': n,
                'idx': i,
                'unit': uo.get(n, '') or uo.get(_up.quote(n, safe=' '), ''),
                'prefix': rp.get(n, '') or rp.get(_up.quote(n, safe=' '), ''),
            } for i, (s, n) in enumerate(found)]
        except Exception:
            items = []

    if request.method == 'POST':
        import urllib.parse as _up2
        new_units = {}
        new_prefixes = {}
        i = 0
        while True:
            raw_name = request.POST.get(f'item_{i}')
            if raw_name is None:
                break
            item_name = _up2.unquote(raw_name.strip())
            uval = (request.POST.get(f'unit_{i}') or '').strip()
            pval = (request.POST.get(f'prefix_{i}') or '').strip()
            if item_name and uval:
                new_units[item_name] = uval
            if item_name and pval:
                new_prefixes[item_name] = pval
            i += 1
        merged_units = dict(backend.units_override or {})
        merged_units.update(new_units)
        backend.units_override = merged_units

        merged_prefixes = dict(backend.repair_prefixes or {})
        merged_prefixes.update(new_prefixes)
        backend.repair_prefixes = merged_prefixes

        backend.save(update_fields=['units_override', 'repair_prefixes', 'updated_at'])
        messages.success(request, "Saved.")
        return redirect('custom_backend_list')

    return render(request, 'accounts/custom_backends/edit_units.html', {
        'backend': backend,
        'items': items,
    })


@login_required
@require_POST
def custom_backend_edit_modules_view(request, backend_id):
    backend = get_object_or_404(UserCustomBackend, pk=backend_id, user=request.user)

    applies_estimate = request.POST.get('applies_estimate') == 'on'
    applies_tempworks = request.POST.get('applies_tempworks') == 'on'
    applies_amc = request.POST.get('applies_amc') == 'on'

    if not (applies_estimate or applies_tempworks or applies_amc):
        messages.error(request, "Select at least one module for this upload to apply to.")
        return redirect('custom_backend_edit_units', backend_id=backend.pk)

    backend.applies_estimate = applies_estimate
    backend.applies_tempworks = applies_tempworks
    backend.applies_amc = applies_amc
    backend.save(update_fields=['applies_estimate', 'applies_tempworks', 'applies_amc', 'updated_at'])
    messages.success(request, "Updated which modules this upload applies to.")
    return redirect('custom_backend_edit_units', backend_id=backend.pk)


@login_required
@require_POST
def custom_backend_delete_view(request, backend_id):
    backend = get_object_or_404(UserCustomBackend, pk=backend_id, user=request.user)
    name = backend.name
    try:
        backend.file.delete(save=False)
    except Exception:
        pass
    backend.delete()
    messages.success(request, f"Deleted '{name}'.")
    return redirect('custom_backend_list')


@login_required
def custom_backend_download_view(request, backend_id):
    backend = get_object_or_404(UserCustomBackend, pk=backend_id, user=request.user)
    filename = os.path.basename(backend.file.name) or f"{backend.name}.xlsx"
    return FileResponse(backend.file.open('rb'), as_attachment=True, filename=filename)


@login_required
@require_POST
def custom_backend_replace_view(request, backend_id):
    backend = get_object_or_404(UserCustomBackend, pk=backend_id, user=request.user)
    uploaded = request.FILES.get('file')

    if not uploaded:
        messages.error(request, "Please choose an Excel (.xlsx) file.")
        return redirect('custom_backend_list')
    if not uploaded.name.lower().endswith('.xlsx'):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect('custom_backend_list')

    try:
        found, _wb = _scan_all_sheets_for_items(uploaded)
    except Exception as e:
        messages.error(request, f"Could not read Excel file: {e}")
        return redirect('custom_backend_list')
    if not found:
        messages.error(
            request,
            "No item blocks detected. Mark item headings with yellow fill + red text in any sheet."
        )
        return redirect('custom_backend_list')
    try:
        uploaded.seek(0)
    except Exception:
        pass

    # Carry over unit/prefix overrides for items still present in the new
    # file; auto-fill defaults for newly detected items; drop the rest.
    from core.utils_excel import _determine_unit_from_heading
    existing_units = dict(backend.units_override or {})
    existing_prefixes = dict(backend.repair_prefixes or {})
    new_item_names = {n for _sheet, n in found}
    units = {n: existing_units.get(n) or _determine_unit_from_heading(n) for n in new_item_names}
    prefixes = {n: p for n, p in existing_prefixes.items() if n in new_item_names}

    try:
        backend.file.delete(save=False)
    except Exception:
        pass
    backend.file = uploaded
    backend.units_override = units
    backend.repair_prefixes = prefixes
    backend.save()

    messages.success(request, f"Replaced file for '{backend.name}'. Detected {len(found)} item(s). Review units below.")
    return redirect('custom_backend_edit_units', backend_id=backend.pk)


@login_required
@require_POST
def custom_backend_toggle_view(request, backend_id):
    backend = get_object_or_404(UserCustomBackend, pk=backend_id, user=request.user)
    backend.is_active = not backend.is_active
    backend.save(update_fields=['is_active', 'updated_at'])
    return redirect('custom_backend_list')
