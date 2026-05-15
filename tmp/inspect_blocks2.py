# Direct inspection of backend blocks and their formulas — no Django needed.
import re
from openpyxl import load_workbook

BACKENDS = [
    r'core/data/electrical.xlsx',
    r'core/data/civil.xlsx',
    r'core/data/temp_electrical.xlsx',
]

def is_yellow_red(cell):
    fill = cell.fill
    if not fill or not fill.patternType or fill.patternType.lower() != 'solid':
        return False
    rgb = getattr(fill.fgColor, 'rgb', None)
    yellow = rgb and str(rgb).upper().endswith('FFFF00')
    font = cell.font
    if not font or not font.color:
        return False
    frgb = getattr(font.color, 'rgb', None)
    red = frgb and str(frgb).upper().endswith('FF0000')
    return yellow and red

def detect_items(ws):
    items = []
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        heading = None
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if is_yellow_red(cell) and str(cell.value or '').strip():
                heading = str(cell.value).strip()
                break
        if heading:
            start = r
            end = max_row
            rr = r + 1
            while rr <= max_row:
                found = any(
                    is_yellow_red(ws.cell(row=rr, column=cc)) and str(ws.cell(row=rr, column=cc).value or '').strip()
                    for cc in range(1, 11)
                )
                if found:
                    end = rr - 1
                    break
                rr += 1
            items.append({'name': heading, 'start_row': start, 'end_row': end})
            r = end + 1
        else:
            r += 1
    return items

ABS_OR_MIXED = re.compile(r'\$[A-Za-z]{1,3}\$?\d+|[A-Za-z]{1,3}\$\d+')
PLAIN_REF = re.compile(r'(?<![!\w\$])([A-Za-z]{1,3})(\d+)(?![A-Za-z\d])')

for path in BACKENDS:
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        print(f'{path}: CANNOT OPEN - {e}')
        continue
    ws = wb['Master Datas']
    items = detect_items(ws)
    print(f'\n========== {path} ({len(items)} items) ==========')

    for idx, it in enumerate(items[:8]):
        s, e, name = it['start_row'], it['end_row'], it['name']
        formulas = []
        cross_block_relative = []
        for r in range(s, e + 1):
            for c in range(1, 11):
                v = ws.cell(row=r, column=c).value
                if not (v and isinstance(v, str) and v.startswith('=')):
                    continue
                formulas.append((r, c, v))
                # Check for relative refs that point OUTSIDE this block
                for m in PLAIN_REF.finditer(v):
                    ref_row = int(m.group(2))
                    if ref_row < s or ref_row > e:
                        cross_block_relative.append((r, c, v, ref_row))

        has_abs = any(ABS_OR_MIXED.search(f) for _, _, f in formulas)
        print(f'\n  Item {idx+1}: {name!r} rows {s}-{e}')
        print(f'    formulas: {len(formulas)}, has_absolute_or_mixed: {has_abs}')
        if cross_block_relative:
            print(f'    CROSS-BLOCK relative refs:')
            for r, c, formula, ref_row in cross_block_relative[:5]:
                print(f'      R{r}C{c}: {formula}  → references row {ref_row} (outside {s}-{e})')
        else:
            print(f'    (no cross-block relative refs)')
        # Show sample formulas
        for r, c, f in formulas[:3]:
            print(f'    sample R{r}C{c}: {f}')
