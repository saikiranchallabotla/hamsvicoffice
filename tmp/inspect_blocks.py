import sys
sys.path.insert(0, r'c:\Users\keesa\Desktop\hamsvicoffice')
import django
import os
os.environ['DJANGO_SETTINGS_MODULE'] = 'estimate_site.settings'

from openpyxl import load_workbook
from core.utils_excel import detect_items

wb = load_workbook('core/data/electrical.xlsx', data_only=False)
ws = wb['Master Datas']
items = detect_items(ws)
print(f'Total items: {len(items)}')

for it in items[:6]:
    name = it['name']
    s = it['start_row']
    e = it['end_row']
    print(f'\n--- Item: {name!r}  rows {s}-{e} ---')
    for r in range(s, e + 1):
        for c in range(1, 11):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and v.startswith('='):
                print(f'  R{r}C{c}: {v}')
