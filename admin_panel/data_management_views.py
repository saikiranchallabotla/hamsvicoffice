# admin_panel/data_management_views.py
"""
Backend data management views - upload/replace Excel files without affecting users.
Supports multiple backends per module (state-wise SOR rates).
"""

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.conf import settings
import pandas as pd

from admin_panel.decorators import admin_required, superadmin_required
from subscriptions.models import Module, ModuleBackend


# Data files location
DATA_DIR = Path(settings.BASE_DIR) / 'core' / 'data'
BACKUP_DIR = Path(settings.BASE_DIR) / 'core' / 'data' / 'backups'
MEDIA_BACKENDS_DIR = Path(settings.MEDIA_ROOT) / 'module_backends'


def get_file_info(filepath):
    """Get file metadata."""
    if not filepath.exists():
        return None
    
    stat = filepath.stat()
    return {
        'name': filepath.name,
        'path': str(filepath),
        'size': stat.st_size,
        'size_readable': format_size(stat.st_size),
        'modified': datetime.fromtimestamp(stat.st_mtime),
    }


def format_size(size_bytes):
    """Format bytes to human readable."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"


def get_excel_preview(filepath, max_rows=10):
    """
    Get preview of Excel file contents.
    Returns dict with sheet names and sample data.
    """
    try:
        xl = pd.ExcelFile(filepath)
        preview = {
            'sheets': [],
            'total_sheets': len(xl.sheet_names)
        }
        
        for sheet_name in xl.sheet_names[:5]:  # Max 5 sheets
            df = pd.read_excel(xl, sheet_name=sheet_name, nrows=max_rows)
            preview['sheets'].append({
                'name': sheet_name,
                'rows': len(df),
                'columns': list(df.columns)[:10],  # First 10 columns
                'total_columns': len(df.columns),
                'sample_data': df.head(5).fillna('').values.tolist()
            })
        
        return preview
    except Exception as e:
        return {'error': str(e)}


@admin_required
def data_management(request):
    """
    Main data management page showing current backend files.
    Now supports multiple backends per module for different states/regions.
    """
    # Ensure directories exist
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    MEDIA_BACKENDS_DIR.mkdir(parents=True, exist_ok=True)
    
    # Get modules that can have their own backend data uploads
    # Note: workslip shares backends with new_estimate module
    # temp_works and amc have their own separate backends
    # estimate module doesn't require backend data
    backend_modules = Module.objects.filter(
        code__in=['new_estimate', 'temp_works', 'amc']
    ).order_by('display_order', 'name')
    
    # Get all module backends grouped by module and category
    module_backends_data = []
    for module in backend_modules:
        electrical_backends = ModuleBackend.objects.filter(
            module=module, category='electrical', is_active=True
        ).order_by('display_order', 'name')
        
        civil_backends = ModuleBackend.objects.filter(
            module=module, category='civil', is_active=True
        ).order_by('display_order', 'name')
        
        module_backends_data.append({
            'module': module,
            'electrical_backends': electrical_backends,
            'civil_backends': civil_backends,
        })
    
    # Define all legacy managed files (for backward compatibility)
    legacy_categories = [
        ('civil', DATA_DIR / 'civil.xlsx', 'Civil Data', 'bi-building', '#d97706', 'civil', 'btn-warning'),
        ('electrical', DATA_DIR / 'electrical.xlsx', 'Electrical Data', 'bi-lightning', '#2563eb', 'electrical', 'btn-primary'),
        ('temp_civil', DATA_DIR / 'temp_civil.xlsx', 'Temp Civil Data', 'bi-building', '#f59e42', 'civil', 'btn-warning'),
        ('temp_electrical', DATA_DIR / 'temp_electrical.xlsx', 'Temp Electrical Data', 'bi-lightning', '#3b82f6', 'electrical', 'btn-primary'),
        ('amc_electrical', DATA_DIR / 'amc_electrical.xlsx', 'AMC Electrical Data', 'bi-tools', '#8b5cf6', 'amc_electrical', 'btn-purple'),
        ('amc_civil', DATA_DIR / 'amc_civil.xlsx', 'AMC Civil Data', 'bi-tools', '#a855f7', 'amc_civil', 'btn-purple'),
    ]

    legacy_files = []
    for key, fpath, title, icon, color, icon_class, btn_class in legacy_categories:
        file_info = get_file_info(fpath)
        legacy_files.append({
            'key': key,
            'title': title,
            'icon': icon,
            'color': color,
            'icon_class': icon_class,
            'btn_class': btn_class,
            'file': file_info,
            'upload_label': f'Upload {title}',
        })

    # Bill templates (always legacy)
    bill_templates = [
        {
            'key': 'ls_form_final',
            'title': 'L.S Form Final',
            'icon': 'bi-file-earmark-spreadsheet',
            'color': '#059669',
            'icon_class': 'civil',
            'btn_class': 'btn-success',
            'file': get_file_info(Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Final.xlsx'),
            'upload_label': 'Upload L.S Form Final',
        },
        {
            'key': 'ls_form_part',
            'title': 'L.S Form Part',
            'icon': 'bi-file-earmark-spreadsheet',
            'color': '#0ea5e9',
            'icon_class': 'electrical',
            'btn_class': 'btn-info',
            'file': get_file_info(Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Part.xlsx'),
            'upload_label': 'Upload L.S Form Part',
        },
    ]

    # Get backups
    backups = []
    for f in sorted(BACKUP_DIR.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)[:20]:
        info = get_file_info(f)
        if info:
            # Parse backup filename: civil_2026-01-04_153022.xlsx
            parts = f.stem.split('_')
            info['category'] = parts[0] if parts else 'unknown'
            backups.append(info)

    # Area Allowance percentages (zone + project location category)
    try:
        from core.models import AreaAllowanceUpload
        area_allowance_current = AreaAllowanceUpload.current()
    except Exception:
        area_allowance_current = None

    context = {
        'module_backends_data': module_backends_data,
        'legacy_files': legacy_files,
        'bill_templates': bill_templates,
        'area_allowance': area_allowance_current,
        'backups': backups,
        'data_dir': str(DATA_DIR),
        'modules': backend_modules,
    }
    return render(request, 'admin_panel/data/management.html', context)


@admin_required
def preview_file(request, category):
    """
    Preview contents of a backend Excel file.
    """
    allowed = ['civil', 'electrical', 'temp_civil', 'temp_electrical', 'amc_electrical', 'amc_civil', 'ls_form_final', 'ls_form_part']
    if category not in allowed:
        messages.error(request, 'Invalid category.')
        return redirect('admin_data_management')

    if category == 'ls_form_final':
        filepath = Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Final.xlsx'
    elif category == 'ls_form_part':
        filepath = Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Part.xlsx'
    else:
        filepath = DATA_DIR / f'{category}.xlsx'

    if not filepath.exists():
        context = {
            'category': category,
            'file_info': None,
            'preview': {'error': f'{category.replace("_", " ").title()} file not found.'},
        }
        return render(request, 'admin_panel/data/preview.html', context)

    preview = get_excel_preview(filepath, max_rows=20)
    file_info = get_file_info(filepath)

    context = {
        'category': category,
        'file_info': file_info,
        'preview': preview,
    }

    return render(request, 'admin_panel/data/preview.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def upload_file(request, category):
    """
    Upload and replace a backend Excel file.
    Automatically backs up the existing file before replacing.
    """
    allowed = ['civil', 'electrical', 'temp_civil', 'temp_electrical', 'amc_electrical', 'amc_civil', 'ls_form_final', 'ls_form_part']
    if category not in allowed:
        messages.error(request, 'Invalid category.')
        return redirect('admin_data_management')

    if category == 'ls_form_final':
        current_file = Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Final.xlsx'
    elif category == 'ls_form_part':
        current_file = Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Part.xlsx'
    else:
        current_file = DATA_DIR / f'{category}.xlsx'
    current_info = get_file_info(current_file)

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            messages.error(request, 'No file uploaded.')
            return redirect('admin_upload_file', category=category)

        # Validate file extension
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload an Excel file (.xlsx or .xls)')
            return redirect('admin_upload_file', category=category)

        # Validate file by trying to read it
        try:
            # Save temporarily to validate
            temp_path = DATA_DIR / f'temp_{category}.xlsx'
            with open(temp_path, 'wb') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)

            # Try to read with pandas - ensure file is closed after reading
            with pd.ExcelFile(temp_path) as xl:
                sheet_count = len(xl.sheet_names)

            # Backup current file if exists
            if current_file.exists():
                timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
                backup_name = f'{category}_{timestamp}.xlsx'
                backup_path = BACKUP_DIR / backup_name
                shutil.copy2(current_file, backup_path)

            # Replace with new file (use copy + delete for Windows compatibility)
            shutil.copy2(temp_path, current_file)
            try:
                temp_path.unlink()
            except PermissionError:
                pass  # File will be cleaned up later

            # Persist file bytes in database (survives Railway/Heroku redeploys)
            if category not in ('ls_form_final', 'ls_form_part'):
                try:
                    from datasets.models import LegacyBackendData
                    with open(current_file, 'rb') as f:
                        file_bytes = f.read()
                    LegacyBackendData.store(category, file_bytes, uploaded_file.name)
                except Exception:
                    pass  # DB persistence is best-effort

            # --- AUDIT LOG ---
            from datasets.models import AuditLog
            AuditLog.log(
                user=request.user,
                action='upload',
                obj=f"BackendData:{category}",  # Use string identifier instead of file object
                changes=None,
                metadata={
                    'category': category,
                    'filename': uploaded_file.name,
                    'sheet_count': sheet_count,
                    'backup': str(backup_path) if current_file.exists() else None,
                },
                request=request
            )

            messages.success(
                request,
                f'{category.replace("_", " ").title()} data updated successfully! '
                f'File contains {sheet_count} sheets. Previous version backed up.'
            )
            return redirect('admin_data_management')

        except Exception as e:
            # Clean up temp file - handle Windows file locking
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except PermissionError:
                pass  # File locked, will be cleaned up later
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('admin_upload_file', category=category)

    # GET - show upload form with preview of current file
    preview = None
    if current_file.exists():
        preview = get_excel_preview(current_file, max_rows=5)

    context = {
        'category': category,
        'current_file': current_info,
        'preview': preview,
    }

    return render(request, 'admin_panel/data/upload.html', context)


@admin_required
def download_file(request, category):
    """
    Download current backend Excel file.
    """
    allowed = ['civil', 'electrical', 'temp_civil', 'temp_electrical', 'amc_electrical', 'amc_civil', 'ls_form_final', 'ls_form_part']
    if category not in allowed:
        messages.error(request, 'Invalid category.')
        return redirect('admin_data_management')

    if category == 'ls_form_final':
        filepath = Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Final.xlsx'
    elif category == 'ls_form_part':
        filepath = Path(settings.BASE_DIR) / 'core' / 'templates' / 'core' / 'bill_templates' / 'LS_Form_Part.xlsx'
    else:
        filepath = DATA_DIR / f'{category}.xlsx'

    if not filepath.exists():
        messages.error(request, f'{category.replace("_", " ").title()} file not found.')
        return redirect('admin_data_management')

    return FileResponse(
        open(filepath, 'rb'),
        as_attachment=True,
        filename=f'{category}.xlsx'
    )


@admin_required
def download_backup(request, filename):
    """
    Download a backup file.
    """
    # Sanitize filename to prevent path traversal
    safe_filename = Path(filename).name
    filepath = BACKUP_DIR / safe_filename
    
    if not filepath.exists() or not str(filepath).startswith(str(BACKUP_DIR)):
        messages.error(request, 'Backup file not found.')
        return redirect('admin_data_management')
    
    return FileResponse(
        open(filepath, 'rb'),
        as_attachment=True,
        filename=safe_filename
    )


@admin_required
@require_POST
def restore_backup(request, filename):
    """
    Restore a backup file to replace current data.
    """
    safe_filename = Path(filename).name
    backup_path = BACKUP_DIR / safe_filename
    
    if not backup_path.exists() or not str(backup_path).startswith(str(BACKUP_DIR)):
        messages.error(request, 'Backup file not found.')
        return redirect('admin_data_management')
    
    # Determine category from filename
    allowed = ['civil', 'electrical', 'temp_civil', 'temp_electrical', 'amc_electrical', 'amc_civil', 'ls_form_final', 'ls_form_part']
    category = None
    for cat in allowed:
        if safe_filename.startswith(cat):
            category = cat
            break
    if not category:
        messages.error(request, 'Cannot determine file category.')
        return redirect('admin_data_management')

    current_file = DATA_DIR / f'{category}.xlsx'

    try:
        # Backup current file first
        if current_file.exists():
            timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
            pre_restore_backup = BACKUP_DIR / f'{category}_pre_restore_{timestamp}.xlsx'
            shutil.copy2(current_file, pre_restore_backup)

        # Restore from backup
        shutil.copy2(backup_path, current_file)

        # --- AUDIT LOG ---
        from datasets.models import AuditLog
        AuditLog.log(
            user=request.user,
            action='restore',
            obj=current_file,
            changes=None,
            metadata={
                'category': category,
                'restored_from': safe_filename,
                'pre_restore_backup': str(pre_restore_backup),
            },
            request=request
        )
        messages.success(request, f'{category.replace("_", " ").title()} data restored from backup: {safe_filename}')
    except Exception as e:
        messages.error(request, f'Error restoring backup: {str(e)}')

    return redirect('admin_data_management')


@admin_required
@require_POST
def delete_backup(request, filename):
    """
    Delete a backup file.
    """
    safe_filename = Path(filename).name
    filepath = BACKUP_DIR / safe_filename
    
    if not filepath.exists() or not str(filepath).startswith(str(BACKUP_DIR)):
        messages.error(request, 'Backup file not found.')
        return redirect('admin_data_management')
    
    try:
        filepath.unlink()
        # --- AUDIT LOG ---
        from datasets.models import AuditLog
        AuditLog.log(
            user=request.user,
            action='delete',
            obj=filepath,
            changes=None,
            metadata={
                'deleted_backup': safe_filename,
            },
            request=request
        )
        messages.success(request, f'Backup deleted: {safe_filename}')
    except Exception as e:
        messages.error(request, f'Error deleting backup: {str(e)}')
    
    return redirect('admin_data_management')


@admin_required
def preview_upload(request):
    """
    AJAX endpoint to preview an uploaded file before confirming.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'error': 'No file provided'}, status=400)
    
    from datasets.models import AuditLog
    try:
        # Read file directly from memory
        xl = pd.ExcelFile(uploaded_file)

        preview = {
            'filename': uploaded_file.name,
            'size': format_size(uploaded_file.size),
            'sheets': []
        }

        for sheet_name in xl.sheet_names[:5]:
            df = pd.read_excel(xl, sheet_name=sheet_name, nrows=5)
            preview['sheets'].append({
                'name': sheet_name,
                'columns': list(df.columns)[:8],
                'row_count': len(pd.read_excel(xl, sheet_name=sheet_name)),
                'sample': df.head(3).fillna('').values.tolist()
            })

        # --- AUDIT LOG ---
        AuditLog.log(
            user=request.user,
            action='preview',
            obj=uploaded_file,
            changes=None,
            metadata={
                'filename': uploaded_file.name,
                'sheet_count': len(xl.sheet_names),
            },
            request=request
        )

        return JsonResponse(preview)

    except Exception as e:
        # --- AUDIT LOG (failure) ---
        AuditLog.log(
            user=request.user,
            action='preview',
            obj=uploaded_file,
            changes=None,
            metadata={
                'filename': uploaded_file.name if uploaded_file else None,
                'error': str(e),
            },
            request=request
        )
        return JsonResponse({'error': 'Failed to preview file.'}, status=400)


# ==============================================================================
# MODULE BACKEND VIEWS (Multi-State SOR Support)
# ==============================================================================

def _validate_backend_workbook(uploaded_file, module=None):
    """
    Validate an uploaded backend workbook BEFORE it is saved.

    The app's load_backend() hard-requires a 'Master Datas' and a 'Groups' sheet;
    without them every page that uses the backend throws and users get an error
    page. So a workbook missing either sheet is REJECTED here (blocking), giving
    the admin an immediate warning to re-check and re-upload.

    Returns (errors, warnings): a non-empty `errors` list must block the save.
    Leaves the file pointer reset for subsequent reads.
    """
    try:
        uploaded_file.seek(0)
        with pd.ExcelFile(uploaded_file) as xl:
            sheets = set(xl.sheet_names)
    except Exception as e:
        return ([f'The file could not be read as an Excel workbook ({e}). '
                 'Please upload a valid .xlsx/.xls file.'], [])
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    errors, warnings = [], []
    missing = [s for s in ('Master Datas', 'Groups') if s not in sheets]
    if missing:
        errors.append(
            'Upload rejected — the workbook is missing required sheet(s): '
            + ', '.join(f'"{m}"' for m in missing)
            + '. A valid backend must contain both a "Master Datas" and a "Groups" '
            'sheet (exact tab names). Found tabs: '
            + (', '.join(f'"{s}"' for s in sorted(sheets)) or '(none)')
            + '. Please fix the file and re-upload.'
        )

    mod_code = (getattr(module, 'code', '') or '').lower()
    if 'temp' in mod_code and 'Daywise Rates' not in sheets:
        warnings.append(
            'Note: no "Daywise Rates" sheet was found — day-based rates may not '
            'work for this Temporary Works backend.'
        )
    return errors, warnings


@admin_required
@require_http_methods(["GET", "POST"])
def add_module_backend(request, module_code):
    """
    Add a new backend (SOR rates file) for a module.
    Example: Add "AP Electrical SOR 2024" for New Estimate module.
    """
    module = get_object_or_404(Module, code=module_code)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        category = request.POST.get('category', '')
        description = request.POST.get('description', '').strip()
        is_default = request.POST.get('is_default') == 'on'
        display_order = request.POST.get('display_order', 0)
        uploaded_file = request.FILES.get('file')
        
        # Validation
        errors = []
        if not name:
            errors.append('Name is required.')
        if category not in ['electrical', 'civil']:
            errors.append('Category must be electrical or civil.')
        if not uploaded_file:
            errors.append('Excel file is required.')
        elif not uploaded_file.name.endswith(('.xlsx', '.xls')):
            errors.append('File must be an Excel file (.xlsx or .xls).')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect('admin_add_module_backend', module_code=module_code)
        
        # Blocking validation: reject a workbook missing the required sheets so
        # a wrong-format file never gets saved and silently breaks the module.
        val_errors, val_warnings = _validate_backend_workbook(uploaded_file, module)
        for w in val_warnings:
            messages.warning(request, w)
        if val_errors:
            for e in val_errors:
                messages.error(request, e)
            return redirect('admin_add_module_backend', module_code=module_code)

        try:
            MEDIA_BACKENDS_DIR.mkdir(parents=True, exist_ok=True)

            # Read file bytes for DB persistence
            uploaded_file.seek(0)
            file_bytes = uploaded_file.read()
            uploaded_file.seek(0)

            # Create backend
            backend = ModuleBackend(
                module=module,
                name=name,
                code=code,
                category=category,
                description=description,
                is_default=is_default,
                display_order=int(display_order) if display_order else 0,
                file=uploaded_file,
                file_data=file_bytes,
                file_name=uploaded_file.name,
            )
            backend.save()
            
            # Audit log
            from datasets.models import AuditLog
            AuditLog.log(
                user=request.user,
                action='create',
                obj=f"ModuleBackend:{backend.pk}",
                changes=None,
                metadata={
                    'module': module.name,
                    'backend_name': name,
                    'category': category,
                    'filename': uploaded_file.name,
                },
                request=request
            )
            
            messages.success(
                request,
                f'Backend "{name}" added successfully for {module.name}!'
            )
            return redirect('admin_data_management')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('admin_add_module_backend', module_code=module_code)
    
    # GET - show form
    context = {
        'module': module,
        'categories': ModuleBackend.CATEGORY_CHOICES,
    }
    return render(request, 'admin_panel/data/add_backend.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def edit_module_backend(request, backend_id):
    """
    Edit an existing module backend.
    """
    backend = get_object_or_404(ModuleBackend, pk=backend_id)
    
    if request.method == 'POST':
        backend.name = request.POST.get('name', '').strip() or backend.name
        backend.code = request.POST.get('code', '').strip()
        backend.description = request.POST.get('description', '').strip()
        backend.is_default = request.POST.get('is_default') == 'on'
        backend.display_order = int(request.POST.get('display_order', 0) or 0)
        backend.is_active = request.POST.get('is_active') == 'on'
        
        # If new file uploaded, replace the old one
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'File must be an Excel file.')
                return redirect('admin_edit_module_backend', backend_id=backend_id)
            
            # Blocking validation: reject a replacement workbook that is missing
            # the required sheets, so the live backend can't be broken by a bad file.
            val_errors, val_warnings = _validate_backend_workbook(uploaded_file, backend.module)
            for w in val_warnings:
                messages.warning(request, w)
            if val_errors:
                for e in val_errors:
                    messages.error(request, e)
                return redirect('admin_edit_module_backend', backend_id=backend_id)

            try:
                # Backup old file before replacing
                if backend.file:
                    try:
                        old_path = Path(backend.file.path)
                        if old_path.exists():
                            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
                            timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
                            safe_name = backend.name.replace(' ', '_').replace('/', '-')
                            backup_name = f'{safe_name}_{backend.category}_{timestamp}.xlsx'
                            shutil.copy2(old_path, BACKUP_DIR / backup_name)
                            old_path.unlink()
                    except Exception:
                        pass  # Backup is best-effort; proceed with replacement
                
                # Save file bytes to DB for persistence
                uploaded_file.seek(0)
                backend.file_data = uploaded_file.read()
                uploaded_file.seek(0)
                backend.file_name = uploaded_file.name
                backend.file = uploaded_file
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                return redirect('admin_edit_module_backend', backend_id=backend_id)

        backend.save()
        
        # Audit log
        from datasets.models import AuditLog
        AuditLog.log(
            user=request.user,
            action='update',
            obj=f"ModuleBackend:{backend.pk}",
            changes=None,
            metadata={
                'backend_name': backend.name,
                'new_file': uploaded_file.name if uploaded_file else None,
            },
            request=request
        )
        
        messages.success(request, f'Backend "{backend.name}" updated successfully!')
        return redirect('admin_data_management')
    
    # GET - show form
    file_info = None
    if backend.file:
        try:
            file_info = get_file_info(Path(backend.file.path))
        except:
            pass
    
    context = {
        'backend': backend,
        'module': backend.module,
        'categories': ModuleBackend.CATEGORY_CHOICES,
        'file_info': file_info,
    }
    return render(request, 'admin_panel/data/edit_backend.html', context)


@admin_required
@require_POST
def delete_module_backend(request, backend_id):
    """
    Delete a module backend.
    """
    backend = get_object_or_404(ModuleBackend, pk=backend_id)
    
    # Store info for message
    name = backend.name
    module_name = backend.module.name
    
    # Delete file
    if backend.file:
        try:
            backend.file.delete(save=False)
        except:
            pass
    
    # Audit log
    from datasets.models import AuditLog
    AuditLog.log(
        user=request.user,
        action='delete',
        obj=f"ModuleBackend:{backend_id}",
        changes=None,
        metadata={
            'backend_name': name,
            'module': module_name,
        },
        request=request
    )
    
    backend.delete()
    messages.success(request, f'Backend "{name}" deleted successfully!')
    return redirect('admin_data_management')


@admin_required
def preview_module_backend(request, backend_id):
    """
    Preview a module backend's Excel file contents.
    """
    backend = get_object_or_404(ModuleBackend, pk=backend_id)
    
    if not backend.file:
        messages.error(request, 'No file associated with this backend.')
        return redirect('admin_data_management')
    
    try:
        # Read file content to support both local and S3 storage
        import tempfile
        file_data = backend.file.read()
        backend.file.seek(0)

        # Write to temp file for preview utilities
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        tmp.write(file_data)
        tmp.close()
        filepath = Path(tmp.name)

        preview = get_excel_preview(filepath, max_rows=20)
        file_info = get_file_info(filepath)

        # Clean up temp file
        try:
            filepath.unlink()
        except:
            pass
        
        context = {
            'category': backend.name,
            'backend': backend,
            'file_info': file_info,
            'preview': preview,
        }
        return render(request, 'admin_panel/data/preview.html', context)
        
    except Exception as e:
        messages.error(request, f'Error previewing file: {str(e)}')
        return redirect('admin_data_management')


@admin_required
def download_module_backend(request, backend_id):
    """
    Download a module backend's Excel file.
    """
    backend = get_object_or_404(ModuleBackend, pk=backend_id)
    
    if not backend.file:
        messages.error(request, 'No file associated with this backend.')
        return redirect('admin_data_management')
    
    try:
        # Generate download filename
        safe_name = backend.name.replace(' ', '_').replace('/', '-')
        download_name = f'{safe_name}_{backend.category}.xlsx'

        return FileResponse(
            backend.file.open('rb'),
            as_attachment=True,
            filename=download_name
        )
    except Exception as e:
        messages.error(request, f'Error downloading file: {str(e)}')
        return redirect('admin_data_management')


@admin_required
@require_POST
def toggle_backend_default(request, backend_id):
    """
    Toggle a backend as default for its module and category.
    """
    backend = get_object_or_404(ModuleBackend, pk=backend_id)
    
    # Toggle default status
    backend.is_default = not backend.is_default
    backend.save()
    
    if backend.is_default:
        messages.success(request, f'"{backend.name}" is now the default for {backend.get_category_display()} in {backend.module.name}.')
    else:
        messages.info(request, f'"{backend.name}" is no longer the default.')
    
    return redirect('admin_data_management')



# ==============================================================================
# AREA ALLOWANCE UPLOAD
# ==============================================================================
# The Area Allowance percentage applied to an estimate depends on the Zone and
# Project Location Category the user selects. Those percentages are maintained
# entirely through the Excel sheet uploaded here -- a new upload replaces the
# previous one outright, and every estimate generated afterwards uses the new
# figures.

AREA_ALLOWANCE_TEMPLATE_COLUMNS = ['Zone', 'Project Location Category', 'Area Allowance Percentage']

# Matches the row of the cross-tab layout that carries the percentages.
AREA_ALLOWANCE_ROW_RE = re.compile(r'area\s*allowance', re.I)


def _parse_area_allowance_workbook(filepath):
    """
    Read and validate an Area Allowance sheet in either accepted layout:

      * the **cross-tab** the SOR publishes -- a ZONE-I/II/III header row,
        one or more heading rows naming each location category beneath its
        zone, and an "Area Allowance" row of percentages; or
      * a flat three-column list (Zone | Project Location Category |
        Area Allowance Percentage).

    Returns ``(rows, errors)``. ``rows`` is the parsed payload for
    :meth:`core.models.AreaAllowanceUpload.replace`; a non-empty ``errors``
    list means the upload must be rejected.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(filepath, data_only=True)
        worksheet = workbook.worksheets[0]
    except Exception as exc:
        return [], [f'Could not read the Excel file: {exc}']

    try:
        if _find_allowance_row(worksheet):
            return _finalize_area_allowance(_parse_area_allowance_matrix(worksheet))
        return _finalize_area_allowance(_parse_area_allowance_columns(filepath))
    finally:
        workbook.close()


def _cell_reader(worksheet):
    """
    Reader that resolves merged cells to their anchor value, so a heading
    merged across two columns reads the same from either of them.
    """
    merged = {}
    for cell_range in worksheet.merged_cells.ranges:
        anchor = worksheet.cell(row=cell_range.min_row, column=cell_range.min_col).value
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for col in range(cell_range.min_col, cell_range.max_col + 1):
                merged[(row, col)] = anchor

    def read(row, col):
        if (row, col) in merged:
            return merged[(row, col)]
        return worksheet.cell(row=row, column=col).value

    return read


def _find_allowance_row(worksheet, search_depth=30):
    """Row number of the "Area Allowance" row in a cross-tab sheet, else None."""
    read = _cell_reader(worksheet)
    for row in range(1, min(worksheet.max_row, search_depth) + 1):
        for col in (1, 2):
            value = read(row, col)
            if isinstance(value, str) and AREA_ALLOWANCE_ROW_RE.search(value):
                return row
    return None


def _parse_area_allowance_matrix(worksheet):
    """Read the cross-tab layout. Returns ``(by_pair, errors)``."""
    from core import zone_policy

    read = _cell_reader(worksheet)
    allowance_row = _find_allowance_row(worksheet)

    # The zone header is the topmost row above the percentages that names a zone.
    zone_row = None
    for row in range(1, allowance_row):
        for col in range(1, worksheet.max_column + 1):
            if zone_policy.normalize_zone(read(row, col)):
                zone_row = row
                break
        if zone_row:
            break
    if zone_row is None:
        return {}, ['Could not find a ZONE-I / ZONE-II / ZONE-III header row above '
                    'the "Area Allowance" row.']

    errors = []
    by_pair = {}
    for col in range(1, worksheet.max_column + 1):
        percent = zone_policy.normalize_percent(read(allowance_row, col))
        if percent is None:
            continue  # Label column, or an empty column past the table

        column_name = worksheet.cell(row=allowance_row, column=col).column_letter
        zone_code = zone_policy.normalize_zone(read(zone_row, col))
        if not zone_code:
            errors.append(f'Column {column_name}: no zone header above the percentage.')
            continue

        # Heading rows between the zone header and the percentages, innermost
        # first -- e.g. ["Upto 16 Kms", "Agency or Tribal Area"]. A category
        # is named either by its own heading alone or by the group heading
        # plus its own, so both readings are tried.
        headings = []
        for row in range(allowance_row - 1, zone_row, -1):
            value = read(row, col)
            text = str(value).strip() if value is not None else ''
            if text and text not in headings:
                headings.append(text)

        candidates = list(headings)
        for outer in headings[1:]:
            candidates.append(f'{outer} {headings[0]}')
        location_code = next(
            (code for code in map(zone_policy.normalize_location, candidates) if code),
            None,
        )
        if not location_code:
            errors.append(
                f'Column {column_name}: unrecognised Project Location Category '
                f'"{" / ".join(headings) or "(blank)"}".'
            )
            continue

        if not zone_policy.is_valid_pair(zone_code, location_code):
            errors.append(
                f'Column {column_name}: "{zone_policy.location_label(location_code)}" '
                f'does not belong to {zone_policy.zone_label(zone_code)}.'
            )
            continue
        if percent < 0:
            errors.append(f'Column {column_name}: percentage cannot be negative.')
            continue

        by_pair[(zone_code, location_code)] = percent

    return by_pair, errors


def _finalize_area_allowance(parsed):
    """Turn ``(by_pair, errors)`` into the stored row payload, rejecting a
    sheet that doesn't cover every zone/location combination."""
    from core import zone_policy

    by_pair, errors = parsed
    if errors:
        return [], errors

    missing_pairs = [pair for pair in zone_policy.expected_pairs() if pair not in by_pair]
    if missing_pairs:
        return [], [
            'Missing entr(y/ies) for: ' + ', '.join(
                f'{zone_policy.zone_label(z)} / {zone_policy.location_label(l)}'
                for z, l in missing_pairs
            ) + '.'
        ]

    rows = [
        zone_policy.build_row(z, l, by_pair[(z, l)])
        for z, l in zone_policy.expected_pairs()
    ]
    return rows, []


def _parse_area_allowance_columns(filepath):
    """Read the flat three-column layout. Returns ``(by_pair, errors)``."""
    from core import zone_policy

    try:
        df = pd.read_excel(filepath)
    except Exception as exc:
        return {}, [f'Could not read the Excel file: {exc}']

    # Map the sheet's headings onto the three fields we need.
    column_for_field = {}
    for column in df.columns:
        field = zone_policy.normalize_header(column)
        if field and field not in column_for_field:
            column_for_field[field] = column

    missing = [
        label for field, label in (
            ('zone', 'Zone'),
            ('location', 'Project Location Category'),
            ('percent', 'Area Allowance Percentage'),
        ) if field not in column_for_field
    ]
    if missing:
        return {}, [
            'Missing required column(s): ' + ', '.join(missing)
            + '. Expected either the SOR cross-tab layout (an "Area Allowance" row '
            + 'under ZONE-I/II/III headers) or the columns: '
            + ', '.join(AREA_ALLOWANCE_TEMPLATE_COLUMNS) + '.'
        ]

    errors = []
    by_pair = {}
    for position, record in enumerate(df.to_dict('records'), start=2):  # +1 for the header row
        raw_zone = record.get(column_for_field['zone'])
        raw_location = record.get(column_for_field['location'])
        raw_percent = record.get(column_for_field['percent'])

        # Skip fully blank rows -- trailing empties are common in hand-edited sheets.
        if all(not str(v).strip() if isinstance(v, str) else pd.isna(v)
               for v in (raw_zone, raw_location, raw_percent)):
            continue

        zone_code = zone_policy.normalize_zone(raw_zone)
        if not zone_code:
            errors.append(f'Row {position}: unrecognised Zone "{raw_zone}".')
            continue

        location_code = zone_policy.normalize_location(raw_location)
        if not location_code:
            errors.append(f'Row {position}: unrecognised Project Location Category "{raw_location}".')
            continue

        if not zone_policy.is_valid_pair(zone_code, location_code):
            errors.append(
                f'Row {position}: "{zone_policy.location_label(location_code)}" does not '
                f'belong to {zone_policy.zone_label(zone_code)}.'
            )
            continue

        percent = zone_policy.normalize_percent(raw_percent)
        if percent is None:
            errors.append(f'Row {position}: "{raw_percent}" is not a valid percentage.')
            continue
        if percent < 0:
            errors.append(f'Row {position}: percentage cannot be negative.')
            continue

        by_pair[(zone_code, location_code)] = percent

    return by_pair, errors


@admin_required
@require_http_methods(["GET", "POST"])
def area_allowance(request):
    """
    Area Allowance Upload page: upload/replace the sheet, see what's currently
    in force, and download the current file back.
    """
    from core import zone_policy
    from core.models import AreaAllowanceUpload

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            messages.error(request, 'No file uploaded.')
            return redirect('admin_area_allowance')

        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload an Excel file (.xlsx or .xls)')
            return redirect('admin_area_allowance')

        DATA_DIR.mkdir(parents=True, exist_ok=True)
        temp_path = DATA_DIR / 'temp_area_allowance.xlsx'
        try:
            file_bytes = uploaded_file.read()
            with open(temp_path, 'wb') as fh:
                fh.write(file_bytes)

            rows, errors = _parse_area_allowance_workbook(temp_path)
            if errors:
                for error in errors[:10]:
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f'...and {len(errors) - 10} more problem(s).')
                return redirect('admin_area_allowance')

            AreaAllowanceUpload.replace(uploaded_file.name, file_bytes, rows, user=request.user)

            try:
                from datasets.models import AuditLog
                AuditLog.log(
                    user=request.user,
                    action='upload',
                    obj='AreaAllowance',
                    changes=None,
                    metadata={'filename': uploaded_file.name, 'rows': len(rows)},
                    request=request,
                )
            except Exception:
                pass  # Audit logging is best-effort

            messages.success(
                request,
                f'Area Allowance updated from "{uploaded_file.name}". '
                f'{len(rows)} zone/location combinations are now in force.'
            )
            return redirect('admin_area_allowance')

        except Exception as exc:
            messages.error(request, f'Error processing file: {exc}')
            return redirect('admin_area_allowance')
        finally:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except OSError:
                pass  # Windows may still hold the handle; harmless

    current = AreaAllowanceUpload.current()
    context = {
        'current': current,
        'rows': current.rows if current else [],
        'zones': zone_policy.ZONES,
        'template_columns': AREA_ALLOWANCE_TEMPLATE_COLUMNS,
    }
    return render(request, 'admin_panel/data/area_allowance.html', context)


@admin_required
def download_area_allowance(request):
    """Download the Area Allowance sheet currently in force."""
    from io import BytesIO
    from core.models import AreaAllowanceUpload

    current = AreaAllowanceUpload.current()
    if not current or not current.file_data:
        messages.error(request, 'No Area Allowance file has been uploaded yet.')
        return redirect('admin_area_allowance')

    return FileResponse(
        BytesIO(bytes(current.file_data)),
        as_attachment=True,
        filename=current.file_name or 'area_allowance.xlsx',
    )
