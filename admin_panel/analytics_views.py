# admin_panel/analytics_views.py
"""
Analytics views for admin panel - user activity, usage stats, revenue trends.
"""

import csv
import json
from io import BytesIO
from datetime import timedelta, date
from collections import defaultdict
from django.shortcuts import render
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q, F
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.utils import timezone
from django.http import JsonResponse, HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from admin_panel.decorators import admin_required
from accounts.models import UserProfile, UserSession
from subscriptions.models import Module, UserModuleSubscription, Payment
from core.models import Estimate, Job, Upload


@admin_required
def analytics_dashboard(request):
    """
    Comprehensive analytics dashboard with charts data.
    """
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    three_months_ago = today - timedelta(days=90)
    
    # -------------------------------------------------------------------------
    # USER STATS
    # -------------------------------------------------------------------------
    total_users = User.objects.filter(is_active=True).count()
    new_users_today = User.objects.filter(date_joined__date=today).count()
    new_users_week = User.objects.filter(date_joined__date__gte=week_ago).count()
    new_users_month = User.objects.filter(date_joined__date__gte=month_ago).count()
    
    # Daily user registrations for last 30 days (for chart)
    user_registrations = (
        User.objects.filter(date_joined__date__gte=month_ago)
        .annotate(day=TruncDate('date_joined'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    registration_chart = {
        'labels': [],
        'data': []
    }
    # Fill in missing days with 0
    date_counts = {item['day']: item['count'] for item in user_registrations}
    for i in range(30, -1, -1):
        d = today - timedelta(days=i)
        registration_chart['labels'].append(d.strftime('%b %d'))
        registration_chart['data'].append(date_counts.get(d, 0))
    
    # -------------------------------------------------------------------------
    # SUBSCRIPTION & MODULE STATS
    # -------------------------------------------------------------------------
    active_subs = UserModuleSubscription.objects.filter(
        status='active',
        expires_at__gt=timezone.now()
    ).count()
    trial_subs = UserModuleSubscription.objects.filter(
        status='trial',
        expires_at__gt=timezone.now()
    ).count()
    expired_subs = UserModuleSubscription.objects.filter(
        status='expired'
    ).count()
    
    # Module popularity (subscribers per module)
    module_stats = (
        Module.objects.filter(is_active=True)
        .annotate(
            active_count=Count(
                'subscriptions',
                filter=Q(subscriptions__status__in=['active', 'trial'], subscriptions__expires_at__gt=timezone.now())
            ),
            total_count=Count('subscriptions')
        )
        .order_by('-active_count')
    )
    
    module_chart = {
        'labels': [m.name for m in module_stats],
        'data': [m.active_count for m in module_stats],
        'colors': [m.color for m in module_stats]
    }
    
    # -------------------------------------------------------------------------
    # USAGE STATS (Estimates, Bills, Jobs)
    # -------------------------------------------------------------------------
    total_estimates = Estimate.objects.count()
    estimates_this_month = Estimate.objects.filter(created_at__date__gte=month_ago).count()
    
    total_jobs = Job.objects.count()
    jobs_this_month = Job.objects.filter(created_at__date__gte=month_ago).count()
    
    # Usage by day for last 30 days
    estimate_usage = (
        Estimate.objects.filter(created_at__date__gte=month_ago)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    job_usage = (
        Job.objects.filter(created_at__date__gte=month_ago)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    
    usage_chart = {
        'labels': [],
        'estimates': [],
        'jobs': []
    }
    estimate_counts = {item['day']: item['count'] for item in estimate_usage}
    job_counts = {item['day']: item['count'] for item in job_usage}
    for i in range(30, -1, -1):
        d = today - timedelta(days=i)
        usage_chart['labels'].append(d.strftime('%b %d'))
        usage_chart['estimates'].append(estimate_counts.get(d, 0))
        usage_chart['jobs'].append(job_counts.get(d, 0))
    
    # -------------------------------------------------------------------------
    # REVENUE STATS
    # -------------------------------------------------------------------------
    total_revenue = Payment.objects.filter(status='success').aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_this_month = Payment.objects.filter(
        status='success',
        created_at__date__gte=month_ago
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_last_month = Payment.objects.filter(
        status='success',
        created_at__date__gte=(month_ago - timedelta(days=30)),
        created_at__date__lt=month_ago
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Monthly revenue for last 6 months
    revenue_by_month = (
        Payment.objects.filter(
            status='success',
            created_at__date__gte=(today - timedelta(days=180))
        )
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Sum('total_amount'))
        .order_by('month')
    )
    
    revenue_chart = {
        'labels': [r['month'].strftime('%b %Y') for r in revenue_by_month],
        'data': [float(r['total']) for r in revenue_by_month]
    }
    
    # -------------------------------------------------------------------------
    # TOP USERS BY USAGE
    # -------------------------------------------------------------------------
    top_users_by_estimates = (
        User.objects.annotate(
            estimate_count=Count('estimates')
        )
        .filter(estimate_count__gt=0)
        .order_by('-estimate_count')[:10]
    )
    
    top_users_by_jobs = (
        User.objects.annotate(
            job_count=Count('jobs')
        )
        .filter(job_count__gt=0)
        .order_by('-job_count')[:10]
    )
    
    # -------------------------------------------------------------------------
    # ACTIVE SESSIONS
    # -------------------------------------------------------------------------
    active_sessions = UserSession.objects.filter(
        is_active=True,
        last_activity__gte=timezone.now() - timedelta(minutes=30)
    ).count()
    
    context = {
        # User stats
        'total_users': total_users,
        'new_users_today': new_users_today,
        'new_users_week': new_users_week,
        'new_users_month': new_users_month,
        'registration_chart': json.dumps(registration_chart),
        
        # Subscription stats
        'active_subscriptions': active_subs,
        'trial_subscriptions': trial_subs,
        'expired_subscriptions': expired_subs,
        'module_stats': module_stats,
        'module_chart': json.dumps(module_chart),
        
        # Usage stats
        'total_estimates': total_estimates,
        'estimates_this_month': estimates_this_month,
        'total_jobs': total_jobs,
        'jobs_this_month': jobs_this_month,
        'usage_chart': json.dumps(usage_chart),
        
        # Revenue stats
        'total_revenue': total_revenue,
        'revenue_this_month': revenue_this_month,
        'revenue_last_month': revenue_last_month,
        'revenue_chart': json.dumps(revenue_chart),
        
        # Top users
        'top_users_by_estimates': top_users_by_estimates,
        'top_users_by_jobs': top_users_by_jobs,
        
        # Sessions
        'active_sessions': active_sessions,
    }
    
    return render(request, 'admin_panel/analytics/dashboard.html', context)


@admin_required
def user_analytics(request, user_id):
    """
    Detailed analytics for a specific user.
    """
    from django.shortcuts import get_object_or_404
    
    user = get_object_or_404(User, id=user_id)
    today = timezone.now().date()
    month_ago = today - timedelta(days=30)
    
    # User's estimates
    estimates = Estimate.objects.filter(user=user).order_by('-created_at')
    estimates_this_month = estimates.filter(created_at__date__gte=month_ago).count()
    
    # User's jobs
    jobs = Job.objects.filter(user=user).order_by('-created_at')
    jobs_this_month = jobs.filter(created_at__date__gte=month_ago).count()
    
    # User's subscriptions
    subscriptions = UserModuleSubscription.objects.filter(user=user).select_related('module')
    
    # User's payments
    payments = Payment.objects.filter(user=user).order_by('-created_at')
    total_paid = payments.filter(status='success').aggregate(total=Sum('amount'))['total'] or 0
    
    # Activity timeline
    activity = []
    for est in estimates[:10]:
        activity.append({
            'type': 'estimate',
            'title': f'Created estimate: {est.work_name}',
            'date': est.created_at
        })
    for job in jobs[:10]:
        activity.append({
            'type': 'job',
            'title': f'Processed job: {job.work_name or "Untitled"}',
            'date': job.created_at
        })
    activity.sort(key=lambda x: x['date'], reverse=True)
    
    context = {
        'target_user': user,
        'estimates': estimates[:20],
        'estimates_count': estimates.count(),
        'estimates_this_month': estimates_this_month,
        'jobs': jobs[:20],
        'jobs_count': jobs.count(),
        'jobs_this_month': jobs_this_month,
        'subscriptions': subscriptions,
        'payments': payments[:10],
        'total_paid': total_paid,
        'activity': activity[:20],
    }
    
    return render(request, 'admin_panel/analytics/user_detail.html', context)


@admin_required
def analytics_api(request):
    """
    API endpoint for live chart data.
    """
    chart_type = request.GET.get('type', 'registrations')
    period = request.GET.get('period', '30')  # days
    
    try:
        days = int(period)
    except ValueError:
        days = 30
    
    today = timezone.now().date()
    start_date = today - timedelta(days=days)
    
    if chart_type == 'registrations':
        data = (
            User.objects.filter(date_joined__date__gte=start_date)
            .annotate(day=TruncDate('date_joined'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )
        date_counts = {item['day']: item['count'] for item in data}
        labels = []
        values = []
        for i in range(days, -1, -1):
            d = today - timedelta(days=i)
            labels.append(d.strftime('%b %d'))
            values.append(date_counts.get(d, 0))
        
        return JsonResponse({'labels': labels, 'data': values})
    
    elif chart_type == 'usage':
        estimates = (
            Estimate.objects.filter(created_at__date__gte=start_date)
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )
        jobs = (
            Job.objects.filter(created_at__date__gte=start_date)
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )
        
        est_counts = {item['day']: item['count'] for item in estimates}
        job_counts = {item['day']: item['count'] for item in jobs}
        
        labels = []
        est_values = []
        job_values = []
        for i in range(days, -1, -1):
            d = today - timedelta(days=i)
            labels.append(d.strftime('%b %d'))
            est_values.append(est_counts.get(d, 0))
            job_values.append(job_counts.get(d, 0))
        
        return JsonResponse({
            'labels': labels,
            'estimates': est_values,
            'jobs': job_values
        })
    
    elif chart_type == 'revenue':
        data = (
            Payment.objects.filter(
                status='success',
                created_at__date__gte=start_date
            )
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(total=Sum('total_amount'))
            .order_by('day')
        )
        date_totals = {item['day']: float(item['total']) for item in data}
        
        labels = []
        values = []
        for i in range(days, -1, -1):
            d = today - timedelta(days=i)
            labels.append(d.strftime('%b %d'))
            values.append(date_totals.get(d, 0))
        
        return JsonResponse({'labels': labels, 'data': values})
    
    return JsonResponse({'error': 'Invalid chart type'}, status=400)


@admin_required
def export_analytics(request, export_type):
    """
    Export analytics data as Excel with multiple sheets and formatting.
    Supports: users, estimates, all
    """
    today = timezone.now()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(bold=True, size=14, color='1F4E79')
    subtitle_font = Font(bold=True, size=12, color='2E75B6')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_adjust_columns(ws):
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def add_data_rows(ws, start_row, data, apply_borders=True):
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                if apply_borders:
                    cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = alt_row_fill
    
    wb = Workbook()
    
    if export_type == 'users':
        ws = wb.active
        ws.title = 'Users'
        
        # Title
        ws['A1'] = 'Users Export'
        ws['A1'].font = title_font
        ws['A2'] = f'Generated: {today.strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(italic=True, color='666666')
        
        # Headers
        headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date Joined', 'Last Login', 'Is Active', 'Is Staff']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        style_header_row(ws, 4, len(headers))
        
        # Data
        users = User.objects.all().order_by('-date_joined')
        data = []
        for user in users:
            data.append([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
                'Yes' if user.is_active else 'No',
                'Yes' if user.is_staff else 'No'
            ])
        add_data_rows(ws, 5, data)
        auto_adjust_columns(ws)
        
        # Freeze panes
        ws.freeze_panes = 'A5'
    
    elif export_type == 'estimates':
        ws = wb.active
        ws.title = 'Estimates'
        
        # Title
        ws['A1'] = 'Estimates Export'
        ws['A1'].font = title_font
        ws['A2'] = f'Generated: {today.strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(italic=True, color='666666')
        
        # Headers
        headers = ['ID', 'Work Name', 'User', 'Email', 'Created At', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        style_header_row(ws, 4, len(headers))
        
        # Data
        estimates = Estimate.objects.all().select_related('user').order_by('-created_at')
        data = []
        for est in estimates:
            data.append([
                est.id,
                est.work_name or 'Untitled',
                est.user.username if est.user else 'N/A',
                est.user.email if est.user else 'N/A',
                est.created_at.strftime('%Y-%m-%d %H:%M:%S') if est.created_at else '',
                getattr(est, 'status', 'N/A')
            ])
        add_data_rows(ws, 5, data)
        auto_adjust_columns(ws)
        ws.freeze_panes = 'A5'
    
    elif export_type == 'all':
        # ========== SUMMARY SHEET ==========
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        
        ws_summary['A1'] = 'Analytics Report'
        ws_summary['A1'].font = Font(bold=True, size=18, color='1F4E79')
        ws_summary['A2'] = f'Generated: {today.strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary['A2'].font = Font(italic=True, color='666666')
        
        row = 4
        
        # User Statistics
        ws_summary.cell(row=row, column=1, value='USER STATISTICS').font = subtitle_font
        row += 1
        stats = [
            ['Total Active Users', User.objects.filter(is_active=True).count()],
            ['New Users (Last 7 Days)', User.objects.filter(date_joined__date__gte=today.date() - timedelta(days=7)).count()],
            ['New Users (Last 30 Days)', User.objects.filter(date_joined__date__gte=today.date() - timedelta(days=30)).count()],
        ]
        for stat in stats:
            ws_summary.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=stat[1])
            row += 1
        row += 1
        
        # Subscription Statistics
        ws_summary.cell(row=row, column=1, value='SUBSCRIPTION STATISTICS').font = subtitle_font
        row += 1
        stats = [
            ['Active Subscriptions', UserModuleSubscription.objects.filter(status='active', expires_at__gt=today).count()],
            ['Trial Subscriptions', UserModuleSubscription.objects.filter(status='trial', expires_at__gt=today).count()],
            ['Expired Subscriptions', UserModuleSubscription.objects.filter(status='expired').count()],
        ]
        for stat in stats:
            ws_summary.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=stat[1])
            row += 1
        row += 1
        
        # Usage Statistics
        ws_summary.cell(row=row, column=1, value='USAGE STATISTICS').font = subtitle_font
        row += 1
        stats = [
            ['Total Estimates', Estimate.objects.count()],
            ['Estimates (Last 30 Days)', Estimate.objects.filter(created_at__date__gte=today.date() - timedelta(days=30)).count()],
            ['Total Jobs', Job.objects.count()],
            ['Jobs (Last 30 Days)', Job.objects.filter(created_at__date__gte=today.date() - timedelta(days=30)).count()],
        ]
        for stat in stats:
            ws_summary.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=stat[1])
            row += 1
        row += 1
        
        # Revenue Statistics
        ws_summary.cell(row=row, column=1, value='REVENUE STATISTICS').font = subtitle_font
        row += 1
        total_revenue = Payment.objects.filter(status='success').aggregate(total=Sum('total_amount'))['total'] or 0
        month_revenue = Payment.objects.filter(status='success', created_at__date__gte=today.date() - timedelta(days=30)).aggregate(total=Sum('total_amount'))['total'] or 0
        stats = [
            ['Total Revenue', f'₹{total_revenue:,.2f}'],
            ['Revenue (Last 30 Days)', f'₹{month_revenue:,.2f}'],
        ]
        for stat in stats:
            ws_summary.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=stat[1])
            row += 1
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # ========== USERS SHEET ==========
        ws_users = wb.create_sheet('Users')
        ws_users['A1'] = 'All Users'
        ws_users['A1'].font = title_font
        
        headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date Joined', 'Last Login', 'Is Active', 'Is Staff']
        for col, header in enumerate(headers, 1):
            ws_users.cell(row=3, column=col, value=header)
        style_header_row(ws_users, 3, len(headers))
        
        users = User.objects.all().order_by('-date_joined')
        data = []
        for user in users:
            data.append([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
                'Yes' if user.is_active else 'No',
                'Yes' if user.is_staff else 'No'
            ])
        add_data_rows(ws_users, 4, data)
        auto_adjust_columns(ws_users)
        ws_users.freeze_panes = 'A4'
        
        # ========== ESTIMATES SHEET ==========
        ws_estimates = wb.create_sheet('Estimates')
        ws_estimates['A1'] = 'All Estimates'
        ws_estimates['A1'].font = title_font
        
        headers = ['ID', 'Work Name', 'User', 'Email', 'Created At']
        for col, header in enumerate(headers, 1):
            ws_estimates.cell(row=3, column=col, value=header)
        style_header_row(ws_estimates, 3, len(headers))
        
        estimates = Estimate.objects.all().select_related('user').order_by('-created_at')
        data = []
        for est in estimates:
            data.append([
                est.id,
                est.work_name or 'Untitled',
                est.user.username if est.user else 'N/A',
                est.user.email if est.user else 'N/A',
                est.created_at.strftime('%Y-%m-%d %H:%M:%S') if est.created_at else ''
            ])
        add_data_rows(ws_estimates, 4, data)
        auto_adjust_columns(ws_estimates)
        ws_estimates.freeze_panes = 'A4'
        
        # ========== JOBS SHEET ==========
        ws_jobs = wb.create_sheet('Jobs')
        ws_jobs['A1'] = 'All Jobs'
        ws_jobs['A1'].font = title_font
        
        headers = ['ID', 'Job Type', 'User', 'Email', 'Created At', 'Status', 'Progress']
        for col, header in enumerate(headers, 1):
            ws_jobs.cell(row=3, column=col, value=header)
        style_header_row(ws_jobs, 3, len(headers))
        
        jobs = Job.objects.all().select_related('user').order_by('-created_at')
        data = []
        for job in jobs:
            data.append([
                job.id,
                job.get_job_type_display() if hasattr(job, 'get_job_type_display') else job.job_type,
                job.user.username if job.user else 'N/A',
                job.user.email if job.user else 'N/A',
                job.created_at.strftime('%Y-%m-%d %H:%M:%S') if job.created_at else '',
                job.get_status_display() if hasattr(job, 'get_status_display') else job.status,
                f'{job.progress}%' if hasattr(job, 'progress') else 'N/A'
            ])
        add_data_rows(ws_jobs, 4, data)
        auto_adjust_columns(ws_jobs)
        ws_jobs.freeze_panes = 'A4'
        
        # ========== SUBSCRIPTIONS SHEET ==========
        ws_subs = wb.create_sheet('Subscriptions')
        ws_subs['A1'] = 'All Subscriptions'
        ws_subs['A1'].font = title_font
        
        headers = ['ID', 'User', 'Email', 'Module', 'Status', 'Started At', 'Expires At']
        for col, header in enumerate(headers, 1):
            ws_subs.cell(row=3, column=col, value=header)
        style_header_row(ws_subs, 3, len(headers))
        
        subs = UserModuleSubscription.objects.all().select_related('user', 'module').order_by('-created_at')
        data = []
        for sub in subs:
            data.append([
                str(sub.id)[:8] if hasattr(sub, 'id') else 'N/A',
                sub.user.username if sub.user else 'N/A',
                sub.user.email if sub.user else 'N/A',
                sub.module.name if sub.module else 'N/A',
                sub.status,
                sub.created_at.strftime('%Y-%m-%d %H:%M:%S') if sub.created_at else '',
                sub.expires_at.strftime('%Y-%m-%d %H:%M:%S') if sub.expires_at else ''
            ])
        add_data_rows(ws_subs, 4, data)
        auto_adjust_columns(ws_subs)
        ws_subs.freeze_panes = 'A4'
        
        # ========== PAYMENTS SHEET ==========
        ws_payments = wb.create_sheet('Payments')
        ws_payments['A1'] = 'All Payments'
        ws_payments['A1'].font = title_font
        
        headers = ['ID', 'User', 'Email', 'Amount', 'Status', 'Payment Date']
        for col, header in enumerate(headers, 1):
            ws_payments.cell(row=3, column=col, value=header)
        style_header_row(ws_payments, 3, len(headers))
        
        payments = Payment.objects.all().select_related('user').order_by('-created_at')
        data = []
        for payment in payments:
            data.append([
                str(payment.id)[:8] if hasattr(payment, 'id') else 'N/A',
                payment.user.username if payment.user else 'N/A',
                payment.user.email if payment.user else 'N/A',
                float(payment.total_amount) if payment.total_amount else 0,
                payment.status,
                payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else ''
            ])
        add_data_rows(ws_payments, 4, data)
        auto_adjust_columns(ws_payments)
        ws_payments.freeze_panes = 'A4'
    
    else:
        return JsonResponse({'error': 'Invalid export type'}, status=400)
    
    # Save workbook to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="analytics_{export_type}_{today.strftime("%Y%m%d")}.xlsx"'
    
    return response
