"""Excel writer: produces a Bill-of-Quantities pivot in legend-table style."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GROUP_FILL = PatternFill("solid", fgColor="DDEBF7")
GROUP_FONT = Font(bold=True, color="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="E7E6E6")
BORDER = Border(*[Side(style="thin", color="888888")] * 4)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _safe_sheet_name(name: str, used: set) -> str:
    cleaned = "".join("_" if c in "[]:*?/\\" else c for c in (name or "Sheet"))
    cleaned = cleaned.strip() or "Sheet"
    base = cleaned[:31]
    candidate = base
    i = 2
    while candidate.lower() in used:
        suffix = f" ({i})"
        candidate = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(candidate.lower())
    return candidate


def _split_key(k: str):
    if "||" in k:
        g, d = k.split("||", 1)
        return g, d
    return "Default", k


def _ensure_unassigned(rows, zones):
    if any("Unassigned" in r["zones"] for r in rows) and "Unassigned" not in zones:
        zones = list(zones) + ["Unassigned"]
    return zones


def _normalize_rows(summary):
    """summary may be either {key: {zone: n}} (key may be 'group||desc') or
    {desc: {zone: n}} (legacy). Return list of {group, desc, zones}."""
    rows = []
    for key, zd in (summary or {}).items():
        g, d = _split_key(key)
        rows.append({"group": g, "desc": d, "zones": dict(zd or {})})
    return rows


def _coerce(v):
    """Counts come through as int; pipe lengths as float. Preserve type so
    Excel renders both naturally (and totals add cleanly)."""
    try:
        fv = float(v)
    except (TypeError, ValueError):
        return 0
    if fv == int(fv):
        return int(fv)
    return round(fv, 3)


def _write_pivot(ws, summary: dict, zones: list):
    """Write a single sheet. Rows are grouped by legend `group` with a
    subtotal row per group when there is more than one group."""
    rows = _normalize_rows(summary)
    zones = _ensure_unassigned(rows, list(zones) if zones else [])

    groups = sorted({r["group"] for r in rows}, key=lambda g: g.lower())
    multi_group = len(groups) > 1

    if multi_group:
        header = ["#", "Category", "Description"] + zones + ["Total"]
        desc_col = 3
    else:
        header = ["#", "Description"] + zones + ["Total"]
        desc_col = 2

    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    grand_total = 0
    zone_totals = {z: 0 for z in zones}
    serial = 0

    if not rows:
        empty = [""] * len(header)
        empty[desc_col - 1] = "(no data)"
        ws.append(empty)

    for g in groups:
        group_rows = sorted(
            (r for r in rows if r["group"] == g),
            key=lambda r: r["desc"].lower(),
        )
        if multi_group:
            # Category header row spanning all data columns.
            ws.append([""] * len(header))
            r = ws.max_row
            ws.cell(row=r, column=2).value = g
            for c in range(1, len(header) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
                cell.alignment = LEFT if c == 2 else CENTER
                cell.border = BORDER

        group_zone_totals = {z: 0 for z in zones}
        group_total = 0
        for r in group_rows:
            serial += 1
            if multi_group:
                row = [serial, "", r["desc"]]
            else:
                row = [serial, r["desc"]]
            row_total = 0
            for z in zones:
                v = _coerce(r["zones"].get(z, 0))
                row.append(v)
                row_total += v
                zone_totals[z] = _coerce(zone_totals[z] + v)
                group_zone_totals[z] = _coerce(group_zone_totals[z] + v)
            row.append(_coerce(row_total))
            group_total = _coerce(group_total + row_total)
            grand_total = _coerce(grand_total + row_total)
            ws.append(row)
            rn = ws.max_row
            ws.cell(row=rn, column=1).alignment = CENTER
            ws.cell(row=rn, column=desc_col).alignment = LEFT
            for c in range(desc_col + 1, len(header) + 1):
                ws.cell(row=rn, column=c).alignment = CENTER
            for c in range(1, len(header) + 1):
                ws.cell(row=rn, column=c).border = BORDER

        if multi_group and group_rows:
            if multi_group:
                totals_row = ["", "", f"Subtotal — {g}"]
            else:
                totals_row = ["", f"Subtotal — {g}"]
            totals_row += [group_zone_totals[z] for z in zones] + [group_total]
            ws.append(totals_row)
            rn = ws.max_row
            for c in range(1, len(header) + 1):
                cell = ws.cell(row=rn, column=c)
                cell.font = Font(bold=True)
                cell.fill = TOTAL_FILL
                cell.alignment = LEFT if c == desc_col else CENTER
                cell.border = BORDER

    if multi_group:
        totals = ["", "", "GRAND TOTAL"] + [zone_totals[z] for z in zones] + [grand_total]
    else:
        totals = ["", "TOTAL"] + [zone_totals[z] for z in zones] + [grand_total]
    ws.append(totals)
    rn = ws.max_row
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=rn, column=c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.alignment = LEFT if c == desc_col else CENTER
        cell.border = BORDER

    ws.column_dimensions["A"].width = 6
    if multi_group:
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 42
    else:
        ws.column_dimensions["B"].width = 42
    for idx in range(desc_col + 1, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.freeze_panes = f"{get_column_letter(desc_col + 1)}2"


def build_takeoff_workbook(name: str, summary: dict, zone_names: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (name or "Takeoff")[:31]
    _write_pivot(ws, summary, list(zone_names) if zone_names else [])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_takeoff_workbook_multi(name: str, per_sheet: dict, zone_names: list) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used: set = set()
    if not per_sheet:
        ws = wb.create_sheet(_safe_sheet_name(name or "Takeoff", used))
        _write_pivot(ws, {}, list(zone_names) if zone_names else [])
    else:
        names = sorted(per_sheet.keys(),
                       key=lambda n: (n.lower().startswith("unassigned"), n.lower()))
        for sheet_name in names:
            ws = wb.create_sheet(_safe_sheet_name(sheet_name, used))
            _write_pivot(ws, per_sheet[sheet_name], list(zone_names) if zone_names else [])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

