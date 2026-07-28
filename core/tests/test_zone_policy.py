"""
Zone / Area-allowance policy: decoding the project-location value, deriving
zone labour rates, and applying both to a real backend item block.
"""

import pytest
from openpyxl import Workbook

from core import zone_policy
from core.utils_excel import (
    apply_policy_to_copied_block,
    compute_block_rate,
    find_labour_rate_rows,
)


# The shape of a real item block from core/data/electrical.xlsx: a labour
# section with four day-rate lines, a per-metre conversion, a 40% area
# allowance, materials, and a 10.615% overhead.
LABOUR_RATES = [775, 650, 650, 680]


def _make_block():
    """Returns (ws_values, ws_formulas) for a one-block sheet spanning rows 1-24."""
    wb_vals = Workbook()
    wb_for = Workbook()
    wv = wb_vals.active
    wf = wb_for.active

    def put(ws, row, col, value):
        ws.cell(row=row, column=col).value = value

    for ws in (wv, wf):
        put(ws, 1, 4, "Concealed PVC Pipe")
        put(ws, 2, 4, "Data")
        put(ws, 3, 4, "Long item description mentioning all labour charges etc., complete.")
        put(ws, 4, 4, "a) Labour charges :")
        put(ws, 4, 9, "Rate")
        for offset, rate in enumerate(LABOUR_RATES):
            row = 5 + offset
            put(ws, row, 2, "Zone-I")   # the marker SOR 2026-27 puts in column B
            put(ws, row, 3, offset + 1)
            put(ws, row, 4, f"Labour line {offset + 1}")
            put(ws, row, 5, 2)       # quantity
            put(ws, row, 7, 1)       # per-unit
            put(ws, row, 9, rate)    # Zone-1 rate
        put(ws, 9, 4, "Labour Rate Per 100 Rm")
        put(ws, 10, 4, "Labour Rate for  1 Mtr")
        put(ws, 11, 4, " Add GHMC allowance@40%")
        put(ws, 12, 4, " Labour Sub  total")
        put(ws, 13, 4, "b) Material")
        put(ws, 14, 4, "Cement")
        put(ws, 14, 5, 50)
        put(ws, 14, 7, 1)
        put(ws, 14, 9, 9)
        put(ws, 19, 4, "Material Rate for 100 Mtrs")
        put(ws, 20, 4, "Material Rate for  1 Mtr")
        put(ws, 21, 4, "Total Cost of Material + Labour Charges")
        put(ws, 22, 4, "Add Overhead @ 10.615% ")
        put(ws, 23, 4, "Total Rate")
        put(ws, 24, 4, " Rounded off rate")

    # Column J: formulas in the formula sheet, cached results in the value sheet.
    formulas = {
        5: "=SUM(E5*I5/G5)", 6: "=SUM(E6*I6/G6)", 7: "=SUM(E7*I7/G7)", 8: "=SUM(E8*I8/G8)",
        9: "=SUM(J5:J8)", 10: "=SUM(J9/100)", 11: "=J10*0.4", 12: "=SUM(J10:J11)",
        14: "=SUM(E14*I14/G14)", 19: "=SUM(J14:J14)", 20: "=SUM(J19/100)",
        21: "=J20+J12", 22: "=(J21)*0.10615", 23: "=SUM(J21:J22)", 24: "=ROUND(J23,0)",
    }
    for row, formula in formulas.items():
        put(wf, row, 10, formula)

    labour_total = sum(2 * r for r in LABOUR_RATES)
    per_metre = labour_total / 100
    allowance = per_metre * 0.4
    material = 450 / 100
    total = per_metre + allowance + material
    cached = {
        5: 1550, 6: 1300, 7: 1300, 8: 1360, 9: labour_total, 10: per_metre,
        11: allowance, 12: per_metre + allowance, 14: 450, 19: 450, 20: material,
        21: total, 22: total * 0.10615, 23: total * 1.10615,
        24: round(total * 1.10615),
    }
    for row, value in cached.items():
        put(wv, row, 10, value)

    return wv, wf


def _expected_rounded(deduction, allowance_percent, overhead=0.10615):
    labour = sum(2 * max(0, r - deduction) for r in LABOUR_RATES) / 100
    labour += labour * (allowance_percent / 100)
    total = labour + 4.5
    return round(total * (1 + overhead))


# ---------------------------------------------------------------------------
# Decoding
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("value,expected", [
    ("municipal", ("zone_1", "ghmc")),
    ("non_municipal", ("zone_1", None)),
    ("", ("zone_1", "ghmc")),
    (None, ("zone_1", "ghmc")),
    ("zone_2:industrial_area", ("zone_2", "industrial_area")),
    ("zone_3:beyond_16_kms", ("zone_3", "beyond_16_kms")),
    # A category that doesn't belong to the zone falls back to the default.
    ("zone_2:ghmc", ("zone_1", "ghmc")),
    ("garbage", ("zone_1", "ghmc")),
])
def test_split_handles_legacy_and_invalid_values(value, expected):
    assert zone_policy.split(value) == expected


def test_legacy_values_keep_their_original_meaning():
    # Municipal left the workbook's own allowance percentage alone...
    assert zone_policy.resolve("municipal")["allowance_percent"] is None
    # ...and non-municipal zeroed it. Neither shifts labour rates.
    assert zone_policy.resolve("non_municipal")["allowance_percent"] == 0.0
    assert zone_policy.resolve("municipal")["labour_deduction"] == 0.0
    assert zone_policy.resolve("non_municipal")["labour_deduction"] == 0.0


def test_default_area_matches_legacy_municipal():
    assert zone_policy.resolve(zone_policy.DEFAULT_AREA) == zone_policy.resolve("municipal")


@pytest.mark.parametrize("area,deduction", [
    ("zone_1:ghmc", 0.0),
    ("zone_2:municipal_area", 40.0),
    ("zone_3:upto_16_kms", 80.0),
])
def test_labour_deduction_per_zone(area, deduction):
    assert zone_policy.resolve(area)["labour_deduction"] == deduction


# ---------------------------------------------------------------------------
# Sheet parsing helpers
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("Zone 1", "zone_1"), ("zone-2", "zone_2"), ("ZONE_3", "zone_3"),
    ("Zone-I", "zone_1"), ("Zone-II", "zone_2"), ("III", "zone_3"),
    ("2", "zone_2"), ("Zone 4", None), ("", None), (None, None),
])
def test_normalize_zone(raw, expected):
    assert zone_policy.normalize_zone(raw) == expected


@pytest.mark.parametrize("raw,expected", [
    ("GHMC", "ghmc"), ("Other than GHMC", "other_than_ghmc"),
    ("other_than_ghmc", "other_than_ghmc"), ("Municipal Area", "municipal_area"),
    ("  tribal area ", "beyond_16_kms"),           # legacy wording, still accepted
    ("Upto 16 Kms", "upto_16_kms"), ("Beyond 16 Kms", "beyond_16_kms"),
    ("Agency or Tribal Area Upto 16 Kms", "upto_16_kms"),
    ("Municipal Area or District HQ", "municipal_area"),
    ("Somewhere Else", None),
])
def test_normalize_location(raw, expected):
    assert zone_policy.normalize_location(raw) == expected


@pytest.mark.parametrize("raw,expected", [
    (40, 40.0), ("40%", 40.0), ("40 %", 40.0), ("40", 40.0),
    (0.4, 40.0),        # percentage-formatted cell
    (12.75, 12.75), ("", None), (None, None), ("abc", None),
])
def test_normalize_percent(raw, expected):
    assert zone_policy.normalize_percent(raw) == expected


def test_normalize_header_accepts_the_documented_columns():
    assert zone_policy.normalize_header("Zone") == "zone"
    assert zone_policy.normalize_header("Project Location Category") == "location"
    assert zone_policy.normalize_header("Area Allowance Percentage") == "percent"
    assert zone_policy.normalize_header("Something Else") is None


# ---------------------------------------------------------------------------
# Rate computation
# ---------------------------------------------------------------------------

def test_find_labour_rate_rows_stops_before_the_material_rates():
    wv, _ = _make_block()
    assert find_labour_rate_rows(wv, 1, 24) == [5, 6, 7, 8]


def test_zone_1_ghmc_reproduces_the_workbook_rate(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    wv, wf = _make_block()
    baseline = compute_block_rate(wv, wf, 1, 24, area="municipal", work_type="repair")
    assert compute_block_rate(wv, wf, 1, 24, area="zone_1:ghmc", work_type="repair") == baseline
    assert baseline == _expected_rounded(0, 40)


def test_zone_2_and_3_deduct_from_every_labour_rate(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    wv, wf = _make_block()
    # No uploaded sheet, so these zones get no allowance.
    assert compute_block_rate(wv, wf, 1, 24, area="zone_2:municipal_area",
                              work_type="repair") == _expected_rounded(40, 0)
    assert compute_block_rate(wv, wf, 1, 24, area="zone_3:upto_16_kms",
                              work_type="repair") == _expected_rounded(80, 0)


def test_uploaded_allowance_is_applied(monkeypatch):
    monkeypatch.setattr(
        zone_policy, "area_allowance_map",
        lambda: {("zone_2", "industrial_area"): 25.0, ("zone_1", "ghmc"): 40.0},
    )
    wv, wf = _make_block()
    assert compute_block_rate(wv, wf, 1, 24, area="zone_2:industrial_area",
                              work_type="repair") == _expected_rounded(40, 25)


def test_original_work_overhead_still_applies_on_top_of_zone(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    wv, wf = _make_block()
    assert compute_block_rate(wv, wf, 1, 24, area="zone_3:beyond_16_kms",
                              work_type="original") == _expected_rounded(80, 0, 0.13615)


def test_block_without_labour_or_policy_rows_is_untouched(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=4).value = "Simple item"
    ws.cell(row=2, column=10).value = 123.45
    assert compute_block_rate(ws, ws, 1, 2, area="zone_3:upto_16_kms") == 123.45


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def test_copied_block_gets_zone_rates_and_allowance_label(monkeypatch):
    monkeypatch.setattr(
        zone_policy, "area_allowance_map",
        lambda: {("zone_2", "industrial_area"): 25.0},
    )
    wv, wf = _make_block()
    # apply_policy_to_copied_block edits the destination in place; the formula
    # sheet stands in for a freshly copied block.
    apply_policy_to_copied_block(wf, 1, 1, 24, "zone_2:industrial_area", "repair")

    assert [wf.cell(row=r, column=9).value for r in (5, 6, 7, 8)] == [735, 610, 610, 640]
    assert [wf.cell(row=r, column=2).value for r in (5, 6, 7, 8)] == ["Zone-II"] * 4
    assert wf.cell(row=11, column=10).value == "=J10*0.25"
    assert "25%" in wf.cell(row=11, column=4).value
    assert not wf.row_dimensions[11].hidden


def test_zero_allowance_hides_the_row(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    wv, wf = _make_block()
    apply_policy_to_copied_block(wf, 1, 1, 24, "zone_3:upto_16_kms", "repair")

    assert [wf.cell(row=r, column=9).value for r in (5, 6, 7, 8)] == [695, 570, 570, 600]
    assert wf.cell(row=11, column=10).value == "=J10*0"
    assert wf.row_dimensions[11].hidden


def test_legacy_municipal_copy_is_a_no_op(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    wv, wf = _make_block()
    apply_policy_to_copied_block(wf, 1, 1, 24, "municipal", "repair")

    assert [wf.cell(row=r, column=9).value for r in (5, 6, 7, 8)] == LABOUR_RATES
    assert [wf.cell(row=r, column=2).value for r in (5, 6, 7, 8)] == ["Zone-I"] * 4
    assert wf.cell(row=11, column=10).value == "=J10*0.4"
    assert not wf.row_dimensions[11].hidden


# ---------------------------------------------------------------------------
# Reading an uploaded Area Allowance sheet
# ---------------------------------------------------------------------------

def _write_sor_crosstab(path, percentages=(0.40, 0.25, 0.20, 0.20, 0.25, 0.40)):
    """The layout the SOR publishes: merged zone headers, merged group
    headings, and an "Area Allowance" row of percent-formatted cells."""
    wb = Workbook()
    ws = wb.active

    ws["A2"] = "Type of Area"
    ws.merge_cells("A2:A3")
    ws["B1"] = "ZONE-I"
    ws.merge_cells("B1:C1")
    ws["D1"] = "ZONE-II"
    ws.merge_cells("D1:E1")
    ws["F1"] = "ZONE-III"
    ws.merge_cells("F1:G1")

    ws["B2"] = "Municipal Corporation"
    ws.merge_cells("B2:C2")
    ws["B3"] = "GHMC"
    ws["C3"] = "Other than GHMC"
    ws["D2"] = "Municipal Area or District HQ"
    ws.merge_cells("D2:D3")
    ws["E2"] = "Industrial Area"
    ws.merge_cells("E2:E3")
    ws["F2"] = "Agency or Tribal Area"
    ws.merge_cells("F2:G2")
    ws["F3"] = "Upto 16 Kms"
    ws["G3"] = "Beyond 16 Kms"

    ws["A4"] = "Area Allowance"
    for col, value in zip("BCDEFG", percentages):
        cell = ws[f"{col}4"]
        cell.value = value
        cell.number_format = "0%"

    wb.save(path)
    return path


def test_sor_crosstab_upload_is_parsed(tmp_path):
    from admin_panel.data_management_views import _parse_area_allowance_workbook

    rows, errors = _parse_area_allowance_workbook(
        _write_sor_crosstab(tmp_path / "area_allowance.xlsx")
    )
    assert errors == []
    assert [(r["zone"], r["location"], r["percent"]) for r in rows] == [
        ("zone_1", "ghmc", 40.0),
        ("zone_1", "other_than_ghmc", 25.0),
        ("zone_2", "municipal_area", 20.0),
        ("zone_2", "industrial_area", 20.0),
        ("zone_3", "upto_16_kms", 25.0),
        ("zone_3", "beyond_16_kms", 40.0),
    ]


def test_crosstab_missing_a_percentage_is_rejected(tmp_path):
    from openpyxl import load_workbook
    from admin_panel.data_management_views import _parse_area_allowance_workbook

    path = _write_sor_crosstab(tmp_path / "area_allowance.xlsx")
    wb = load_workbook(path)
    wb.active["G4"] = None
    wb.save(path)

    rows, errors = _parse_area_allowance_workbook(path)
    assert rows == []
    assert "Beyond 16 Kms" in errors[0]


def test_crosstab_with_an_unknown_category_is_rejected(tmp_path):
    from openpyxl import load_workbook
    from admin_panel.data_management_views import _parse_area_allowance_workbook

    path = _write_sor_crosstab(tmp_path / "area_allowance.xlsx")
    wb = load_workbook(path)
    wb.active["C3"] = "Somewhere Else"
    wb.save(path)

    rows, errors = _parse_area_allowance_workbook(path)
    assert rows == []
    assert "unrecognised Project Location Category" in errors[0]


def test_flat_three_column_layout_is_still_accepted(tmp_path):
    from admin_panel.data_management_views import _parse_area_allowance_workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Zone", "Project Location Category", "Area Allowance Percentage"])
    for zone_label, location_label, percent in [
        ("Zone 1", "GHMC", "40%"),
        ("Zone 1", "Other than GHMC", "25%"),
        ("Zone 2", "Municipal Area or District HQ", "20%"),
        ("Zone 2", "Industrial Area", "20%"),
        # Deliberately the older wording -- it still resolves via aliases.
        ("Zone 3", "Rural Area", "25%"),
        ("Zone 3", "Tribal Area", "40%"),
    ]:
        ws.append([zone_label, location_label, percent])
    path = tmp_path / "flat.xlsx"
    wb.save(path)

    rows, errors = _parse_area_allowance_workbook(path)
    assert errors == []
    assert [(r["zone"], r["location"], r["percent"]) for r in rows] == [
        ("zone_1", "ghmc", 40.0),
        ("zone_1", "other_than_ghmc", 25.0),
        ("zone_2", "municipal_area", 20.0),
        ("zone_2", "industrial_area", 20.0),
        ("zone_3", "upto_16_kms", 25.0),
        ("zone_3", "beyond_16_kms", 40.0),
    ]


def test_a_sheet_in_neither_layout_is_rejected(tmp_path):
    from admin_panel.data_management_views import _parse_area_allowance_workbook

    wb = Workbook()
    wb.active.append(["something", "unrelated"])
    path = tmp_path / "junk.xlsx"
    wb.save(path)

    rows, errors = _parse_area_allowance_workbook(path)
    assert rows == []
    assert "Missing required column(s)" in errors[0]


# ---------------------------------------------------------------------------
# The allowance row's label names the selected project location
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("label,expected", [
    # Every spelling the backends actually use, with spacing/casing preserved.
    ("Add GHMC Allowance @40%", "Add Industrial Area Allowance @20%"),
    ("Add GHMC allowance @40%", "Add Industrial Area allowance @20%"),
    ("Add GHMC allowance@40%", "Add Industrial Area allowance@20%"),
    (" Add GHMC Allowance @ 40 %", " Add Industrial Area Allowance @ 20 %"),
    # Not an allowance row -- left alone.
    ("Labour Sub  total", "Labour Sub  total"),
    # A differently-worded allowance the GHMC matcher never picks up anyway.
    ("Add Muncipality area allowance (Mahbubnagar)@20%",
     "Add Muncipality area allowance (Mahbubnagar)@20%"),
])
def test_rewrite_allowance_label(label, expected):
    from core.utils_excel import rewrite_allowance_label
    assert rewrite_allowance_label(label, "Industrial Area", 20.0) == expected


def test_rewrite_allowance_label_without_a_location_keeps_the_name():
    from core.utils_excel import rewrite_allowance_label
    assert rewrite_allowance_label("Add GHMC Allowance @40%", None, 0) == \
        "Add GHMC Allowance @0%"


@pytest.mark.parametrize("area,expected_label", [
    # Zone 1 / GHMC is the one case the backend already names correctly.
    ("zone_1:ghmc", " Add GHMC allowance@40%"),
    ("zone_1:other_than_ghmc", " Add Other than GHMC allowance@25%"),
    ("zone_2:municipal_area", " Add Municipal Area or District HQ allowance@20%"),
    ("zone_2:industrial_area", " Add Industrial Area allowance@20%"),
    ("zone_3:upto_16_kms", " Add Agency or Tribal Area (Upto 16 Kms) allowance@25%"),
    ("zone_3:beyond_16_kms", " Add Agency or Tribal Area (Beyond 16 Kms) allowance@40%"),
])
def test_copied_block_allowance_row_names_the_selected_location(monkeypatch, area, expected_label):
    monkeypatch.setattr(zone_policy, "area_allowance_map", lambda: {
        ("zone_1", "ghmc"): 40.0,
        ("zone_1", "other_than_ghmc"): 25.0,
        ("zone_2", "municipal_area"): 20.0,
        ("zone_2", "industrial_area"): 20.0,
        ("zone_3", "upto_16_kms"): 25.0,
        ("zone_3", "beyond_16_kms"): 40.0,
    })
    _, wf = _make_block()
    apply_policy_to_copied_block(wf, 1, 1, 24, area, "repair")
    assert wf.cell(row=11, column=4).value == expected_label


def test_legacy_non_municipal_leaves_the_label_name_alone(monkeypatch):
    monkeypatch.setattr(zone_policy, "area_allowance_map", dict)
    _, wf = _make_block()
    apply_policy_to_copied_block(wf, 1, 1, 24, "non_municipal", "repair")
    # No location category to name, and the row is hidden anyway.
    assert wf.cell(row=11, column=4).value == " Add GHMC allowance@0%"
    assert wf.row_dimensions[11].hidden
