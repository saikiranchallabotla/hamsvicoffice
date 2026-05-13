import os
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from django.db.utils import OperationalError, ProgrammingError

from .models import BackendWorkbook


def normalize_text(text):
    """
    Normalize text by replacing special unicode characters with standard ASCII equivalents.
    Fixes encoding issues like em-dash (—) appearing as â€" in output.
    """
    if text is None:
        return None
    if not isinstance(text, str):
        return text
    # Replace various unicode dashes with standard hyphen
    text = text.replace('—', '-')  # em dash
    text = text.replace('–', '-')  # en dash
    text = text.replace('−', '-')  # minus sign
    text = text.replace('‐', '-')  # hyphen
    text = text.replace('‑', '-')  # non-breaking hyphen
    text = text.replace('‒', '-')  # figure dash
    # Replace other common problematic characters
    text = text.replace(''', "'")  # left single quote
    text = text.replace(''', "'")  # right single quote
    text = text.replace('"', '"')  # left double quote
    text = text.replace('"', '"')  # right double quote
    text = text.replace('…', '...')  # ellipsis
    text = text.replace('\u00a0', ' ')  # non-breaking space
    return text


# scan columns A..J
SCAN_COL_START = 1
SCAN_COL_END = 10


# ---------- Color detection helpers ----------
def cell_is_yellow(cell):
    fill = cell.fill
    if not fill or not fill.patternType or fill.patternType.lower() != "solid":
        return False

    rgb = getattr(fill.fgColor, "rgb", None)
    if rgb and str(rgb).upper().endswith("FFFF00"):
        return True

    if getattr(fill.fgColor, "type", None) == "theme":
        if getattr(fill.fgColor, "theme", None) in (4, 5, 6):
            return True

    if getattr(fill.fgColor, "indexed", None) == 6:
        return True

    return False


def cell_is_red_text(cell):
    font = cell.font
    if not font or not font.color:
        return False

    rgb = getattr(font.color, "rgb", None)
    if rgb and str(rgb).upper().endswith("FF0000"):
        return True

    if getattr(font.color, "type", None) == "theme":
        return True

    if getattr(font.color, "indexed", None) == 3:
        return True

    return False


def _is_yellow_and_red(cell):
    return cell_is_yellow(cell) and cell_is_red_text(cell)


# ---------- Detect items from "Master Datas" ----------
def detect_items(ws):
    """
    Detect item blocks:
    - A heading row is a row where ANY cell in A..J is yellow+red.
    - The block continues until the row before the next heading.
    """
    items = []
    max_row = ws.max_row
    r = 1

    while r <= max_row:
        heading_name = None
        for c in range(SCAN_COL_START, SCAN_COL_END + 1):
            cell = ws.cell(row=r, column=c)
            if _is_yellow_and_red(cell) and str(cell.value or "").strip():
                heading_name = str(cell.value).strip()
                break

        if heading_name:
            start_row = r

            # Find next heading
            end_row = max_row
            rr = r + 1
            while rr <= max_row:
                found_next = False
                for c in range(SCAN_COL_START, SCAN_COL_END + 1):
                    cell = ws.cell(row=rr, column=c)
                    if _is_yellow_and_red(cell) and str(cell.value or "").strip():
                        found_next = True
                        break
                if found_next:
                    end_row = rr - 1
                    break
                rr += 1

            items.append({
                "name": heading_name,
                "start_row": start_row,
                "end_row": end_row
            })

            r = end_row + 1
        else:
            r += 1

    return items


# ---------- Upload item block parsing helpers ----------
def _is_valid_item_block(ws_src, start_row, end_row):
    """Check if this block looks like a valid item block (has rate data in column J)."""
    for r in range(start_row, min(end_row + 1, start_row + 50)):
        val = ws_src.cell(row=r, column=10).value
        if val not in (None, "") and str(val).strip():
            return True
    return False


def _find_item_block_end(ws_src, start_row, max_row):
    """Find the true end of an item block by looking for the rate row in column J."""
    next_heading_row = max_row + 1
    for rr in range(start_row + 1, max_row + 1):
        for c in range(1, 11):
            cell = ws_src.cell(row=rr, column=c)
            if _is_yellow_and_red(cell) and str(cell.value or "").strip():
                next_heading_row = rr
                break
        if next_heading_row <= max_row:
            break

    potential_end = next_heading_row - 1
    last_rate_row = start_row
    for r in range(start_row, potential_end + 1):
        val = ws_src.cell(row=r, column=10).value
        if val not in (None, "") and str(val).strip():
            last_rate_row = r

    return last_rate_row, next_heading_row


def _extract_items_from_sheet(ws_src):
    """Extract all item blocks from a single sheet."""
    fetched_items = []
    item_blocks = {}

    max_row = ws_src.max_row
    r = 1

    while r <= max_row:
        heading_name = None
        for c in range(1, 11):
            cell = ws_src.cell(row=r, column=c)
            if _is_yellow_and_red(cell) and str(cell.value or "").strip():
                heading_name = str(cell.value).strip()
                break

        if heading_name:
            start_row = r
            end_row, next_heading_row = _find_item_block_end(ws_src, start_row, max_row)
            if _is_valid_item_block(ws_src, start_row, end_row):
                fetched_items.append(heading_name)
                item_blocks[heading_name] = (start_row, end_row)
            r = next_heading_row if next_heading_row <= max_row else end_row + 1
        else:
            r += 1

    return fetched_items, item_blocks


def _determine_unit_from_heading(heading_name, upload_units_map=None):
    """Determine unit from Groups sheet units_map first, then fall back to heuristic."""
    if upload_units_map and heading_name in upload_units_map:
        return upload_units_map[heading_name]

    heading_lower = heading_name.lower()

    if "light point" in heading_lower or "fan point" in heading_lower:
        return "Pts"
    light_fan_keywords = ["light", "fan", "bulb", "fixture", "downlight", "spotlight", "batten"]
    for keyword in light_fan_keywords:
        if keyword in heading_lower:
            return "Nos"
    pipe_keywords = ["pipe", "wire", "cable", "conduit", "duct", "channel", "rod", "bar", "rail", "tube"]
    for keyword in pipe_keywords:
        if keyword in heading_lower:
            return "Mtr"
    if "point" in heading_lower or "pts" in heading_lower:
        return "Pts"

    return "Nos"


# ---------- Read groups from "Groups" ----------
def read_groups(ws_groups):
    """
    Column A = Item Name, Column B = Group, Column C = Prefix, Column D = Unit (optional)
    Return tuple: (groups_dict, units_dict)
      - groups_dict: { group_name: [item1, item2, ...] }
      - units_dict: { item_name: unit }
    """
    groups = {}
    units = {}
    max_row = ws_groups.max_row
    for r in range(2, max_row + 1):
        item = ws_groups.cell(row=r, column=1).value
        group = ws_groups.cell(row=r, column=2).value
        unit = ws_groups.cell(row=r, column=4).value  # Column D for unit
        if item and group:
            item = str(item).strip()
            group = str(group).strip()
            groups.setdefault(group, []).append(item)
            # Store unit if provided
            if unit:
                units[item] = str(unit).strip()
    return groups, units


# ---------- Load workbook and sheets ----------
def load_backend(category, base_dir, backend_id=None, module_code=None, user=None):
    """
    Load backend Excel data.
    
    Returns: (items_list, groups_map, units_map, ws_data, filepath)

    Resolution priority:
    1. ModuleBackend by ID (if backend_id provided)
    2. User's preferred backend (if user provided)
    3. ModuleBackend default for module+category (if module_code provided)
    4. BackendWorkbook (legacy)
    5. Static file_map (core/data/*.xlsx)
    
    Args:
        category: Work type like 'electrical', 'civil', 'temp_electrical', etc.
        base_dir: Django BASE_DIR
        backend_id: Specific ModuleBackend ID to use (optional)
        module_code: Module code like 'estimate', 'workslip' (optional)
        user: User object (optional, for user preferences)
    """
    category_key = str(category).lower().strip()
    filepath = None
    
    # Map category to base category (remove temp_, amc_ prefixes for matching)
    base_category = category_key
    if category_key.startswith('temp_'):
        base_category = category_key.replace('temp_', '')
    elif category_key.startswith('amc_'):
        base_category = category_key.replace('amc_', '')
    
    # Helper to resolve backend file path (with DB and S3 fallback)
    def _resolve_backend_path(backend_obj):
        """Try disk path first, then S3 download, then restore from DB."""
        if not backend_obj or not backend_obj.file:
            return None
        # Try local disk path first
        try:
            fpath = backend_obj.file.path
            if os.path.exists(fpath):
                return fpath
        except (NotImplementedError, Exception):
            # NotImplementedError is raised by S3 storage for .path
            pass
        # For S3/cloud storage: download to a temp file
        try:
            import tempfile
            data = backend_obj.file.read()
            backend_obj.file.seek(0)
            if data:
                ext = os.path.splitext(backend_obj.file.name)[1] or '.xlsx'
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
                tmp.write(data)
                tmp.close()
                return tmp.name
        except Exception:
            pass
        # Try restoring from DB via get_file_bytes()
        if hasattr(backend_obj, 'get_file_bytes'):
            try:
                data = backend_obj.get_file_bytes()
                if data:
                    import tempfile
                    ext = os.path.splitext(backend_obj.file.name)[1] or '.xlsx'
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
                    tmp.write(data)
                    tmp.close()
                    return tmp.name
            except Exception:
                pass
        return None

    # 1. Try ModuleBackend by specific ID
    if backend_id:
        try:
            from subscriptions.models import ModuleBackend
            backend = ModuleBackend.objects.filter(pk=backend_id, is_active=True).first()
            filepath = _resolve_backend_path(backend)
        except (OperationalError, ProgrammingError, Exception):
            pass

    # 2. Try user's preferred backend
    if not filepath and user and user.is_authenticated and module_code:
        try:
            from accounts.models import UserBackendPreference
            backend = UserBackendPreference.get_user_backend(user, module_code, base_category)
            filepath = _resolve_backend_path(backend)
        except (OperationalError, ProgrammingError, Exception):
            pass

    # 3. Try ModuleBackend default for module + category
    module_backend_checked = False
    if not filepath and module_code:
        try:
            from subscriptions.models import ModuleBackend
            module_backend_checked = True
            backend = ModuleBackend.get_for_module(module_code, base_category)
            filepath = _resolve_backend_path(backend)
        except (OperationalError, ProgrammingError, Exception):
            pass
    
    # 4. Try BackendWorkbook (legacy) - only if module_code was NOT provided
    # When module_code is provided, we rely on ModuleBackend system only
    if not filepath and not module_code:
        try:
            backend = BackendWorkbook.objects.filter(
                category=category_key,
                is_active=True
            ).order_by('-is_default', '-uploaded_at').first()
            if backend and backend.file:
                filepath = backend.file.path
        except (OperationalError, ProgrammingError, Exception):
            pass
    
    # 5. Fall back to static file_map - only if module_code was NOT provided
    # When module_code is provided but no backend exists, we should NOT use static files
    # This allows showing "Coming Soon" for modules without configured backends
    if not filepath and not module_code:
        file_map = {
            "electrical":      os.path.join(base_dir, "core", "data", "electrical.xlsx"),
            "civil":           os.path.join(base_dir, "core", "data", "civil.xlsx"),
            "temp_electrical": os.path.join(base_dir, "core", "data", "temp_electrical.xlsx"),
            "temp_civil":      os.path.join(base_dir, "core", "data", "temp_civil.xlsx"),
            "amc_electrical":  os.path.join(base_dir, "core", "data", "amc_electrical.xlsx"),
            "amc_civil":       os.path.join(base_dir, "core", "data", "amc_civil.xlsx"),
        }
        filepath = file_map.get(category_key)

        # Restore from DB if admin uploaded a newer version (survives redeploys)
        if filepath:
            try:
                from datasets.models import LegacyBackendData
                LegacyBackendData.restore_if_needed(category_key, filepath)
            except Exception:
                pass

    if not filepath or not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found for category: {category_key}")

    # Keep formulas (needed for rates, etc.)
    wb = load_workbook(filepath, data_only=False)

    if "Master Datas" not in wb.sheetnames:
        raise ValueError("Sheet 'Master Datas' missing in backend Excel.")
    if "Groups" not in wb.sheetnames:
        raise ValueError("Sheet 'Groups' missing in backend Excel.")

    ws_data = wb["Master Datas"]
    ws_groups = wb["Groups"]

    items_list = detect_items(ws_data)
    groups_map, units_map = read_groups(ws_groups)

    # ---- Merge in user's custom backends (per-user uploaded items/groups) ----
    if user is not None and getattr(user, 'is_authenticated', False) and module_code in ('new_estimate', 'temp_works', 'amc'):
        try:
            from accounts.models import UserCustomBackend
            customs = UserCustomBackend.for_user_module(user, module_code, base_category)
            for cb in customs:
                cb_path = _resolve_backend_path(cb)
                if not cb_path or not os.path.exists(cb_path):
                    continue
                try:
                    cb_wb = load_workbook(cb_path, data_only=False)
                    # data_only workbook for cached numeric rates
                    try:
                        cb_wb_vals = load_workbook(cb_path, data_only=True)
                    except Exception:
                        cb_wb_vals = None

                    cb_group = (cb.group_name or cb.name or 'Custom').strip()
                    cb_units = dict(cb.units_override or {})
                    custom_item_names = []

                    # Scan EVERY sheet for item blocks
                    for sheet_name in cb_wb.sheetnames:
                        try:
                            cb_ws_data = cb_wb[sheet_name]
                            cb_items = detect_items(cb_ws_data)
                        except Exception:
                            continue
                        if not cb_items:
                            continue
                        cb_ws_vals = None
                        if cb_wb_vals is not None and sheet_name in cb_wb_vals.sheetnames:
                            try:
                                cb_ws_vals = cb_wb_vals[sheet_name]
                            except Exception:
                                cb_ws_vals = None

                        for it in cb_items:
                            it['_source_ws'] = cb_ws_data
                            it['_source_filepath'] = cb_path
                            it['_source_wb'] = cb_wb
                            it['_source_sheet'] = sheet_name
                            it['_is_custom'] = True
                            # cache description (row+2 col D)
                            try:
                                d = cb_ws_data.cell(row=it['start_row'] + 2, column=4).value
                                it['_cached_desc'] = str(d).strip() if d else ''
                            except Exception:
                                it['_cached_desc'] = ''
                            # cache rate
                            rate_val = ''
                            ws_for_rate = cb_ws_vals if cb_ws_vals is not None else cb_ws_data
                            try:
                                for r in range(it['end_row'], it['start_row'] - 1, -1):
                                    v = ws_for_rate.cell(row=r, column=10).value
                                    if v not in (None, ''):
                                        rate_val = v
                                        break
                            except Exception:
                                pass
                            it['_cached_rate'] = rate_val
                            items_list.append(it)
                            custom_item_names.append(it['name'])

                    # Append all custom items under user-specified group name
                    if custom_item_names:
                        existing = groups_map.setdefault(cb_group, [])
                        for n in custom_item_names:
                            if n not in existing:
                                existing.append(n)

                    # Apply user-supplied units (UI takes precedence)
                    for itname, unit in cb_units.items():
                        if unit:
                            units_map[itname] = unit
                except Exception:
                    continue
        except (OperationalError, ProgrammingError, Exception):
            pass

    return items_list, groups_map, units_map, ws_data, filepath


def get_ws_for_item(item_info, default_ws):
    """Return the worksheet that owns this item — custom item's own ws if present, else default."""
    if isinstance(item_info, dict) and item_info.get('_source_ws') is not None:
        return item_info['_source_ws']
    return default_ws


def get_filepath_for_item(item_info, default_filepath):
    """Return the file path of the workbook that owns this item."""
    if isinstance(item_info, dict) and item_info.get('_source_filepath'):
        return item_info['_source_filepath']
    return default_filepath


def get_available_backends_for_module(module_code, category):
    """
    Get list of available backends for a module and category.
    
    Returns list of dicts: [{'id': 1, 'name': 'Telangana SOR', 'is_default': True}, ...]
    """
    category_key = str(category).lower().strip()
    
    try:
        from subscriptions.models import ModuleBackend
        
        backends = ModuleBackend.objects.filter(
            module__code=module_code,
            category=category_key,
            is_active=True
        ).order_by('display_order', 'name')
        
        return [
            {'id': b.pk, 'name': b.name, 'is_default': b.is_default}
            for b in backends
        ]
    except (OperationalError, ProgrammingError, Exception):
        return []



# ---------- Extract full cell objects for a block ----------
def extract_item_block(ws_data, item_info):
    ws = get_ws_for_item(item_info, ws_data)
    rows = []
    for r in range(item_info["start_row"], item_info["end_row"] + 1):
        row_cells = [ws.cell(row=r, column=c) for c in range(SCAN_COL_START, SCAN_COL_END + 1)]
        rows.append(row_cells)
    return rows


# ---------- Copy with styles, merges, widths, heights & formula translation ----------
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from copy import copy

from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell

# ---------------------------------------------------------------------------
# Single-pass formula reference remapper for item-data block copying.
#
# Matches every plain cell reference  $?COL  $?ROW  inside a formula.
# Four reference types are handled correctly in one pass:
#
#   $COL$ROW  – both absolute
#   $COL ROW  – absolute col, relative row
#    COL$ROW  – relative col, absolute row
#    COL ROW  – both relative
#
# Lookbehind: (?<![!\w\$])
#   The `!` blocks col-letters that directly follow a sheet-ref separator.
#   The `\w` blocks letters that are part of a longer identifier/name.
#   The `\$` is the critical addition: blocks `J$131` from being matched
#   when it is the tail of `!$J$131` — the `$` of the absolute col-ref is
#   not consumed (lookbehind blocked it with `!`), so the match would
#   otherwise start at `J`, leaving an "orphaned $" that turns
#   `LEAD!$J$131` into `LEAD!$J$14` instead of leaving it unchanged.
#
# Lookahead: (?![A-Za-z\d])
#   Prevents partial matches on longer identifiers, e.g. `A1` in `A1B2`.
# ---------------------------------------------------------------------------
_CELL_REF_RE = re.compile(
    r'(?<![!\w\$])(\$?)([A-Za-z]{1,3})(\$?)(\d+)(?![A-Za-z\d])'
)


def _remap_formula(formula, src_r, src_c, dst_r, dst_c,
                   src_min_row, src_max_row, col_start, col_end):
    """
    Remap every cell reference in *formula* for moving the containing cell
    from (src_r, src_c) to (dst_r, dst_c).

    Each item block is treated as a self-contained unit. Only references
    that point inside the block's own row range [src_min_row, src_max_row]
    and column range [col_start, col_end] are shifted. Everything outside
    those bounds (other item blocks, global headers, cross-sheet refs) is
    left unchanged.

    Remapping rules applied to each $?COL$?ROW token:

      Relative component (no $):
          Always shift by the position delta (dst - src).

      Absolute component ($):
        Column ($COL): shift if col_start <= COL <= col_end.
        Row ($ROW):    shift if src_min_row <= ROW <= src_max_row.

    For tempworks blocks where only a subset of rows is physically copied,
    pass the full logical block span as block_max_row to
    copy_block_with_styles_and_formulas so that src_max_row here reflects
    the true block boundary, not the truncated copy boundary.

    Cross-sheet refs are excluded by the regex lookbehind and left intact.
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    row_delta = dst_r - src_r
    col_delta = dst_c - src_c
    if row_delta == 0 and col_delta == 0:
        return formula

    def _sub(m):
        col_dollar = m.group(1)   # '$' or ''
        col_str    = m.group(2)
        row_dollar = m.group(3)   # '$' or ''
        row_str    = m.group(4)

        col_num = column_index_from_string(col_str)
        row_num = int(row_str)

        # --- Column ---
        if col_dollar:
            # Absolute col: shift only within block's column range
            new_col = col_num + col_delta if (col_start <= col_num <= col_end) else col_num
        else:
            new_col = col_num + col_delta   # relative: always shift

        # --- Row ---
        if row_dollar:
            # Absolute row: shift only if inside this block's row span.
            # References to other item blocks, global headers, or rows before
            # this block are left unchanged — each block is remapped independently.
            new_row = row_num + row_delta if (src_min_row <= row_num <= src_max_row) else row_num
        else:
            new_row = row_num + row_delta   # relative: always shift

        new_col = max(1, min(new_col, 16384))
        new_row = max(1, min(new_row, 1048576))
        return f'{col_dollar}{get_column_letter(new_col)}{row_dollar}{new_row}'

    return _CELL_REF_RE.sub(_sub, formula)


def copy_block_with_styles_and_formulas(
    ws_src,
    ws_dst,
    src_min_row,
    src_max_row,
    col_start,
    col_end,
    dst_start_row,
    dst_start_col=1,
    external_sheets=None,
    block_max_row=None,
):
    """
    Copies a rectangular block including:
      - values / formulas (✅ translated to new position)
      - styles
      - row heights / col widths
      - merged ranges (replicated with offset)

    Safely handles merged cells (does not write into MergedCell).

    external_sheets: kept for API compatibility; cross-sheet refs are excluded
      automatically by the regex lookbehind in _remap_formula.

    block_max_row: the logical end of the item block (row before next heading).
      Defaults to src_max_row.  Pass the full block span here when only a
      subset of rows is being physically copied (e.g. tempworks rate-row
      truncation) so that absolute refs to the un-copied tail rows still get
      shifted correctly.
    """

    row_offset = dst_start_row - src_min_row
    col_offset = dst_start_col - col_start

    # Full logical block end for formula-ref bound checking.
    # This must be >= src_max_row so that absolute refs to rows between
    # src_max_row and block_max_row are recognised as intra-block.
    _formula_max_row = block_max_row if (block_max_row is not None and block_max_row >= src_max_row) else src_max_row

    # Precompile self-sheet qualifier pattern once for this call.
    # Formulas in the backend sheet may reference their own sheet by name
    # (e.g. ='Master Datas'!$B$105).  The regex lookbehind in _remap_formula
    # rightly treats those as cross-sheet refs and leaves them alone.  We must
    # strip the qualifier FIRST so the remapper sees them as plain local refs.
    _self_sheet_name = ws_src.title
    _self_sheet_re = re.compile(
        r'(?:\[\d+\])?\'?' + re.escape(_self_sheet_name) + r'\'?!'
    )

    # 1) Column widths
    for c in range(col_start, col_end + 1):
        letter = get_column_letter(c)
        dst_letter = get_column_letter(c + col_offset)
        src_dim = ws_src.column_dimensions.get(letter)
        if src_dim and src_dim.width:
            ws_dst.column_dimensions[dst_letter].width = src_dim.width

    # 2) Row heights
    for r in range(src_min_row, src_max_row + 1):
        src_dim = ws_src.row_dimensions.get(r)
        if src_dim and src_dim.height:
            ws_dst.row_dimensions[r + row_offset].height = src_dim.height

    # 3) Prepare merged ranges mapping inside the block.
    merged_map = []  # list of (min_row, min_col, max_row, max_col)
    for merged in ws_src.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if (
            min_row >= src_min_row and max_row <= src_max_row and
            min_col >= col_start and max_col <= col_end
        ):
            merged_map.append((min_row, min_col, max_row, max_col))

    def find_merged_top(r, c):
        for (mr1, mc1, mr2, mc2) in merged_map:
            if mr1 <= r <= mr2 and mc1 <= c <= mc2:
                return mr1, mc1
        return None

    # 4) Copy cell values + styles (✅ translate formulas)
    for r in range(src_min_row, src_max_row + 1):
        for c in range(col_start, col_end + 1):
            # If this cell is inside a merged range, use the top-left src cell as source
            top = find_merged_top(r, c)
            if top:
                src_r, src_c = top
                # For non-top-left cells in merged regions, skip value/style copy
                # (they'll be part of the merge later)
                if (r, c) != top:
                    continue
            else:
                src_r, src_c = r, c

            src_cell = ws_src.cell(row=src_r, column=src_c)
            dst_cell = ws_dst.cell(row=r + row_offset, column=c + col_offset)
            
            # Skip if source is a MergedCell (shouldn't happen with our logic, but safety check)
            if isinstance(src_cell, MergedCell):
                continue

            v = src_cell.value

            # Remap all cell references in formula to the new block position.
            if isinstance(v, str) and v.startswith("="):
                try:
                    # Strip self-sheet qualifiers BEFORE remapping.
                    # e.g. ='Master Datas'!$B$105  →  =$B$105
                    # Without this the lookbehind in _remap_formula treats these
                    # as cross-sheet refs and skips them; then fix_cross_sheet_refs
                    # strips the qualifier later but the row is never shifted.
                    v = _self_sheet_re.sub('', v)

                    v = _remap_formula(
                        v,
                        src_r=src_r, src_c=src_c,
                        dst_r=r + row_offset, dst_c=c + col_offset,
                        src_min_row=src_min_row, src_max_row=_formula_max_row,
                        col_start=col_start, col_end=col_end,
                    )
                except Exception:
                    pass
            elif isinstance(v, str):
                # Normalize text to fix encoding issues (em-dash -> hyphen, etc.)
                v = normalize_text(v)

            dst_cell.value = v

            # Always copy styles - don't rely on has_style which can be unreliable
            # Copy each style attribute individually for maximum compatibility
            try:
                if src_cell.font:
                    dst_cell.font = copy(src_cell.font)
            except Exception:
                pass
            
            try:
                if src_cell.fill:
                    dst_cell.fill = copy(src_cell.fill)
            except Exception:
                pass
            
            try:
                if src_cell.border:
                    dst_cell.border = copy(src_cell.border)
            except Exception:
                pass
            
            try:
                if src_cell.alignment:
                    dst_cell.alignment = copy(src_cell.alignment)
            except Exception:
                pass
            
            try:
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
            except Exception:
                pass
            
            try:
                if src_cell.protection:
                    dst_cell.protection = copy(src_cell.protection)
            except Exception:
                pass

    # 5) Now replicate merged cells inside the block (after copying values/styles)
    for (min_row, min_col, max_row, max_col) in merged_map:
        ws_dst.merge_cells(
            start_row=min_row + row_offset,
            end_row=max_row + row_offset,
            start_column=min_col + col_offset,
            end_column=max_col + col_offset,
        )


# Matches sheet-qualified refs in a formula:  [N]'Sheet Name'!A1  or  Sheet!$A$1
# Sheet name is captured from group(1) (quoted) or group(2) (unquoted).
_SHEET_REF_RE = re.compile(
    r"(?:\[\d+\])?(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_\. ]*))!\$?[A-Za-z]+\$?\d+"
)

# Same but no trailing cell ref — used when rewriting formulas (we substitute
# only the "Sheet!" portion and let the rest of the formula stay put).
_SHEET_PREFIX_RE = re.compile(
    r"(?:'((?:\[\d+\])?[^']+)'|((?:\[\d+\])?[A-Za-z_][A-Za-z0-9_\. ]*))!"
)


def _resolve_sheet_name_in_wb(captured, wb):
    """
    Map a captured sheet-name token (possibly prefixed with [N], possibly
    case-mismatched, possibly carrying extra qualifier text like
    'lead 2025-26') to a real sheet that exists in `wb`. Returns the actual
    sheet name from wb.sheetnames, or None.
    """
    if not captured:
        return None
    s = re.sub(r'^\[\d+\]', '', str(captured)).strip()
    if not s:
        return None
    lookup = {n.lower(): n for n in wb.sheetnames}
    low = s.lower()
    if low in lookup:
        return lookup[low]
    for tok in re.split(r'[\s\-_]+', low):
        if tok and tok in lookup:
            return lookup[tok]
    return None


def find_referenced_sheets(ws, src_min_row, src_max_row, col_start, col_end, exclude=None):
    """
    Scan formulas in the given block of `ws` and return the set of sheet names
    referenced from other sheets (e.g. INPUT, LEAD, 'My Sheet'). External or
    indexed refs (e.g. '[1]lead 2025-26'!) are resolved to a same-named sheet
    in ws.parent when possible.
    Self-references (to ws.title) and any name in `exclude` are filtered out.
    """
    exclude = set(exclude or ())
    exclude.add(ws.title)
    wb = ws.parent
    found = set()
    for r in range(src_min_row, src_max_row + 1):
        for c in range(col_start, col_end + 1):
            try:
                v = ws.cell(row=r, column=c).value
            except Exception:
                continue
            if not (isinstance(v, str) and v.startswith('=')):
                continue
            for m in _SHEET_REF_RE.finditer(v):
                name = m.group(1) or m.group(2)
                resolved = _resolve_sheet_name_in_wb(name, wb)
                if resolved and resolved not in exclude:
                    found.add(resolved)
    return found


def _scan_sheet_for_sheet_refs(ws):
    """Return the set of sheet names (resolved against ws.parent) referenced
    by formulas anywhere in `ws`."""
    wb = ws.parent
    found = set()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith('=')):
                continue
            for m in _SHEET_REF_RE.finditer(v):
                name = m.group(1) or m.group(2)
                resolved = _resolve_sheet_name_in_wb(name, wb)
                if resolved:
                    found.add(resolved)
    return found


def normalize_external_sheet_refs(wb):
    """
    For every formula in every sheet of `wb`, rewrite external/indexed sheet
    refs (e.g. '[1]lead 2025-26'!M31) to local refs (LEAD!M31) when the
    captured name resolves to a sheet that exists in `wb`. Leaves
    already-local refs untouched.
    """
    def _replace(m):
        captured = m.group(1) or m.group(2)
        real = _resolve_sheet_name_in_wb(captured, wb)
        if not real or real == captured:
            return m.group(0)
        if re.search(r"[\s\-!']", real):
            return "'" + real.replace("'", "''") + "'!"
        return real + "!"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith('=')):
                    continue
                new_v = _SHEET_PREFIX_RE.sub(_replace, v)
                if new_v != v:
                    cell.value = new_v


def expand_referenced_sheets_transitively(wb, initial_sheets, exclude=None):
    """
    Given an initial set of sheet names referenced by item blocks, walk the
    workbook to find every sheet they reference (and that those sheets
    reference, recursively). Returns the closure as a set of sheet names that
    actually exist in `wb`. `exclude` names are removed from the result and
    are not traversed (typical: {'Master Datas'}).
    """
    exclude = set(exclude or ())
    closure = set()
    pending = [n for n in initial_sheets if n in wb.sheetnames and n not in exclude]
    while pending:
        name = pending.pop()
        if name in closure or name in exclude:
            continue
        closure.add(name)
        try:
            for ref in _scan_sheet_for_sheet_refs(wb[name]):
                if ref in wb.sheetnames and ref not in closure and ref not in exclude:
                    pending.append(ref)
        except Exception:
            continue
    return closure


_XLSX_MAX_COL = 16384  # Excel column limit (XFD)
_XLSX_MAX_ROW = 1048576


def trim_to_xlsx_limits(wb):
    """
    Defensively clamp every sheet to its actual populated content range and to
    Excel's hard limits. Source backends often carry a default <col min=1
    max=16384> that openpyxl materializes into 16,384 ColumnDimension entries;
    when written out, viewers like Trio Office (which has a smaller column
    cap than Excel) warn about exceeding the column limit. Strip phantom
    cells / column / row dimensions so the saved sheet stays tight.
    """
    from openpyxl.utils import column_index_from_string

    for ws in wb.worksheets:
        # 1) Drop cells beyond Excel's limits and find the true max column /
        #    row that actually carries a value.
        max_data_col = 1
        max_data_row = 1
        try:
            for key in list(ws._cells.keys()):
                r, c = key
                if r > _XLSX_MAX_ROW or c > _XLSX_MAX_COL:
                    del ws._cells[key]
                    continue
                cell = ws._cells.get(key)
                if cell is not None and cell.value not in (None, ""):
                    if c > max_data_col:
                        max_data_col = c
                    if r > max_data_row:
                        max_data_row = r
        except Exception:
            pass

        # Small buffer so we don't accidentally clip a column the user is
        # about to type into / a sheet's right-most blank padding column.
        col_keep = min(max_data_col + 4, _XLSX_MAX_COL)
        row_keep = min(max_data_row + 4, _XLSX_MAX_ROW)

        # 2) Drop column_dimensions whose index exceeds the kept range, and
        #    clamp `min`/`max` of the surviving ones. A single ColumnDimension
        #    with min=1 max=16384 (openpyxl's default for backends) will
        #    otherwise write out as a giant <col> range and trigger Excel's
        #    "we found a problem with some content" repair prompt.
        try:
            for letter in list(ws.column_dimensions.keys()):
                try:
                    idx = column_index_from_string(letter)
                except Exception:
                    del ws.column_dimensions[letter]
                    continue
                if idx > col_keep:
                    del ws.column_dimensions[letter]
                    continue
                dim = ws.column_dimensions[letter]
                try:
                    if dim.max is not None and dim.max > col_keep:
                        dim.max = col_keep
                    if dim.min is not None and dim.min > col_keep:
                        del ws.column_dimensions[letter]
                except Exception:
                    pass
        except Exception:
            pass

        # 3) Drop row_dimensions beyond kept range.
        try:
            for r in list(ws.row_dimensions.keys()):
                if r > row_keep:
                    del ws.row_dimensions[r]
        except Exception:
            pass

        # 4) Drop merged_cells ranges that extend beyond kept range — leaving
        #    them can re-introduce phantom dimension references.
        try:
            for mr in list(ws.merged_cells.ranges):
                try:
                    if mr.max_col > col_keep or mr.max_row > row_keep:
                        ws.merged_cells.ranges.remove(mr)
                except Exception:
                    try:
                        ws.merged_cells.ranges.remove(mr)
                    except Exception:
                        pass
        except Exception:
            pass

        # 5) Reset auto_filter / print_area refs that may span the full sheet.
        try:
            if ws.auto_filter and ws.auto_filter.ref:
                ref = str(ws.auto_filter.ref)
                if ':' in ref:
                    end = ref.split(':')[1]
                    end_col = ''.join(ch for ch in end if ch.isalpha())
                    if end_col and column_index_from_string(end_col) > col_keep:
                        ws.auto_filter.ref = None
        except Exception:
            pass
        try:
            ws.print_area = None
        except Exception:
            pass

    # 6) Strip workbook-level defined names whose value is broken
    #    (#REF!, external [N] workbook indexes). Backend Excel files
    #    (e.g. core/data/civil.xlsx, amc_civil.xlsx) ship dozens of such
    #    dangling names; when carried into the output they make Excel pop
    #    the "we found a problem with some content" recovery prompt.
    try:
        import re as _re
        _bad_name_re = _re.compile(r'#REF!|\[\d+\]')
        try:
            _names = list(wb.defined_names)
        except Exception:
            _names = []
        for _dn in _names:
            try:
                _val = wb.defined_names[_dn].value
            except Exception:
                _val = None
            if not _val or _bad_name_re.search(str(_val)):
                try:
                    del wb.defined_names[_dn]
                except Exception:
                    pass
    except Exception:
        pass


def fix_cross_sheet_refs(ws, src_sheet_name='Master Datas'):
    """
    After copying blocks from a source sheet into a destination sheet (e.g. Output/ItemBlocks),
    fix two classes of broken formula references:
      1. 'src_sheet_name'!CellRef  — the block is now local, so strip the sheet qualifier.
      2. [N]SheetName!CellRef      — external-workbook index artifacts; strip the [N] prefix.
    """
    import re
    quoted = f"'{src_sheet_name}'!"
    unquoted = f"{src_sheet_name}!"
    for row in ws.iter_rows():
        for cell in row:
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                continue
            v = cell.value
            v = v.replace(quoted, '').replace(unquoted, '')
            v = re.sub(r'\[\d+\]', '', v)
            if v != cell.value:
                cell.value = v


def copy_sheet_to_workbook(src_wb, sheet_name, dst_wb):
    """Copy a full sheet from src_wb into dst_wb, preserving values, formulas, styles, and dimensions."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import column_index_from_string
    if sheet_name not in src_wb.sheetnames:
        return None

    ws_src = src_wb[sheet_name]
    ws_dst = dst_wb.create_sheet(sheet_name)

    # Determine actual data extent of the source sheet so we don't carry
    # phantom merged ranges / column dimensions that span the full sheet
    # (Excel flags those as content corruption when opening).
    _src_max_col = 1
    _src_max_row = 1
    try:
        for (r, c), cell in ws_src._cells.items():
            if cell is not None and cell.value not in (None, ""):
                if c > _src_max_col:
                    _src_max_col = c
                if r > _src_max_row:
                    _src_max_row = r
    except Exception:
        _src_max_col = ws_src.max_column or 1
        _src_max_row = ws_src.max_row or 1
    _col_cap = min(_src_max_col + 4, _XLSX_MAX_COL)
    _row_cap = min(_src_max_row + 4, _XLSX_MAX_ROW)

    # Column dimensions — skip phantom ranges and clamp `max`.
    for col_letter, dim in ws_src.column_dimensions.items():
        if not dim.width:
            continue
        try:
            idx = column_index_from_string(col_letter)
        except Exception:
            continue
        if idx > _col_cap:
            continue
        new_dim = ws_dst.column_dimensions[col_letter]
        new_dim.width = dim.width
        try:
            if dim.max is not None:
                new_dim.max = min(dim.max, _col_cap)
            if dim.min is not None:
                new_dim.min = min(dim.min, _col_cap)
        except Exception:
            pass

    # Row dimensions
    for row_idx, dim in ws_src.row_dimensions.items():
        if dim.height and row_idx <= _row_cap:
            ws_dst.row_dimensions[row_idx].height = dim.height

    # Merged cells (must be done before writing values) — filter to data extent.
    for merged in ws_src.merged_cells.ranges:
        try:
            if merged.max_col > _col_cap or merged.max_row > _row_cap:
                continue
        except Exception:
            continue
        ws_dst.merge_cells(str(merged))

    # Cell values and styles
    for row in ws_src.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            dst_cell = ws_dst.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            try:
                if cell.font:
                    dst_cell.font = copy(cell.font)
            except Exception:
                pass
            try:
                if cell.fill:
                    dst_cell.fill = copy(cell.fill)
            except Exception:
                pass
            try:
                if cell.border:
                    dst_cell.border = copy(cell.border)
            except Exception:
                pass
            try:
                if cell.alignment:
                    dst_cell.alignment = copy(cell.alignment)
            except Exception:
                pass
            try:
                if cell.number_format:
                    dst_cell.number_format = cell.number_format
            except Exception:
                pass

    return ws_dst



def get_item_description_and_rate(ws_data, item_info):
    """
    Extracts:
      - Description from 2 rows below header in col D
      - Rate as last non-empty cell in column J (inside block only)
    """
    if isinstance(item_info, dict) and item_info.get('_is_custom'):
        return item_info.get('_cached_desc', ''), item_info.get('_cached_rate', '')
    ws = get_ws_for_item(item_info, ws_data)
    desc = ws.cell(item_info["start_row"] + 2, 4).value
    desc = str(desc).strip() if desc else ""

    rate = ""
    # iterate from end_row down to start_row (inclusive)
    for r in range(item_info["end_row"], item_info["start_row"] - 1, -1):
        val = ws.cell(r, 10).value
        if val not in (None, ""):
            rate = val
            break

    return desc, rate
# core/utils_excel.py

import json

import re

def _norm_item_name(s: str) -> str:
    """Normalize item names so dict keys match UI values."""
    s = "" if s is None else str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

import re

def _parse_number_maybe(v):
    """
    Converts values like:
      1016
      1,016
      ₹ 1,016.00
      "1016.00"
    into float.
    If v is a formula string like "=G148*I148", returns None.
    """
    if v is None or v == "":
        return None

    # numeric already
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()

    # formula -> cannot evaluate here (use data_only=True sheet to get cached result)
    if s.startswith("="):
        return None

    # remove currency / commas / non-numeric except dot and minus
    s = s.replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)

    if s in ("", "-", ".", "-."):
        return None

    try:
        return float(s)
    except Exception:
        return None

import re
import ast
import operator as op

# ---- safe eval for very simple excel formulas like: =G148*I148 or =J10*2 ----
_ALLOWED_OPS = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.USub: op.neg,
    ast.UAdd: op.pos,
}

_CELL_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)", re.IGNORECASE)

def _to_number(v):
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except Exception:
        return 0.0

def _safe_eval_expr(expr: str) -> float:
    """
    Evaluate expression containing only numbers and + - * / ().
    """
    def _eval(node):
        if isinstance(node, ast.Num):  # py<3.8
            return float(node.n)
        if isinstance(node, ast.Constant):  # py>=3.8
            if isinstance(node.value, (int, float)):
                return float(node.value)
            return 0.0
        if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_OPS:
            return _ALLOWED_OPS[type(node.op)](_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_OPS:
            return _ALLOWED_OPS[type(node.op)](_eval(node.operand))
        raise ValueError("Unsafe/unsupported expression")
    tree = ast.parse(expr, mode="eval")
    return float(_eval(tree.body))

def _eval_excel_formula_cell(ws, formula: str, current_row: int) -> float:
    """
    Supports formulas like:
      =G148*I148
      =$G$148*$I$148
      =G148*2
      =G148+I148
    We replace cell refs with their numeric values from the SAME SHEET.
    """
    if not isinstance(formula, str):
        return 0.0

    f = formula.strip()
    if not f.startswith("="):
        return _to_number(f)

    f = f[1:]  # remove '='

    def repl(m):
        col = m.group(1).upper()
        row = int(m.group(2))
        v = ws[f"{col}{row}"].value

        # if referenced cell is itself a formula, we try ONE level of evaluation
        if isinstance(v, str) and v.strip().startswith("="):
            return str(_eval_excel_formula_cell(ws, v, row))
        return str(_to_number(v))

    expr = _CELL_REF_RE.sub(repl, f)

    # remove any leftover illegal characters
    # (keep digits, dot, operators, parentheses, spaces)
    expr = re.sub(r"[^0-9\.\+\-\*\/\(\)\s]", "", expr)

    try:
        return _safe_eval_expr(expr)
    except Exception:
        return 0.0

import re
import math
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def _safe_float(x):
    try:
        if x is None or x == "":
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).replace(",", "").strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


_cell_ref_re = re.compile(r"\b(\$?[A-Z]{1,3}\$?\d+)\b")


def _get_cell_value(ws_vals, ws_formulas, addr, _depth=0):
    """
    Return numeric value for a cell reference like 'J148'.
    Uses cached values first; if missing and formula exists, tries to evaluate recursively.
    """
    if _depth > 25:
        return 0.0  # prevent infinite loops

    col_letter, row = coordinate_from_string(addr.replace("$", ""))
    col = column_index_from_string(col_letter)
    r = int(row)

    v = ws_vals.cell(row=r, column=col).value
    num = _safe_float(v)
    if num is not None:
        return num

    # if cached missing, try formula evaluation
    f = ws_formulas.cell(row=r, column=col).value
    if isinstance(f, str) and f.startswith("="):
        return _eval_excel_formula(f, ws_vals, ws_formulas, _depth=_depth + 1)

    return 0.0


def _round_excel(x, ndigits):
    # Excel ROUND halves away from zero? Python round is bankers.
    # We'll implement "away from zero" to be closer.
    try:
        nd = int(ndigits)
    except Exception:
        nd = 0
    factor = 10 ** nd
    if factor == 0:
        return x
    if x >= 0:
        return math.floor(x * factor + 0.5) / factor
    return math.ceil(x * factor - 0.5) / factor


def _eval_excel_formula(formula, ws_vals, ws_formulas, _depth=0):
    """
    Minimal evaluator for formulas like:
      =ROUND(G148*I148,0)
      =G148*I148
      =(G148+I148)/2
    Supports: + - * / ( ) and ROUND(x,n)
    """
    if _depth > 25:
        return 0.0

    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:].strip()

    # Normalize Excel function names
    # Handle ROUND(...)
    # We'll convert ROUND(a,b) -> ROUND(a,b) and process ourselves.
    expr_up = expr.upper()

    # Replace cell refs with numeric values
    def repl(m):
        addr = m.group(1)
        val = _get_cell_value(ws_vals, ws_formulas, addr, _depth=_depth + 1)
        return str(val)

    expr2 = _cell_ref_re.sub(repl, expr_up)

    # Handle ROUND(...) manually
    # We will repeatedly evaluate innermost ROUND.
    round_re = re.compile(r"ROUND\(([^()]+?),([^()]+?)\)")
    while True:
        m = round_re.search(expr2)
        if not m:
            break
        a = m.group(1).strip()
        b = m.group(2).strip()
        try:
            aval = _safe_eval_arith(a)
            bval = _safe_eval_arith(b)
            rep = str(_round_excel(aval, bval))
        except Exception:
            rep = "0.0"
        expr2 = expr2[:m.start()] + rep + expr2[m.end():]

    # Now evaluate remaining arithmetic safely
    try:
        return float(_safe_eval_arith(expr2))
    except Exception:
        return 0.0


def _safe_eval_arith(s):
    """
    Safe arithmetic evaluator for numbers and + - * / parentheses only.
    Uses AST parsing instead of eval() for security.
    """
    s = s.strip()
    if s == "":
        return 0.0

    # Reject anything unexpected - only allow digits, decimals, operators, parentheses, spaces
    if re.search(r"[^0-9\.\+\-\*\/\(\)\s]", s):
        raise ValueError("Unsafe expression")

    # Use AST-based safe evaluation instead of eval()
    try:
        return _safe_eval_expr(s)
    except Exception:
        raise ValueError("Invalid arithmetic expression")


def build_temp_day_rates(filepath, items_list):
    """
    Backend TEMP structure (your final clarification):
      - Day numbers are in Column C
      - Rates are in Column J
      - Column J often has formulas like ROUND()
    We load:
      - ws_vals (data_only=True) for cached values
      - ws_formulas (data_only=False) for the formula text
    """
    wb_vals = load_workbook(filepath, data_only=True)
    wb_for = load_workbook(filepath, data_only=False)

    if "Master Datas" not in wb_vals.sheetnames or "Master Datas" not in wb_for.sheetnames:
        return {}

    ws_vals = wb_vals["Master Datas"]
    ws_for = wb_for["Master Datas"]

    day_rates = {}

    for it in items_list:
        name = (it.get("name") or "").strip()
        sr = int(it.get("start_row") or 0)
        er = int(it.get("end_row") or 0)
        if not name or sr <= 0 or er <= 0:
            continue

        per_item = {}

        for r in range(sr, er + 1):
            day_cell = ws_vals.cell(row=r, column=3).value  # Column C
            if day_cell in (None, ""):
                continue

            try:
                day_no = int(float(day_cell))
            except Exception:
                continue

            if day_no <= 0:
                continue

            # Try cached numeric from column J
            rate_cached = _safe_float(ws_vals.cell(row=r, column=10).value)
            if rate_cached is not None and rate_cached > 0:
                per_item[day_no] = rate_cached  # Use int key
                continue

            # Fallback: evaluate formula from ws_for J
            f = ws_for.cell(row=r, column=10).value
            if isinstance(f, str) and f.startswith("="):
                rate_calc = _eval_excel_formula(f, ws_vals, ws_for)
                if rate_calc and rate_calc > 0:
                    per_item[day_no] = float(rate_calc)  # Use int key
                    continue
            
            # Last fallback: try column I (sometimes rates are there)
            rate_col_i = _safe_float(ws_vals.cell(row=r, column=9).value)
            if rate_col_i is not None and rate_col_i > 0:
                per_item[day_no] = rate_col_i

        if per_item:
            day_rates[_norm_item_name(name)] = per_item

    return day_rates
