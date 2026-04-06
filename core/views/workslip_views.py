# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import (_apply_print_settings, _format_indian_number,
    _number_to_words_rupees, _get_current_financial_year, _get_current_date_formatted,
    _get_letter_settings, get_org_from_request, check_org_access, create_job_for_excel,
    enqueue_excel_task, _format_date_to_ddmmyyyy)

@login_required(login_url='login')
def workslip(request):
    from core.utils_excel import get_available_backends_for_module
    
    # ---- Clear session data on fresh page load (GET without preserve param) ----
    if request.method == "GET" and not request.GET.get("preserve") and not request.GET.get("group"):
        # Clear all workslip session data for a fresh start
        request.session["ws_estimate_rows"] = []
        request.session["ws_exec_map"] = {}
        request.session["ws_rate_map"] = {}  # Map of row_key -> custom rate (for user-modified rates)
        request.session["ws_tp_percent"] = 0.0
        request.session["ws_tp_type"] = "Excess"
        request.session["ws_supp_items"] = []
        request.session["ws_estimate_grand_total"] = 0.0
        request.session["ws_work_name"] = ""
        request.session["ws_current_phase"] = 1
        request.session["ws_previous_phases"] = []
        request.session["ws_previous_supp_items"] = []
        request.session["ws_previous_ae_data"] = []
        request.session["ws_metadata"] = {}  # Clear workslip metadata
        request.session["ws_deduct_old_material"] = 0.0
        request.session["ws_lc_percent"] = 0.0
        request.session["ws_qc_percent"] = 0.0
        request.session["ws_nac_percent"] = 0.0
        request.session["ws_selected_backend_id"] = None  # Clear backend selection
        request.session["ws_work_type"] = None  # Clear work type selection
        request.session["ws_work_mode"] = None  # Clear work mode selection
        request.session["ws_category"] = None  # Clear category selection
        request.session["current_saved_work_id"] = None  # Clear saved work link so new estimate doesn't update an old work
        request.session["ws_parent_work_id"] = None  # Clear parent work link
    
    # Handle work type, work mode and category from URL parameters (from workslip_home)
    url_work_type = request.GET.get("work_type")
    url_work_mode = request.GET.get("work_mode")
    url_category = request.GET.get("category")
    
    if url_work_type:
        request.session["ws_work_type"] = url_work_type
    if url_work_mode:
        request.session["ws_work_mode"] = url_work_mode
    if url_category:
        request.session["ws_category"] = url_category
    
    # Get work type, work mode and category from session
    ws_work_type = request.session.get("ws_work_type", "new_estimate") or "new_estimate"
    ws_work_mode = request.session.get("ws_work_mode", "original") or "original"
    ws_category = request.session.get("ws_category", "electrical") or "electrical"
    
    # Handle backend switching via URL parameter
    url_backend_id = request.GET.get("backend_id")
    if url_backend_id:
        try:
            bid = int(url_backend_id)
            request.session["ws_selected_backend_id"] = bid
            # Persist choice to DB so it survives re-login/redeployment
            if request.user.is_authenticated:
                try:
                    from subscriptions.models import ModuleBackend
                    from accounts.models import UserBackendPreference
                    _backend = ModuleBackend.objects.filter(pk=bid, is_active=True).first()
                    if _backend:
                        UserBackendPreference.set_user_backend(request.user, _backend)
                except Exception:
                    pass
        except (ValueError, TypeError):
            pass

    ws_selected_backend_id = request.session.get("ws_selected_backend_id")

    # Initialize from user's saved preference if session has no selection
    if ws_selected_backend_id is None and request.user.is_authenticated:
        try:
            from accounts.models import UserBackendPreference
            _base_cat = ws_category  # 'electrical' or 'civil'
            _mod_code = 'new_estimate'
            if request.session.get("ws_work_type") == 'amc':
                _mod_code = 'amc'
            elif request.session.get("ws_work_type") == 'tempworks':
                _mod_code = 'temp_works'
            _pref_backend = UserBackendPreference.get_user_backend(request.user, _mod_code, _base_cat)
            if _pref_backend:
                ws_selected_backend_id = _pref_backend.pk
                request.session["ws_selected_backend_id"] = ws_selected_backend_id
        except Exception:
            pass

    # ---- session state ----
    ws_estimate_rows = request.session.get("ws_estimate_rows", []) or []
    ws_exec_map = request.session.get("ws_exec_map", {}) or {}
    ws_rate_map = request.session.get("ws_rate_map", {}) or {}  # Custom rate changes by user
    ws_tp_percent = request.session.get("ws_tp_percent", 0.0)
    ws_tp_type = request.session.get("ws_tp_type", "Excess")
    ws_supp_items = request.session.get("ws_supp_items", []) or []
    ws_estimate_grand_total = request.session.get("ws_estimate_grand_total", 0.0)
    ws_work_name = request.session.get("ws_work_name", "") or ""
    ws_deduct_old_material = request.session.get("ws_deduct_old_material", 0.0)
    
    # Log session values for debugging
    logger.info(f"[WORKSLIP] Session: work_type={ws_work_type}, category={ws_category}, tp_percent={ws_tp_percent}, tp_type={ws_tp_type}, deduct={ws_deduct_old_material}")
    
    # Multi-phase workslip tracking
    ws_current_phase = request.session.get("ws_current_phase", 1)  # Current phase number (1, 2, 3, etc.)
    ws_previous_phases = request.session.get("ws_previous_phases", [])  # List of previous phase exec_maps
    ws_previous_supp_items = request.session.get("ws_previous_supp_items", [])  # Supplemental items from previous phases

    # Determine module code and categories based on work type
    # - backend_db_category: category stored in ModuleBackend table (electrical/civil)
    # - load_category: category prefix used by load_backend for file paths (amc_electrical, temp_electrical, etc.)
    if ws_work_type == 'amc':
        module_code = 'amc'
        backend_db_category = ws_category  # stored as 'electrical' or 'civil' in DB
        load_category = f'amc_{ws_category}'  # used for file loading
    elif ws_work_type == 'tempworks':
        module_code = 'temp_works'
        backend_db_category = ws_category  # stored as 'electrical' or 'civil' in DB
        load_category = f'temp_{ws_category}'  # used for file loading
    else:  # new_estimate (default)
        module_code = 'new_estimate'
        backend_db_category = ws_category
        load_category = ws_category
    
    category = load_category  # Use load_category for loading backend files

    # Get available backends for dropdown (use DB category format)
    available_backends = get_available_backends_for_module(module_code, backend_db_category)

    # -------- 0) Load backend (groups/items + Master Datas) ----------
    desc_to_item = {}
    item_name_to_desc = {}  # Maps yellow header name → row+2 description from Master Datas
    try:
        items_list, groups_map, units_map, ws_data, filepath = load_backend(
            category, settings.BASE_DIR,
            backend_id=ws_selected_backend_id,
            module_code=module_code,
            user=request.user
        )
        if ws_data is not None:
            for info in items_list:
                item_name = info["name"]
                start_row = info["start_row"]
                desc_cell = ws_data.cell(row=start_row + 2, column=4).value
                desc_text = str(desc_cell or "").strip()
                if desc_text:
                    desc_to_item.setdefault(desc_text, item_name)
                    item_name_to_desc[item_name] = desc_text
    except Exception:
        items_list, groups_map, ws_data, filepath = [], {}, None, ""

    groups = sorted(groups_map.keys(), key=lambda s: s.lower()) if groups_map else []
    current_group = request.GET.get("group") or (groups[0] if groups else "")

    group_items = groups_map.get(current_group, []) if current_group else []
    detected_names = {i["name"] for i in items_list}
    items_in_group = [name for name in group_items if name in detected_names]

    # Build item subtypes map: items with ":" are subtypes
    # Group subtypes by their parent name (part before ":")
    import re as _re
    _colon_re = _re.compile(r'\s*:\s*')

    def _has_colon(name):
        return bool(_colon_re.search(name))

    def _split_parent(name):
        return _colon_re.split(name, 1)[0].strip()

    item_subtypes = {}
    parent_items_set = set()

    for name in items_in_group:
        if _has_colon(name):
            parent_name = _split_parent(name)
            if parent_name not in item_subtypes:
                item_subtypes[parent_name] = []
            item_subtypes[parent_name].append(name)
            parent_items_set.add(parent_name)

    items_info = []
    seen_parents = set()
    for name in items_in_group:
        if _has_colon(name):
            parent_name = _split_parent(name)
            if parent_name not in seen_parents:
                subtypes_list = item_subtypes.get(parent_name, [])
                items_info.append({
                    "name": parent_name,
                    "has_subtypes": True,
                    "subtypes": json.dumps(subtypes_list),
                    "subtypes_count": len(subtypes_list),
                })
                seen_parents.add(parent_name)
        else:
            items_info.append({
                "name": name,
                "has_subtypes": False,
                "subtypes": "[]",
                "subtypes_count": 0,
            })

    # helper: group â†’ units
    item_to_group = {}
    for grp_name, item_list_grp in groups_map.items():
        for nm in item_list_grp:
            item_to_group.setdefault(nm, grp_name)

    def units_for(name):
        # Priority 1: Use unit from Groups sheet (Column D) via units_map
        backend_unit = units_map.get(name, "") if units_map else ""
        if backend_unit:
            bu = backend_unit.lower()
            if bu in ("mtrs", "mtr", "metre", "meters"):
                return ("Mtrs", "Mtr")
            elif bu in ("pts", "pt", "point", "points"):
                return ("Pts", "Pt")
            elif bu in ("nos", "no"):
                return ("Nos", "No")
            else:
                return (backend_unit, backend_unit)
        # Fallback: group-based
        grp = item_to_group.get(name, "")
        if grp in ("Piping", "Wiring & Cables"):
            return ("Mtrs", "Mtr")
        elif grp == "Points":
            return ("Pts", "Pt")
        else:
            return ("Nos", "No")

    # ------------- 1) POST actions -------------
    if request.method == "POST":
        action = request.POST.get("action")

        # A) Upload Estimate file and parse items
        if action == "upload_estimate":
            file = request.FILES.get("estimate_file")
            if not file:
                return render(request, "core/workslip.html", {
                    "error": "Please upload an Estimate Excel file.",
                    "category": category,
                    "groups": groups,
                    "current_group": current_group,
                    "items_in_group": items_in_group, "items_info": items_info,
                    "ws_estimate_rows": ws_estimate_rows,
                    "preview_rows": [],
                    "tp_percent": ws_tp_percent if ws_tp_percent else "",
                    "tp_type": ws_tp_type,
                    "supp_items_selected": ws_supp_items,
                    "work_name": ws_work_name,
                })

            try:
                # read once, create two workbooks: formulas + values
                excel_bytes = file.read()
                wb_est = load_workbook(BytesIO(excel_bytes), data_only=False)
                wb_est_vals = load_workbook(BytesIO(excel_bytes), data_only=True)
            except Exception as e:
                return render(request, "core/workslip.html", {
                    "error": f"Couldn't read uploaded Estimate file: {e}",
                    "category": category,
                    "groups": groups,
                    "current_group": current_group,
                    "items_in_group": items_in_group, "items_info": items_info,
                    "ws_estimate_rows": ws_estimate_rows,
                    "preview_rows": [],
                    "tp_percent": ws_tp_percent if ws_tp_percent else "",
                    "tp_type": ws_tp_type,
                    "supp_items_selected": ws_supp_items,
                    "work_name": ws_work_name,
                })

            # ---- detect our Estimate-format sheet (ignore sheet name) ----
            def looks_like_our_estimate_sheet(sh):
                """
                Heuristic: row 3 headers should look like our Estimate:
                A: Sl.No, B: Quantity (Unit), D: Item Description, E: Rate, H: Amount
                """
                row = 3
                a = str(sh.cell(row=row, column=1).value or "").strip().lower()
                b = str(sh.cell(row=row, column=2).value or "").strip().lower()
                d = str(sh.cell(row=row, column=4).value or "").strip().lower()
                e = str(sh.cell(row=row, column=5).value or "").strip().lower()
                h = str(sh.cell(row=row, column=8).value or "").strip().lower()

                score = 0
                if "sl" in a and "no" in a:
                    score += 1
                if "quantity" in b:
                    score += 1
                if "item" in d and "description" in d:
                    score += 1
                if "rate" in e:
                    score += 1
                if "amount" in h:
                    score += 1
                return score >= 3  # tolerant

            estimate_sheets = [sh for sh in wb_est.worksheets if looks_like_our_estimate_sheet(sh)]

            if not estimate_sheets:
                return render(request, "core/workslip.html", {
                    "error": "No sheet in the uploaded workbook matches the Estimate format.",
                    "category": category,
                    "groups": groups,
                    "current_group": current_group,
                    "items_in_group": items_in_group, "items_info": items_info,
                    "ws_estimate_rows": ws_estimate_rows,
                    "preview_rows": [],
                    "tp_percent": ws_tp_percent if ws_tp_percent else "",
                    "tp_type": ws_tp_type,
                    "supp_items_selected": ws_supp_items,
                    "work_name": ws_work_name,
                })

            # âœ… Only use the FIRST matching estimate sheet (no multi-sheet aggregation)
            ws_est_sheet = estimate_sheets[0]
            ws_est_vals_sheet = wb_est_vals[ws_est_sheet.title]

            # ---------- Find a sheet with yellow+red item headers (Item Blocks) ---------- #
            def get_heading_name(sheet, row_index: int):
                """
                Returns the heading text in this row (yellow fill + red font in A..J),
                or None if not a heading row.
                """
                for col in range(1, 10 + 1):  # A..J
                    cell = sheet.cell(row=row_index, column=col)
                    fill = getattr(cell, "fill", None)
                    font = getattr(cell, "font", None)

                    pattern = getattr(fill, "patternType", None)
                    fg = getattr(getattr(fill, "fgColor", None), "rgb", "")
                    color = getattr(getattr(font, "color", None), "rgb", "")

                    is_yellow = (
                        fill
                        and pattern
                        and str(pattern).lower() == "solid"
                        and fg
                        and str(fg).upper().endswith("FFFF00")
                    )
                    is_red = (
                        font
                        and getattr(font, "color", None)
                        and color
                        and str(color).upper().endswith("FF0000")
                    )

                    if is_yellow and is_red and str(cell.value or "").strip():
                        return str(cell.value).strip()
                return None

            # Try to find any sheet (except the estimate sheet) that contains such yellow headers
            blocks_sheet = None
            for sh in wb_est.worksheets:
                if sh.title == ws_est_sheet.title:
                    continue
                found = False
                for r in range(1, min(sh.max_row, 200) + 1):
                    if get_heading_name(sh, r):
                        found = True
                        break
                if found:
                    blocks_sheet = sh
                    break

            # Ordered list of item NAMES from the blocks sheet (yellow headers)
            heading_names = []
            if blocks_sheet is not None:
                for r in range(1, blocks_sheet.max_row + 1):
                    nm = get_heading_name(blocks_sheet, r)
                    if nm:
                        heading_names.append(nm)
            # If not found, we'll just fall back in parsing

            # prepare backend rate lookup (Master Datas  -  correct numeric rate)
            item_to_info = {it["name"]: it for it in items_list}
            wb_backend_vals = None
            ws_backend_vals = None
            if filepath and os.path.exists(filepath):
                try:
                    wb_backend_vals = load_workbook(filepath, data_only=True)
                    ws_backend_vals = wb_backend_vals["Master Datas"]
                except Exception:
                    wb_backend_vals = None
                    ws_backend_vals = None

            def backend_rate_for_item(name):
                info = item_to_info.get(name)
                if not info or ws_backend_vals is None:
                    return 0.0
                start_row = info["start_row"]
                end_row = info["end_row"]
                for r in range(end_row, start_row - 1, -1):
                    v = ws_backend_vals.cell(row=r, column=10).value  # col J
                    if v not in (None, ""):
                        try:
                            return float(v)
                        except Exception:
                            return 0.0
                return 0.0

            def to_number(v):
                try:
                    return float(v)
                except Exception:
                    return 0.0

            # ---- Dynamic column detection for estimate sheet ----
            # Detect column positions from header row (row 3)
            header_row = 3
            col_desc = 4      # D - Description (default)
            col_qty = 2       # B - Quantity (default)
            col_unit = 3      # C - Unit (default)
            col_rate = 5      # E - Rate (default)
            col_amount = 8    # H - Amount (default)
            
            # Scan header row to find actual column positions
            for c in range(1, 15):  # Scan columns A to O
                header_val = str(ws_est_sheet.cell(row=header_row, column=c).value or "").strip().lower()
                if not header_val:
                    continue
                    
                if "description" in header_val or "item" in header_val:
                    col_desc = c
                elif "quantity" in header_val or "qty" in header_val:
                    # Check if it's execution qty or estimate qty
                    if "exec" not in header_val and "execution" not in header_val:
                        col_qty = c
                elif "unit" in header_val and "per" not in header_val:
                    col_unit = c
                elif "rate" in header_val:
                    # Check if it's execution rate or estimate rate - prefer estimate
                    if "exec" not in header_val and "execution" not in header_val:
                        col_rate = c
                elif "amount" in header_val or "amt" in header_val:
                    if "exec" not in header_val and "execution" not in header_val:
                        col_amount = c
            
            logger.info(f"[WORKSLIP DEBUG] Column detection: desc={col_desc}, qty={col_qty}, unit={col_unit}, rate={col_rate}, amount={col_amount}")

            # ---- parse this Estimate sheet ----
            parsed_rows = []
            max_row = ws_est_sheet.max_row
            r = 4
            grand_total_val = 0.0
            
            # DEBUG: Log sheet info
            logger.info(f"[WORKSLIP DEBUG] Parsing sheet: {ws_est_sheet.title}, max_row={max_row}")

            # "Name of the work" from row 2
            work_name_local = ""
            name_cell = ws_est_sheet.cell(row=2, column=1).value
            if name_cell:
                text = str(name_cell)
                parts = text.split(":", 1)
                if len(parts) > 1:
                    work_name_local = parts[1].strip()
                else:
                    work_name_local = text.strip()

            heading_idx = 0  # to walk through heading_names in order

            while r <= max_row:
                desc = ws_est_sheet.cell(row=r, column=col_desc).value  # Dynamic column
                desc_str = str(desc or "").strip()
                desc_upper = desc_str.upper()

                # Rate may be formula; get value from data_only sheet
                rate_formula = ws_est_sheet.cell(row=r, column=col_rate).value   # Dynamic column (formula or value)
                rate_value = ws_est_vals_sheet.cell(row=r, column=col_rate).value  # Dynamic column (cached value)
                rate_is_empty = (rate_formula is None or str(rate_formula).strip() == "")

                # Quantity may also be formula â†’ use data_only workbook first
                qty_formula = ws_est_sheet.cell(row=r, column=col_qty).value   # Dynamic column
                qty_value = ws_est_vals_sheet.cell(row=r, column=col_qty).value  # Dynamic column value
                qty_is_empty = (qty_formula is None or str(qty_formula).strip() == "")
                
                # DEBUG: Log each row's data
                if r <= 10:  # Only first 10 rows to avoid flooding logs
                    logger.info(f"[WORKSLIP DEBUG] Row {r}: desc='{desc_str[:30] if desc_str else 'None'}', "
                               f"rate_formula={rate_formula}, rate_value={rate_value}, "
                               f"qty_formula={qty_formula}, qty_value={qty_value}")

                # If we see any totals keywords â†’ end of items
                if desc_str:
                    if any(
                        kw in desc_upper
                        for kw in ("SUB TOTAL", "SUBTOTAL", "TOTAL", "ECV")
                    ):
                        break

                # completely blank line
                if desc is None and ws_est_sheet.cell(row=r, column=1).value is None:
                    r += 1
                    continue

                # If row has NO rate AND NO quantity, it's a heading/section label â†’ skip it
                if rate_is_empty and qty_is_empty:
                    r += 1
                    continue

                # Process only rows that have BOTH description AND (rate or quantity)
                if desc_str != "" and not (rate_is_empty and qty_is_empty):
                    # Quantity conversion
                    if isinstance(qty_value, (int, float)):
                        qty_num = float(qty_value)
                    else:
                        qty_num = to_number(qty_formula)

                    unit = ws_est_sheet.cell(row=r, column=col_unit).value  # Dynamic column

                    # backend item name from desc (for rate lookup, etc.)
                    backend_item_name = desc_to_item.get(desc_str, desc_str)

                    # display name from yellow header list (for UI only)
                    if heading_idx < len(heading_names):
                        display_name = heading_names[heading_idx]
                    else:
                        display_name = backend_item_name or desc_str
                    heading_idx += 1

                    # rate: prefer cached numeric from data_only workbook,
                    # fallback to direct numeric or backend
                    if isinstance(rate_value, (int, float)) and rate_value != 0:
                        rate_num = float(rate_value)
                    else:
                        rate_num = to_number(rate_formula)
                        # Fallback: try to get rate from backend using item name or display name
                        if rate_num == 0.0:
                            if backend_item_name in item_to_info:
                                rate_num = backend_rate_for_item(backend_item_name)
                            elif display_name in item_to_info:
                                rate_num = backend_rate_for_item(display_name)
                            # Also try matching by description
                            elif desc_str in item_to_info:
                                rate_num = backend_rate_for_item(desc_str)
                    
                    # DEBUG: Log rate source
                    logger.info(f"[WORKSLIP DEBUG] Row {r} rate: formula={rate_formula}, value={rate_value}, final={rate_num}")

                    # item_desc: full specification description for Excel output.
                    # Collect all candidate descriptions and pick the longest (most detailed) one.
                    desc_candidates = [desc_str]  # always include estimate description

                    backend_desc = (
                        item_name_to_desc.get(display_name, '')
                        or item_name_to_desc.get(backend_item_name, '')
                    )
                    if backend_desc:
                        desc_candidates.append(backend_desc)

                    # Peek at the next 1-2 rows for a description-only row (no rate/qty)
                    # that carries the full specification text (the "row+2 of item name" pattern)
                    for peek_offset in (1, 2):
                        peek_row = r + peek_offset
                        if peek_row > max_row:
                            break
                        peek_desc = ws_est_sheet.cell(row=peek_row, column=col_desc).value
                        peek_desc_str = str(peek_desc or "").strip()
                        peek_rate = ws_est_sheet.cell(row=peek_row, column=col_rate).value
                        peek_qty = ws_est_sheet.cell(row=peek_row, column=col_qty).value
                        peek_rate_empty = (peek_rate is None or str(peek_rate).strip() == "")
                        peek_qty_empty = (peek_qty is None or str(peek_qty).strip() == "")
                        if peek_desc_str and peek_rate_empty and peek_qty_empty:
                            desc_candidates.append(peek_desc_str)
                            break

                    # Pick the longest candidate as the most detailed description
                    item_desc = max(desc_candidates, key=len) if desc_candidates else desc_str

                    parsed_rows.append({
                        "key": f"{ws_est_sheet.title}_row{r}",
                        "excel_row": r,
                        "item_name": backend_item_name,      # backend / mapping name
                        "display_name": display_name,        # yellow header for UI
                        "desc": desc_str,                    # full description from Estimate
                        "item_desc": item_desc,              # row+2 content from Master Datas for Excel output
                        "qty_est": qty_num,
                        "unit": str(unit or "").strip(),
                        "rate": rate_num,
                    })
                    # DEBUG: Log each parsed row
                    logger.info(f"[WORKSLIP DEBUG] Parsed row {r}: desc={desc_str[:50]}, qty={qty_num}, rate={rate_num}")

                r += 1

            # ---- find GRAND TOTAL *below* items block if present ----
            for rr in range(r, max_row + 1):
                d2 = str(ws_est_sheet.cell(row=rr, column=col_desc).value or "").strip().upper()
                if "GRAND TOTAL" in d2:
                    grand_total_val = to_number(ws_est_vals_sheet.cell(row=rr, column=col_amount).value)
                    break

            # store in session
            ws_estimate_rows = parsed_rows
            ws_exec_map = {}
            ws_supp_items = []
            ws_work_name = work_name_local
            
            # DEBUG: Log total parsed rows before saving to session
            logger.info(f"[WORKSLIP DEBUG] Total parsed rows: {len(parsed_rows)}, grand_total={grand_total_val}")

            # Parse additional metadata from rows 2-7 of the Estimate sheet
            base_metadata = {
                "work_name": work_name_local,
                "estimate_amount": "",
                "admin_sanction": "",
                "tech_sanction": "",
                "agreement": "",
                "agency_name": "",
                "grand_total": grand_total_val,
            }
            for meta_row in range(2, 8):
                cell_val = str(ws_est_sheet.cell(row=meta_row, column=1).value or "").strip()
                cell_lower = cell_val.lower()
                
                if ":" in cell_val:
                    parts = cell_val.split(":", 1)
                    extracted_value = parts[1].strip() if len(parts) > 1 else ""
                    
                    # If no value after colon, check other columns
                    if not extracted_value:
                        for c in range(2, 11):
                            val = ws_est_sheet.cell(row=meta_row, column=c).value
                            if val and str(val).strip():
                                extracted_value = str(val).strip()
                                break
                    
                    if "estimate amount" in cell_lower:
                        base_metadata["estimate_amount"] = extracted_value
                    elif "administrative" in cell_lower or "admin" in cell_lower:
                        base_metadata["admin_sanction"] = extracted_value
                    elif "technical" in cell_lower or "tech" in cell_lower:
                        base_metadata["tech_sanction"] = extracted_value
                    elif "agreement" in cell_lower:
                        base_metadata["agreement"] = extracted_value
                    elif "agency" in cell_lower:
                        base_metadata["agency_name"] = extracted_value

            request.session["ws_estimate_rows"] = ws_estimate_rows
            request.session["ws_exec_map"] = ws_exec_map
            request.session["ws_supp_items"] = ws_supp_items
            request.session["ws_estimate_grand_total"] = grand_total_val
            request.session["ws_work_name"] = ws_work_name
            request.session["ws_metadata"] = base_metadata  # Store metadata from base estimate
            # Reset phase tracking for new estimate upload
            request.session["ws_current_phase"] = 1
            request.session["ws_previous_phases"] = []
            
            # DEBUG: Verify session save
            logger.info(f"[WORKSLIP DEBUG] Session saved. ws_estimate_rows length: {len(request.session.get('ws_estimate_rows', []))}")

            return redirect(reverse('workslip_main') + '?preserve=1')

        # B) (optional) update preview
        elif action == "update_preview":
            exec_str = request.POST.get("exec_map", "")
            rate_str = request.POST.get("rate_map", "")
            tp_percent_str = request.POST.get("tp_percent", "")
            tp_type = request.POST.get("tp_type", "Excess")

            new_exec_map = {}
            if exec_str:
                try:
                    raw = json.loads(exec_str)
                    if isinstance(raw, dict):
                        for k, v in raw.items():
                            try:
                                new_exec_map[str(k)] = float(v)
                            except Exception:
                                continue
                except Exception:
                    pass

            # Parse rate_map for custom rate changes
            new_rate_map = {}
            if rate_str:
                try:
                    raw = json.loads(rate_str)
                    if isinstance(raw, dict):
                        for k, v in raw.items():
                            try:
                                new_rate_map[str(k)] = float(v)
                            except Exception:
                                continue
                except Exception:
                    pass

            ws_exec_map.update(new_exec_map)
            ws_rate_map.update(new_rate_map)
            try:
                ws_tp_percent = float(tp_percent_str) if tp_percent_str != "" else 0.0
            except Exception:
                ws_tp_percent = 0.0
            ws_tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

            request.session["ws_exec_map"] = ws_exec_map
            request.session["ws_rate_map"] = ws_rate_map
            request.session["ws_tp_percent"] = ws_tp_percent
            request.session["ws_tp_type"] = ws_tp_type

            return redirect(reverse('workslip_main') + '?preserve=1')

        # C) Add Supplemental items
        elif action == "add_supplemental":
            selected = request.POST.getlist("supp_items")
            new_list = ws_supp_items[:]
            for nm in selected:
                if nm not in new_list:
                    new_list.append(nm)
            ws_supp_items = new_list

            # merge exec_map, rate_map & TP coming from hidden fields
            exec_str = request.POST.get("exec_map", "")
            rate_str = request.POST.get("rate_map", "")
            tp_percent_str = request.POST.get("tp_percent", "")
            tp_type = request.POST.get("tp_type", "Excess")

            ws_exec_map_session = request.session.get("ws_exec_map", {}) or {}
            ws_rate_map_session = request.session.get("ws_rate_map", {}) or {}
            new_exec_map = {}
            if exec_str:
                try:
                    raw = json.loads(exec_str)
                    if isinstance(raw, dict):
                        for k, v in raw.items():
                            try:
                                new_exec_map[str(k)] = float(v)
                            except Exception:
                                continue
                except Exception:
                    pass

            # Parse rate_map for custom rate changes
            new_rate_map = {}
            if rate_str:
                try:
                    raw = json.loads(rate_str)
                    if isinstance(raw, dict):
                        for k, v in raw.items():
                            try:
                                new_rate_map[str(k)] = float(v)
                            except Exception:
                                continue
                except Exception:
                    pass

            ws_exec_map = ws_exec_map_session.copy()
            ws_exec_map.update(new_exec_map)
            
            ws_rate_map = ws_rate_map_session.copy()
            ws_rate_map.update(new_rate_map)

            try:
                ws_tp_percent = float(tp_percent_str) if tp_percent_str != "" else 0.0
            except Exception:
                ws_tp_percent = 0.0
            ws_tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

            # Filter exec_map: keep all base items, and only supplemental items that are still selected
            filtered_exec_map = {}
            for k, v in ws_exec_map.items():
                if not k.startswith("supp:") or k[5:] in ws_supp_items:
                    filtered_exec_map[k] = v

            request.session["ws_supp_items"] = ws_supp_items
            request.session["ws_exec_map"] = filtered_exec_map
            request.session["ws_rate_map"] = ws_rate_map
            request.session["ws_tp_percent"] = ws_tp_percent
            request.session["ws_tp_type"] = ws_tp_type

            return redirect(reverse('workslip_main') + '?preserve=1')

        # D-COMBINED) Upload Both Estimate + Previous Workslips Together (Dynamic Multi-Workslip)
        elif action == "upload_combined":
            estimate_file = request.FILES.get("estimate_file")
            
            # Get target workslip from form
            target_workslip_str = request.POST.get("target_workslip", "1")
            try:
                target_workslip = int(target_workslip_str)
                if target_workslip < 1:
                    target_workslip = 1
                if target_workslip > 10:
                    target_workslip = 10
            except (ValueError, TypeError):
                target_workslip = 1
            
            logger.info(f"[MULTI-WORKSLIP] Target workslip: {target_workslip}")
            
            # Collect all previous workslip files based on target
            workslip_files = []
            for i in range(1, target_workslip):
                ws_file = request.FILES.get(f"workslip_file_{i}")
                if ws_file:
                    workslip_files.append((i, ws_file))
                    logger.info(f"[MULTI-WORKSLIP] Found workslip file {i}: {ws_file.name}")
            
            # Legacy support: also check for old "previous_workslip_file" field
            legacy_ws_file = request.FILES.get("previous_workslip_file")
            if legacy_ws_file and not workslip_files:
                workslip_files.append((1, legacy_ws_file))
                logger.info(f"[MULTI-WORKSLIP] Using legacy workslip file: {legacy_ws_file.name}")
            
            # Helper to get heading names from yellow+red cells
            def get_heading_name_from_sheet(sheet, row_index: int):
                for col in range(1, 10 + 1):
                    cell = sheet.cell(row=row_index, column=col)
                    fill = getattr(cell, "fill", None)
                    font = getattr(cell, "font", None)
                    pattern = getattr(fill, "patternType", None)
                    fg = getattr(getattr(fill, "fgColor", None), "rgb", "")
                    color = getattr(getattr(font, "color", None), "rgb", "")
                    is_yellow = (fill and pattern and str(pattern).lower() == "solid" and fg and str(fg).upper().endswith("FFFF00"))
                    is_red = (font and getattr(font, "color", None) and color and str(color).upper().endswith("FF0000"))
                    if is_yellow and is_red and str(cell.value or "").strip():
                        return str(cell.value).strip()
                return None
            
            # Validation based on target workslip
            if target_workslip == 1:
                # For Workslip-1, only estimate is required
                if not estimate_file:
                    return render(request, "core/workslip.html", {
                        "error": "Please upload an Estimate file for Workslip-1.",
                        "category": category, "groups": groups, "current_group": current_group,
                        "items_in_group": items_in_group, "items_info": items_info, "ws_estimate_rows": ws_estimate_rows,
                        "preview_rows": [], "tp_percent": ws_tp_percent if ws_tp_percent else "",
                        "tp_type": ws_tp_type, "supp_items_selected": ws_supp_items,
                        "work_name": ws_work_name, "current_phase": ws_current_phase,
                        "previous_phases": ws_previous_phases, "target_workslip": target_workslip,
                    })
            else:
                # For Workslip-2 and above: Estimate is OPTIONAL if previous workslips are provided
                # Item descriptions can be extracted from previous workslip files
                if not estimate_file and not workslip_files:
                    return render(request, "core/workslip.html", {
                        "error": f"Please upload either an Estimate file OR previous Workslip files for Workslip-{target_workslip}.",
                        "category": category, "groups": groups, "current_group": current_group,
                        "items_in_group": items_in_group, "items_info": items_info, "ws_estimate_rows": ws_estimate_rows,
                        "preview_rows": [], "tp_percent": ws_tp_percent if ws_tp_percent else "",
                        "tp_type": ws_tp_type, "supp_items_selected": ws_supp_items,
                        "work_name": ws_work_name, "current_phase": ws_current_phase,
                        "previous_phases": ws_previous_phases, "target_workslip": target_workslip,
                    })
                
                # Check if at least the most recent previous workslip is uploaded
                # (e.g., for Workslip-3, at least Workslip-2 should be uploaded)
                if workslip_files:
                    uploaded_nums = {num for num, _ in workslip_files}
                    most_recent_required = target_workslip - 1
                    if most_recent_required not in uploaded_nums:
                        return render(request, "core/workslip.html", {
                            "error": f"Please upload at least Workslip-{most_recent_required} to generate Workslip-{target_workslip}.",
                            "category": category, "groups": groups, "current_group": current_group,
                            "items_in_group": items_in_group, "items_info": items_info, "ws_estimate_rows": ws_estimate_rows,
                            "preview_rows": [], "tp_percent": ws_tp_percent if ws_tp_percent else "",
                            "tp_type": ws_tp_type, "supp_items_selected": ws_supp_items,
                            "work_name": ws_work_name, "current_phase": ws_current_phase,
                            "previous_phases": ws_previous_phases, "target_workslip": target_workslip,
                        })
            
            # Parse Estimate file for display names (yellow+red rows from Items Blocks sheet)
            estimate_display_names = []
            est_bytes = None  # Store bytes for reuse later
            if estimate_file:
                try:
                    est_bytes = estimate_file.read()
                    wb_est = load_workbook(BytesIO(est_bytes), data_only=False)
                    
                    # Find Items Blocks sheet (sheet with yellow+red headers)
                    blocks_sheet = None
                    for sh in wb_est.worksheets:
                        found = False
                        for r in range(1, min(sh.max_row, 200) + 1):
                            if get_heading_name_from_sheet(sh, r):
                                found = True
                                break
                        if found:
                            blocks_sheet = sh
                            break
                    
                    if blocks_sheet:
                        for r in range(1, blocks_sheet.max_row + 1):
                            nm = get_heading_name_from_sheet(blocks_sheet, r)
                            if nm:
                                estimate_display_names.append(nm)
                        logger.info(f"[COMBINED UPLOAD] Found {len(estimate_display_names)} item names from estimate Items Blocks sheet")
                except Exception as e:
                    logger.warning(f"[COMBINED UPLOAD] Could not parse estimate for display names: {e}")
            
            # If no estimate file, extract item names from the most recent workslip file
            # This allows generating Workslip-3 using only Workslip-2 (without original estimate)
            if not estimate_file and workslip_files:
                # Use the most recent workslip file for item names
                most_recent_ws = max(workslip_files, key=lambda x: x[0])
                ws_file_num, workslip_file_for_names = most_recent_ws
                logger.info(f"[COMBINED UPLOAD] No estimate file - extracting item names from Workslip-{ws_file_num}")
                
                try:
                    workslip_file_for_names.seek(0)  # Reset file pointer
                    ws_bytes_for_names = workslip_file_for_names.read()
                    workslip_file_for_names.seek(0)  # Reset again for later use
                    wb_ws_for_names = load_workbook(BytesIO(ws_bytes_for_names), data_only=False)
                    
                    # Find Items Blocks sheet in workslip (same structure as estimate)
                    ws_blocks_sheet = None
                    for sh in wb_ws_for_names.worksheets:
                        found = False
                        for r in range(1, min(sh.max_row or 0, 200) + 1):
                            if get_heading_name_from_sheet(sh, r):
                                found = True
                                break
                        if found:
                            ws_blocks_sheet = sh
                            break
                    
                    if ws_blocks_sheet:
                        for r in range(1, (ws_blocks_sheet.max_row or 0) + 1):
                            nm = get_heading_name_from_sheet(ws_blocks_sheet, r)
                            if nm and nm not in estimate_display_names:
                                estimate_display_names.append(nm)
                        logger.info(f"[COMBINED UPLOAD] Extracted {len(estimate_display_names)} item names from Workslip-{ws_file_num}")
                    else:
                        logger.warning(f"[COMBINED UPLOAD] No Items Blocks sheet found in Workslip-{ws_file_num}")
                except Exception as e:
                    logger.warning(f"[COMBINED UPLOAD] Could not extract item names from workslip: {e}")
            
            # Process multiple workslip files for multi-workslip upload
            if workslip_files:
                # Sort workslip files by phase number to process in order
                workslip_files.sort(key=lambda x: x[0])
                
                # Initialize combined data structures for all phases
                all_phase_exec_maps = []  # List of exec_maps, one per phase
                all_phase_ae_data = []    # List of ae_data, one per phase
                all_previous_supp_items = []  # Combined supplemental items from all phases
                parsed_items = None  # Will hold the base items (from the last/most recent workslip)
                ws_metadata = None  # Will hold metadata from the most recent workslip
                
                # Helper function to detect phase and column structure
                def detect_workslip_phase_and_columns(ws):
                    header_row = 8
                    for r in range(1, 15):
                        cell_val = str(ws.cell(row=r, column=1).value or "").strip().lower()
                        if "sl" in cell_val:
                            header_row = r
                            break
                    
                    phase_count = 0
                    col_map = {"header_row": header_row}
                    
                    for c in range(1, 30):
                        header = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
                        if ("execution" in header or "exec" in header or "workslip" in header) and ("qty" in header or "quantity" in header):
                            phase_count += 1
                            col_map[f"exec_qty_phase_{phase_count}"] = c
                        elif ("execution" in header or "exec" in header or "workslip" in header) and ("amount" in header or "amt" in header):
                            col_map[f"exec_amt_phase_{phase_count}"] = c
                        elif "qty" in header and ("est" in header or "estimate" in header):
                            col_map["est_qty"] = c
                        elif "description" in header or "desc" in header:
                            col_map["desc"] = c
                        elif "rate" in header and ("est" in header or "estimate" in header):
                            col_map["est_rate"] = c
                    
                    if phase_count == 0:
                        phase_count = 1
                        col_map["exec_qty_phase_1"] = 7
                        col_map["exec_amt_phase_1"] = 9
                    
                    return phase_count, col_map
                
                # Process each workslip file
                for ws_file_num, workslip_file in workslip_files:
                    logger.info(f"[MULTI-WORKSLIP] Processing Workslip-{ws_file_num}")
                    try:
                        ws_bytes = workslip_file.read()
                        wb_ws = load_workbook(BytesIO(ws_bytes), data_only=True)
                        wb_ws_formulas = load_workbook(BytesIO(ws_bytes), data_only=False)
                    except Exception as e:
                        return render(request, "core/workslip.html", {
                            "error": f"Couldn't read Workslip-{ws_file_num} file: {e}",
                            "category": category, "groups": groups, "current_group": current_group,
                            "items_in_group": items_in_group, "items_info": items_info, "ws_estimate_rows": ws_estimate_rows,
                            "preview_rows": [], "tp_percent": ws_tp_percent if ws_tp_percent else "",
                            "tp_type": ws_tp_type, "supp_items_selected": ws_supp_items,
                            "work_name": ws_work_name, "current_phase": ws_current_phase,
                            "previous_phases": ws_previous_phases, "target_workslip": target_workslip,
                        })
                    
                    # Find WorkSlip sheet (for quantities)
                    ws_sheet = None
                    for sh in wb_ws.worksheets:
                        if "workslip" in sh.title.lower() or "working estimate" in str(sh.cell(row=1, column=1).value or "").lower():
                            ws_sheet = sh
                            logger.info(f"[MULTI-WORKSLIP] Workslip-{ws_file_num}: Found workslip sheet: '{sh.title}'")
                            break
                    if not ws_sheet:
                        ws_sheet = wb_ws.active
                        logger.info(f"[MULTI-WORKSLIP] Workslip-{ws_file_num}: Using active sheet: '{ws_sheet.title}'")
                    
                    # Find Workslip Items Blocks sheet (for supplemental item names - yellow+red rows)
                    ws_blocks_sheet = None
                    for sh in wb_ws_formulas.worksheets:
                        if "items" in sh.title.lower() and "block" in sh.title.lower():
                            found = False
                            for r in range(1, min(sh.max_row, 200) + 1):
                                if get_heading_name_from_sheet(sh, r):
                                    found = True
                                    break
                            if found:
                                ws_blocks_sheet = sh
                                break
                    
                    # Parse Workslip Items Blocks for supplemental item names
                    ws_supp_item_names = []  # Supplemental items from this workslip
                    if ws_blocks_sheet:
                        in_supplemental_section = False
                        for r in range(1, ws_blocks_sheet.max_row + 1):
                            nm = get_heading_name_from_sheet(ws_blocks_sheet, r)
                            if nm:
                                nm_upper = nm.upper()
                                if "SUPPLEMENTAL" in nm_upper:
                                    in_supplemental_section = True
                                    continue
                                if in_supplemental_section:
                                    ws_supp_item_names.append(nm)
                        logger.info(f"[MULTI-WORKSLIP] Workslip-{ws_file_num}: Found {len(ws_supp_item_names)} supplemental items")
                    
                    phase_count, col_map = detect_workslip_phase_and_columns(ws_sheet)
                    header_row = col_map.get("header_row", 8)
                    
                    # Extract metadata from workslip header rows (rows 2-7) - only from last file
                    file_metadata = {
                        "work_name": "",
                        "estimate_amount": "",
                        "admin_sanction": "",
                        "tech_sanction": "",
                        "agreement": "",
                        "agency_name": "",
                        "tp_percent": 0.0,
                        "tp_type": "Excess",
                        "grand_total": 0.0,
                        "deduct_old_material": 0.0,
                        "lc_percent": 0.0,
                        "qc_percent": 0.0,
                        "nac_percent": 0.0,
                    }
                    
                    for r in range(2, 8):
                        cell_val = str(ws_sheet.cell(row=r, column=1).value or "").strip()
                        cell_lower = cell_val.lower()
                    
                        # Check if this is a label row and extract value
                        extracted_value = ""
                        if ":" in cell_val:
                            parts = cell_val.split(":", 1)
                            extracted_value = parts[1].strip() if len(parts) > 1 else ""
                        
                            # If no value after colon, check columns 2-10 for value
                            if not extracted_value:
                                for c in range(2, 11):
                                    val = ws_sheet.cell(row=r, column=c).value
                                    if val and str(val).strip():
                                        extracted_value = str(val).strip()
                                        break
                            
                            if "name of the work" in cell_lower or "work name" in cell_lower:
                                file_metadata["work_name"] = extracted_value
                            elif "estimate amount" in cell_lower:
                                file_metadata["estimate_amount"] = extracted_value
                                import re
                                num_match = re.search(r'[\d,]+\.?\d*', extracted_value.replace(',', ''))
                                if num_match:
                                    try:
                                        file_metadata["grand_total"] = float(num_match.group().replace(',', ''))
                                    except:
                                        pass
                            elif "administrative" in cell_lower or "admin" in cell_lower:
                                file_metadata["admin_sanction"] = extracted_value
                            elif "technical" in cell_lower or "tech" in cell_lower:
                                file_metadata["tech_sanction"] = extracted_value
                            elif "agreement" in cell_lower:
                                file_metadata["agreement"] = extracted_value
                            elif "agency" in cell_lower:
                                file_metadata["agency_name"] = extracted_value
                    
                    # Scan footer section for T.P, Grand Total, Deduct, LC, QC, NAC values
                    desc_col = col_map.get("desc", 2)
                    for r in range(header_row + 1, ws_sheet.max_row + 1):
                        desc_text = str(ws_sheet.cell(row=r, column=desc_col).value or "").strip()
                        desc_upper = desc_text.upper()
                        
                        if ("ADD" in desc_upper or "DEDUCT" in desc_upper) and ("T.P" in desc_upper or "T P" in desc_upper) and "UNUSED" not in desc_upper:
                            import re
                            tp_match = re.search(r'@\s*([\d.]+)\s*%', desc_text)
                            if tp_match:
                                try:
                                    file_metadata["tp_percent"] = float(tp_match.group(1))
                                except:
                                    pass
                            if "EXCESS" in desc_upper:
                                file_metadata["tp_type"] = "Excess"
                            elif "LESS" in desc_upper:
                                file_metadata["tp_type"] = "Less"
                        elif "GRAND TOTAL" in desc_upper:
                            for c in range(ws_sheet.max_column, 0, -1):
                                val = ws_sheet.cell(row=r, column=c).value
                                if val:
                                    try:
                                        file_metadata["grand_total"] = float(val)
                                        break
                                    except:
                                        continue
                        elif "DEDUCT" in desc_upper and "OLD" in desc_upper:
                            for c in range(ws_sheet.max_column, 0, -1):
                                val = ws_sheet.cell(row=r, column=c).value
                                if val:
                                    try:
                                        file_metadata["deduct_old_material"] = abs(float(val))
                                        break
                                    except:
                                        continue
                        elif "L.C" in desc_upper or "LC @" in desc_upper or "ADD LC" in desc_upper:
                            import re
                            lc_match = re.search(r'@\s*([\d.]+)\s*%', desc_text)
                            if lc_match:
                                try:
                                    file_metadata["lc_percent"] = float(lc_match.group(1))
                                except:
                                    pass
                        elif "Q.C" in desc_upper or "QC @" in desc_upper or "ADD QC" in desc_upper:
                            import re
                            qc_match = re.search(r'@\s*([\d.]+)\s*%', desc_text)
                            if qc_match:
                                try:
                                    file_metadata["qc_percent"] = float(qc_match.group(1))
                                except:
                                    pass
                        elif "NAC" in desc_upper:
                            import re
                            nac_match = re.search(r'@\s*([\d.]+)\s*%', desc_text)
                            if nac_match:
                                try:
                                    file_metadata["nac_percent"] = float(nac_match.group(1))
                                except:
                                    pass
                    
                    # Parse main items from this workslip file
                    file_parsed_items = []
                    file_phase_exec_maps = [{} for _ in range(phase_count)]
                    file_phase_ae_data = [{} for _ in range(phase_count)]
                    last_base_key = None
                    display_name_idx = 0
                    
                    for r in range(header_row + 1, ws_sheet.max_row + 1):
                        desc = str(ws_sheet.cell(row=r, column=desc_col).value or "").strip()
                        
                        if not desc:
                            continue
                        
                        if any(kw in desc.upper() for kw in ("SUB TOTAL", "SUBTOTAL", "TOTAL", "GRAND TOTAL", "DEDUCT", "SUPPLEMENTAL")):
                            break
                        
                        desc_upper = desc.upper().strip()
                        is_ae_row = (desc_upper.startswith("AE") and len(desc_upper) >= 2 and 
                                     (len(desc_upper) == 2 or desc_upper[2:].isdigit() or desc_upper[2] == ' '))
                        
                        est_qty = ws_sheet.cell(row=r, column=col_map.get("est_qty", 4)).value or 0
                        est_rate = ws_sheet.cell(row=r, column=col_map.get("est_rate", 5)).value or 0
                        unit = ws_sheet.cell(row=r, column=3).value or ""
                        
                        try:
                            est_qty = float(est_qty)
                        except:
                            est_qty = 0.0
                        try:
                            est_rate = float(est_rate)
                        except:
                            est_rate = 0.0
                        
                        if is_ae_row and last_base_key:
                            for p in range(1, phase_count + 1):
                                exec_qty_col = col_map.get(f"exec_qty_phase_{p}")
                                if exec_qty_col:
                                    exec_qty = ws_sheet.cell(row=r, column=exec_qty_col).value or 0
                                    try:
                                        exec_qty = float(exec_qty)
                                    except:
                                        exec_qty = 0.0
                                    if exec_qty > 0:
                                        if last_base_key in file_phase_exec_maps[p-1]:
                                            file_phase_exec_maps[p-1][last_base_key] += exec_qty
                                        else:
                                            file_phase_exec_maps[p-1][last_base_key] = exec_qty
                                        ae_key = f"{last_base_key}:ae:{desc}"
                                        file_phase_ae_data[p-1][ae_key] = exec_qty
                        else:
                            row_key = f"ws{ws_file_num}_row_{r}"
                            last_base_key = row_key
                            
                            for p in range(1, phase_count + 1):
                                exec_qty_col = col_map.get(f"exec_qty_phase_{p}")
                                if exec_qty_col:
                                    exec_qty = ws_sheet.cell(row=r, column=exec_qty_col).value or 0
                                    try:
                                        exec_qty = float(exec_qty)
                                    except:
                                        exec_qty = 0.0
                                    if exec_qty > 0:
                                        file_phase_exec_maps[p-1][row_key] = exec_qty
                            
                            if estimate_display_names and display_name_idx < len(estimate_display_names):
                                display_name = estimate_display_names[display_name_idx]
                            else:
                                display_name = desc
                            display_name_idx += 1
                            
                            file_parsed_items.append({
                                "key": row_key,
                                "excel_row": r,
                                "item_name": desc,
                                "display_name": display_name,
                                "desc": desc,
                                "qty_est": est_qty,
                                "unit": str(unit).strip(),
                                "rate": est_rate,
                            })
                    
                    # Parse supplemental items from this workslip
                    file_supp_items = []
                    in_supp_section = False
                    current_supp_section = 0  # Track which supplemental section we're in
                    supp_item_idx = 0
                    for r in range(header_row + 1, ws_sheet.max_row + 1):
                        desc = str(ws_sheet.cell(row=r, column=desc_col).value or "").strip()
                        
                        # Check if this is a supplemental section header (e.g., "Supplemental Items-1", "Supplemental Items-2")
                        if "SUPPLEMENTAL" in desc.upper():
                            in_supp_section = True
                            # Extract the section number from header like "Supplemental Items-1" or "Supplemental Items-2"
                            import re
                            supp_match = re.search(r'(\d+)\s*$', desc)
                            if supp_match:
                                current_supp_section = int(supp_match.group(1))
                            else:
                                current_supp_section += 1  # Increment for unnamed sections
                            logger.info(f"[MULTI-WORKSLIP] Found supplemental section {current_supp_section}: {desc}")
                            continue
                        
                        if in_supp_section and desc:
                            if any(kw in desc.upper() for kw in ("SUB TOTAL", "SUBTOTAL", "TOTAL", "GRAND TOTAL", "DEDUCT")):
                                break
                            
                            if supp_item_idx < len(ws_supp_item_names):
                                supp_display_name = ws_supp_item_names[supp_item_idx]
                            else:
                                supp_display_name = desc
                            supp_item_idx += 1
                            
                            # Get unit for supplemental item
                            supp_unit = str(ws_sheet.cell(row=r, column=3).value or "").strip()
                            
                            # Try to get rate from est_rate column first
                            supp_rate_col = col_map.get("est_rate", 5)
                            supp_rate = ws_sheet.cell(row=r, column=supp_rate_col).value or 0
                            try:
                                supp_rate = float(supp_rate)
                            except:
                                supp_rate = 0.0
                            
                            for p in range(1, phase_count + 1):
                                exec_qty_col = col_map.get(f"exec_qty_phase_{p}")
                                exec_amt_col = col_map.get(f"exec_amt_phase_{p}")
                                if exec_qty_col:
                                    exec_qty = ws_sheet.cell(row=r, column=exec_qty_col).value or 0
                                    try:
                                        exec_qty = float(exec_qty)
                                    except:
                                        exec_qty = 0.0
                                    
                                    # If rate is 0, try to calculate from exec amount / qty
                                    if supp_rate == 0 and exec_amt_col and exec_qty > 0:
                                        exec_amt = ws_sheet.cell(row=r, column=exec_amt_col).value or 0
                                        try:
                                            exec_amt = float(exec_amt)
                                            if exec_amt > 0:
                                                supp_rate = round(exec_amt / exec_qty, 2)
                                        except:
                                            pass
                                    
                                    if exec_qty > 0:
                                        supp_amount = exec_qty * supp_rate
                                        file_supp_items.append({
                                            "name": supp_display_name,
                                            "qty": exec_qty,
                                            "phase": p,  # Use the workslip phase number (1, 2, etc.) from the column
                                            "supp_section": current_supp_section,  # Which supplemental section (1 or 2)
                                            "desc": desc,
                                            "unit": supp_unit,
                                            "rate": supp_rate,
                                            "amount": supp_amount,
                                        })
                    
                    # Accumulate data from this workslip file into the combined structures
                    # For base items, use the most recent workslip's items
                    if file_parsed_items:
                        parsed_items = file_parsed_items
                    
                    # Accumulate execution maps from all phases in this file
                    for exec_map in file_phase_exec_maps:
                        if exec_map:  # Only add non-empty maps
                            all_phase_exec_maps.append(exec_map)
                    
                    for ae_data in file_phase_ae_data:
                        if ae_data:
                            all_phase_ae_data.append(ae_data)
                    
                    # Accumulate supplemental items
                    all_previous_supp_items.extend(file_supp_items)
                    
                    # Keep metadata from the most recent workslip
                    ws_metadata = file_metadata
                    
                    logger.info(f"[MULTI-WORKSLIP] Workslip-{ws_file_num}: Parsed {len(file_parsed_items)} items, {len(file_supp_items)} supp items, {phase_count} phases")
                
                # End of workslip files loop - save accumulated data to session
                if parsed_items:
                    ws_estimate_rows = parsed_items
                    request.session["ws_estimate_rows"] = ws_estimate_rows
                    request.session["ws_previous_phases"] = all_phase_exec_maps
                    request.session["ws_previous_ae_data"] = all_phase_ae_data
                    request.session["ws_previous_supp_items"] = all_previous_supp_items
                    request.session["ws_current_phase"] = target_workslip
                    request.session["ws_exec_map"] = {}
                    request.session["ws_target_workslip"] = target_workslip
                    
                    if ws_metadata:
                        request.session["ws_metadata"] = ws_metadata
                        if ws_metadata.get("work_name"):
                            request.session["ws_work_name"] = ws_metadata["work_name"]
                        request.session["ws_tp_percent"] = ws_metadata.get("tp_percent", 0.0)
                        request.session["ws_tp_type"] = ws_metadata.get("tp_type", "Excess")
                        request.session["ws_deduct_old_material"] = ws_metadata.get("deduct_old_material", 0.0)
                        request.session["ws_lc_percent"] = ws_metadata.get("lc_percent", 0.0)
                        request.session["ws_qc_percent"] = ws_metadata.get("qc_percent", 0.0)
                        request.session["ws_nac_percent"] = ws_metadata.get("nac_percent", 0.0)
                        if ws_metadata.get("grand_total", 0) > 0:
                            request.session["ws_estimate_grand_total"] = ws_metadata.get("grand_total", 0.0)
                    
                    logger.info(f"[MULTI-WORKSLIP] Complete: Loaded {len(all_phase_exec_maps)} phases, {len(parsed_items)} items, {len(all_previous_supp_items)} supp items for Workslip-{target_workslip}")
                
                return redirect(reverse('workslip_main') + '?preserve=1')
            
            # Only estimate file was uploaded (no previous workslip) - process as new Workslip-1
            elif estimate_file and est_bytes:
                try:
                    # Use the already-read estimate bytes (from display names extraction)
                    wb_est_parse = load_workbook(BytesIO(est_bytes), data_only=False)
                    wb_est_vals_parse = load_workbook(BytesIO(est_bytes), data_only=True)
                    
                    # Find estimate sheet
                    def looks_like_our_estimate_sheet_local(sh):
                        row = 3
                        a = str(sh.cell(row=row, column=1).value or "").strip().lower()
                        b = str(sh.cell(row=row, column=2).value or "").strip().lower()
                        d = str(sh.cell(row=row, column=4).value or "").strip().lower()
                        e = str(sh.cell(row=row, column=5).value or "").strip().lower()
                        h = str(sh.cell(row=row, column=8).value or "").strip().lower()
                        score = 0
                        if "sl" in a and "no" in a: score += 1
                        if "quantity" in b: score += 1
                        if "item" in d and "description" in d: score += 1
                        if "rate" in e: score += 1
                        if "amount" in h: score += 1
                        return score >= 3
                    
                    estimate_sheets = [sh for sh in wb_est_parse.worksheets if looks_like_our_estimate_sheet_local(sh)]
                    if not estimate_sheets:
                        return render(request, "core/workslip.html", {
                            "error": "No sheet in the uploaded workbook matches the Estimate format.",
                            "category": category, "groups": groups, "current_group": current_group,
                            "items_in_group": items_in_group, "items_info": items_info, "ws_estimate_rows": ws_estimate_rows,
                            "preview_rows": [], "tp_percent": ws_tp_percent if ws_tp_percent else "",
                            "tp_type": ws_tp_type, "supp_items_selected": ws_supp_items,
                            "work_name": ws_work_name, "current_phase": ws_current_phase,
                            "previous_phases": ws_previous_phases,
                        })
                    
                    ws_est_sheet = estimate_sheets[0]
                    ws_est_vals_sheet = wb_est_vals_parse[ws_est_sheet.title]
                    
                    # Use display names from estimate if we found them earlier
                    heading_names_local = estimate_display_names if estimate_display_names else []
                    
                    def to_number_local(v):
                        try:
                            return float(v)
                        except:
                            return 0.0
                    
                    # Parse rows
                    parsed_rows = []
                    max_row = ws_est_sheet.max_row
                    r = 4
                    heading_idx = 0
                    
                    work_name_local = ""
                    name_cell = ws_est_sheet.cell(row=2, column=1).value
                    if name_cell:
                        text = str(name_cell)
                        parts = text.split(":", 1)
                        work_name_local = parts[1].strip() if len(parts) > 1 else text.strip()
                    
                    while r <= max_row:
                        desc = ws_est_sheet.cell(row=r, column=4).value
                        desc_str = str(desc or "").strip()
                        desc_upper = desc_str.upper()
                        
                        rate_formula = ws_est_sheet.cell(row=r, column=5).value
                        rate_value = ws_est_vals_sheet.cell(row=r, column=5).value
                        rate_is_empty = (rate_formula is None or str(rate_formula).strip() == "")
                        
                        qty_formula = ws_est_sheet.cell(row=r, column=2).value
                        qty_value = ws_est_vals_sheet.cell(row=r, column=2).value
                        qty_is_empty = (qty_formula is None or str(qty_formula).strip() == "")
                        
                        if desc_str and any(kw in desc_upper for kw in ("SUB TOTAL", "SUBTOTAL", "TOTAL", "ECV")):
                            break
                        if desc is None and ws_est_sheet.cell(row=r, column=1).value is None:
                            r += 1
                            continue
                        if rate_is_empty and qty_is_empty:
                            r += 1
                            continue
                        
                        if desc_str != "" and not (rate_is_empty and qty_is_empty):
                            qty_num = float(qty_value) if isinstance(qty_value, (int, float)) else to_number_local(qty_formula)
                            unit = ws_est_sheet.cell(row=r, column=3).value
                            backend_item_name = desc_to_item.get(desc_str, desc_str)
                            
                            if heading_idx < len(heading_names_local):
                                display_name = heading_names_local[heading_idx]
                            else:
                                display_name = backend_item_name or desc_str
                            heading_idx += 1
                            
                            rate_num = float(rate_value) if isinstance(rate_value, (int, float)) else to_number_local(rate_formula)
                            
                            parsed_rows.append({
                                "key": f"{ws_est_sheet.title}_row{r}",
                                "excel_row": r,
                                "item_name": backend_item_name,
                                "display_name": display_name,
                                "item_desc": item_name_to_desc.get(display_name, '') or item_name_to_desc.get(backend_item_name, '') or desc_str,
                                "desc": desc_str,
                                "qty_est": qty_num,
                                "unit": str(unit or "").strip(),
                                "rate": rate_num,
                            })
                        r += 1
                    
                    # Store in session
                    request.session["ws_estimate_rows"] = parsed_rows
                    request.session["ws_exec_map"] = {}
                    request.session["ws_supp_items"] = []
                    request.session["ws_work_name"] = work_name_local
                    request.session["ws_current_phase"] = target_workslip  # Use target workslip number
                    request.session["ws_target_workslip"] = target_workslip
                    request.session["ws_previous_phases"] = []
                    request.session["ws_previous_supp_items"] = []
                    request.session["ws_previous_ae_data"] = []
                    logger.info(f"[COMBINED UPLOAD - ESTIMATE ONLY] Parsed {len(parsed_rows)} items from estimate for Workslip-{target_workslip}")
                    
                except Exception as e:
                    logger.warning(f"[COMBINED UPLOAD] Error parsing estimate only: {e}")
                    return render(request, "core/workslip.html", {
                        "error": f"Error parsing estimate file: {e}",
                        "category": category, "groups": groups, "current_group": current_group,
                        "items_in_group": items_in_group, "items_info": items_info, "ws_estimate_rows": ws_estimate_rows,
                        "preview_rows": [], "tp_percent": ws_tp_percent if ws_tp_percent else "",
                        "tp_type": ws_tp_type, "supp_items_selected": ws_supp_items,
                        "work_name": ws_work_name, "current_phase": ws_current_phase,
                        "previous_phases": ws_previous_phases,
                    })
                
                return redirect(reverse('workslip_main') + '?preserve=1')
            
            return redirect(reverse('workslip_main') + '?preserve=1')

        # D) Clear Everything
        elif action == "clear_all":
            ws_estimate_rows = []
            ws_exec_map = {}
            ws_tp_percent = 0.0
            ws_tp_type = "Excess"
            ws_supp_items = []
            ws_work_name = ""
            request.session["ws_estimate_rows"] = []
            request.session["ws_exec_map"] = {}
            request.session["ws_tp_percent"] = 0.0
            request.session["ws_tp_type"] = "Excess"
            request.session["ws_supp_items"] = []
            request.session["ws_estimate_grand_total"] = 0.0
            request.session["ws_work_name"] = ""
            request.session["ws_current_phase"] = 1
            request.session["ws_target_workslip"] = 1
            request.session["ws_previous_phases"] = []
            request.session["ws_previous_supp_items"] = []
            request.session["ws_previous_ae_data"] = []
            return redirect("workslip_main")

        # D2) Upload Previous Workslip for Next Phase
        elif action == "upload_previous_workslip":
            file = request.FILES.get("previous_workslip_file")
            if not file:
                return render(request, "core/workslip.html", {
                    "error": "Please upload a previous Workslip Excel file.",
                    "category": category,
                    "groups": groups,
                    "current_group": current_group,
                    "items_in_group": items_in_group, "items_info": items_info,
                    "ws_estimate_rows": ws_estimate_rows,
                    "preview_rows": [],
                    "tp_percent": ws_tp_percent if ws_tp_percent else "",
                    "tp_type": ws_tp_type,
                    "supp_items_selected": ws_supp_items,
                    "work_name": ws_work_name,
                    "current_phase": ws_current_phase,
                    "previous_phases": ws_previous_phases,
                })

            try:
                excel_bytes = file.read()
                wb_ws = load_workbook(BytesIO(excel_bytes), data_only=True)
            except Exception as e:
                return render(request, "core/workslip.html", {
                    "error": f"Couldn't read uploaded Workslip file: {e}",
                    "category": category,
                    "groups": groups,
                    "current_group": current_group,
                    "items_in_group": items_in_group, "items_info": items_info,
                    "ws_estimate_rows": ws_estimate_rows,
                    "preview_rows": [],
                    "tp_percent": ws_tp_percent if ws_tp_percent else "",
                    "tp_type": ws_tp_type,
                    "supp_items_selected": ws_supp_items,
                    "work_name": ws_work_name,
                    "current_phase": ws_current_phase,
                    "previous_phases": ws_previous_phases,
                })

            # Find WorkSlip sheet
            ws_sheet = None
            for sh in wb_ws.worksheets:
                if "workslip" in sh.title.lower() or "working estimate" in str(sh.cell(row=1, column=1).value or "").lower():
                    ws_sheet = sh
                    break
            
            if not ws_sheet:
                ws_sheet = wb_ws.active

            # Detect phase number from workslip (look for "Phase X" in title or just count columns)
            # Parse the workslip to extract: item descriptions, estimated qty, and all phase execution data
            
            def detect_workslip_phase_and_columns(ws):
                """
                Detect the phase number and column structure of a workslip.
                Returns: (phase_number, column_map) where column_map has exec_qty columns per phase
                """
                # Find header row (usually row 8)
                header_row = 8
                for r in range(1, 15):
                    cell_val = str(ws.cell(row=r, column=1).value or "").strip().lower()
                    if "sl" in cell_val:
                        header_row = r
                        break
                
                # Count execution columns to determine phase
                # Standard: Sl, Desc, Unit, Est Qty, Est Rate, Est Amt, Exec Qty, Exec Rate, Exec Amt, More, Less, Remarks
                # Multi-phase: ... Phase1 Qty, Phase1 Amt, Phase2 Qty, Phase2 Amt, ...
                
                phase_count = 0
                col_map = {"header_row": header_row}
                
                for c in range(1, 30):
                    header = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
                    if "execution" in header or "exec" in header:
                        if "qty" in header or "quantity" in header:
                            phase_count += 1
                            col_map[f"exec_qty_phase_{phase_count}"] = c
                        elif "amount" in header or "amt" in header:
                            col_map[f"exec_amt_phase_{phase_count}"] = c
                    elif "qty" in header and "est" in header:
                        col_map["est_qty"] = c
                    elif "description" in header or "desc" in header:
                        col_map["desc"] = c
                    elif "rate" in header and "est" in header:
                        col_map["est_rate"] = c
                
                # If no explicit phase columns found, assume standard single-phase layout
                if phase_count == 0:
                    phase_count = 1
                    col_map["exec_qty_phase_1"] = 7  # Column G
                    col_map["exec_amt_phase_1"] = 9  # Column I
                
                return phase_count, col_map
            
            phase_count, col_map = detect_workslip_phase_and_columns(ws_sheet)
            header_row = col_map.get("header_row", 8)
            
            # Parse rows from the workslip
            parsed_items = []
            new_estimate_rows = []
            phase_exec_maps = [{} for _ in range(phase_count)]  # One exec_map per phase
            phase_ae_data = [{} for _ in range(phase_count)]  # Track AE data separately per phase
            
            last_base_key = None  # Track the last base item for merging AE quantities
            
            for r in range(header_row + 1, ws_sheet.max_row + 1):
                desc_col = col_map.get("desc", 2)
                desc = str(ws_sheet.cell(row=r, column=desc_col).value or "").strip()
                
                if not desc:
                    continue
                    
                # Skip total/subtotal rows
                if any(kw in desc.upper() for kw in ("SUB TOTAL", "SUBTOTAL", "TOTAL", "GRAND TOTAL", "DEDUCT", "SUPPLEMENTAL")):
                    break
                
                # Check if this is an AE row (AE1, AE2, etc.)
                desc_upper = desc.upper().strip()
                is_ae_row = (desc_upper.startswith("AE") and len(desc_upper) >= 2 and 
                             (len(desc_upper) == 2 or desc_upper[2:].isdigit() or desc_upper[2] == ' '))
                
                # Get estimate qty and rate
                est_qty = ws_sheet.cell(row=r, column=col_map.get("est_qty", 4)).value or 0
                est_rate = ws_sheet.cell(row=r, column=col_map.get("est_rate", 5)).value or 0
                unit = ws_sheet.cell(row=r, column=3).value or ""
                
                try:
                    est_qty = float(est_qty)
                except:
                    est_qty = 0.0
                try:
                    est_rate = float(est_rate)
                except:
                    est_rate = 0.0
                
                if is_ae_row and last_base_key:
                    # This is an AE row - add its execution quantities to the previous base item
                    for p in range(1, phase_count + 1):
                        exec_qty_col = col_map.get(f"exec_qty_phase_{p}")
                        if exec_qty_col:
                            exec_qty = ws_sheet.cell(row=r, column=exec_qty_col).value or 0
                            try:
                                exec_qty = float(exec_qty)
                            except:
                                exec_qty = 0.0
                            if exec_qty > 0:
                                # Add AE qty to base item's exec_map
                                if last_base_key in phase_exec_maps[p-1]:
                                    phase_exec_maps[p-1][last_base_key] += exec_qty
                                else:
                                    phase_exec_maps[p-1][last_base_key] = exec_qty
                                # Also store the AE quantity separately for Excel output
                                ae_key = f"{last_base_key}:ae:{desc}"
                                phase_ae_data[p-1][ae_key] = exec_qty
                else:
                    # This is a base item row
                    row_key = f"phase_row_{r}"
                    last_base_key = row_key
                    
                    # Collect execution data from all phases
                    for p in range(1, phase_count + 1):
                        exec_qty_col = col_map.get(f"exec_qty_phase_{p}")
                        if exec_qty_col:
                            exec_qty = ws_sheet.cell(row=r, column=exec_qty_col).value or 0
                            try:
                                exec_qty = float(exec_qty)
                            except:
                                exec_qty = 0.0
                            if exec_qty > 0:
                                phase_exec_maps[p-1][row_key] = exec_qty
                    
                    new_estimate_rows.append({
                        "key": row_key,
                        "excel_row": r,
                        "item_name": desc,
                        "display_name": desc,
                        "item_desc": item_name_to_desc.get(desc, '') or desc,
                        "desc": desc,
                        "qty_est": est_qty,
                        "unit": str(unit).strip(),
                        "rate": est_rate,
                    })
            
            if new_estimate_rows:
                # Store the estimate rows
                ws_estimate_rows = new_estimate_rows
                request.session["ws_estimate_rows"] = ws_estimate_rows
                
                # Store previous phases' execution data (with AE merged into base items)
                ws_previous_phases = phase_exec_maps
                request.session["ws_previous_phases"] = ws_previous_phases
                
                # Store previous phases' AE data separately for Excel output
                request.session["ws_previous_ae_data"] = phase_ae_data
                
                # Set current phase to next phase
                ws_current_phase = phase_count + 1
                request.session["ws_current_phase"] = ws_current_phase
                
                # Clear current exec_map for new phase entry
                request.session["ws_exec_map"] = {}
                
                logger.info(f"[WORKSLIP PHASE] Parsed workslip with {phase_count} phases, {len(new_estimate_rows)} items. Now phase {ws_current_phase}")
            
            return redirect(reverse('workslip_main') + '?preserve=1')

        # E) Download Workslip
        elif action == "download_workslip":
            exec_str = request.POST.get("exec_map", "")
            rate_str = request.POST.get("rate_map", "")
            tp_percent_str = request.POST.get("tp_percent", "")
            tp_type = request.POST.get("tp_type", "Excess")
            deduct_old_material_str = request.POST.get("deduct_old_material", "")
            
            # Get metadata from form
            ws_work_name_form = request.POST.get("ws_work_name", "")
            ws_estimate_amount_form = request.POST.get("ws_estimate_amount", "")
            ws_agency_name_form = request.POST.get("ws_agency_name", "")
            ws_admin_sanction_form = request.POST.get("ws_admin_sanction", "")
            ws_tech_sanction_form = request.POST.get("ws_tech_sanction", "")
            ws_agreement_form = request.POST.get("ws_agreement", "")

            # merge UI exec_map into session map
            ws_exec_map_session = request.session.get("ws_exec_map", {}) or {}
            new_exec_map = {}
            if exec_str:
                try:
                    raw = json.loads(exec_str)
                    if isinstance(raw, dict):
                        for k, v in raw.items():
                            try:
                                new_exec_map[str(k)] = float(v)
                            except Exception:
                                continue
                except Exception:
                    pass
            ws_exec_map = ws_exec_map_session.copy()
            ws_exec_map.update(new_exec_map)

            # merge UI rate_map into session map
            ws_rate_map_session = request.session.get("ws_rate_map", {}) or {}
            new_rate_map = {}
            if rate_str:
                try:
                    raw = json.loads(rate_str)
                    if isinstance(raw, dict):
                        for k, v in raw.items():
                            try:
                                new_rate_map[str(k)] = float(v)
                            except Exception:
                                continue
                except Exception:
                    pass
            ws_rate_map = ws_rate_map_session.copy()
            ws_rate_map.update(new_rate_map)

            # Use form values if provided, otherwise fall back to session values (from uploaded workslip)
            try:
                ws_tp_percent = float(tp_percent_str) if tp_percent_str != "" else request.session.get("ws_tp_percent", 0.0)
            except Exception:
                ws_tp_percent = request.session.get("ws_tp_percent", 0.0)
            ws_tp_type = tp_type if tp_type in ("Less", "Excess") else request.session.get("ws_tp_type", "Excess")
            
            # Parse Deduct Old Material Cost - use form value or fall back to session
            try:
                ws_deduct_old_material = float(deduct_old_material_str) if deduct_old_material_str != "" else request.session.get("ws_deduct_old_material", 0.0)
            except Exception:
                ws_deduct_old_material = request.session.get("ws_deduct_old_material", 0.0)

            request.session["ws_exec_map"] = ws_exec_map
            request.session["ws_rate_map"] = ws_rate_map
            request.session["ws_tp_percent"] = ws_tp_percent
            request.session["ws_tp_type"] = ws_tp_type
            request.session["ws_deduct_old_material"] = ws_deduct_old_material
            
            # Update metadata in session from form values
            ws_metadata_session = request.session.get("ws_metadata", {}) or {}
            if ws_work_name_form:
                ws_metadata_session["work_name"] = ws_work_name_form
            if ws_estimate_amount_form:
                ws_metadata_session["estimate_amount"] = ws_estimate_amount_form
            if ws_agency_name_form:
                ws_metadata_session["agency_name"] = ws_agency_name_form
            if ws_admin_sanction_form:
                ws_metadata_session["admin_sanction"] = ws_admin_sanction_form
            if ws_tech_sanction_form:
                ws_metadata_session["tech_sanction"] = ws_tech_sanction_form
            if ws_agreement_form:
                ws_metadata_session["agreement"] = ws_agreement_form
            # Also update TP values in metadata
            ws_metadata_session["tp_percent"] = ws_tp_percent
            ws_metadata_session["tp_type"] = ws_tp_type
            request.session["ws_metadata"] = ws_metadata_session
            request.session.modified = True

            # helper to safely fetch execution quantity for base estimate rows
            def get_exec_qty_for_base(row_key, item_name, desc):
                candidates = [
                    f"base:{row_key}",
                    row_key,
                    item_name or "",
                    desc or "",
                ]
                for k in candidates:
                    k = str(k).strip()
                    if not k:
                        continue
                    if k in ws_exec_map:
                        try:
                            return float(ws_exec_map[k])
                        except Exception:
                            return 0.0
                return 0.0

            # helper to get rate for a row, considering custom rates from ws_rate_map
            def get_rate_for_row(row_key, original_rate):
                """Get the rate for a row, using custom rate from ws_rate_map if available."""
                if row_key in ws_rate_map:
                    try:
                        return float(ws_rate_map[row_key])
                    except Exception:
                        return original_rate
                return original_rate

            # ---------- build supplemental description+rate from backend ----------
            item_to_info = {it["name"]: it for it in items_list}
            wb_backend_vals = None
            ws_backend_vals = None
            if filepath and os.path.exists(filepath):
                try:
                    wb_backend_vals = load_workbook(filepath, data_only=True)
                    ws_backend_vals = wb_backend_vals["Master Datas"]
                except Exception:
                    wb_backend_vals = None
                    ws_backend_vals = None

            supp_desc_map = {}
            supp_rate_map = {}
            for name in ws_supp_items:
                info = item_to_info.get(name)
                if not info or ws_backend_vals is None or ws_data is None:
                    continue
                start_row = info["start_row"]
                end_row = info["end_row"]
                # Description: 2nd row below yellow header in col D
                desc_cell = ws_data.cell(row=start_row + 2, column=4).value
                supp_desc_map[name] = str(desc_cell or "").strip()
                # Rate from Master Datas col J
                rate_val = 0.0
                for r in range(end_row, start_row - 1, -1):
                    v = ws_backend_vals.cell(row=r, column=10).value
                    if v not in (None, ""):
                        try:
                            rate_val = float(v)
                        except Exception:
                            rate_val = 0.0
                        break
                supp_rate_map[name] = rate_val

            # Load prefix mapping for repair mode (used in both ItemBlocks and WorkSlip sheets)
            item_to_prefix_ws = {}
            if ws_work_mode == 'repair':
                from core.saved_works_views import load_prefix_map
                item_to_prefix_ws = load_prefix_map(
                    ws_category, backend_id=ws_selected_backend_id, user=request.user
                )

            # ---------- create workbook ----------
            wb_out = Workbook()

            # Sheet 1: ItemBlocks (only if supplemental items exist)
            if ws_supp_items:
                ws_blocks = wb_out.active
                ws_blocks.title = "ItemBlocks"
                # Add "Name of Work" header
                ws_blocks.merge_cells("A1:J1")
                hdr = ws_blocks["A1"]
                hdr.value = f"Name of the work : {ws_work_name}" if ws_work_name else "Name of the work : "
                hdr.font = Font(bold=True, size=11)
                hdr.alignment = Alignment(horizontal="left", vertical="center")
                current_row = 3  # start blocks after header + blank row
                data_serial_blocks = 1
                if ws_data is not None:
                    for name in ws_supp_items:
                        info = item_to_info.get(name)
                        if not info:
                            continue
                        start_row = info["start_row"]
                        end_row = info["end_row"]
                        copy_block_with_styles_and_formulas(
                            ws_src=ws_data,
                            ws_dst=ws_blocks,
                            src_min_row=start_row,
                            src_max_row=end_row,
                            dst_start_row=current_row,
                            col_start=1,
                            col_end=10,
                        )
                        # Add Data serial number to column A of block header row
                        ws_blocks.cell(row=current_row, column=1).value = f"Data {data_serial_blocks}"
                        data_serial_blocks += 1
                        # Apply repair prefix to description cell (row+2, col D)
                        if ws_work_mode == 'repair' and item_to_prefix_ws:
                            prefix = item_to_prefix_ws.get(name, "")
                            if prefix:
                                desc_cell_block = ws_blocks.cell(row=current_row + 2, column=4)
                                base_val = desc_cell_block.value
                                base_str = str(base_val).strip() if base_val not in (None, "") else ""
                                desc_cell_block.value = f"{prefix} {base_str}" if base_str else prefix
                        current_row += (end_row - start_row + 1)  # No blank row between blocks
            else:
                # no supplemental items â†’ remove default sheet so only WorkSlip will exist
                default_sheet = wb_out.active
                wb_out.remove(default_sheet)

            # Sheet 2 (or 1 if no ItemBlocks): WorkSlip
            ws_ws = wb_out.create_sheet("WorkSlip")

            thin = Side(border_style="thin", color="000000")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
            header_fill = PatternFill("solid", fgColor="FFC8C8C8")
            subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")
            supp_fill = PatternFill("solid", fgColor="FFF5E1")
            phase_fill = PatternFill("solid", fgColor="FFFEF3C7")  # Amber for previous phases
            current_phase_fill = PatternFill("solid", fgColor="FFDBEAFE")  # Blue for current phase
            
            # Get phase data
            ws_previous_phases = request.session.get("ws_previous_phases", []) or []
            ws_current_phase = request.session.get("ws_current_phase", 1)
            num_previous_phases = len(ws_previous_phases)
            
            # Calculate total columns: Base 11 + 2 per previous phase (Qty + Amount)
            # Columns: Sl, Desc, Unit, Est Qty, Est Rate, Est Amt, [Phase1 Qty, Phase1 Amt, ...], Curr Qty, Curr Amt, More, Less, Remarks
            extra_phase_cols = num_previous_phases * 2
            total_cols = 11 + extra_phase_cols
            
            # Build column letter for last column
            def col_letter(n):
                result = ""
                while n > 0:
                    n, remainder = divmod(n - 1, 26)
                    result = chr(65 + remainder) + result
                return result
            
            last_col_letter = col_letter(total_cols)

            # Top main heading - merge across all columns
            ws_ws.merge_cells(f"A1:{last_col_letter}1")
            c = ws_ws["A1"]
            phase_title = f"WORKING ESTIMATE-{ws_current_phase}" if ws_current_phase > 1 else "WORKING ESTIMATE"
            c.value = phase_title
            c.font = Font(bold=True, size=14)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Get stored metadata from previous workslip
            ws_metadata = request.session.get("ws_metadata", {})
            
            # 6 merged rows below heading - use values from uploaded workslip if available
            work_name_val = ws_metadata.get("work_name", "") or ws_work_name or ""
            estimate_amount_val = ws_metadata.get("estimate_amount", "") or ""
            admin_sanction_val = ws_metadata.get("admin_sanction", "") or ""
            tech_sanction_val = ws_metadata.get("tech_sanction", "") or ""
            agreement_val = ws_metadata.get("agreement", "") or ""
            agency_name_val = ws_metadata.get("agency_name", "") or ""
            
            headings = [
                f"Name of the work : {work_name_val}" if work_name_val else "Name of the work :",
                f"Estimate Amount : {estimate_amount_val}" if estimate_amount_val else "Estimate Amount :",
                f"Ref. to Administrative sanction : {admin_sanction_val}" if admin_sanction_val else "Ref. to Administrative sanction :",
                f"Ref. to Technical sanction : {tech_sanction_val}" if tech_sanction_val else "Ref. to Technical sanction :",
                f"Ref. to Agreement : {agreement_val}" if agreement_val else "Ref. to Agreement :",
                f"Name of the Agency : {agency_name_val}" if agency_name_val else "Name of the Agency :",
            ]
            for i, text in enumerate(headings, start=2):
                merge_range = f"A{i}:{last_col_letter}{i}"
                ws_ws.merge_cells(merge_range)
                cell = ws_ws[f"A{i}"]
                cell.value = text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Apply borders to heading area
            for row_idx in range(1, 8):
                for col_idx in range(1, total_cols + 1):
                    cell = ws_ws.cell(row=row_idx, column=col_idx)
                    cell.border = border_all

            # Table header - dynamic based on phases
            header_row = 8
            # Base columns (first 6)
            base_cols = ["Sl.No", "Description of Item", "Unit", "Qty (Estimate)", "Rate (Estimate)", "Amount (Estimate)"]
            
            # Previous phase columns
            phase_cols = []
            for p in range(1, num_previous_phases + 1):
                phase_cols.extend([f"Workslip-{p} Qty", f"Workslip-{p} Amt"])
            
            # Current phase columns (last 5 - no Rate since it's same as Estimate Rate)
            if ws_current_phase > 1:
                current_cols = [f"Workslip-{ws_current_phase} Qty", f"Workslip-{ws_current_phase} Amt", "More", "Less", "Remarks"]
            else:
                current_cols = ["Qty (Execution)", "Amount (Execution)", "More", "Less", "Remarks"]
            
            all_header_cols = base_cols + phase_cols + current_cols
            
            for col_idx, text in enumerate(all_header_cols, start=1):
                cell = ws_ws.cell(row=header_row, column=col_idx, value=text)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_all
                cell.fill = header_fill
                # Color phase columns differently
                if col_idx > 6 and col_idx <= 6 + extra_phase_cols:
                    cell.fill = phase_fill
                elif col_idx > 6 + extra_phase_cols and col_idx <= 6 + extra_phase_cols + 2:
                    if ws_current_phase > 1:
                        cell.fill = current_phase_fill

            # Column widths
            ws_ws.column_dimensions["A"].width = 6
            ws_ws.column_dimensions["B"].width = 70
            ws_ws.column_dimensions["C"].width = 10
            ws_ws.column_dimensions["D"].width = 14
            ws_ws.column_dimensions["E"].width = 12
            ws_ws.column_dimensions["F"].width = 14
            
            # Previous phase columns
            for i in range(extra_phase_cols):
                col_l = col_letter(7 + i)
                ws_ws.column_dimensions[col_l].width = 12
            
            # Current execution columns (no Rate - uses Estimate Rate)
            exec_start_col = 7 + extra_phase_cols
            ws_ws.column_dimensions[col_letter(exec_start_col)].width = 14      # Qty
            ws_ws.column_dimensions[col_letter(exec_start_col + 1)].width = 14  # Amount
            ws_ws.column_dimensions[col_letter(exec_start_col + 2)].width = 10  # More
            ws_ws.column_dimensions[col_letter(exec_start_col + 3)].width = 10  # Less
            ws_ws.column_dimensions[col_letter(exec_start_col + 4)].width = 25  # Remarks

            # remark helper
            def remark_for_item(q_est, q_exec, is_supp=False, has_ae_split=False):
                if is_supp:
                    return "Proposed as per site condition"
                if q_exec == 0:
                    return "Deleted"
                # If we're splitting into base + AE rows, base row should not say "Excess"
                if has_ae_split:
                    return ""  # Base row gets no remark when there's an AE row
                if q_exec > q_est:
                    return "Excess as per estimated"
                if 0 < q_exec < q_est:
                    return "Less as per estimated"
                return ""

            data_start = header_row + 1
            out_row = data_start
            sl_counter = 1
            ae_counter = 1
            
            # Column indices for data (no COL_CURR_RATE - uses Estimate Rate)
            COL_SL = 1
            COL_DESC = 2
            COL_UNIT = 3
            COL_EST_QTY = 4
            COL_EST_RATE = 5
            COL_EST_AMT = 6
            COL_PHASE_START = 7  # Previous phases start here
            COL_CURR_QTY = 7 + extra_phase_cols
            COL_CURR_AMT = 8 + extra_phase_cols
            COL_MORE = 9 + extra_phase_cols
            COL_LESS = 10 + extra_phase_cols
            COL_REMARKS = 11 + extra_phase_cols

            # Get previous phases' AE data and supplemental items
            ws_previous_ae_data = request.session.get("ws_previous_ae_data", [])
            ws_previous_supp_items = request.session.get("ws_previous_supp_items", [])
            
            # ---- Base Estimate items with row-splitting ----
            for row in ws_estimate_rows:
                row_key = row["key"]
                qty_est = round(float(row.get("qty_est", 0) or 0), 2)

                qty_exec = get_exec_qty_for_base(
                    row_key=row_key,
                    item_name=row.get("item_name"),
                    desc=row.get("desc"),
                )

                unit = row.get("unit") or ""
                original_rate = round(float(row.get("rate", 0) or 0), 2)
                rate = round(get_rate_for_row(row_key, original_rate), 2)  # Apply custom rate if user modified it
                desc_est = row.get("item_desc") or row.get("desc") or row.get("item_name") or ""
                # If desc_est is just the header name, try to get the real row+2 description from backend
                _display = row.get("display_name") or row.get("item_name") or ""
                if desc_est == _display and _display in item_name_to_desc:
                    desc_est = item_name_to_desc[_display]
                # Apply repair prefix to base estimate items
                if ws_work_mode == 'repair' and item_to_prefix_ws:
                    _item_name_for_prefix = row.get("item_name") or row.get("display_name") or ""
                    _prefix = item_to_prefix_ws.get(_item_name_for_prefix, "")
                    if _prefix:
                        desc_est = f"{_prefix} {desc_est}" if desc_est else _prefix
                
                # Get previous phases' execution quantities for this row (AE already merged)
                prev_phase_qtys = []
                for phase_map in ws_previous_phases:
                    prev_qty = phase_map.get(row_key, 0)
                    try:
                        prev_qty = round(float(prev_qty), 2)
                    except:
                        prev_qty = 0.0
                    prev_phase_qtys.append(prev_qty)
                
                # Calculate excess for each previous phase
                prev_phase_excess = []
                for p_qty in prev_phase_qtys:
                    excess = round(max(0, p_qty - qty_est), 2) if qty_est > 0 else 0
                    prev_phase_excess.append(excess)
                
                # Calculate base qty for each previous phase (capped at estimate)
                prev_base_qtys = []
                for p_qty in prev_phase_qtys:
                    base_qty = round(min(p_qty, qty_est), 2) if qty_est > 0 else round(p_qty, 2)
                    prev_base_qtys.append(base_qty)
                
                # Calculate current phase excess
                current_excess = round(max(0, qty_exec - qty_est), 2) if qty_est > 0 else 0
                current_base_qty = round(min(qty_exec, qty_est), 2) if qty_est > 0 else round(qty_exec, 2)
                
                # Check if any phase (previous or current) has excess - if so, we need ONE AE row
                has_any_excess = any(e > 0 for e in prev_phase_excess) or current_excess > 0

                # FIRST: Always write the base row
                ws_ws.cell(out_row, COL_SL, sl_counter)
                ws_ws.cell(out_row, COL_DESC, desc_est)
                ws_ws.cell(out_row, COL_UNIT, unit)
                ws_ws.cell(out_row, COL_EST_QTY, qty_est)
                ws_ws.cell(out_row, COL_EST_RATE, rate)
                ws_ws.cell(out_row, COL_EST_AMT, f"={col_letter(COL_EST_QTY)}{out_row}*{col_letter(COL_EST_RATE)}{out_row}")
                
                # Previous phases' data - show base qty (capped at estimate)
                for p_idx, p_qty in enumerate(prev_base_qtys):
                    phase_qty_col = COL_PHASE_START + (p_idx * 2)
                    phase_amt_col = phase_qty_col + 1
                    ws_ws.cell(out_row, phase_qty_col, p_qty)
                    ws_ws.cell(out_row, phase_amt_col, f"={col_letter(phase_qty_col)}{out_row}*{col_letter(COL_EST_RATE)}{out_row}")
                    ws_ws.cell(out_row, phase_qty_col).fill = phase_fill
                    ws_ws.cell(out_row, phase_amt_col).fill = phase_fill
                
                # Current execution (base qty capped at estimate if there's excess)
                ws_ws.cell(out_row, COL_CURR_QTY, current_base_qty)
                ws_ws.cell(out_row, COL_CURR_AMT, f"={col_letter(COL_CURR_QTY)}{out_row}*{col_letter(COL_EST_RATE)}{out_row}")
                ws_ws.cell(out_row, COL_MORE, f"=IF({col_letter(COL_CURR_AMT)}{out_row}>{col_letter(COL_EST_AMT)}{out_row},{col_letter(COL_CURR_AMT)}{out_row}-{col_letter(COL_EST_AMT)}{out_row},\"\")")
                ws_ws.cell(out_row, COL_LESS, f"=IF({col_letter(COL_EST_AMT)}{out_row}>{col_letter(COL_CURR_AMT)}{out_row},{col_letter(COL_EST_AMT)}{out_row}-{col_letter(COL_CURR_AMT)}{out_row},\"\")")
                ws_ws.cell(out_row, COL_REMARKS, remark_for_item(qty_est, qty_exec, is_supp=False, has_ae_split=has_any_excess))

                for cidx in range(1, total_cols + 1):
                    cell = ws_ws.cell(out_row, cidx)
                    cell.border = border_all
                    if cidx in (COL_DESC, COL_REMARKS):
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                out_row += 1
                sl_counter += 1
                base_row_for_rate = out_row - 1  # Save for rate reference
                
                # SECOND: If any phase has excess, write ONE AE row with all phase excess values
                if has_any_excess:
                    ae_label = f"AE{ae_counter}"
                    ae_counter += 1

                    ws_ws.cell(out_row, COL_SL, "")          # no serial number
                    ws_ws.cell(out_row, COL_DESC, ae_label)
                    ws_ws.cell(out_row, COL_UNIT, unit)
                    ws_ws.cell(out_row, COL_EST_QTY, None)
                    ws_ws.cell(out_row, COL_EST_RATE, None)
                    ws_ws.cell(out_row, COL_EST_AMT, None)
                    
                    # Previous phases - show excess for each phase in respective columns
                    for pi, excess_qty in enumerate(prev_phase_excess):
                        phase_qty_col = COL_PHASE_START + (pi * 2)
                        phase_amt_col = phase_qty_col + 1
                        if excess_qty > 0:
                            ws_ws.cell(out_row, phase_qty_col, excess_qty)
                            ws_ws.cell(out_row, phase_amt_col, f"={col_letter(phase_qty_col)}{out_row}*{col_letter(COL_EST_RATE)}{base_row_for_rate}")
                        else:
                            ws_ws.cell(out_row, phase_qty_col, None)
                            ws_ws.cell(out_row, phase_amt_col, None)
                        ws_ws.cell(out_row, phase_qty_col).fill = phase_fill
                        ws_ws.cell(out_row, phase_amt_col).fill = phase_fill
                    
                    # Current phase excess
                    if current_excess > 0:
                        ws_ws.cell(out_row, COL_CURR_QTY, current_excess)
                        ws_ws.cell(out_row, COL_CURR_AMT, f"={col_letter(COL_CURR_QTY)}{out_row}*{col_letter(COL_EST_RATE)}{base_row_for_rate}")
                    else:
                        ws_ws.cell(out_row, COL_CURR_QTY, None)
                        ws_ws.cell(out_row, COL_CURR_AMT, None)
                    
                    ws_ws.cell(out_row, COL_MORE, f"=IF({col_letter(COL_CURR_AMT)}{out_row}>{col_letter(COL_EST_AMT)}{out_row},{col_letter(COL_CURR_AMT)}{out_row}-{col_letter(COL_EST_AMT)}{out_row},\"\")")
                    ws_ws.cell(out_row, COL_LESS, "")
                    ws_ws.cell(out_row, COL_REMARKS, "Excess as per estimated")

                    for cidx in range(1, total_cols + 1):
                        cell = ws_ws.cell(out_row, cidx)
                        cell.border = border_all
                        if cidx in (COL_DESC, COL_REMARKS):
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    out_row += 1

            # ---- Previous Workslip Supplemental items (per phase) ----
            if ws_previous_supp_items:
                # Group previous supplemental items by phase
                supp_by_phase = {}
                for supp in ws_previous_supp_items:
                    phase_num = supp.get("phase", 1)
                    if phase_num not in supp_by_phase:
                        supp_by_phase[phase_num] = []
                    supp_by_phase[phase_num].append(supp)
                
                # Output each phase's supplemental items
                for phase_num in sorted(supp_by_phase.keys()):
                    phase_supps = supp_by_phase[phase_num]
                    
                    # Heading row for this phase's supplemental items
                    supp_phase_header = f"Supplemental Items-{phase_num}"
                    supp_cell = ws_ws.cell(out_row, COL_DESC, supp_phase_header)
                    supp_cell.font = Font(bold=True, color="FF0000")  # Red text
                    for col in range(1, total_cols + 1):
                        cell = ws_ws.cell(out_row, col)
                        cell.border = border_all
                        cell.fill = supp_fill
                        if col == COL_DESC:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    out_row += 1
                    
                    # Output each supplemental item from this phase
                    for supp in phase_supps:
                        supp_name = supp.get("name", "")
                        supp_qty = round(float(supp.get("qty", 0) or 0), 2)
                        # Use description from row+2 (stored in 'desc'), fallback to item name
                        supp_desc = supp.get("desc", supp_name) or supp_name
                        # Apply repair prefix to previous phase supplemental items
                        if ws_work_mode == 'repair' and item_to_prefix_ws:
                            prefix = item_to_prefix_ws.get(supp_name, "")
                            if prefix:
                                supp_desc = f"{prefix} {supp_desc}" if supp_desc else prefix
                        supp_unit = supp.get("unit", "-") or "-"

                        # Apply custom rate if user modified it
                        prev_supp_key = f"prev_supp:{phase_num}:{supp_name}"
                        original_supp_rate = round(float(supp.get("rate", 0) or 0), 2)
                        supp_rate = round(get_rate_for_row(prev_supp_key, original_supp_rate), 2)
                        supp_amount = round(supp.get("amount", supp_qty * supp_rate), 2)
                        # Recalculate amount if rate was modified
                        if prev_supp_key in ws_rate_map:
                            supp_amount = round(supp_qty * supp_rate, 2)
                        
                        ws_ws.cell(out_row, COL_SL, sl_counter)
                        # Use supp_desc (row+2 description) instead of supp_name
                        ws_ws.cell(out_row, COL_DESC, supp_desc)
                        ws_ws.cell(out_row, COL_UNIT, supp_unit)
                        ws_ws.cell(out_row, COL_EST_QTY, None)
                        ws_ws.cell(out_row, COL_EST_RATE, supp_rate if supp_rate > 0 else None)
                        ws_ws.cell(out_row, COL_EST_AMT, None)
                        
                        # Put quantity in the correct phase column
                        for p_idx in range(num_previous_phases):
                            phase_qty_col = COL_PHASE_START + (p_idx * 2)
                            phase_amt_col = phase_qty_col + 1
                            if (p_idx + 1) == phase_num:
                                ws_ws.cell(out_row, phase_qty_col, supp_qty)
                                # Amount calculated from rate
                                ws_ws.cell(out_row, phase_amt_col, supp_amount if supp_rate > 0 else None)
                            else:
                                ws_ws.cell(out_row, phase_qty_col, None)
                                ws_ws.cell(out_row, phase_amt_col, None)
                            ws_ws.cell(out_row, phase_qty_col).fill = phase_fill
                            ws_ws.cell(out_row, phase_amt_col).fill = phase_fill
                        
                        # Check if user entered current workslip quantity for this previous supp item
                        # prev_supp_key is already defined above when getting rate
                        prev_supp_curr_qty = round(float(ws_exec_map.get(prev_supp_key, 0) or 0), 2)

                        if prev_supp_curr_qty > 0:
                            prev_supp_curr_amt = round(prev_supp_curr_qty * supp_rate, 2)
                            ws_ws.cell(out_row, COL_CURR_QTY, prev_supp_curr_qty)
                            ws_ws.cell(out_row, COL_CURR_AMT, prev_supp_curr_amt)
                            ws_ws.cell(out_row, COL_MORE, prev_supp_curr_amt)  # All extra work for prev supp items goes to More
                            ws_ws.cell(out_row, COL_LESS, None)
                            ws_ws.cell(out_row, COL_REMARKS, "Proposed as per site condition")
                        else:
                            ws_ws.cell(out_row, COL_CURR_QTY, None)
                            ws_ws.cell(out_row, COL_CURR_AMT, None)
                            ws_ws.cell(out_row, COL_MORE, None)
                            ws_ws.cell(out_row, COL_LESS, None)
                            ws_ws.cell(out_row, COL_REMARKS, f"From Workslip-{phase_num}")
                        
                        for cidx in range(1, total_cols + 1):
                            cell = ws_ws.cell(out_row, cidx)
                            cell.border = border_all
                            if cidx in (COL_DESC, COL_REMARKS):
                                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        out_row += 1
                        sl_counter += 1
            
            # ---- Current Phase Supplemental items block ----
            if ws_supp_items:
                # heading row with phase-specific name
                supp_header_text = f"Supplemental Items-{ws_current_phase}" if ws_current_phase > 1 else "Supplemental Items"
                supp_cell = ws_ws.cell(out_row, COL_DESC, supp_header_text)
                supp_cell.font = Font(bold=True)
                for col in range(1, total_cols + 1):
                    cell = ws_ws.cell(out_row, col)
                    cell.border = border_all
                    cell.fill = supp_fill
                    if col == COL_DESC:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                out_row += 1

                # actual supplemental rows
                for name in ws_supp_items:
                    desc_supp = supp_desc_map.get(name, name)
                    if ws_work_mode == 'repair' and item_to_prefix_ws:
                        prefix = item_to_prefix_ws.get(name, "")
                        if prefix:
                            desc_supp = f"{prefix} {desc_supp}" if desc_supp else prefix
                    unit_pl, _ = units_for(name)
                    rate = round(float(supp_rate_map.get(name, 0.0) or 0.0), 2)
                    key = f"supp:{name}"
                    qty_exec = round(float(ws_exec_map.get(key, 0) or 0), 2)

                    ws_ws.cell(out_row, COL_SL, sl_counter)
                    ws_ws.cell(out_row, COL_DESC, desc_supp)
                    ws_ws.cell(out_row, COL_UNIT, unit_pl)
                    ws_ws.cell(out_row, COL_EST_QTY, None)
                    ws_ws.cell(out_row, COL_EST_RATE, rate)  # Put supp rate in Est Rate column for amount calculation
                    ws_ws.cell(out_row, COL_EST_AMT, None)
                    
                    # Previous phases empty for current supplemental
                    for p_idx in range(num_previous_phases):
                        phase_qty_col = COL_PHASE_START + (p_idx * 2)
                        phase_amt_col = phase_qty_col + 1
                        ws_ws.cell(out_row, phase_qty_col, None)
                        ws_ws.cell(out_row, phase_amt_col, None)
                        ws_ws.cell(out_row, phase_qty_col).fill = phase_fill
                        ws_ws.cell(out_row, phase_amt_col).fill = phase_fill
                    
                    ws_ws.cell(out_row, COL_CURR_QTY, qty_exec)
                    ws_ws.cell(out_row, COL_CURR_AMT, f"={col_letter(COL_CURR_QTY)}{out_row}*{col_letter(COL_EST_RATE)}{out_row}")
                    ws_ws.cell(out_row, COL_MORE, f"=IF({col_letter(COL_CURR_AMT)}{out_row}>{col_letter(COL_EST_AMT)}{out_row},{col_letter(COL_CURR_AMT)}{out_row}-{col_letter(COL_EST_AMT)}{out_row},\"\")")
                    ws_ws.cell(out_row, COL_LESS, f"=IF({col_letter(COL_EST_AMT)}{out_row}>{col_letter(COL_CURR_AMT)}{out_row},{col_letter(COL_EST_AMT)}{out_row}-{col_letter(COL_CURR_AMT)}{out_row},\"\")")
                    ws_ws.cell(out_row, COL_REMARKS, remark_for_item(0, qty_exec, is_supp=True))

                    for cidx in range(1, total_cols + 1):
                        cell = ws_ws.cell(out_row, cidx)
                        cell.border = border_all
                        if cidx in (COL_DESC, COL_REMARKS):
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    out_row += 1
                    sl_counter += 1

            # ---- Sub Total row (over all items) ----
            sub_row = out_row
            ws_ws.cell(sub_row, COL_DESC, "Sub Total Amount")
            ws_ws.cell(sub_row, COL_EST_AMT, f"=SUM({col_letter(COL_EST_AMT)}{data_start}:{col_letter(COL_EST_AMT)}{sub_row-1})")
            
            # Previous phases' subtotals
            for p_idx in range(num_previous_phases):
                phase_amt_col = COL_PHASE_START + (p_idx * 2) + 1  # Amount column for this phase
                phase_amt_letter = col_letter(phase_amt_col)
                ws_ws.cell(sub_row, phase_amt_col, f"=SUM({phase_amt_letter}{data_start}:{phase_amt_letter}{sub_row-1})")
                ws_ws.cell(sub_row, phase_amt_col).fill = phase_fill
            
            ws_ws.cell(sub_row, COL_CURR_AMT, f"=SUM({col_letter(COL_CURR_AMT)}{data_start}:{col_letter(COL_CURR_AMT)}{sub_row-1})")
            # More / Less for Sub Total row
            ws_ws.cell(sub_row, COL_MORE, f"=SUM({col_letter(COL_MORE)}{data_start}:{col_letter(COL_MORE)}{sub_row-1})")
            ws_ws.cell(sub_row, COL_LESS, f"=SUM({col_letter(COL_LESS)}{data_start}:{col_letter(COL_LESS)}{sub_row-1})")

            for col in range(1, total_cols + 1):
                cell = ws_ws.cell(sub_row, col)
                cell.font = Font(bold=True)
                cell.border = border_all
                cell.fill = subtotal_fill
                if col == COL_DESC:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ---- Rows below Sub Total ----
            # Determine if we need a deduct row
            current_row = sub_row
            deduct_row = None
            if ws_deduct_old_material > 0:
                current_row += 1
                deduct_row = current_row
            
            tp_row    = current_row + 1
            sub1_row  = current_row + 2
            lc_row    = current_row + 3
            qc_row    = current_row + 4
            nac_row   = current_row + 5
            sub2_row  = current_row + 6
            gst_row   = current_row + 7
            unused_row= current_row + 8
            ls_row    = current_row + 9
            grand_row = current_row + 10
            
            # Column letters for dynamic formulas
            EST_AMT_COL = col_letter(COL_EST_AMT)
            CURR_AMT_COL = col_letter(COL_CURR_AMT)
            MORE_COL = col_letter(COL_MORE)
            LESS_COL = col_letter(COL_LESS)
            
            # Helper to get phase amount column letter
            def phase_amt_letter(p_idx):
                return col_letter(COL_PHASE_START + (p_idx * 2) + 1)
            
            # Add Deduct Old Material Cost row (if applicable)
            if deduct_row:
                ws_ws.cell(deduct_row, COL_DESC, "Deduct Old Material Cost")
                ws_ws.cell(deduct_row, COL_EST_AMT, round(-ws_deduct_old_material, 2))  # Estimate - negative
                # Previous phases - same deduction
                for p_idx in range(num_previous_phases):
                    phase_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                    ws_ws.cell(deduct_row, phase_amt_col, round(-ws_deduct_old_material, 2))
                    ws_ws.cell(deduct_row, phase_amt_col).fill = phase_fill
                ws_ws.cell(deduct_row, COL_CURR_AMT, round(-ws_deduct_old_material, 2))  # Execution - negative
                ws_ws.cell(deduct_row, COL_MORE, "")  # More
                ws_ws.cell(deduct_row, COL_LESS, "")  # Less

            # i) Add / Deduct T.P
            tp_label_prefix = "Add" if ws_tp_type == "Excess" else "Deduct"
            ws_ws.cell(tp_row, COL_DESC, f"{tp_label_prefix} T.P @ {ws_tp_percent} % {ws_tp_type}")
            ws_ws.cell(tp_row, COL_EST_AMT, None)  # Estimate MUST be empty

            # Previous phases TP
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                if deduct_row:
                    if ws_tp_type == "Excess":
                        ws_ws.cell(tp_row, p_amt_col, f"=({p_amt_letter}{sub_row}+{p_amt_letter}{deduct_row})*{ws_tp_percent}/100")
                    else:
                        ws_ws.cell(tp_row, p_amt_col, f"=-({p_amt_letter}{sub_row}+{p_amt_letter}{deduct_row})*{ws_tp_percent}/100")
                else:
                    if ws_tp_type == "Excess":
                        ws_ws.cell(tp_row, p_amt_col, f"={p_amt_letter}{sub_row}*{ws_tp_percent}/100")
                    else:
                        ws_ws.cell(tp_row, p_amt_col, f"=-{p_amt_letter}{sub_row}*{ws_tp_percent}/100")
                ws_ws.cell(tp_row, p_amt_col).fill = phase_fill

            # Current Amount: positive if Excess, negative if Less
            if deduct_row:
                if ws_tp_type == "Excess":
                    ws_ws.cell(tp_row, COL_CURR_AMT, f"=({CURR_AMT_COL}{sub_row}+{CURR_AMT_COL}{deduct_row})*{ws_tp_percent}/100")
                else:
                    ws_ws.cell(tp_row, COL_CURR_AMT, f"=-({CURR_AMT_COL}{sub_row}+{CURR_AMT_COL}{deduct_row})*{ws_tp_percent}/100")
            else:
                if ws_tp_type == "Excess":
                    ws_ws.cell(tp_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub_row}*{ws_tp_percent}/100")
                else:
                    ws_ws.cell(tp_row, COL_CURR_AMT, f"=-{CURR_AMT_COL}{sub_row}*{ws_tp_percent}/100")

            # More / Less for TP row
            ws_ws.cell(tp_row, COL_MORE, f"=IF({CURR_AMT_COL}{tp_row}>{EST_AMT_COL}{tp_row},{CURR_AMT_COL}{tp_row}-{EST_AMT_COL}{tp_row},\"\")")
            ws_ws.cell(tp_row, COL_LESS, f"=IF({EST_AMT_COL}{tp_row}>{CURR_AMT_COL}{tp_row},{EST_AMT_COL}{tp_row}-{CURR_AMT_COL}{tp_row},\"\")")

            # ii) Sub Total 1 - includes deduction if present
            ws_ws.cell(sub1_row, COL_DESC, "Sub Total 1")
            if deduct_row:
                ws_ws.cell(sub1_row, COL_EST_AMT, f"={EST_AMT_COL}{sub_row}+{EST_AMT_COL}{deduct_row}")
                # Previous phases Sub Total 1
                for p_idx in range(num_previous_phases):
                    p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                    p_amt_letter = phase_amt_letter(p_idx)
                    ws_ws.cell(sub1_row, p_amt_col, f"={p_amt_letter}{sub_row}+{p_amt_letter}{deduct_row}+{p_amt_letter}{tp_row}")
                    ws_ws.cell(sub1_row, p_amt_col).fill = phase_fill
                ws_ws.cell(sub1_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub_row}+{CURR_AMT_COL}{deduct_row}+{CURR_AMT_COL}{tp_row}")
            else:
                ws_ws.cell(sub1_row, COL_EST_AMT, f"={EST_AMT_COL}{sub_row}")
                # Previous phases Sub Total 1
                for p_idx in range(num_previous_phases):
                    p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                    p_amt_letter = phase_amt_letter(p_idx)
                    ws_ws.cell(sub1_row, p_amt_col, f"={p_amt_letter}{sub_row}+{p_amt_letter}{tp_row}")
                    ws_ws.cell(sub1_row, p_amt_col).fill = phase_fill
                ws_ws.cell(sub1_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub_row}+{CURR_AMT_COL}{tp_row}")

            # iii) Add LC @ 1%
            ws_ws.cell(lc_row, COL_DESC, "Add LC @ 1%")
            ws_ws.cell(lc_row, COL_EST_AMT, f"={EST_AMT_COL}{sub1_row}*0.01")
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                ws_ws.cell(lc_row, p_amt_col, f"={p_amt_letter}{sub1_row}*0.01")
                ws_ws.cell(lc_row, p_amt_col).fill = phase_fill
            ws_ws.cell(lc_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub1_row}*0.01")
            ws_ws.cell(lc_row, COL_MORE, f"=IF({CURR_AMT_COL}{lc_row}>{EST_AMT_COL}{lc_row},{CURR_AMT_COL}{lc_row}-{EST_AMT_COL}{lc_row},\"\")")
            ws_ws.cell(lc_row, COL_LESS, f"=IF({EST_AMT_COL}{lc_row}>{CURR_AMT_COL}{lc_row},{EST_AMT_COL}{lc_row}-{CURR_AMT_COL}{lc_row},\"\")")

            # iv) Add QC @ 1%
            ws_ws.cell(qc_row, COL_DESC, "Add QC @ 1%")
            ws_ws.cell(qc_row, COL_EST_AMT, f"={EST_AMT_COL}{sub1_row}*0.01")
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                ws_ws.cell(qc_row, p_amt_col, f"={p_amt_letter}{sub1_row}*0.01")
                ws_ws.cell(qc_row, p_amt_col).fill = phase_fill
            ws_ws.cell(qc_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub1_row}*0.01")
            ws_ws.cell(qc_row, COL_MORE, f"=IF({CURR_AMT_COL}{qc_row}>{EST_AMT_COL}{qc_row},{CURR_AMT_COL}{qc_row}-{EST_AMT_COL}{qc_row},\"\")")
            ws_ws.cell(qc_row, COL_LESS, f"=IF({EST_AMT_COL}{qc_row}>{CURR_AMT_COL}{qc_row},{EST_AMT_COL}{qc_row}-{CURR_AMT_COL}{qc_row},\"\")")

            # v) Add NAC chargers @ 0.1%
            ws_ws.cell(nac_row, COL_DESC, "Add NAC chargers @ 0.1 %")
            ws_ws.cell(nac_row, COL_EST_AMT, f"={EST_AMT_COL}{sub1_row}*0.001")
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                ws_ws.cell(nac_row, p_amt_col, f"={p_amt_letter}{sub1_row}*0.001")
                ws_ws.cell(nac_row, p_amt_col).fill = phase_fill
            ws_ws.cell(nac_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub1_row}*0.001")
            ws_ws.cell(nac_row, COL_MORE, f"=IF({CURR_AMT_COL}{nac_row}>{EST_AMT_COL}{nac_row},{CURR_AMT_COL}{nac_row}-{EST_AMT_COL}{nac_row},\"\")")
            ws_ws.cell(nac_row, COL_LESS, f"=IF({EST_AMT_COL}{nac_row}>{CURR_AMT_COL}{nac_row},{EST_AMT_COL}{nac_row}-{CURR_AMT_COL}{nac_row},\"\")")

            # vi) Sub Total 2
            ws_ws.cell(sub2_row, COL_DESC, "Sub Total 2")
            ws_ws.cell(sub2_row, COL_EST_AMT, f"={EST_AMT_COL}{sub1_row}+{EST_AMT_COL}{lc_row}+{EST_AMT_COL}{qc_row}+{EST_AMT_COL}{nac_row}")
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                ws_ws.cell(sub2_row, p_amt_col, f"={p_amt_letter}{sub1_row}+{p_amt_letter}{lc_row}+{p_amt_letter}{qc_row}+{p_amt_letter}{nac_row}")
                ws_ws.cell(sub2_row, p_amt_col).fill = phase_fill
            ws_ws.cell(sub2_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub1_row}+{CURR_AMT_COL}{lc_row}+{CURR_AMT_COL}{qc_row}+{CURR_AMT_COL}{nac_row}")
            # (NO More/Less formulas in Sub Total 2 as per requirement)

            # vii) Add GST @ 18%
            ws_ws.cell(gst_row, COL_DESC, "Add GST @ 18%")
            ws_ws.cell(gst_row, COL_EST_AMT, f"={EST_AMT_COL}{sub2_row}*0.18")
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                ws_ws.cell(gst_row, p_amt_col, f"={p_amt_letter}{sub2_row}*0.18")
                ws_ws.cell(gst_row, p_amt_col).fill = phase_fill
            ws_ws.cell(gst_row, COL_CURR_AMT, f"={CURR_AMT_COL}{sub2_row}*0.18")
            ws_ws.cell(gst_row, COL_MORE, f"=IF({CURR_AMT_COL}{gst_row}>{EST_AMT_COL}{gst_row},{CURR_AMT_COL}{gst_row}-{EST_AMT_COL}{gst_row},\"\")")
            ws_ws.cell(gst_row, COL_LESS, f"=IF({EST_AMT_COL}{gst_row}>{CURR_AMT_COL}{gst_row},{EST_AMT_COL}{gst_row}-{CURR_AMT_COL}{gst_row},\"\")")

            # viii) Unused T.P @ % on ECV (Estimate empty, Execution uses Estimate of Sub Total row)
            ws_ws.cell(unused_row, COL_DESC, f"Unused T.P @ {ws_tp_percent} % on ECV")
            ws_ws.cell(unused_row, COL_EST_AMT, None)   # Estimate MUST be empty
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                ws_ws.cell(unused_row, p_amt_col, f"={EST_AMT_COL}{sub_row}*{ws_tp_percent}/100")
                ws_ws.cell(unused_row, p_amt_col).fill = phase_fill
            ws_ws.cell(unused_row, COL_CURR_AMT, f"={EST_AMT_COL}{sub_row}*{ws_tp_percent}/100")
            ws_ws.cell(unused_row, COL_MORE, f"=IF({CURR_AMT_COL}{unused_row}>{EST_AMT_COL}{unused_row},{CURR_AMT_COL}{unused_row}-{EST_AMT_COL}{unused_row},\"\")")
            ws_ws.cell(unused_row, COL_LESS, f"=IF({EST_AMT_COL}{unused_row}>{CURR_AMT_COL}{unused_row},{EST_AMT_COL}{unused_row}-{CURR_AMT_COL}{unused_row},\"\")")

            # ix) L.S. provision row
            ws_ws.cell(ls_row, COL_DESC, "L.S provision towards unforeseen items")
            ws_ws.cell(ls_row, COL_EST_AMT, f"={EST_AMT_COL}{grand_row}-{EST_AMT_COL}{unused_row}-{EST_AMT_COL}{gst_row}-{EST_AMT_COL}{sub2_row}")
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                p_amt_letter = phase_amt_letter(p_idx)
                ws_ws.cell(ls_row, p_amt_col, f"={p_amt_letter}{grand_row}-{p_amt_letter}{unused_row}-{p_amt_letter}{gst_row}-{p_amt_letter}{sub2_row}")
                ws_ws.cell(ls_row, p_amt_col).fill = phase_fill
            ws_ws.cell(ls_row, COL_CURR_AMT, f"={CURR_AMT_COL}{grand_row}-{CURR_AMT_COL}{unused_row}-{CURR_AMT_COL}{gst_row}-{CURR_AMT_COL}{sub2_row}")
            ws_ws.cell(ls_row, COL_MORE, f"=IF({CURR_AMT_COL}{ls_row}>{EST_AMT_COL}{ls_row},{CURR_AMT_COL}{ls_row}-{EST_AMT_COL}{ls_row},\"\")")
            ws_ws.cell(ls_row, COL_LESS, f"=IF({EST_AMT_COL}{ls_row}>{CURR_AMT_COL}{ls_row},{EST_AMT_COL}{ls_row}-{CURR_AMT_COL}{ls_row},\"\")")

            # x) Grand Total = Grand Total of uploaded Estimate (both Estimate & Execution same)
            grand_total_val = round(float(request.session.get("ws_estimate_grand_total", 0.0) or 0.0), 2)
            ws_ws.cell(grand_row, COL_DESC, "Grand Total")
            ws_ws.cell(grand_row, COL_EST_AMT, grand_total_val)
            for p_idx in range(num_previous_phases):
                p_amt_col = COL_PHASE_START + (p_idx * 2) + 1
                ws_ws.cell(grand_row, p_amt_col, grand_total_val)
                ws_ws.cell(grand_row, p_amt_col).fill = phase_fill
            ws_ws.cell(grand_row, COL_CURR_AMT, grand_total_val)
            # More / Less in Grand Total row = sum of Sub Total â†’ LS rows
            ws_ws.cell(grand_row, COL_MORE, f"=SUM({MORE_COL}{sub_row}:{MORE_COL}{ls_row})")
            ws_ws.cell(grand_row, COL_LESS, f"=SUM({LESS_COL}{sub_row}:{LESS_COL}{ls_row})")

            # style all total rows
            rows_to_style = [tp_row, sub1_row, lc_row, qc_row, nac_row,
                        sub2_row, gst_row, unused_row, ls_row, grand_row]
            if deduct_row:
                rows_to_style.insert(0, deduct_row)  # Add deduct row at the beginning
            for r_i in rows_to_style:
                for col in range(1, total_cols + 1):
                    cell = ws_ws.cell(r_i, col)
                    cell.font = Font(bold=True)
                    cell.border = border_all
                    cell.fill = subtotal_fill
                    if col == COL_DESC:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # reset row heights auto
            for r in range(1, ws_ws.max_row + 1):
                ws_ws.row_dimensions[r].height = None

            # Apply number_format to numeric columns - use #,##0.00 for amounts/rates, #,##0.## for quantities
            fmt_money = '#,##0.00'
            fmt_qty = '#,##0.##'
            amt_cols = {COL_EST_RATE, COL_EST_AMT, COL_CURR_AMT, COL_MORE, COL_LESS}
            qty_cols = {COL_EST_QTY, COL_CURR_QTY}
            # Add phase columns
            for p_idx in range(num_previous_phases):
                qty_cols.add(COL_PHASE_START + (p_idx * 2))      # phase qty
                amt_cols.add(COL_PHASE_START + (p_idx * 2) + 1)  # phase amt
            for r in range(data_start, ws_ws.max_row + 1):
                for c in amt_cols:
                    ws_ws.cell(r, c).number_format = fmt_money
                for c in qty_cols:
                    ws_ws.cell(r, c).number_format = fmt_qty

            # Reorder sheets: WorkSlip first, then ItemBlocks
            if "WorkSlip" in wb_out.sheetnames:
                ws_idx = wb_out.sheetnames.index("WorkSlip")
                if ws_idx > 0:
                    wb_out.move_sheet("WorkSlip", offset=-ws_idx)

            # Apply print settings: Landscape, A4, fit columns, Times New Roman
            _apply_print_settings(wb_out, landscape=True)

            # return file
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="WorkSlip.xlsx"'
            wb_out.save(resp)
            return resp

    # ------------- 2) Build preview_rows for UI (GET or redirect) -------------
    preview_rows = []
    
    # DEBUG: Log session state on GET
    logger.info(f"[WORKSLIP DEBUG] GET - ws_estimate_rows from session: {len(ws_estimate_rows)} rows")
    if ws_estimate_rows and len(ws_estimate_rows) > 0:
        first_row = ws_estimate_rows[0]
        logger.info(f"[WORKSLIP DEBUG] First row structure: {first_row}")
        logger.info(f"[WORKSLIP DEBUG] First row qty_est={first_row.get('qty_est')}, rate={first_row.get('rate')}")

    # base items
    for idx, row in enumerate(ws_estimate_rows, start=1):
        row_key = row["key"]
        # try same candidate keys as in download
        candidates = [
            f"base:{row_key}",
            row_key,
            row.get("item_name") or "",
            row.get("desc") or "",
        ]
        qty_preview = ""
        for k in candidates:
            k = str(k).strip()
            if k and k in ws_exec_map:
                qty_preview = ws_exec_map[k]
                break

        preview_rows.append({
            "row_type": "base",
            "key": row_key,
            "sl": idx,
            # ðŸ‘‡ UI shows item NAME (yellow header) if present
            "desc": row.get("display_name") or row.get("item_name") or row.get("desc"),
            "qty_est": row.get("qty_est", 0),
            "unit": row.get("unit", ""),
            "rate": row.get("rate", 0),
            "qty_exec": qty_preview,
        })

    # supplemental preview with rates from backend
    item_to_info = {it["name"]: it for it in items_list}
    wb_vals = None
    ws_vals = None
    if filepath and os.path.exists(filepath):
        try:
            wb_vals = load_workbook(filepath, data_only=True)
            ws_vals = wb_vals["Master Datas"]
        except Exception:
            wb_vals = None
            ws_vals = None

    supp_details = []
    if ws_vals is not None:
        for name in ws_supp_items:
            info = item_to_info.get(name)
            if not info:
                continue
            start_row = info["start_row"]
            end_row = info["end_row"]
            rate_val = 0.0
            for r in range(end_row, start_row - 1, -1):
                v = ws_vals.cell(row=r, column=10).value
                if v not in (None, ""):
                    try:
                        rate_val = float(v)
                    except Exception:
                        rate_val = 0.0
                    break
            unit_pl, _ = units_for(name)
            key = f"supp:{name}"
            supp_details.append({
                "name": name,
                "unit": unit_pl,
                "rate": rate_val,
                "qty_exec": ws_exec_map.get(key, ""),
            })

    # Add Previous Workslip Supplemental Items BEFORE current supplemental items
    if ws_previous_supp_items:
        # Group items by supplemental section AND consolidate quantities by workslip phase
        # Structure: { supp_section: { item_name: { phase1_qty, phase2_qty, ... } } }
        supp_by_section = {}
        
        for supp in ws_previous_supp_items:
            section_num = supp.get("supp_section", supp.get("phase", 1))  # Fallback to phase for backward compat
            item_name = supp.get("name", "")
            phase_num = supp.get("phase", 1)
            
            if section_num not in supp_by_section:
                supp_by_section[section_num] = {}
            
            if item_name not in supp_by_section[section_num]:
                supp_by_section[section_num][item_name] = {
                    "name": item_name,
                    "desc": supp.get("desc", item_name),
                    "unit": supp.get("unit", "-") or "-",
                    "rate": supp.get("rate", 0) or 0,
                    "phase_quantities": {},  # { phase_num: qty }
                    "phase_amounts": {},     # { phase_num: amount }
                }
            
            # Add this phase's quantity to the item
            supp_by_section[section_num][item_name]["phase_quantities"][phase_num] = supp.get("qty", 0)
            supp_by_section[section_num][item_name]["phase_amounts"][phase_num] = supp.get("amount", 0)
            # Update rate if not set
            if supp_by_section[section_num][item_name]["rate"] == 0:
                supp_by_section[section_num][item_name]["rate"] = supp.get("rate", 0) or 0
        
        base_count = len(preview_rows)
        preview_rows.append({
            "row_type": "heading",
            "label": "Previous Workslip Supplemental Items",
        })
        
        for section_num in sorted(supp_by_section.keys()):
            section_items = supp_by_section[section_num]
            preview_rows.append({
                "row_type": "heading",
                "label": f"Supplemental Items-{section_num}",
            })
            for i, (item_name, item_data) in enumerate(section_items.items(), start=1):
                supp_key = f"prev_supp:{section_num}:{item_name}"
                
                # Build previous phases execution data for display
                # Must include ALL phases to match the table columns
                num_prev_phases = len(ws_previous_phases) if ws_previous_phases else 0
                prev_phases_exec = []
                for phase_idx in range(1, num_prev_phases + 1):
                    prev_phases_exec.append({
                        "phase": phase_idx,
                        "qty": item_data["phase_quantities"].get(phase_idx, 0),
                        "amount": item_data["phase_amounts"].get(phase_idx, 0),
                    })
                
                # Apply custom rate from ws_rate_map if user modified it
                prev_supp_rate = item_data["rate"]
                if supp_key in ws_rate_map:
                    prev_supp_rate = ws_rate_map[supp_key]
                
                prev_supp_row = {
                    "row_type": "prev_supp",
                    "key": supp_key,
                    "name": item_data["name"],
                    "sl": base_count + i,
                    "desc": item_data["name"],
                    "qty_est": "-",
                    "unit": item_data["unit"],
                    "rate": prev_supp_rate,
                    "qty_exec": ws_exec_map.get(supp_key, ""),
                    "supp_section": section_num,
                    "previous_phases_exec": prev_phases_exec,  # List of {phase, qty, amount}
                }
                preview_rows.append(prev_supp_row)
                base_count += 1

    # Add Current Phase Supplemental Items
    if supp_details:
        base_count = len(preview_rows)
        preview_rows.append({
            "row_type": "heading",
            "label": f"Supplemental Items{'-' + str(ws_current_phase) if ws_current_phase > 1 else ''}",
        })
        for i, s in enumerate(supp_details, start=1):
            key = f"supp:{s['name']}"
            preview_rows.append({
                "row_type": "supp",
                "key": key,
                "name": s["name"],
                "sl": base_count + i,
                "desc": s["name"],      # UI shows supplemental item name
                "qty_est": 0,
                "unit": s["unit"],
                "rate": s["rate"],
                "qty_exec": ws_exec_map.get(key, ""),
            })

    tp_percent_str = "" if not ws_tp_percent else str(ws_tp_percent)
    ws_deduct_old_material = request.session.get("ws_deduct_old_material", 0.0) or 0.0
    ws_lc_percent = request.session.get("ws_lc_percent", 0.0) or 0.0
    ws_qc_percent = request.session.get("ws_qc_percent", 0.0) or 0.0
    ws_nac_percent = request.session.get("ws_nac_percent", 0.0) or 0.0
    ws_metadata = request.session.get("ws_metadata", {}) or {}
    
    # Add phase data to preview_rows for display (skip prev_supp rows which already have their data)
    for row in preview_rows:
        if row.get("row_type") not in ("heading", "prev_supp"):
            row_key = row.get("key", "")
            # Add previous phases' execution data
            row["previous_phases_exec"] = []
            for phase_idx, phase_map in enumerate(ws_previous_phases):
                phase_exec = phase_map.get(row_key, 0)
                row["previous_phases_exec"].append({
                    "phase": phase_idx + 1,
                    "qty": phase_exec,
                    "amount": phase_exec * row.get("rate", 0) if phase_exec else 0
                })

    return render(request, "core/workslip.html", {
        "category": category,
        "work_type": ws_work_type,
        "work_type_display": {
            'new_estimate': 'New Estimate',
            'amc': 'AMC',
            'tempworks': 'Temporary Works'
        }.get(ws_work_type, 'New Estimate'),
        "work_mode": ws_work_mode,
        "work_mode_display": {
            'original': 'Original',
            'repair': 'Repair'
        }.get(ws_work_mode, 'Original'),
        "category_display": ws_category.title() if ws_category else 'Electrical',
        "module_code": module_code,
        "groups": groups,
        "current_group": current_group,
        "items_in_group": items_in_group, "items_info": items_info,
        "ws_estimate_rows": ws_estimate_rows,
        "preview_rows": preview_rows,
        "tp_percent": tp_percent_str,
        "tp_type": ws_tp_type,
        "supp_items_selected": ws_supp_items,
        "work_name": ws_work_name,
        "deduct_old_material": ws_deduct_old_material if ws_deduct_old_material > 0 else "",
        "lc_percent": ws_lc_percent if ws_lc_percent > 0 else "",
        "qc_percent": ws_qc_percent if ws_qc_percent > 0 else "",
        "nac_percent": ws_nac_percent if ws_nac_percent > 0 else "",
        "current_phase": ws_current_phase,
        "target_workslip": request.session.get("ws_target_workslip", 1) or 1,
        "previous_phases": ws_previous_phases,
        "total_phases": len(ws_previous_phases) if ws_previous_phases else 0,
        "previous_supp_items": ws_previous_supp_items,
        "ws_metadata": ws_metadata,
        "available_backends": available_backends,
        "selected_backend_id": ws_selected_backend_id,
    })


# -----------------------
# WORKSLIP AJAX TOGGLE SUPPLEMENTAL ITEM
# -----------------------
@login_required(login_url='login')
def workslip_ajax_toggle_supp(request):
    """
    AJAX endpoint to toggle a supplemental item in workslip without page reload.
    POST with JSON: { "item": "item_name", "action": "add" or "remove" }
    Returns JSON: { "status": "ok", "supp_items": [...], "action_taken": "added" or "removed", "item_info": {...} }
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST required"}, status=405)
    
    try:
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body.decode("utf-8") or "{}")
        else:
            data = request.POST
        
        item = data.get("item", "").strip()
        action = data.get("action", "toggle")  # "add", "remove", or "toggle"
        
        if not item:
            return JsonResponse({"status": "error", "message": "No item specified"}, status=400)
        
        ws_supp_items = request.session.get("ws_supp_items", []) or []
        action_taken = None
        
        if action == "add":
            if item not in ws_supp_items:
                ws_supp_items.append(item)
                action_taken = "added"
            else:
                action_taken = "already_exists"
        elif action == "remove":
            if item in ws_supp_items:
                ws_supp_items.remove(item)
                action_taken = "removed"
                # Also remove from exec_map
                ws_exec_map = request.session.get("ws_exec_map", {}) or {}
                key = f"supp:{item}"
                if key in ws_exec_map:
                    del ws_exec_map[key]
                request.session["ws_exec_map"] = ws_exec_map
            else:
                action_taken = "not_found"
        else:  # toggle
            if item in ws_supp_items:
                ws_supp_items.remove(item)
                action_taken = "removed"
                # Also remove from exec_map
                ws_exec_map = request.session.get("ws_exec_map", {}) or {}
                key = f"supp:{item}"
                if key in ws_exec_map:
                    del ws_exec_map[key]
                request.session["ws_exec_map"] = ws_exec_map
            else:
                ws_supp_items.append(item)
                action_taken = "added"
        
        request.session["ws_supp_items"] = ws_supp_items
        
        # Get item info (rate, unit) for newly added items
        item_info = None
        if action_taken == "added":
            try:
                # Load backend for the default category
                category = "original"
                items_list, groups_map, units_map, ws_data, filepath = load_backend(category, settings.BASE_DIR)
                
                # Get rate
                wb_vals = load_workbook(filepath, data_only=True)
                ws_vals = wb_vals["Master Datas"]
                
                item_rate = None
                for info in items_list:
                    if info["name"] == item:
                        start_row = info["start_row"]
                        end_row = info["end_row"]
                        for r in range(end_row, start_row - 1, -1):
                            val = ws_vals.cell(row=r, column=10).value  # column J
                            if val not in (None, ""):
                                item_rate = val
                                break
                        break
                
                # Get unit from backend units_map (Column D of Groups sheet)
                unit = units_map.get(item, "Nos")
                
                item_info = {
                    "name": item,
                    "rate": item_rate,
                    "unit": unit
                }
                
                wb_vals.close()
            except Exception as e:
                # If we can't get item info, just return without it
                item_info = {"name": item, "rate": None, "unit": "Nos"}
        
        return JsonResponse({
            "status": "ok",
            "supp_items": ws_supp_items,
            "action_taken": action_taken,
            "item": item,
            "item_info": item_info
        })
        
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)


# Provide placeholder views for any names expected by URLConf but not yet
# implemented during the ongoing refactor. This uses module-level
# __getattr__ (PEP 562) so imports like `from core import views` succeed.
_placeholder_views = {}

