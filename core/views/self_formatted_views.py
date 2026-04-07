# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import get_org_from_request

def _extract_value_part_from_line(s: str) -> str:
    """
    For a line like 'Name of the work : Construction of XYZ',
    returns 'Construction of XYZ'.
    Also handles OCR artifacts like leading numbers/asterisks.
    """
    s = (s or "").strip()
    # Remove leading line numbers like "1 ", "4 ", etc.
    s = re.sub(r'^\d+\s+', '', s)
    
    if ":" in s:
        val = s.split(":", 1)[1].strip()
        # Clean OCR artifacts
        val = val.lstrip('*').strip()
        return val
    if " - " in s:
        return s.split(" - ", 1)[1].strip()
    return ""


def _collect_multiline_value(lines, start_idx, max_lines=4):
    """
    Collect value that spans multiple lines (common in OCR output).
    Stops when hitting a new numbered line or label pattern.
    """
    collected = []
    for i in range(start_idx, min(start_idx + max_lines, len(lines))):
        line = str(lines[i]).strip()
        if not line:
            continue
        # Stop if we hit a new numbered item (like "2 Ref to...")
        if re.match(r'^\d+\s+[A-Z]', line):
            break
        # Stop if we hit another label with colon
        if re.search(r'^[A-Za-z\s]+\s*:\s*$', line):
            break
        collected.append(line)
    return " ".join(collected).strip()


def _extract_labels_from_lines(lines):
    """
    Fuzzy-read important labels from list of text lines.
    Handles OCR output from scanned documents like Forwarding Slips, Estimates, Bills.
    """
    labels = {
        "name_of_work": "",
        "agreement": "",
        "admin_sanction": "",
        "admin_sanction_amount": "",
        "tech_sanction": "",
        "tech_sanction_amount": "",
        "agency": "",
        "contractor": "",
        "contractor_address": "",
        "mb_details": "",
        "tp_details": "",
        "cc_header": "",
        "amount": "",
        "estimate_amount": "",
        "bond_no": "",
        "nit_no": "",
        "tender_premium": "",
        "period_of_completion": "",
        "earnest_money": "",
        "security_deposit": "",
        "work_order_no": "",
        "work_order_date": "",
        "date_of_commencement": "",
        "date_of_completion": "",
        "doi": "",
        "doc": "",
        "domr": "",
        "dobr": "",
        "amount_paid": "",
    }

    for idx, raw in enumerate(lines):
        s = str(raw or "").strip()
        if not s:
            continue
        low = s.lower()

        # i. Name of the work (various formats)
        if not labels["name_of_work"]:
            # Standard patterns - "Name of Work:" or "Name of the Work:"
            if "name of work" in low or "name of the work" in low:
                val = _extract_value_part_from_line(s)
                # If value is empty or too short, collect from next lines
                if not val or len(val) < 5:
                    val = _collect_multiline_value(lines, idx + 1, max_lines=5)
                if val:
                    # Clean up OCR artifacts like "ast :"
                    val = re.sub(r'^[a-z]+\s*:\s*', '', val, flags=re.I)
                    labels["name_of_work"] = val
                    continue
            
            # Look for "for the work of" pattern (common in Forwarding Slips)
            if "for the work of" in low:
                # Extract text after "for the work of"
                match = re.search(r'for the work of\s*["\']?\s*(.+?)(?:["\']?\s*duly|$)', s, re.I)
                if match:
                    val = match.group(1).strip().strip('"\'')
                    if val and len(val) > 5:
                        labels["name_of_work"] = val
                        continue
                # Value might be empty here (""), collect from next lines
                if idx + 1 < len(lines):
                    val = _collect_multiline_value(lines, idx + 1, max_lines=3)
                    if val and len(val) > 10:
                        labels["name_of_work"] = val
                        continue
            
            # Look for work description keywords (maintenance, servicing, generator, etc.)
            work_keywords = ['maintenance', 'servicing', 'generator', 'diesel', 'kva', 'annual']
            if any(kw in low for kw in work_keywords) and len(s) > 30:
                # Skip if line starts with fragment like "rk:" or is too short
                if not re.match(r'^[a-z]{1,3}:', low) and not low.startswith('rk'):
                    # Collect this and following lines
                    work_parts = [s]
                    for j in range(idx + 1, min(idx + 5, len(lines))):
                        next_line = str(lines[j]).strip()
                        if not next_line or len(next_line) < 3:
                            continue
                        next_low = next_line.lower()
                        if any(x in next_low for x in ['sanction', 'contractor', 'nit', 'accord', 'permit', 'memo']):
                            break
                        if len(next_line) > 5:
                            work_parts.append(next_line)
                    full_work = " ".join(work_parts)
                    # Add "Providing A" prefix if work starts with "nual" (fragment of "Annual")
                    if full_work.lower().startswith('nual'):
                        full_work = "Providing A" + full_work
                    labels["name_of_work"] = full_work
                    continue
            
            # Look for "Providing" or "Supply" at start of line (common work name patterns)
            # Also handle OCR fragments like "viding" (from "Providing")
            if (low.startswith("providing") or low.startswith("supply") or low.startswith("construction") or
                low.startswith("viding") or "viding annual" in low or "viding " in low):
                # Skip fragments like "rk: Providing"
                if re.match(r'^[a-z]{1,3}:', low):
                    continue
                # This line and following lines might be the work name - collect them
                if len(s) > 10:
                    # Collect multiple lines for fragmented OCR
                    work_parts = [s]
                    for j in range(idx + 1, min(idx + 8, len(lines))):
                        next_line = str(lines[j]).strip()
                        if not next_line or len(next_line) < 3:
                            continue
                        next_low = next_line.lower()
                        # Stop if we hit a new section
                        if any(x in next_low for x in ['sanction', 'contractor', 'nit', 'tender', 'amount', 'period']):
                            break
                        # Stop if line looks like a label
                        if ':' in next_line and len(next_line.split(':')[0]) < 25:
                            break
                        work_parts.append(next_line)
                    # Join and clean up the work name
                    full_work = " ".join(work_parts)
                    # Add "Pro" prefix if it starts with "viding"
                    if full_work.lower().startswith("viding"):
                        full_work = "Pro" + full_work
                    labels["name_of_work"] = full_work
                    continue
            
            # Check if previous line was just "Name of Work" label (table format)
            if idx > 0:
                prev_line = str(lines[idx - 1]).strip().lower()
                if prev_line in ("name of work", "name of the work", "name of work:"):
                    if len(s) > 10 and not any(x in low for x in ['sanction', 'amount', 'contractor', 'nit']):
                        labels["name_of_work"] = s
                        continue
                    continue

        # Bond No / Agreement (from header like "Suppl.Agreement Bond No. 30/2024-2025")
        if not labels["agreement"]:
            if "bond no" in low or "agreement bond" in low or "agreement" in low:
                # Always prefer full text after colon (most reliable)
                val = _extract_value_part_from_line(s)
                if val:
                    labels["agreement"] = val
                    if len(val) < 20 and idx + 1 < len(lines):
                        next_val = _collect_multiline_value(lines, idx + 1, max_lines=2)
                        if next_val:
                            labels["agreement"] = val + " " + next_val
                    continue
                elif idx + 1 < len(lines):
                    next_val = _collect_multiline_value(lines, idx + 1, max_lines=3)
                    if next_val:
                        labels["agreement"] = next_val
                        continue

        # Admin Sanction / Sanctioned Estimate (TA.No pattern or Ref. to Administrative sanction)
        if not labels["admin_sanction"]:
            if ("administrative sanction" in low or "admin sanction" in low or
                ("a)" in low and "sanction" in low) or
                ("ref" in low and "admin" in low)):
                # Always prefer full text after colon (most reliable)
                val = _extract_value_part_from_line(s)
                if val:
                    labels["admin_sanction"] = val
                    # If value seems incomplete, collect from next lines
                    if len(val) < 20 and idx + 1 < len(lines):
                        next_val = _collect_multiline_value(lines, idx + 1, max_lines=2)
                        if next_val:
                            labels["admin_sanction"] = val + " " + next_val
                elif idx + 1 < len(lines):
                    # No colon on this line — value is on next line
                    next_val = _collect_multiline_value(lines, idx + 1, max_lines=3)
                    if next_val:
                        labels["admin_sanction"] = next_val
                # Also try to get amount if present on this or next line
                amount_match = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if amount_match and not labels["admin_sanction_amount"]:
                    labels["admin_sanction_amount"] = "Rs." + amount_match[-1] + "/-"
                continue
        
        # Check for "b)Administrative sanction Amount" or "Technical Sanction Amount" patterns
        if not labels["admin_sanction_amount"]:
            if ("admin" in low or "a)" in low) and "sanction" in low and "amount" in low:
                amount_match = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if amount_match:
                    labels["admin_sanction_amount"] = "Rs." + amount_match[-1] + "/-"
                else:
                    # Value might be on next line
                    if idx + 1 < len(lines):
                        next_line = str(lines[idx + 1]).strip()
                        amt_match = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', next_line, re.I)
                        if amt_match:
                            labels["admin_sanction_amount"] = "Rs." + amt_match[-1] + "/-"
                continue
        
        # Tech Sanction Amount specifically
        if not labels["tech_sanction_amount"]:
            if ("tech" in low or "b)" in low) and "sanction" in low and "amount" in low:
                amount_match = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if amount_match:
                    labels["tech_sanction_amount"] = "Rs." + amount_match[-1] + "/-"
                else:
                    # Value might be on next line
                    if idx + 1 < len(lines):
                        next_line = str(lines[idx + 1]).strip()
                        amt_match = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', next_line, re.I)
                        if amt_match:
                            labels["tech_sanction_amount"] = "Rs." + amt_match[-1] + "/-"
                continue

        # Tech Sanction - handle DR.NO pattern and value on next line
        if not labels["tech_sanction"]:
            if ("tech" in low and "sanc" in low) or "b)technical" in low or "technical sanction" in low:
                # Always prefer full text after colon
                val = _extract_value_part_from_line(s)
                if val:
                    labels["tech_sanction"] = val
                    if len(val) < 20 and idx + 1 < len(lines):
                        next_val = _collect_multiline_value(lines, idx + 1, max_lines=2)
                        if next_val:
                            labels["tech_sanction"] = val + " " + next_val
                elif idx + 1 < len(lines):
                    next_val = _collect_multiline_value(lines, idx + 1, max_lines=3)
                    if next_val:
                        labels["tech_sanction"] = next_val
                # Get amount if present
                amount_match = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if amount_match and not labels["tech_sanction_amount"]:
                    labels["tech_sanction_amount"] = "Rs." + amount_match[-1] + "/-"
                continue

        # Contractor / Agency - treat as same thing
        # "Name of the Contractor" = "Name of the Agency"
        if not labels["agency"]:
            if "contractor" in low or "agency" in low or "name of the contractor" in low:
                # Look for M/s. pattern first (common for contractor names) - case insensitive
                match = re.search(r'M/s\.?\s*([A-Za-z][A-Za-z0-9\s\.\&\-]+)', s, re.I)
                if match:
                    labels["agency"] = "M/s." + match.group(1).strip()
                    continue
                # Otherwise use value after colon
                val = _extract_value_part_from_line(s)
                if val and len(val) > 2 and not val.lower().startswith('yes') and not val.lower().startswith('no'):
                    # Skip if value looks like an address (starts with number pattern like 1-6-620)
                    if re.match(r'^\d+[\-\/]\d+', val):
                        # This is an address, not contractor name
                        labels["contractor_address"] = val
                        continue
                    # Check if val itself contains M/s pattern
                    ms_match = re.search(r'(?:M/s|Mis|M\.s)\.?\s*(.+)', val, re.I)
                    if ms_match:
                        labels["agency"] = "M/s." + ms_match.group(1).strip()
                    else:
                        labels["agency"] = val
                    continue
                # Value might be on next line
                if idx + 1 < len(lines):
                    next_line = str(lines[idx + 1]).strip()
                    # Skip if next line looks like address
                    if re.match(r'^\d+[\-\/]\d+', next_line):
                        labels["contractor_address"] = next_line
                        continue
                    match = re.search(r'(?:M/s|Mis|M\.s)\.?\s*([A-Za-z][A-Za-z0-9\s\.\&\-]+)', next_line, re.I)
                    if match:
                        labels["agency"] = "M/s." + match.group(1).strip()
                    if match:
                        labels["agency"] = "M/s." + match.group(1).strip()
                        continue
                    elif next_line and len(next_line) > 5:
                        labels["agency"] = next_line
                        continue
        
        # Also detect standalone M/s. lines (contractor name without label)
        if not labels["agency"]:
            # Match M/s patterns including OCR variations like "M/s.", "Mis.", "M/S", etc.
            if low.startswith("m/s") or low.startswith("m/s.") or low.startswith("mis.") or low.startswith("m.s"):
                # This line is likely a contractor name
                match = re.search(r'(?:M/s|Mis|M\.s)\.?\s*(.+)', s, re.I)
                if match:
                    labels["agency"] = "M/s." + match.group(1).strip()
                    continue

        # Also check specifically for "Name of the Contractor" with value on same line
        if not labels["agency"]:
            if "name of the contractor" in low or "name of contractor" in low or "contractor name" in low:
                val = _extract_value_part_from_line(s)
                if val:
                    ms_match = re.search(r'(?:M/s|Mis|M\.s)\.?\s*(.+)', val, re.I)
                    if ms_match:
                        labels["agency"] = "M/s." + ms_match.group(1).strip()
                    else:
                        labels["agency"] = val
                    continue
        
        # Contractor Address - separate field (look for address patterns)
        if not labels["contractor_address"]:
            # Detect address-like patterns: "1-6-620/1 Near..." or contains locality names
            if "contractor" in low and "address" in low:
                val = _extract_value_part_from_line(s)
                if val:
                    labels["contractor_address"] = val
                    continue
            # Standalone address line pattern (e.g., "1-6-620/1 Near Ramalayam...")
            address_match = re.match(r'^(\d+[\-\/]\d+[\-\/]?\d*\s+(?:Near|Opp|Behind|At|H\.?No)?\s*.+)', s, re.I)
            if address_match and not labels["contractor_address"]:
                # This looks like an address - but only capture if we already have contractor
                if labels["agency"]:
                    labels["contractor_address"] = s
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["contractor_address"] = val
                    # Address often spans multiple lines - collect more
                    if idx + 1 < len(lines):
                        addr_parts = [val]
                        for j in range(idx + 1, min(idx + 4, len(lines))):
                            addr_line = str(lines[j]).strip()
                            # Stop if we hit a new label/section
                            if ':' in addr_line[:20] or re.match(r'^\\d+[\\)\\.]', addr_line):
                                break
                            if any(x in addr_line.lower() for x in ['estimate', 'period', 'sanction', 'nit']):
                                break
                            if addr_line and len(addr_line) > 3:
                                addr_parts.append(addr_line)
                        labels["contractor_address"] = " ".join(addr_parts)
                    continue
                # Value might be on next line
                if idx + 1 < len(lines):
                    next_line = str(lines[idx + 1]).strip()
                    if next_line and len(next_line) > 10:
                        labels["contractor_address"] = next_line
                        # Collect more address lines
                        addr_parts = [next_line]
                        for j in range(idx + 2, min(idx + 5, len(lines))):
                            addr_line = str(lines[j]).strip()
                            if ':' in addr_line[:20] or re.match(r'^\\d+[\\)\\.]', addr_line):
                                break
                            if any(x in addr_line.lower() for x in ['estimate', 'period', 'sanction', 'nit']):
                                break
                            if addr_line and len(addr_line) > 3:
                                addr_parts.append(addr_line)
                        labels["contractor_address"] = " ".join(addr_parts)
                        continue

        # Amount of Estimate / Estimate Amount (value often on next line like "*2,04,798/-" or "Rs. 102820.48")
        if not labels["estimate_amount"]:
            if "amount of estimate" in low or "estimate amount" in low or ("estimate" in low and "amount" in low):
                val = _extract_value_part_from_line(s)
                # Check for Rs. pattern in the value
                if val:
                    rs_match = re.search(r'Rs\\.?\\s*([\\d,]+(?:\\.\\d+)?)', val, re.I)
                    if rs_match:
                        labels["estimate_amount"] = "Rs. " + rs_match.group(1)
                    else:
                        labels["estimate_amount"] = val.lstrip('*').rstrip('/-').strip()
                    continue
                if not val and idx + 1 < len(lines):
                    # Value is on next line - extract number
                    next_line = str(lines[idx + 1]).strip()
                    # Match patterns like "Rs. 102820.48" or "*2,04,798/-" or "2,04,798"
                    rs_match = re.search(r'Rs\\.?\\s*([\\d,]+(?:\\.\\d+)?)', next_line, re.I)
                    if rs_match:
                        labels["estimate_amount"] = "Rs. " + rs_match.group(1)
                        continue
                    match = re.search(r'[\\*]?([\\d,]+(?:\\.\\d+)?)', next_line)
                    if match:
                        val = match.group(1)
                        labels["estimate_amount"] = "Rs. " + val.lstrip('*').rstrip('/-').strip()
                        continue

        # MB Details - Measurement Book details
        if not labels["mb_details"]:
            # Various MB patterns
            if "mb" in low and ("no" in low or "details" in low or "page" in low or "sl" in low):
                val = _extract_value_part_from_line(s)
                if val:
                    labels["mb_details"] = val
                    continue
            # Look for patterns like "MB No. 123, Page 45-67"
            if "mb" in low:
                match = re.search(r'MB\s*(?:No\.?|No)?\s*:?\s*(\d+)[,\s]*(?:Page|Pg\.?|P)?\s*(?:No\.?)?\s*:?\s*(\d+(?:\s*[-to]+\s*\d+)?)', s, re.I)
                if match:
                    mb_no = match.group(1)
                    pages = match.group(2) if match.group(2) else ""
                    labels["mb_details"] = f"MB No. {mb_no}, Page {pages}" if pages else f"MB No. {mb_no}"
                    continue
            # Look for "Sl. No." or "Serial No." patterns (often have MB reference)
            if ("sl." in low or "serial" in low) and "no" in low:
                val = _extract_value_part_from_line(s)
                if val:
                    labels["mb_details"] = val
                    continue

        # T.P Details (like "T.P @7.86 % Less")
        if not labels["tp_details"]:
            if "t.p" in low or "tp@" in low or "tp %" in low:
                # Look for percentage pattern
                match = re.search(r'(T\.?P\.?\s*@?\s*[\d\.]+\s*%?\s*(?:less|more)?)', s, re.I)
                if match:
                    labels["tp_details"] = match.group(1).strip()
                else:
                    val = _extract_value_part_from_line(s)
                    if val:
                        labels["tp_details"] = val
                continue

        # CC Header - Bill Type (L.S. Bill, 1st Part Bill, Final Bill, etc.)
        if not labels["cc_header"]:
            # Look for "Bill of L.S." or "L.S. Bill" pattern
            if "bill of l.s" in low or "l.s. bill" in low or "l.s bill" in low:
                labels["cc_header"] = "L.S. Bill"
                continue
            # Look for Part Bills
            part_match = re.search(r'(\d+(?:st|nd|rd|th)?\s*part\s*bill)', low, re.I)
            if part_match:
                labels["cc_header"] = part_match.group(1).title()
                continue
            # Look for Final Bill
            if "final bill" in low or "final part" in low:
                labels["cc_header"] = "Final Bill"
                continue
            # Look for generic CC Bill pattern
            if "cc" in low and "bill" in low:
                labels["cc_header"] = s.strip()
                continue

        # NIT No. / Tender No.
        if not labels["nit_no"]:
            if "nit" in low or ("tender" in low and "no" in low):
                # Look for NIT number pattern - capture full reference including date
                # Pattern like: NIT.No14/EE/GI/DB/HD/TA1/2025-26, Dt:04.04.2025
                match = re.search(r'(NIT\.?\s*No\.?\s*[A-Za-z0-9\/\-]+(?:\s*,?\s*Dt[:\.]?\s*[\d\.\/\-]+)?)', s, re.I)
                if match:
                    labels["nit_no"] = match.group(1).strip()
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["nit_no"] = val
                    # Value might continue on next line (date part)
                    if idx + 1 < len(lines):
                        next_line = str(lines[idx + 1]).strip()
                        if next_line.lower().startswith('dt') or re.match(r'^[\d\.\-\/]+', next_line):
                            labels["nit_no"] = val + ", " + next_line
                    continue

        # Tender Premium
        if not labels["tender_premium"]:
            if "tender premium" in low or ("premium" in low and ("%" in s or "er" in low)):
                # Look for percentage pattern
                match = re.search(r'([\d\.]+\s*%?\s*(?:ER|less|more|above|below)?)', s, re.I)
                if match:
                    labels["tender_premium"] = match.group(1).strip()
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["tender_premium"] = val
                    continue

        # Period of Completion
        if not labels["period_of_completion"]:
            if "period of completion" in low or "completion period" in low or "stipulated period" in low or "period of compl" in low:
                # Look for months/days pattern on this line
                match = re.search(r'(\d+\s*(?:months?|days?|years?|weeks?))', s, re.I)
                if match:
                    labels["period_of_completion"] = match.group(1).strip()
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["period_of_completion"] = val
                    continue
                # Value might be on next line
                if idx + 1 < len(lines):
                    next_line = str(lines[idx + 1]).strip()
                    match = re.search(r'(\d+\s*(?:months?|days?|years?|weeks?))', next_line, re.I)
                    if match:
                        labels["period_of_completion"] = match.group(1).strip()
                        continue

        # Earnest Money Deposit
        if not labels["earnest_money"]:
            if "earnest money" in low or "emd" in low:
                nums = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if nums:
                    labels["earnest_money"] = nums[0]
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["earnest_money"] = val
                    continue

        # Security Deposit
        if not labels["security_deposit"]:
            if "security deposit" in low or "s.d" in low:
                nums = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if nums:
                    labels["security_deposit"] = nums[0]
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["security_deposit"] = val
                    continue

        # Work Order No. / Date
        if not labels["work_order_no"]:
            if "work order" in low:
                match = re.search(r'No\.?\s*[:\.]?\s*([\w\/\-]+)', s, re.I)
                if match:
                    labels["work_order_no"] = match.group(1).strip()
                # Also look for date
                date_match = re.search(r'(?:Dt|Date|dated)[:\.]?\s*([\d\.\/\-]+)', s, re.I)
                if date_match:
                    labels["work_order_date"] = date_match.group(1).strip()
                continue

        # Date of Commencement
        if not labels["date_of_commencement"]:
            if "commencement" in low or "start date" in low:
                date_match = re.search(r'([\d]{1,2}[\.\-\/][\d]{1,2}[\.\-\/][\d]{2,4})', s)
                if date_match:
                    labels["date_of_commencement"] = date_match.group(1)
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["date_of_commencement"] = val
                    continue

        # Date of Completion
        if not labels["date_of_completion"]:
            if "date of completion" in low or "completion date" in low:
                date_match = re.search(r'([\d]{1,2}[\.\-\/][\d]{1,2}[\.\-\/][\d]{2,4})', s)
                if date_match:
                    labels["date_of_completion"] = date_match.group(1)
                    continue
                val = _extract_value_part_from_line(s)
                if val:
                    labels["date_of_completion"] = val
                    continue

        # DOI / DOC / DOMR / DOBR (from bill header line like "DOI : 01.01.2025    DOC : ...")
        if not labels["doi"] or not labels["doc"] or not labels["domr"] or not labels["dobr"]:
            if "doi" in low and "doc" in low:
                doi_m = re.search(r'DOI\s*:\s*([\d\.\/\-]+)', s, re.I)
                doc_m = re.search(r'DOC\s*:\s*([\d\.\/\-]+)', s, re.I)
                domr_m = re.search(r'DOMR\s*:\s*([\d\.\/\-]+)', s, re.I)
                dobr_m = re.search(r'DOBR\s*:\s*([\d\.\/\-]+)', s, re.I)
                if doi_m:
                    labels["doi"] = doi_m.group(1).strip()
                if doc_m:
                    labels["doc"] = doc_m.group(1).strip()
                if domr_m:
                    labels["domr"] = domr_m.group(1).strip()
                if dobr_m:
                    labels["dobr"] = dobr_m.group(1).strip()
                continue

        # Total / Amount (with Rs. prefix) - must be a significant amount
        # Prefer bill amounts (1,10,000) over estimate amounts (102820)
        if not labels["amount"]:
            # Look for bill amount, total amount, sanction amount patterns
            if ("total" in low or "bill amount" in low or "amount of bill" in low or 
                "sanction amount" in low or "approximate value" in low):
                nums = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)', s, re.I)
                if nums:
                    # Filter out small amounts like security deposits, stamps etc
                    for num in reversed(nums):
                        clean_num = num.replace(',', '')
                        try:
                            val = float(clean_num)
                            # Prefer amounts like 1,10,000 (>100000) over amounts like 102820.48
                            if val >= 100000:
                                labels["amount"] = num
                                break
                        except:
                            labels["amount"] = num
                            break
                    if labels["amount"]:
                        continue
            
            # Look for standalone Rs. amounts on sanction amount lines
            if "sanction" in low and "amount" in low:
                nums = re.findall(r'Rs\.?\s*([\d,]+(?:\.\d+)?)/?-?', s, re.I)
                for num in nums:
                    clean_num = num.replace(',', '')
                    try:
                        if float(clean_num) >= 100000:
                            labels["amount"] = num
                            break
                    except:
                        pass

    # Use estimate_amount as amount if no amount found
    if not labels["amount"] and labels["estimate_amount"]:
        labels["amount"] = labels["estimate_amount"]
    
    # Use contractor as agency if agency not found
    if not labels["agency"] and labels["contractor"]:
        labels["agency"] = labels["contractor"]
    
    # FIX COMMON OCR SPELLING ERRORS
    ocr_corrections = {
        # Maintenance
        "mala tenance": "maintenance",
        "malatenance": "maintenance",
        "maintenace": "maintenance",
        "maintainance": "maintenance",
        "maintanance": "maintenance",
        # Monthly
        "mouthly": "monthly",
        "monthely": "monthly",
        "montly": "monthly",
        # Servicing
        "servlcing": "servicing",
        "serviclng": "servicing",
        # Annual
        "annuai": "annual",
        "annuat": "annual",
        # Generator
        "generater": "generator",
        "genertor": "generator",
        # Providing
        "provlding": "providing",
        "providlng": "providing",
        # Located
        "Iocated": "located",
        # Guest house
        "bow se": "house",
        "bowse": "house",
        # Complex
        "complec": "complex",
        "compex": "complex",
        # Begumpet
        "Bite pad": "Begumpet",
        "Bitépad": "Begumpet",
        # Hyderabad
        "Hydera bad": "Hyderabad",
        # Common OCR errors - general
        "sanctlon": "sanction",
        "sanchon": "sanction",
        "Techical": "Technical",
        "Addinistrative": "Administrative",
        "ot cated": "located",
        "st ate": "state",
        # Common l/1/I confusions
        "Estimale": "Estimate",
        "estimale": "estimate",
        "estlmate": "estimate",
        "eslimate": "estimate",
        "Admlnistrative": "Administrative",
        "admlnistrative": "administrative",
        "Technlcal": "Technical",
        "technlcal": "technical",
        # o/0 confusions
        "c0ntract": "contract",
        "c0ntractor": "contractor",
        "c0mpletion": "completion",
        # Common spacing issues
        "sub division": "sub-division",
        "Sub Division": "Sub-Division",
        "work order": "work order",
        "Agree ment": "Agreement",
        "agree ment": "agreement",
        "Ad ministrative": "Administrative",
        "ad ministrative": "administrative",
        "con tractor": "contractor",
        "san ction": "sanction",
        "comple tion": "completion",
        # Common document terms
        "Supplylng": "Supplying",
        "supplylng": "supplying",
        "Constructlon": "Construction",
        "constructlon": "construction",
        "Electrlcal": "Electrical",
        "electrlcal": "electrical",
        "lnstallation": "installation",
        "Installatlon": "Installation",
        "Erectl0n": "Erection",
        "erecti0n": "erection",
    }

    # Apply corrections to ALL text fields (not just name_of_work)
    text_fields = [
        "name_of_work", "agreement", "admin_sanction", "tech_sanction",
        "agency", "contractor_name", "contractor_address",
        "nit_number", "period_of_completion",
    ]
    for field in text_fields:
        if labels.get(field):
            val = labels[field]
            for wrong, correct in ocr_corrections.items():
                val = re.sub(re.escape(wrong), correct, val, flags=re.I)
            labels[field] = val
    
    # SECOND PASS: Handle table-format where labels and values are on separate lines
    # This handles OCR from table documents where "Name of Work" is one row and value is next row
    label_to_key = {
        "name of work": "name_of_work",
        "name of the work": "name_of_work",
        "administrative sanction": "admin_sanction",
        "admin sanction": "admin_sanction",
        "administrative sanction amount": "admin_sanction_amount",
        "admin sanction amount": "admin_sanction_amount",
        "admin sanction ref": "admin_sanction",
        "technical sanction": "tech_sanction",
        "tech sanction": "tech_sanction",
        "technical sanction amount": "tech_sanction_amount",
        "tech sanction amount": "tech_sanction_amount",
        "nit no": "nit_no",
        "nit no. & date": "nit_no",
        "nit no & date": "nit_no",
        "contractor name": "agency",
        "name of contractor": "agency",
        "name of the contractor": "agency",
        "contractor address": "contractor_address",
        "estimate amount": "estimate_amount",
        "amount of estimate": "estimate_amount",
        "period of completion": "period_of_completion",
        "period of compl": "period_of_completion",
        "tender premium": "tender_premium",
    }
    
    for idx, raw in enumerate(lines):
        s = str(raw or "").strip()
        if not s:
            continue
        low = s.lower().rstrip(':').strip()
        
        # Check if this line is just a label (no value after it)
        if low in label_to_key:
            key = label_to_key[low]
            # Only fill if not already set
            if not labels.get(key) and idx + 1 < len(lines):
                next_val = str(lines[idx + 1]).strip()
                # Make sure next line isn't another label
                if next_val and next_val.lower().rstrip(':') not in label_to_key:
                    # Clean up Rs. format for amount fields
                    if key in ("admin_sanction_amount", "tech_sanction_amount", "estimate_amount"):
                        rs_match = re.search(r'Rs\.?\s*([\d,\.]+)', next_val, re.I)
                        if rs_match:
                            labels[key] = "Rs." + rs_match.group(1) + "/-"
                        else:
                            labels[key] = next_val
                    # Handle contractor/agency with M/s. pattern
                    elif key == "agency":
                        ms_match = re.search(r'M/s\.?\s*(.+)', next_val, re.I)
                        if ms_match:
                            labels[key] = "M/s." + ms_match.group(1).strip()
                        else:
                            labels[key] = next_val
                    # For multiline values like name_of_work, collect more lines
                    elif key == "name_of_work" and len(next_val) < 100:
                        full_val = _collect_multiline_value(lines, idx + 1, max_lines=5)
                        labels[key] = full_val if full_val else next_val
                    else:
                        labels[key] = next_val

    return labels


def _preprocess_image_for_ocr(img):
    """
    ULTIMATE OCR preprocessing for best accuracy on government documents.
    Handles blurred, skewed, low-quality scanned images with advanced techniques.
    
    Enhancement Pipeline:
    1. Auto-orientation detection
    2. Deskewing (straightening tilted text)
    3. Smart upscaling for small images
    4. Adaptive contrast enhancement
    5. Noise reduction with edge preservation
    6. Sharpening for blurred text
    7. Adaptive binarization (Otsu)
    8. Morphological cleanup
    """
    try:
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps
        import numpy as np
    except ImportError:
        return img  # Return original if dependencies not available

    import logging
    logger = logging.getLogger(__name__)

    try:
        original_mode = img.mode
        
        # 1. Auto-rotate based on EXIF data (for photos)
        try:
            img = ImageOps.exif_transpose(img)
        except Exception:
            pass

        # Convert to RGB for processing
        if img.mode not in ('RGB', 'L'):
            img = img.convert('RGB')

        # 2. Smart upscaling for better OCR on small/low-res images
        width, height = img.size
        target_dpi_size = 2500  # Target minimum dimension
        if width < target_dpi_size or height < target_dpi_size:
            scale_factor = max(target_dpi_size / width, target_dpi_size / height)
            # Limit scale factor to prevent excessive memory use
            scale_factor = min(scale_factor, 4.0)
            new_width = int(width * scale_factor)
            new_height = int(height * scale_factor)
            img = img.resize((new_width, new_height), Image.LANCZOS)
            logger.debug(f"OCR: Upscaled image from {width}x{height} to {new_width}x{new_height}")

        # 3. Convert to grayscale
        gray = img.convert('L')

        # 4. Deskew detection and correction
        try:
            gray = _deskew_image(gray)
        except Exception as e:
            logger.debug(f"OCR: Deskew failed: {e}")

        # 5. Adaptive contrast enhancement (CLAHE-like approach)
        try:
            img_array = np.array(gray)
            # Apply local contrast enhancement
            from PIL import ImageFilter
            
            # High-pass filter for local contrast
            blurred = gray.filter(ImageFilter.GaussianBlur(radius=50))
            blurred_array = np.array(blurred)
            
            # Unsharp masking with gentle settings
            enhanced = img_array.astype(np.float32) - 0.5 * blurred_array.astype(np.float32)
            enhanced = np.clip(enhanced, 0, 255).astype(np.uint8)
            gray = Image.fromarray(enhanced)
        except Exception:
            pass

        # 6. Auto-levels (stretch histogram for better contrast)
        try:
            gray = ImageOps.autocontrast(gray, cutoff=0.5)
        except Exception:
            pass

        # 7. Enhanced sharpness for blurred documents
        gray_rgb = gray.convert('RGB')
        enhancer = ImageEnhance.Sharpness(gray_rgb)
        gray_rgb = enhancer.enhance(2.5)
        
        # Apply unsharp mask for additional deblurring
        gray_rgb = gray_rgb.filter(ImageFilter.UnsharpMask(radius=2, percent=200, threshold=2))
        gray = gray_rgb.convert('L')

        # 8. Noise reduction with edge preservation
        try:
            # Bilateral-like filtering using median + edge detection
            gray = gray.filter(ImageFilter.MedianFilter(size=3))
        except Exception:
            pass

        # 9. Otsu-style adaptive thresholding for binarization
        try:
            img_array = np.array(gray)
            
            # Apply slight Gaussian blur before thresholding
            blurred = Image.fromarray(img_array).filter(ImageFilter.GaussianBlur(radius=1))
            img_array = np.array(blurred)
            
            # Calculate Otsu threshold
            hist, _ = np.histogram(img_array.ravel(), bins=256, range=(0, 256))
            total = img_array.size
            sum_all = np.sum(np.arange(256) * hist)
            sum_bg, w_bg = 0.0, 0
            max_var, threshold = 0.0, 128
            
            for t in range(256):
                w_bg += hist[t]
                if w_bg == 0:
                    continue
                w_fg = total - w_bg
                if w_fg == 0:
                    break
                sum_bg += t * hist[t]
                mean_bg = sum_bg / w_bg
                mean_fg = (sum_all - sum_bg) / w_fg
                var_between = w_bg * w_fg * (mean_bg - mean_fg) ** 2
                if var_between > max_var:
                    max_var = var_between
                    threshold = t

            # Apply threshold with slight bias towards preserving text (lower threshold)
            adjusted_threshold = max(threshold - 10, 50)
            img_array = np.where(img_array > adjusted_threshold, 255, 0).astype(np.uint8)
            gray = Image.fromarray(img_array)
        except Exception:
            pass

        # 10. Morphological operations to clean up text
        try:
            img_array = np.array(gray)
            # Slight dilation to thicken thin text (helps with degraded scans)
            gray = Image.fromarray(img_array).filter(ImageFilter.MinFilter(size=3))
            # Then slight erosion to remove noise
            gray = gray.filter(ImageFilter.MaxFilter(size=3))
        except Exception:
            pass

        logger.debug(f"OCR: Preprocessing complete, final size: {gray.size}")
        return gray

    except Exception as e:
        logger.warning(f"OCR preprocessing failed: {e}")
        return img  # Return original on error


def _deskew_image(img):
    """
    Detect and correct skew angle in document images.
    Uses projection profile analysis for accurate skew detection.
    Returns the deskewed image.
    """
    import numpy as np
    from PIL import Image
    import logging
    logger = logging.getLogger(__name__)

    try:
        img_array = np.array(img)
        
        # Convert to binary for analysis
        threshold = np.mean(img_array)
        binary = (img_array < threshold).astype(np.uint8)  # Text is dark
        
        # Test rotation angles from -15 to +15 degrees
        best_angle = 0
        best_score = 0
        
        for angle in np.arange(-15, 15.5, 0.5):
            # Rotate image
            rotated = img.rotate(angle, fillcolor=255, resample=Image.BICUBIC)
            rot_array = np.array(rotated)
            rot_binary = (rot_array < threshold).astype(np.uint8)
            
            # Calculate horizontal projection profile
            projection = np.sum(rot_binary, axis=1)
            
            # Score: variance of projection (higher = more aligned text)
            score = np.var(projection)
            
            if score > best_score:
                best_score = score
                best_angle = angle
        
        if abs(best_angle) > 0.5:  # Only deskew if significant angle detected
            logger.debug(f"OCR: Deskewing by {best_angle} degrees")
            return img.rotate(best_angle, fillcolor=255, resample=Image.BICUBIC, expand=True)
        
        return img
        
    except Exception as e:
        logger.debug(f"Deskew failed: {e}")
        return img


def _ocr_with_multiple_configs(img, lang='eng'):
    """
    ULTIMATE OCR extraction with multiple Tesseract configurations for maximum accuracy.
    Tries multiple PSM modes, OEM engines, and language combinations.
    Automatically detects and uses available languages.
    Applies advanced image preprocessing before OCR.
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        import pytesseract
    except ImportError:
        logger.error("pytesseract not installed. OCR unavailable.")
        return ""

    # Preprocess the image for better OCR accuracy
    preprocessed_img = _preprocess_image_for_ocr(img)

    # Detect available languages (fallback to eng if detection fails)
    available_langs = ['eng']
    try:
        langs_output = pytesseract.get_languages()
        available_langs = [l for l in langs_output if l not in ('osd', 'snum')]
        logger.debug(f"OCR: Available languages: {available_langs}")
    except Exception:
        pass

    # Build language string - prioritize English, add Hindi/Telugu if available for Indian docs
    lang_priority = ['eng']
    for indian_lang in ['hin', 'tel', 'tam', 'kan', 'mar', 'guj', 'ben']:
        if indian_lang in available_langs:
            lang_priority.append(indian_lang)
    
    lang_str = '+'.join(lang_priority[:3])  # Max 3 languages for speed
    logger.debug(f"OCR: Using languages: {lang_str}")

    # Extended configurations to try - ordered by typical effectiveness for documents
    configs = [
        ('--oem 3 --psm 6', 'PSM6-block'),       # Uniform block of text (best for documents)
        ('--oem 1 --psm 6', 'LSTM-PSM6'),        # LSTM neural net, block mode
        ('--oem 3 --psm 3', 'PSM3-auto'),        # Fully automatic page segmentation
        ('--oem 3 --psm 4', 'PSM4-column'),      # Single column of text
        ('--oem 3 --psm 11', 'PSM11-sparse'),    # Sparse text - find as much text as possible
        ('--oem 1 --psm 3', 'LSTM-PSM3'),        # LSTM with auto segmentation
    ]

    best_result = ""
    best_score = 0
    
    # Track all results for ensemble approach
    all_results = []

    for config, label in configs:
        try:
            result = pytesseract.image_to_string(preprocessed_img, lang=lang_str, config=config)
            
            # Apply domain-specific corrections
            result = _apply_domain_corrections(result)
            
            # Score based on content quality metrics
            result_lines = [ln.strip() for ln in result.splitlines() if ln.strip() and len(ln.strip()) > 2]
            alnum_chars = sum(1 for c in result if c.isalnum())
            total_chars = len(result) or 1
            
            # Advanced scoring
            # - More lines is better
            # - Higher alphanumeric ratio is better
            # - Presence of key document terms is better
            key_terms_found = sum(1 for term in ['sanction', 'estimate', 'contractor', 'work', 'amount', 'Rs'] 
                                  if term.lower() in result.lower())
            
            score = (
                len(result_lines) * 10 +           # Line count
                alnum_chars +                       # Character count
                (alnum_chars / total_chars) * 50 + # Alphanumeric ratio
                key_terms_found * 30                # Domain relevance
            )
            
            logger.debug(f"OCR {label}: {len(result_lines)} lines, {key_terms_found} keywords, score={score:.0f}")
            
            all_results.append((score, result, label))
            
            if score > best_score:
                best_score = score
                best_result = result
                
        except Exception as e:
            logger.debug(f"OCR {label} failed: {e}")

    # If preprocessed results are poor, try original image as fallback
    if best_score < 100:
        try:
            result = pytesseract.image_to_string(img, lang=lang_str, config='--oem 3 --psm 6')
            result = _apply_domain_corrections(result)
            result_lines = [ln.strip() for ln in result.splitlines() if ln.strip() and len(ln.strip()) > 2]
            alnum_chars = sum(1 for c in result if c.isalnum())
            total_chars = len(result) or 1
            key_terms_found = sum(1 for term in ['sanction', 'estimate', 'contractor', 'work', 'amount', 'Rs'] 
                                  if term.lower() in result.lower())
            score = len(result_lines) * 10 + alnum_chars + (alnum_chars / total_chars) * 50 + key_terms_found * 30
            
            if score > best_score:
                best_result = result
                best_score = score
                logger.debug(f"Original image OCR was better: score={score:.0f}")
        except Exception:
            pass

    logger.info(f"OCR completed: {len(best_result)} chars, {len(best_result.splitlines())} lines, score={best_score:.0f}")
    return best_result


# Domain-specific word corrections for government & engineering documents
_DOMAIN_CORRECTIONS = {
    # Common OCR errors in government documents
    'sanctoin': 'sanction',
    'sancfion': 'sanction',
    'sanct1on': 'sanction',
    'estlmate': 'estimate',
    'est1mate': 'estimate',
    'esimate': 'estimate',
    'estmate': 'estimate',
    'contractar': 'contractor',
    'contracter': 'contractor',
    'contractor': 'contractor',
    'contracor': 'contractor',
    'arnount': 'amount',
    'amoun+': 'amount',
    'amcunt': 'amount',
    'adminisirative': 'administrative',
    'administratve': 'administrative',
    'administrat1ve': 'administrative',
    'techmical': 'technical',
    'technlcal': 'technical',
    'technica1': 'technical',
    'tander': 'tender',
    'tendar': 'tender',
    'prernium': 'premium',
    'premiurn': 'premium',
    'prem1um': 'premium',
    'cornpletion': 'completion',
    'completlon': 'completion',
    'complet1on': 'completion',
    'perlod': 'period',
    'pericd': 'period',
    'per1od': 'period',
    'electricai': 'electrical',
    'electrlcal': 'electrical',
    'electrica1': 'electrical',
    'c1vil': 'civil',
    'civll': 'civil',
    'maintenace': 'maintenance',
    'maintainence': 'maintenance',
    'maintenanoe': 'maintenance',
    'providng': 'providing',
    'provlding': 'providing',
    'provid1ng': 'providing',
    'generatar': 'generator',
    'generater': 'generator',
    'generat0r': 'generator',
    'servlcing': 'servicing',
    'serviclng': 'servicing',
    'annuai': 'annual',
    'annua1': 'annual',
    'forwaroing': 'forwarding',
    'forwardlng': 'forwarding',
    # NIT/Agreement patterns
    'NlT': 'NIT',
    'N1T': 'NIT',
    'Agreernent': 'Agreement',
    'Agreemen+': 'Agreement',
    # Common unit errors
    'Rs,': 'Rs.',
    'Rs ': 'Rs.',
    'R5.': 'Rs.',
    'Rs-': 'Rs.',
    '/-': '/-',
    '/ -': '/-',
    '/- ': '/-',
}


def _apply_domain_corrections(text):
    """Apply domain-specific word corrections for government documents."""
    if not text:
        return text
    
    # Apply word-level corrections (case-insensitive with proper replacement)
    for wrong, correct in _DOMAIN_CORRECTIONS.items():
        # Use word boundaries for whole word replacement
        pattern = re.compile(re.escape(wrong), re.IGNORECASE)
        text = pattern.sub(correct, text)
    
    return text


# Common OCR character confusions applied to all text
_OCR_CHAR_FIXES = {
    '|': 'l',     # pipe → lowercase L
    '¢': 'c',     # cent sign → c
    '©': 'c',     # copyright → c
    '®': 'R',     # registered → R
    '—': '-',     # em dash → hyphen
    '–': '-',     # en dash → hyphen
    '\u2018': "'", # left single quote  
    '\u2019': "'", # right single quote
    '\u201c': '"', # left double quote
    '\u201d': '"', # right double quote
    '„': '"',     # low double quote
    '»': '"',     # right angle quote
    '«': '"',     # left angle quote
}


def _fix_ocr_text(text):
    """Apply character-level OCR fixes and clean common artifacts."""
    if not text:
        return text
    
    # Apply character fixes
    for wrong, correct in _OCR_CHAR_FIXES.items():
        text = text.replace(wrong, correct)
    
    # Fix common OCR patterns
    text = re.sub(r'\bRs\s*[.,]\s*', 'Rs. ', text)  # Fix "Rs," or "Rs ." to "Rs. "
    text = re.sub(r'\b0f\b', 'of', text, flags=re.I)  # 0f → of
    text = re.sub(r'\b1n\b', 'in', text, flags=re.I)  # 1n → in
    text = re.sub(r'\bthe\s+the\b', 'the', text, flags=re.I)  # Remove duplicate "the the"
    text = re.sub(r'\btbe\b', 'the', text, flags=re.I)  # tbe → the
    text = re.sub(r'\bvvork\b', 'work', text, flags=re.I)  # vvork → work
    text = re.sub(r'\bworK\b', 'work', text, flags=re.I)  # worK → work
    
    # Fix number/letter confusions in specific contexts
    text = re.sub(r'(\d),(\d{3})', r'\1,\2', text)  # Ensure proper comma in numbers
    text = re.sub(r'Rs\.?\s*([0-9,]+)\s*/-', r'Rs. \1/-', text)  # Standardize Rs format
    
    # Apply domain corrections
    text = _apply_domain_corrections(text)
    
    return text


def _extract_labels_from_source_file(uploaded_file):
    """
    Read uploaded file (Excel / Word / PDF / Image / text)
    and return:
      - labels: dict from _extract_labels_from_lines
      - lines:  flattened list of text lines (for custom placeholders)
    
    Supports: .xlsx, .xlsm, .docx, .pdf, .jpg, .jpeg, .png, .bmp, .tiff, .gif, .txt, .csv
    Uses advanced image preprocessing for blurred/low-quality scanned documents.
    """
    filename = uploaded_file.name or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    lines = []

    if ext in ("xlsx", "xlsm"):
        wb = load_workbook(uploaded_file, data_only=True)
        for ws in wb.worksheets:
            max_r = min(ws.max_row or 0, 150)
            max_c = min(ws.max_column or 0, 20)
            for r in range(1, max_r + 1):
                vals = []
                for c in range(1, max_c + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        vals.append(str(v).strip())
                if vals:
                    # For 2-column table format (Label | Value), create "Label: Value" format
                    if len(vals) == 2:
                        label = vals[0].lower().strip()
                        # Check if first column looks like a label
                        if any(x in label for x in ['name', 'work', 'sanction', 'amount', 'contractor', 
                                                      'nit', 'estimate', 'period', 'address', 'premium']):
                            lines.append(f"{vals[0]}: {vals[1]}")
                        else:
                            lines.append(" ".join(vals))
                    # For single cell or multi-column, just join
                    else:
                        lines.append(" ".join(vals))

    elif ext == "docx":
        doc = Document(uploaded_file)
        for p in doc.paragraphs:
            t = (p.text or "").strip()
            if t:
                lines.append(t)
        for table in doc.tables:
            for row in table.rows:
                vals = [(cell.text or "").strip() for cell in row.cells]
                vals = [v for v in vals if v]
                if vals:
                    # For 2-column table format (Label | Value), create "Label: Value" format
                    if len(vals) == 2:
                        label = vals[0].lower().strip()
                        # Check if first column looks like a label
                        if any(x in label for x in ['name', 'work', 'sanction', 'amount', 'contractor', 
                                                      'nit', 'estimate', 'period', 'address', 'premium']):
                            lines.append(f"{vals[0]}: {vals[1]}")
                        else:
                            lines.append(" ".join(vals))
                    else:
                        lines.append(" ".join(vals))

    elif ext == "pdf":
        import logging
        logger = logging.getLogger(__name__)
        
        # First try PyPDF2 for text-based PDFs
        try:
            import PyPDF2
            uploaded_file.seek(0)  # Reset file pointer
            reader = PyPDF2.PdfReader(uploaded_file)
            for page in reader.pages:
                txt = page.extract_text() or ""
                # PDF text often has inconsistent spacing/newlines
                for ln in txt.splitlines():
                    ln = ln.strip()
                    if ln:
                        # Handle cases where multiple fields are on one line
                        if ':' in ln and len(ln) > 50:
                            parts = re.split(r'(?<=[a-zA-Z])\s*:\s*(?=[A-Z])', ln)
                            if len(parts) > 1:
                                for p in parts:
                                    if p.strip():
                                        lines.append(p.strip())
                                continue
                        lines.append(ln)
        except Exception as e:
            logger.warning(f"PyPDF2 extraction failed: {e}")
        
        # If no text extracted or very little text, try OCR for scanned/blurred PDFs
        if len(lines) < 3:
            try:
                from pdf2image import convert_from_bytes
                import pytesseract
                from PIL import Image

                logger.info("Attempting OCR for scanned PDF...")
                uploaded_file.seek(0)
                pdf_bytes = uploaded_file.read()

                # Convert PDF pages to images at 300 DPI for better OCR accuracy
                images = convert_from_bytes(pdf_bytes, dpi=300)

                ocr_lines = []

                for idx, img in enumerate(images):
                    logger.info(f"Processing page {idx + 1} with OCR...")

                    # Multi-config OCR with preprocessing
                    txt = _ocr_with_multiple_configs(img, lang='eng')
                    # Apply character-level OCR fixes
                    txt = _fix_ocr_text(txt)

                    for ln in txt.splitlines():
                        ln = ln.strip()
                        if ln and len(ln) > 1:
                            ocr_lines.append(ln)

                # Use OCR results if they have more content
                if len(ocr_lines) > len(lines):
                    lines = ocr_lines
                    logger.info(f"OCR extracted {len(lines)} lines from scanned PDF")
                    
            except ImportError as e:
                logger.warning(f"OCR libraries not available: {e}. Install pdf2image, pytesseract, and Pillow.")
            except Exception as e:
                logger.warning(f"OCR extraction failed: {e}")

    elif ext in ("jpg", "jpeg", "png", "bmp", "tiff", "tif", "gif", "webp"):
        # Direct image file support
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            from PIL import Image
            import pytesseract
            
            logger.info(f"Processing image file: {filename}")
            uploaded_file.seek(0)
            img = Image.open(uploaded_file)

            # Multi-config OCR with preprocessing
            txt = _ocr_with_multiple_configs(img, lang='eng')
            # Apply character-level OCR fixes
            txt = _fix_ocr_text(txt)

            for ln in txt.splitlines():
                ln = ln.strip()
                if ln and len(ln) > 1:  # Skip single characters
                    lines.append(ln)
            
            logger.info(f"OCR extracted {len(lines)} lines from image")
            
        except ImportError as e:
            logger.warning(f"Image OCR libraries not available: {e}. Install Pillow and pytesseract.")
        except Exception as e:
            logger.warning(f"Image OCR extraction failed: {e}")

    else:
        # plain text / csv / unknown  -  treat as text
        try:
            content = uploaded_file.read()
        except Exception:
            content = b""
        try:
            text = content.decode("utf-8", errors="ignore")
        except Exception:
            text = str(content)
        for ln in text.splitlines():
            ln = ln.strip()
            if ln:
                lines.append(ln)

    # Debug logging - show what was extracted from source file
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"=== SOURCE FILE LINES ({len(lines)} total) ===")
    for i, line in enumerate(lines[:30]):  # First 30 lines
        logger.info(f"  [{i}] {line[:100]}...")  # First 100 chars
    logger.info("=== END SOURCE FILE LINES ===")

    labels = _extract_labels_from_lines(lines)
    
    # Debug logging - show extracted labels
    logger.info("=== EXTRACTED LABELS ===")
    for key, val in labels.items():
        if val:
            logger.info(f"  {key}: {val[:80]}..." if len(str(val)) > 80 else f"  {key}: {val}")
    logger.info("=== END EXTRACTED LABELS ===")
    
    return labels, lines


def _extract_labels_per_work(uploaded_file):
    """
    Like _extract_labels_from_source_file but returns a *list* of works.
    For Excel files with multiple sheets, each sheet is treated as a separate
    work.  For all other file types, returns a single-element list.

    Also extracts bill-specific fields for Progress Report:
      - amount_paid: rightmost column, bottommost numeric value
      - cc_header: bill type from the title row (for Remarks)
      - doi, doc, domr, dobr: dates from header row 9
      - mb_details: from header row 8

    Returns: list of (source_name, labels) tuples
    """
    filename = uploaded_file.name or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext in ("xlsx", "xlsm"):
        wb = load_workbook(uploaded_file, data_only=True)
        results = []
        for ws in wb.worksheets:
            lines = []
            max_r = min(ws.max_row or 0, 150)
            max_c = min(ws.max_column or 0, 20)
            for r in range(1, max_r + 1):
                vals = []
                for c in range(1, max_c + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        vals.append(str(v).strip())
                if vals:
                    if len(vals) == 2:
                        label = vals[0].lower().strip()
                        if any(x in label for x in ['name', 'work', 'sanction', 'amount', 'contractor',
                                                      'nit', 'estimate', 'period', 'address', 'premium']):
                            lines.append(f"{vals[0]}: {vals[1]}")
                        else:
                            lines.append(" ".join(vals))
                    else:
                        lines.append(" ".join(vals))
            if lines:
                labels = _extract_labels_from_lines(lines)
                source_name = f"{filename} — {ws.title}"

                # --- Bill-specific extraction from structured header rows ---
                _extract_bill_specific_fields(ws, labels, max_r, max_c)

                results.append((source_name, labels))
        return results if results else [(filename, {})]
    else:
        # Non-Excel: delegate to existing function (one work per file)
        labels, lines = _extract_labels_from_source_file(uploaded_file)
        return [(filename, labels)]


def _extract_bill_specific_fields(ws, labels, max_r, max_c):
    """
    Extract bill-specific fields directly from the worksheet structure:
    - amount_paid: bottommost numeric value in the rightmost data column
    - cc_header: bill type from the title (row 1 typically)
    - doi, doc, domr, dobr: from the "DOI : ... DOC : ..." row
    - mb_details: from the "M.B.No Details" row
    """
    # --- Amount Paid: scan rightmost column bottom-up for last number ---
    amount_paid = ""
    # Find the actual rightmost column with data (typically column H=8 for bills)
    for r in range(max_r, 0, -1):
        for c in range(max_c, 0, -1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                sv = str(v).strip()
                # Check if it's a number (remove commas, Rs., /- etc.)
                cleaned = re.sub(r'^Rs\.?\s*', '', sv, flags=re.I)
                cleaned = cleaned.replace(',', '').rstrip('/-').strip()
                try:
                    num_val = float(cleaned)
                    if num_val > 0:
                        # Round to 2 decimals
                        num_val = round(num_val, 2)
                        if num_val == int(num_val):
                            amount_paid = str(int(num_val))
                        else:
                            amount_paid = f"{num_val:.2f}"
                        break
                except (ValueError, TypeError):
                    continue
        if amount_paid:
            break
    labels["amount_paid"] = amount_paid

    # --- Scan first ~15 rows for structured bill header fields ---
    scan_rows = min(max_r, 15)
    for r in range(1, scan_rows + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is None:
            continue
        s = str(cell_val).strip()
        low = s.lower()

        # cc_header / bill type from title row (row 1 typically)
        if not labels.get("cc_header"):
            if ("bill" in low or "workslip" in low or "work slip" in low or
                "work-slip" in low or "estimate" in low):
                # Extract the bill type
                if "workslip" in low or "work slip" in low or "work-slip" in low:
                    labels["cc_header"] = "Workslip"
                elif "final" in low and "bill" in low:
                    labels["cc_header"] = "Final Bill"
                else:
                    part_match = re.search(r'(\d+(?:st|nd|rd|th)?\s*(?:&\s*)?\w*\s*(?:part\s*)?bill)', low, re.I)
                    if part_match:
                        labels["cc_header"] = part_match.group(1).strip().title()
                    elif "bill" in low:
                        labels["cc_header"] = s.strip()

        # DOI / DOC / DOMR / DOBR row (format: "DOI : xx   DOC : xx   DOMR : xx   DOBR : xx")
        if "doi" in low and "doc" in low:
            doi_match = re.search(r'DOI\s*:\s*([^\s]+(?:\s*[^\s]+)?)', s, re.I)
            doc_match = re.search(r'DOC\s*:\s*([^\s]+(?:\s*[^\s]+)?)', s, re.I)
            domr_match = re.search(r'DOMR\s*:\s*([^\s]+(?:\s*[^\s]+)?)', s, re.I)
            dobr_match = re.search(r'DOBR\s*:\s*([^\s]+(?:\s*[^\s]+)?)', s, re.I)

            if doi_match:
                # Clean: stop at next label keyword
                val = doi_match.group(1).strip()
                val = re.split(r'\s{2,}|DOC|DOMR|DOBR', val, flags=re.I)[0].strip()
                labels["doi"] = val
            if doc_match:
                val = doc_match.group(1).strip()
                val = re.split(r'\s{2,}|DOMR|DOBR', val, flags=re.I)[0].strip()
                labels["doc"] = val
            if domr_match:
                val = domr_match.group(1).strip()
                val = re.split(r'\s{2,}|DOBR', val, flags=re.I)[0].strip()
                labels["domr"] = val
            if dobr_match:
                labels["dobr"] = dobr_match.group(1).strip()

        # M.B.No Details row
        if "m.b" in low or "mb" in low and ("no" in low or "detail" in low):
            if not labels.get("mb_details"):
                # Use the entire cell value (it's the full MB details line)
                mb_val = s.strip()
                # Remove the label prefix if present
                mb_val = re.sub(r'^M\.?B\.?\s*(?:No\.?)?\s*Details?\s*:?\s*', '', mb_val, flags=re.I).strip()
                if mb_val:
                    labels["mb_details"] = mb_val


# -------------------------------------------
#  SELF FORMATTED MODULE  -  HELPERS
# -------------------------------------------

# imports consolidated at top of file
# from .utils import _number_to_words_rupees, _extract_labels_from_source_file
# or similar  -  just make sure they are imported somewhere above.


def _fuzzy_find_from_lines(lines, label_hint: str, threshold: float = 0.55) -> str:
    """
    Given a list of text lines (strings) and a label hint (e.g. "Name of the work"),
    find the line that best matches the label text using fuzzy similarity.

    Used by _build_placeholder_map() to map a custom KEY to the nearest
    line in the source document.
    """
    label = (label_hint or "").strip()
    if not label:
        return ""

    if not lines:
        return ""

    candidates = []

    # Normalize lines into plain strings
    for ln in lines:
        # If it's a list/tuple from some extractor, join elements
        if isinstance(ln, (list, tuple)):
            text = " ".join(str(x) for x in ln if x is not None)
        else:
            text = str(ln)
        text = text.strip()
        if not text:
            continue
        candidates.append(text)

    if not candidates:
        return ""

    best_line = ""
    best_score = 0.0
    label_low = label.lower()

    for text in candidates:
        score = SequenceMatcher(None, label_low, text.lower()).ratio()
        if score > best_score:
            best_score = score
            best_line = text

    # If nothing is similar enough, return empty string
    if best_score < threshold:
        return ""

    return best_line


def _extract_last_number(text: str) -> str:
    """
    From a string like 'Grand Total          105000.00',
    return '105000.00'. Returns "" if no number.
    """
    if not text:
        return ""
    cleaned = text.replace(",", "")
    nums = re.findall(r"[-+]?\d+(?:\.\d+)?", cleaned)
    return nums[-1] if nums else ""


def _build_placeholder_map(labels, lines, custom_text: str):
    """
    labels: from _extract_labels_from_source_file()
    lines : flattened text lines
    custom_text: textarea content "KEY = label text"

    Returns: dict { "{{PLACEHOLDER}}": "value" }
    """

    # -------- AMOUNT + INDIAN WORDS HANDLING --------
    # 'amount' is expected to be in labels (e.g. "56,000.00")
    raw_amount = (labels.get("amount") or "").strip()
    # remove common currency prefixes like 'Rs.', 'INR', or the rupee symbol
    raw_amount = re.sub(r'^(?:rs\.?|inr|â‚¹)\s*[:\- - â€”]?\s*', '', raw_amount, flags=re.I)
    raw_amount = raw_amount.replace(",", "").strip()
    total_amount = 0.0
    amount_in_words = ""

    if raw_amount:
        try:
            total_amount = float(raw_amount)
        except ValueError:
            total_amount = 0.0

    if total_amount:
        try:
            # Indian-style: Crores, Lakhs, Thousands, Rupees, Paise
            amount_in_words = _number_to_words_rupees(total_amount)
        except Exception:
            amount_in_words = ""

    # -------- ESTIMATE GRAND TOTAL (special case) --------
    est_grand_total = labels.get("est_grand_total", "")

    # If extractor did not supply it, try to guess from lines
    if not est_grand_total:
        grand_line = _fuzzy_find_from_lines(lines, "Grand Total")
        if grand_line:
            est_grand_total = _extract_last_number(grand_line)

    # As a fallback, use generic "amount"
    if not est_grand_total:
        est_grand_total = labels.get("amount", "")

    # prepare formatted amount strings
    formatted_amount = ""
    amount_raw = (labels.get("amount") or "").strip()
    if total_amount:
        try:
            if "." in amount_raw:
                formatted_amount = f"{total_amount:,.2f}"
            else:
                formatted_amount = f"{int(total_amount):,}"
        except Exception:
            formatted_amount = str(total_amount)

    # -------- BUILT-IN PLACEHOLDERS --------
    placeholder_map = {
        "{{NAME_OF_WORK}}":     labels.get("name_of_work", ""),
        "{{REF_OF_AGREEMENT}}": labels.get("agreement", ""),
        "{{AGREEMENT_REF}}":    labels.get("agreement", ""),
        "{{BOND_NO}}":          labels.get("agreement", ""),  # Alias for agreement/bond
        "{{ADMIN_SANCTION}}":   labels.get("admin_sanction", ""),
        "{{TECH_SANCTION}}":    labels.get("tech_sanction", ""),
        "{{NAME_OF_AGENCY}}":   labels.get("agency", ""),
        "{{AGENCY_NAME}}":      labels.get("agency", ""),
        "{{CONTRACTOR}}":       labels.get("agency", ""),  # Alias for agency/contractor
        "{{NAME_OF_CONTRACTOR}}": labels.get("agency", ""),

        # MB details content only (label part stripped in _extract_labels_from_source_file)
        "{{MB_DETAILS}}":       labels.get("mb_details", ""),

        "{{TP_DETAILS}}":       labels.get("tp_details", ""),
        "{{TENDER_PREMIUM}}":   labels.get("tender_premium", ""),
        "{{CC_HEADER}}":        labels.get("cc_header", ""),

        # Estimate amount (from "Amount of Estimate" line)
        "{{ESTIMATE_AMOUNT}}":  labels.get("estimate_amount", ""),
        
        # Admin/Tech Sanction amounts
        "{{ADMIN_SANCTION_AMOUNT}}": labels.get("admin_sanction_amount", ""),
        "{{TECH_SANCTION_AMOUNT}}": labels.get("tech_sanction_amount", ""),
        
        # NIT / Tender details
        "{{NIT_NO}}":           labels.get("nit_no", ""),
        "{{TENDER_NO}}":        labels.get("nit_no", ""),
        
        # Contractor address
        "{{CONTRACTOR_ADDRESS}}": labels.get("contractor_address", ""),
        
        # Period and Dates
        "{{PERIOD_OF_COMPLETION}}": labels.get("period_of_completion", ""),
        "{{STIPULATED_PERIOD}}": labels.get("period_of_completion", ""),
        "{{DATE_OF_COMMENCEMENT}}": labels.get("date_of_commencement", ""),
        "{{DATE_OF_COMPLETION}}": labels.get("date_of_completion", ""),
        "{{WORK_ORDER_NO}}":    labels.get("work_order_no", ""),
        "{{WORK_ORDER_DATE}}":  labels.get("work_order_date", ""),
        
        # Money related
        "{{EARNEST_MONEY}}":    labels.get("earnest_money", ""),
        "{{EMD}}":              labels.get("earnest_money", ""),
        "{{SECURITY_DEPOSIT}}": labels.get("security_deposit", ""),
        "{{SD}}":               labels.get("security_deposit", ""),

        # Generic amount from source (Bill / Estimate / etc.)
        # {{AMOUNT}}: formatted with commas (e.g. 1,23,456) when possible
        "{{AMOUNT}}":           formatted_amount or labels.get("amount", ""),
        "{{TOTAL_AMOUNT}}":     formatted_amount or labels.get("amount", ""),
        # raw numeric value without commas/currency
        "{{AMOUNT_RAW}}":       amount_raw.replace(",", ""),
        # convenience: amount with Rs. prefix
        "{{AMOUNT_WITH_RS}}":   ("Rs. " + (formatted_amount or labels.get("amount", ""))).strip(),

        # For your estimate covering letter  -  maps to Grand Total
        "{{EST_GRAND_TOTAL}}":  est_grand_total,

        "{{AMOUNT_IN_WORDS}}":  amount_in_words,
    }

    # -------- CUSTOM PLACEHOLDERS (from textarea) --------
    # Syntax: MY_PLACEHOLDER = label text
    custom_raw = (custom_text or "").strip()
    if custom_raw:
        for line in custom_raw.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue

            key, label_hint = line.split("=", 1)
            key = key.strip()
            label_hint = label_hint.strip()
            if not key or not label_hint:
                continue

            # normalise to {{KEY}} format
            if not key.startswith("{{"):
                key = "{{" + key
            if not key.endswith("}}"):
                key = key + "}}"

            match_line = _fuzzy_find_from_lines(lines, label_hint)
            if not match_line:
                continue

            # Try to be smart: if there is a number on that line, use the number;
            # otherwise use the full line.
            value_num = _extract_last_number(match_line)
            value = value_num or match_line

            placeholder_map[key] = value

    return placeholder_map


def _replace_placeholders_in_docx_xml(xml_str, safe_map):
    """
    Replace {{PLACEHOLDER}} values in DOCX XML text.
    Handles placeholders split across multiple <w:t> elements within a paragraph.
    """
    # Pass 1: direct replacement (handles placeholders inside a single <w:t>)
    for ph, val in safe_map.items():
        xml_str = xml_str.replace(ph, val)

    # Pass 2: handle placeholders split across <w:t> elements within <w:p>
    t_pattern = re.compile(r'(<w:t[^>]*>)(.*?)(</w:t>)', re.DOTALL)

    def _process_paragraph(match):
        para = match.group(0)
        t_matches = list(t_pattern.finditer(para))
        if len(t_matches) < 2:
            return para  # single or no text runs, nothing to join

        # Concatenate text across all runs in this paragraph
        combined = ''.join(m.group(2) for m in t_matches)

        # Check if any placeholder still exists in the combined text
        new_combined = combined
        changed = False
        for ph, val in safe_map.items():
            if ph in new_combined:
                new_combined = new_combined.replace(ph, val)
                changed = True

        if not changed:
            return para

        # Rebuild: put all text in the first <w:t>, empty the rest
        result = para
        for i, m in enumerate(reversed(t_matches)):
            idx = len(t_matches) - 1 - i
            if idx == 0:
                # First run gets the full replaced text; preserve space
                open_tag = m.group(1)
                if 'xml:space' not in open_tag:
                    open_tag = open_tag.replace('<w:t', '<w:t xml:space="preserve"', 1)
                replacement = f'{open_tag}{new_combined}{m.group(3)}'
            else:
                replacement = f'{m.group(1)}{m.group(3)}'
            result = result[:m.start()] + replacement + result[m.end():]

        return result

    xml_str = re.sub(r'<w:p[ >].*?</w:p>', _process_paragraph, xml_str, flags=re.DOTALL)
    return xml_str


def _fill_template_file(template_file, placeholder_map):
    """
    Apply placeholders to template_file and return a HttpResponse with the
    filled file.
    Uses ZIP-level XML replacement for DOCX to preserve all document features
    (themes, fonts, styles, images, headers/footers, compatibility settings).
    """
    import zipfile
    from xml.sax.saxutils import escape as xml_escape

    template_name = template_file.name or "template"
    ext = template_name.lower().rsplit(".", 1)[-1] if "." in template_name else ""

    # -------- DOCX --------
    if ext == "docx":
        # Build XML-safe replacement map
        # Strip characters illegal in XML 1.0 (OCR can produce control chars)
        _xml_illegal = re.compile(
            r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]'
        )
        safe_map = {}
        for ph, val in placeholder_map.items():
            clean = _xml_illegal.sub('', str(val) if val is not None else "")
            safe_map[ph] = xml_escape(clean)

        # Read original file bytes
        original_data = template_file.read()

        # Process DOCX as ZIP — only touch XML text, preserve everything else
        input_buf = io.BytesIO(original_data)
        output_buf = io.BytesIO()

        with zipfile.ZipFile(input_buf, 'r') as zin:
            with zipfile.ZipFile(output_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)

                    # Only modify XML parts inside word/ that may contain text
                    if item.filename.startswith('word/') and item.filename.endswith('.xml'):
                        try:
                            xml_str = data.decode('utf-8')
                            xml_str = _replace_placeholders_in_docx_xml(xml_str, safe_map)
                            data = xml_str.encode('utf-8')
                        except Exception:
                            pass  # Not a text XML, skip

                    zout.writestr(item, data)

        output_buf.seek(0)
        resp = HttpResponse(
            output_buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        resp["Content-Disposition"] = f'attachment; filename="Filled_{os.path.basename(template_name)}"'
        return resp

    # -------- Excel (XLSX / XLSM) --------
    if ext in ("xlsx", "xlsm"):
        wb = load_workbook(template_file)
        for ws in wb.worksheets:
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell.value, str):
                        txt = cell.value
                        changed = False
                        for ph, val in placeholder_map.items():
                            if ph in txt:
                                txt = txt.replace(ph, str(val) if val is not None else "")
                                changed = True
                        if changed:
                            cell.value = txt

        _apply_print_settings(wb)
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="Filled_{os.path.basename(template_name)}"'
        wb.save(resp)
        return resp

    # -------- Text / CSV --------
    if ext in ("txt", "csv"):
        try:
            content = template_file.read()
        except Exception:
            content = b""
        try:
            text = content.decode("utf-8", errors="ignore")
        except Exception:
            text = str(content)

        for ph, val in placeholder_map.items():
            text = text.replace(ph, str(val) if val is not None else "")

        resp = HttpResponse(text, content_type="text/plain; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="Filled_{os.path.basename(template_name)}"'
        return resp

    return HttpResponse(
        f"Unsupported template type .{ext}. Use DOCX / XLSX / XLSM / TXT / CSV.",
        status=400,
    )


# ============================================
#  SELF-FORMATTED FORMS  -  VIEWS
# ============================================

