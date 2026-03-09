# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import (_apply_print_settings, get_org_from_request,
    create_job_for_excel, _format_date_to_ddmmyyyy)
from .bill_parsing import (parse_estimate_items, parse_workslip_items,
    _extract_header_data_from_sheet, _extract_header_data_fuzzy_from_wb,
    find_estimate_sheet_and_header_row, find_all_estimate_sheets_and_header_rows,
    find_workslip_sheet, find_all_workslip_sheets,
    parse_first_bill_for_nth, parse_nth_bill_for_next,
    read_tp_from_sheet, _detect_bill_format, extract_tp_from_workslip)
from .bill_excel import (create_first_bill_sheet, build_first_bill_wb,
    build_nth_bill_wb, _populate_nth_bill_sheet)

@login_required(login_url='login')
def bill(request):
    """
    Bill generator with four sections:

    SECTION 1: Bill from Estimate
        action = "estimate_first_part"   â†’ CC First & Part Bill
        action = "estimate_first_final"  â†’ CC First & Final Bill

    SECTION 2: Bill from WorkSlip
        action = "workslip_first_part"   â†’ CC First & Part Bill
        action = "workslip_first_final"  â†’ CC First & Final Bill

    SECTION 3: Nth & Part / 2nd & Final from First & Part Bill
        action = "firstpart_nth_part"    â†’ CC Nth & Part Bill
        action = "firstpart_2nd_final"   â†’ CC Second & Final Bill
        - Prev Qty (F) = Column B of uploaded First Bill
        - Prev Amt (G) = Column H of uploaded First Bill
                         (if H is formula with no value â†’ recompute as B * E)

    SECTION 4: Nth & Part / Nth & Final from Nth & Part Bill
        action = "nth_nth_part"          â†’ CC Nth & Part Bill
        action = "nth_nth_final"         â†’ CC Nth & Final Bill
        - Prev Qty (F) = Column C (Quantity Till Date)
        - Prev Amt (G) = Column E (Total Value till date)
                         (if E has no value â†’ recompute as C * D)
    """

    print("DEBUG: Entered bill view; method=", getattr(request, 'method', None))
    # If GET, render the page. For POST we currently return a safe stub
    # (temporary) so the view never returns None while refactoring continues.
    method = getattr(request, 'method', '').upper()
    if method == 'GET':
        # ---- Clear stale Saved Works session data on direct navigation ----
        # When the bill page is accessed directly (not via Saved Works redirect),
        # clear any lingering bill session keys so the module works independently.
        if not request.GET.get('from_saved'):
            for key in [
                'bill_from_workslip', 'bill_ws_rows', 'bill_ws_exec_map',
                'bill_ws_tp_percent', 'bill_ws_tp_type', 'bill_source_work_name',
                'bill_ws_metadata', 'bill_target_number', 'bill_source_work_id',
                'bill_source_work_type', 'bill_parent_work_id',
                'bill_previous_bill_id', 'bill_previous_bill_number',
                'bill_sequence_number', 'bill_type',
            ]:
                request.session.pop(key, None)
            request.session.modified = True

        # Get user's document templates for display
        from core.template_views import get_user_template
        covering_letter_template = get_user_template(request.user, 'covering_letter')
        movement_slip_template = get_user_template(request.user, 'movement_slip')

        # Check if session has pre-loaded data from Saved Works
        bill_from_workslip = request.session.get('bill_from_workslip', False)
        bill_from_saved = False
        bill_preload_items = []
        bill_preload_source = ''
        bill_preload_tp_percent = 0.0
        bill_preload_tp_type = 'Excess'
        bill_preload_count = 0
        bill_preload_total = 0.0

        if bill_from_workslip:
            ws_rows = request.session.get('bill_ws_rows', [])
            ws_exec_map = request.session.get('bill_ws_exec_map', {}) or {}
            if ws_rows:
                bill_from_saved = True
                bill_preload_source = request.session.get('bill_source_work_name', 'WorkSlip')
                bill_preload_tp_percent = float(request.session.get('bill_ws_tp_percent', 0) or 0)
                bill_preload_tp_type = request.session.get('bill_ws_tp_type', 'Excess')
                for idx, row in enumerate(ws_rows):
                    key = row.get('key', f'saved_{idx}')
                    # Try exec_map first, fall back to qty_est if exec_map is empty
                    exec_qty = ws_exec_map.get(key, 0)
                    try:
                        exec_qty = float(exec_qty) if exec_qty else 0.0
                    except (ValueError, TypeError):
                        exec_qty = 0.0
                    # Fallback: if no exec_qty and ws_exec_map is empty,
                    # use qty_est so the bill still shows items
                    if exec_qty <= 0 and not ws_exec_map:
                        try:
                            exec_qty = float(row.get('qty_est', 0) or 0)
                        except (ValueError, TypeError):
                            exec_qty = 0.0
                    if exec_qty <= 0:
                        continue
                    rate = float(row.get('rate', 0) or 0)
                    amount = exec_qty * rate
                    bill_preload_total += amount
                    bill_preload_count += 1
                    # Pick the longest (most detailed) description
                    _desc_candidates = [v for v in [row.get('item_desc', ''), row.get('desc', ''), row.get('display_name', ''), row.get('item_name', '')] if v]
                    _best_desc = max(_desc_candidates, key=len) if _desc_candidates else row.get('item_name', '')
                    bill_preload_items.append({
                        'sl': bill_preload_count,
                        'desc': _best_desc,
                        'unit': row.get('unit', 'Nos'),
                        'qty': exec_qty,
                        'rate': rate,
                        'amount': round(amount, 2),
                    })

        return render(request, "core/bill.html", {
            'covering_letter_template': covering_letter_template,
            'movement_slip_template': movement_slip_template,
            'bill_from_saved': bill_from_saved,
            'bill_preload_items': bill_preload_items,
            'bill_preload_source': bill_preload_source,
            'bill_preload_tp_percent': bill_preload_tp_percent,
            'bill_preload_tp_type': bill_preload_tp_type,
            'bill_preload_count': bill_preload_count,
            'bill_preload_total': round(bill_preload_total, 2),
            'bill_target_number': request.session.get('bill_target_number', 1) if bill_from_saved else 0,
        })

    if method == 'POST':
        action = str(request.POST.get('action') or '').strip()
        bill_type = str(request.POST.get('bill_type') or '').strip()
        uploaded = request.FILES.get('bill_file') or request.FILES.get('file')

        # ── SESSION-BASED BILL GENERATION (from Saved Works) ──
        # When coming from Saved Works, data is pre-loaded in session.
        # If no file is uploaded AND session data exists, generate from session.
        if not uploaded and request.session.get('bill_from_workslip'):
            ws_rows = request.session.get('bill_ws_rows', [])
            ws_exec_map = request.session.get('bill_ws_exec_map', {}) or {}
            ws_tp_percent = float(request.session.get('bill_ws_tp_percent', 0) or 0)
            ws_tp_type = request.session.get('bill_ws_tp_type', 'Excess')
            ws_metadata = request.session.get('bill_ws_metadata', {}) or {}

            if not ws_rows:
                return JsonResponse({"error": "No workslip data found in session. Please upload a file."}, status=400)

            # Convert session rows + exec_map into items format for create_first_bill_sheet
            items = []
            bill_ae_counter = 0
            for idx, row in enumerate(ws_rows):
                key = row.get('key', f'saved_{idx}')
                exec_qty = ws_exec_map.get(key, 0)
                try:
                    exec_qty = float(exec_qty) if exec_qty else 0.0
                except (ValueError, TypeError):
                    exec_qty = 0.0
                # Fallback: use qty_est when exec_map is empty
                if exec_qty <= 0 and not ws_exec_map:
                    try:
                        exec_qty = float(row.get('qty_est', 0) or 0)
                    except (ValueError, TypeError):
                        exec_qty = 0.0
                if exec_qty <= 0:
                    continue
                rate = float(row.get('rate', 0) or 0)
                if rate == 0:
                    continue
                # Pick the longest (most detailed) description available
                _desc_opts = [v for v in [row.get('item_desc', ''), row.get('desc', ''), row.get('display_name', ''), row.get('item_name', '')] if v]
                desc = max(_desc_opts, key=len) if _desc_opts else row.get('item_name', '')
                unit = row.get('unit', 'Nos')
                is_ae = str(desc).lower().startswith('ae')
                item_dict = {
                    'qty': exec_qty,
                    'unit': unit,
                    'desc': desc,
                    'rate': rate,
                    'is_ae': is_ae,
                }
                if is_ae:
                    # Extract AE number from description (e.g. "AE1" -> 1, "AE2" -> 2)
                    import re as _re
                    _ae_match = _re.match(r'[Aa][Ee]\s*(\d+)', desc.strip())
                    if _ae_match:
                        item_dict['ae_number'] = int(_ae_match.group(1))
                    else:
                        bill_ae_counter += 1
                        item_dict['ae_number'] = bill_ae_counter
                items.append(item_dict)

            if not items:
                return JsonResponse({"error": "No executed items found (all quantities are zero)."}, status=400)

            # Extract MB details and dates from POST
            mb_measure_no = str(request.POST.get('mb_measure_no') or '').strip()
            mb_measure_p_from = str(request.POST.get('mb_measure_p_from') or '').strip()
            mb_measure_p_to = str(request.POST.get('mb_measure_p_to') or '').strip()
            mb_abs_no = str(request.POST.get('mb_abstract_no') or '').strip()
            mb_abs_p_from = str(request.POST.get('mb_abstract_p_from') or '').strip()
            mb_abs_p_to = str(request.POST.get('mb_abstract_p_to') or '').strip()
            doi = _format_date_to_ddmmyyyy(request.POST.get('doi') or '')
            doc = _format_date_to_ddmmyyyy(request.POST.get('doc') or '')
            domr = _format_date_to_ddmmyyyy(request.POST.get('domr') or '')
            dobr = _format_date_to_ddmmyyyy(request.POST.get('dobr') or '')

            # Determine bill title from action
            if action == 'workslip_first_final':
                title_text = 'CC First & Final Bill'
            else:
                title_text = 'CC First & Part Bill'

            # Build header_data from session metadata
            header_data = {
                'name_of_work': ws_metadata.get('name_of_work', ''),
                'estimate_amount': ws_metadata.get('estimate_amount', ''),
                'admin_sanction': ws_metadata.get('admin_sanction', ''),
                'tech_sanction': ws_metadata.get('tech_sanction', ''),
                'agreement': ws_metadata.get('agreement', ''),
                'agency': ws_metadata.get('agency', ''),
            }

            wb_out = Workbook()
            create_first_bill_sheet(
                wb_out,
                sheet_name='Bill',
                items=items,
                header_data=header_data,
                title_text=title_text,
                tp_percent=ws_tp_percent,
                tp_type=ws_tp_type,
                mb_measure_no=mb_measure_no,
                mb_measure_p_from=mb_measure_p_from,
                mb_measure_p_to=mb_measure_p_to,
                mb_abs_no=mb_abs_no,
                mb_abs_p_from=mb_abs_p_from,
                mb_abs_p_to=mb_abs_p_to,
                doi=doi,
                doc=doc,
                domr=domr,
                dobr=dobr,
            )

            _apply_print_settings(wb_out)
            resp = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            resp['Content-Disposition'] = 'attachment; filename="Bill_from_WorkSlip.xlsx"'
            wb_out.save(resp)
            print(f"DEBUG: Generated bill from session data ({len(items)} items)")
            return resp

        if not uploaded:
            return JsonResponse({"error": "no uploaded file"}, status=400)

        # Extract MB details from POST request
        mb_measure_no = str(request.POST.get('mb_measure_no') or '').strip()
        mb_measure_p_from = str(request.POST.get('mb_measure_p_from') or '').strip()
        mb_measure_p_to = str(request.POST.get('mb_measure_p_to') or '').strip()
        mb_abs_no = str(request.POST.get('mb_abstract_no') or '').strip()
        mb_abs_p_from = str(request.POST.get('mb_abstract_p_from') or '').strip()
        mb_abs_p_to = str(request.POST.get('mb_abstract_p_to') or '').strip()
        
        # Extract dates from POST request and convert from yyyy-mm-dd to dd-mm-yyyy format
        doi = _format_date_to_ddmmyyyy(request.POST.get('doi') or '')
        doc = _format_date_to_ddmmyyyy(request.POST.get('doc') or '')
        domr = _format_date_to_ddmmyyyy(request.POST.get('domr') or '')
        dobr = _format_date_to_ddmmyyyy(request.POST.get('dobr') or '')

        try:
            wb = load_workbook(uploaded, data_only=True)
        except Exception as e:
            return JsonResponse({"error": f"failed to read uploaded workbook: {e}"}, status=400)

        # Helper function to generate bill title (used by both estimate and workslip)
        def get_bill_title(action, idx, total):
            """Generate appropriate bill title based on action (button clicked)."""
            if not action:
                # No action, just use "Bill" without numbers
                return "CC Bill"
            
            # Map action to bill type descriptions with CC prefix
            type_map = {
                "estimate_first_part": "CC First & Part Bill",
                "estimate_first_final": "CC First & Final Bill",
                "workslip_first_part": "CC First & Part Bill",
                "workslip_first_final": "CC First & Final Bill",
                "firstpart_nth_part": "CC Nth & Part Bill",
                "firstpart_2nd_final": "CC 2nd & Final Bill",
                "nth_nth_part": "CC Nth & Part Bill",
                "nth_nth_final": "CC Nth & Final Bill",
            }
            
            # Get the bill type name from the action
            bill_type_name = type_map.get(action, "CC Bill")
            
            # If only one bill, return type name without numbering
            if total <= 1:
                return bill_type_name
            
            # Otherwise return type name (no numbers added, just the type)
            return bill_type_name

        # Support estimate -> first bill (minimal)
        if action.startswith('estimate_'):
            # Try to detect obvious estimate sheets first
            pairs = find_all_estimate_sheets_and_header_rows(wb)

            # If detection failed, fallback to trying every worksheet with a
            # conservative header guess. We'll keep only sheets that parse
            # into at least one item.
            if not pairs:
                candidates = []
                for ws in wb.worksheets:
                    items_try = parse_estimate_items(ws, 3)
                    if items_try:
                        candidates.append((ws, 3))
                if not candidates:
                    # final fallback: try header_row=2 and 4 as last resort
                    for ws in wb.worksheets:
                        for hr in (2, 4):
                            items_try = parse_estimate_items(ws, hr)
                            if items_try:
                                candidates.append((ws, hr))
                                break
                pairs = candidates

            if not pairs:
                return JsonResponse({"error": "no estimate-like sheets found"}, status=400)

            # Build output workbook with one Bill sheet per parsed estimate sheet.
            wb_out = Workbook()
            created = 0
            for idx, (ws, header_row) in enumerate(pairs, start=1):
                items = parse_estimate_items(ws, header_row)
                if not items:
                    continue
                created += 1
                
                # Extract header data from this specific sheet
                sheet_header_data = _extract_header_data_from_sheet(ws)
                
                # Extract T.P percentage and type from workslip
                tp_percent, tp_type = extract_tp_from_workslip(ws)
                
                # Generate appropriate title
                title_text = get_bill_title(action, idx, len(pairs))
                
                # Only add sheet numbering if there are multiple sheets
                if len(pairs) > 1:
                    sheet_name = f"Bill-{created}"
                else:
                    sheet_name = "Bill"
                
                create_first_bill_sheet(
                    wb_out,
                    sheet_name=sheet_name,
                    items=items,
                    header_data=sheet_header_data,
                    title_text=title_text,
                    tp_percent=None,
                    tp_type=None,
                    mb_measure_no=mb_measure_no, 
                    mb_measure_p_from=mb_measure_p_from, 
                    mb_measure_p_to=mb_measure_p_to,
                    mb_abs_no=mb_abs_no, 
                    mb_abs_p_from=mb_abs_p_from, 
                    mb_abs_p_to=mb_abs_p_to,
                    doi=doi, 
                    doc=doc, 
                    domr=domr, 
                    dobr=dobr
                )

            if created == 0:
                return JsonResponse({"error": "no items parsed from upload"}, status=400)

            _apply_print_settings(wb_out)
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="Bills.xlsx"'
            wb_out.save(resp)
            return resp

        # Support workslip -> first bill
        if action.startswith('workslip_'):
            # Try to detect workslip sheets
            workslip_sheets = find_all_workslip_sheets(wb)
            # Defaults in case extraction fails
            tp_percent = 0.0
            tp_type = "Excess"
            
            # Build output workbook with one Bill sheet per workslip sheet
            wb_out = Workbook()
            created = 0
            for idx, ws in enumerate(workslip_sheets, start=1):
                items = parse_workslip_items(ws)
                if not items:
                    continue
                created += 1
                
                # Extract header data from this specific sheet
                sheet_header_data = _extract_header_data_from_sheet(ws)
                
                # Extract T.P percentage and type from workslip
                try:
                    tp_percent, tp_type = extract_tp_from_workslip(ws)
                    print(f"DEBUG: Workslip '{ws.title}' TP extracted percent={tp_percent} type={tp_type}")
                except Exception as e:
                    print(f"DEBUG: Failed to extract TP from workslip '{ws.title}': {e}")
                    tp_percent, tp_type = 0.0, "Excess"
                
                # Generate appropriate title
                title_text = get_bill_title(action, idx, len(workslip_sheets))
                
                # Only add sheet numbering if there are multiple sheets
                if len(workslip_sheets) > 1:
                    sheet_name = f"Bill-{created}"
                else:
                    sheet_name = "Bill"
                
                create_first_bill_sheet(
                    wb_out,
                    sheet_name=sheet_name,
                    items=items,
                    header_data=sheet_header_data,
                    title_text=title_text,
                    tp_percent=tp_percent,
                    tp_type=tp_type,
                    mb_measure_no=mb_measure_no, 
                    mb_measure_p_from=mb_measure_p_from, 
                    mb_measure_p_to=mb_measure_p_to,
                    mb_abs_no=mb_abs_no, 
                    mb_abs_p_from=mb_abs_p_from, 
                    mb_abs_p_to=mb_abs_p_to,
                    doi=doi, 
                    doc=doc, 
                    domr=domr, 
                    dobr=dobr
                )

            if created == 0:
                return JsonResponse({"error": "no workslip items parsed from upload"}, status=400)

            _apply_print_settings(wb_out)
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="Bills.xlsx"'
            wb_out.save(resp)
            return resp

        # Support Nth/2nd bills from First Bill (temporary shim)
        # Support Nth/2nd bills from First Bill (multi-sheet support)
        if action in ("firstpart_nth_part", "firstpart_2nd_final"):
            # Find all bill sheets (from multiple estimates)
            bill_sheets = [ws for ws in wb.worksheets if ws.title.startswith("Bill")]
            print(f"DEBUG: Found {len(bill_sheets)} sheets starting with 'Bill': {[ws.title for ws in bill_sheets]}")
            
            if not bill_sheets:
                # Fallback: find all sheets that look like estimate/bill sheets
                bill_sheets = []
                for ws in wb.worksheets:
                    print(f"DEBUG: Checking sheet '{ws.title}'...")
                    for r in range(1, 30):
                        a = str(ws.cell(row=r, column=1).value or "").strip().lower()
                        b = str(ws.cell(row=r, column=2).value or "").strip().lower()
                        d = str(ws.cell(row=r, column=4).value or "").strip().lower()
                        if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
                            if ws not in bill_sheets:
                                bill_sheets.append(ws)
                            print(f"DEBUG: Found bill-like sheet '{ws.title}' with header at row {r}")
                            break
                
                print(f"DEBUG: Fallback found {len(bill_sheets)} bill sheets: {[ws.title for ws in bill_sheets]}")
            
            if not bill_sheets:
                # Last fallback: try all non-empty sheets
                bill_sheets = [ws for ws in wb.worksheets if ws.max_row > 1]
                print(f"DEBUG: No bill sheets found by fallback, using all non-empty sheets: {[ws.title for ws in bill_sheets]}")
            
            if not bill_sheets:
                bill_sheets = [wb.worksheets[0]]  # fallback to first sheet
                print(f"DEBUG: No bill sheets found, using first sheet '{bill_sheets[0].title}'")
            
            print(f"DEBUG: Processing {len(bill_sheets)} bill sheets total: {[ws.title for ws in bill_sheets]}")
            
            # Extract Nth number
            if action == "firstpart_nth_part":
                nth_str = request.POST.get("nth_number", "").strip()
                try:
                    nth_val = int(nth_str)
                except Exception:
                    nth_val = 2
                if nth_val < 2:
                    nth_val = 2
                ord_word = ordinal_word(nth_val)
                # Prefer bill_type (dropdown) if present, else action
                which = bill_type if bill_type in ("firstpart_nth_part", "firstpart_2nd_final") else action
                if which == "firstpart_nth_part":
                    title_text = f"CC {ord_word} & Part Bill"
                else:
                    title_text = f"CC {ord_word} & Final Bill"
            else:
                ord_word = ordinal_word(2)
                title_text = f"CC {ord_word} & Final Bill"

            # Build output workbook with one Nth bill sheet per input bill sheet
            wb_out = Workbook()
            created = 0
            
            for idx, ws_first in enumerate(bill_sheets, start=1):
                print(f"DEBUG: Processing sheet {idx}/{len(bill_sheets)}: '{ws_first.title}'")
                
                # Find header row in this sheet
                header_row = None
                for r in range(1, 30):
                    a = str(ws_first.cell(row=r, column=1).value or "").strip().lower()
                    b = str(ws_first.cell(row=r, column=2).value or "").strip().lower()
                    d = str(ws_first.cell(row=r, column=4).value or "").strip().lower()
                    if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
                        header_row = r
                        print(f"DEBUG: Found header row {r} in sheet '{ws_first.title}'")
                        break
                
                if header_row is None:
                    header_row = 10  # default fallback
                    print(f"DEBUG: No header found, using default row 10")
                
                header_data = _extract_header_data_from_sheet(ws_first)
                items = parse_first_bill_for_nth(ws_first, header_row)
                print(f"DEBUG: Parsed {len(items)} items from sheet '{ws_first.title}'")

                if not items:
                    print(f"DEBUG: No items found, skipping sheet '{ws_first.title}'")
                    continue
                created += 1

                tp_percent, tp_type = read_tp_from_sheet(ws_first)
                if tp_percent is None:
                    tp_percent = 0.0
                if tp_type is None:
                    tp_type = "Excess"

                # Determine sheet name
                if len(bill_sheets) > 1:
                    sheet_name = f"Bill-{created}"
                else:
                    sheet_name = "Bill"

                # Keep a single title text even when multiple sheets are generated
                sheet_title = title_text

                # Create Nth bill sheet
                if created == 1:
                    ws_nth = wb_out.active
                    ws_nth.title = sheet_name
                    print(f"DEBUG: Using active sheet for output, naming it '{sheet_name}'")
                else:
                    ws_nth = wb_out.create_sheet(title=sheet_name)
                    print(f"DEBUG: Created new sheet '{sheet_name}'")

                # Populate this Nth bill sheet
                print(f"DEBUG: Populating sheet '{sheet_name}' with {len(items)} items")
                _populate_nth_bill_sheet(
                    ws_nth,
                    items=items,
                    header_data=header_data,
                    title_text=sheet_title,
                    tp_percent=tp_percent,
                    tp_type=tp_type,
                    mb_measure_no=mb_measure_no,
                    mb_measure_p_from=mb_measure_p_from,
                    mb_measure_p_to=mb_measure_p_to,
                    mb_abs_no=mb_abs_no,
                    mb_abs_p_from=mb_abs_p_from,
                    mb_abs_p_to=mb_abs_p_to,
                    doi=doi,
                    doc=doc,
                    domr=domr,
                    dobr=dobr,
                )
                print(f"DEBUG: Populated sheet '{sheet_name}' successfully")

            print(f"DEBUG: Created {created} output sheets total")

            if created == 0:
                return JsonResponse({"error": "no items parsed from any First Bill sheets"}, status=400)

            _apply_print_settings(wb_out)
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = "Nth_Bill_from_FirstPart.xlsx" if action == "firstpart_nth_part" else "Second_Final_from_FirstPart.xlsx"
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb_out.save(resp)
            return resp

        # Support Nth from Nth bill (temporary shim)
        # Support Nth from Nth bill (multi-sheet support)
        if action in ("nth_nth_part", "nth_nth_final"):
            # Find all Nth bill sheets
            bill_sheets = [ws for ws in wb.worksheets if ws.title.startswith("Bill")]
            print(f"DEBUG (Nthâ†’Nth): Found {len(bill_sheets)} sheets starting with 'Bill': {[ws.title for ws in bill_sheets]}")
            
            if not bill_sheets:
                # Fallback: find sheets with "Quantity Till Date" header
                bill_sheets = []
                for ws in wb.worksheets:
                    print(f"DEBUG (Nthâ†’Nth): Checking sheet '{ws.title}'...")
                    for r in range(1, 30):
                        c = str(ws.cell(row=r, column=3).value or "").strip().lower()
                        if "quantity till date" in c:
                            if ws not in bill_sheets:
                                bill_sheets.append(ws)
                            print(f"DEBUG (Nthâ†’Nth): Found Nth-bill-like sheet '{ws.title}' with header at row {r}")
                            break
                
                print(f"DEBUG (Nthâ†’Nth): Fallback found {len(bill_sheets)} bill sheets: {[ws.title for ws in bill_sheets]}")
            
            if not bill_sheets:
                # Last fallback: try all non-empty sheets
                bill_sheets = [ws for ws in wb.worksheets if ws.max_row > 1]
                print(f"DEBUG (Nthâ†’Nth): No bill sheets found by fallback, using all non-empty sheets: {[ws.title for ws in bill_sheets]}")
            
            if not bill_sheets:
                bill_sheets = [wb.worksheets[0]]  # fallback to first sheet
                print(f"DEBUG (Nthâ†’Nth): No bill sheets found, using first sheet '{bill_sheets[0].title}'")
            
            print(f"DEBUG (Nthâ†’Nth): Processing {len(bill_sheets)} bill sheets total: {[ws.title for ws in bill_sheets]}")
            
            # Extract Nth number
            nth_str = request.POST.get("nth_number", "").strip()
            try:
                nth_val = int(nth_str)
            except Exception:
                nth_val = 2
            if nth_val < 2:
                nth_val = 2
            ord_word = ordinal_word(nth_val)

            # Prefer bill_type (dropdown) if present, else action
            which = bill_type if bill_type in ("nth_nth_part", "nth_nth_final") else action
            if which == "nth_nth_part":
                title_text = f"CC {ord_word} & Part Bill"
            else:
                title_text = f"CC {ord_word} & Final Bill"

            # Build output workbook with one Nth bill sheet per input sheet
            wb_out = Workbook()
            created = 0
            
            print(f"DEBUG (Nthâ†’Nth): Processing {len(bill_sheets)} input bill sheets")
            
            for idx, ws_nth in enumerate(bill_sheets, start=1):
                print(f"DEBUG (Nthâ†’Nth): Processing sheet {idx}/{len(bill_sheets)}: '{ws_nth.title}'")
                
                # Find header row in this sheet
                header_row = None
                for r in range(1, 30):
                    a = str(ws_nth.cell(row=r, column=1).value or "").strip().lower()
                    c = str(ws_nth.cell(row=r, column=3).value or "").strip().lower()
                    if "sl" in a and "quantity till date" in c:
                        header_row = r
                        print(f"DEBUG (Nthâ†’Nth): Found header row {r} in sheet '{ws_nth.title}'")
                        break
                
                if header_row is None:
                    header_row = 10  # default fallback
                    print(f"DEBUG (Nthâ†’Nth): No header found, using default row 10")
                
                header_data = _extract_header_data_from_sheet(ws_nth)
                items = parse_nth_bill_for_next(ws_nth, header_row)
                print(f"DEBUG (Nthâ†’Nth): Parsed {len(items)} items from sheet '{ws_nth.title}'")

                if not items:
                    print(f"DEBUG (Nthâ†’Nth): No items found, skipping sheet '{ws_nth.title}'")
                    continue
                created += 1

                tp_percent, tp_type = read_tp_from_sheet(ws_nth)
                if tp_percent is None:
                    tp_percent = 0.0
                if tp_type is None:
                    tp_type = "Excess"

                # Determine sheet name
                if len(bill_sheets) > 1:
                    sheet_name = f"Bill-{created}"
                else:
                    sheet_name = "Bill"

                # Keep a single title text even when multiple sheets are generated
                sheet_title = title_text

                # Create Nth bill sheet
                if created == 1:
                    ws_out = wb_out.active
                    ws_out.title = sheet_name
                    print(f"DEBUG (Nthâ†’Nth): Using active sheet for output, naming it '{sheet_name}'")
                else:
                    ws_out = wb_out.create_sheet(title=sheet_name)
                    print(f"DEBUG (Nthâ†’Nth): Created new sheet '{sheet_name}'")

                # Populate this Nth bill sheet
                print(f"DEBUG (Nthâ†’Nth): Populating sheet '{sheet_name}' with {len(items)} items")
                _populate_nth_bill_sheet(
                    ws_out,
                    items=items,
                    header_data=header_data,
                    title_text=sheet_title,
                    tp_percent=tp_percent,
                    tp_type=tp_type,
                    mb_measure_no=mb_measure_no,
                    mb_measure_p_from=mb_measure_p_from,
                    mb_measure_p_to=mb_measure_p_to,
                    mb_abs_no=mb_abs_no,
                    mb_abs_p_from=mb_abs_p_from,
                    mb_abs_p_to=mb_abs_p_to,
                    doi=doi,
                    doc=doc,
                    domr=domr,
                    dobr=dobr,
                )
                print(f"DEBUG (Nthâ†’Nth): Populated sheet '{sheet_name}' successfully")

            print(f"DEBUG (Nthâ†’Nth): Created {created} output sheets total")
            
            if created == 0:
                return JsonResponse({"error": "no items parsed from any Nth Bill sheets"}, status=400)

            _apply_print_settings(wb_out)
            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = "Nth_Bill_from_NthPart.xlsx" if action == "nth_nth_part" else "Nth_Final_from_NthPart.xlsx"
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb_out.save(resp)
            return resp

        return JsonResponse({"error": "action not supported in temporary handler"}, status=501)

    return HttpResponseNotAllowed(['GET', 'POST'])
