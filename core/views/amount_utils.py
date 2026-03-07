# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import (_apply_print_settings, _format_indian_number,
    _number_to_words_rupees, _format_date_to_ddmmyyyy)
from .bill_parsing import (_detect_bill_format, _extract_total_amount_from_bill_wb,
    to_number, is_merged_cell, find_estimate_sheet_and_header_row,
    find_workslip_sheet, read_tp_from_sheet, extract_tp_from_workslip)

def _to_number_amt(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def _is_merged_cell_amt(ws, row, col):
    """
    Local helper for amount extraction:
    True if (row, col) lies in any merged range on this sheet.
    """
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return True
    return False


def _find_estimate_sheet_header_for_amount(wb):
    """
    Same header logic as your find_estimate_sheet_and_header_row:
      A: 'Sl', B: 'Quantity', D: 'Item'/'Description'
    """
    for ws in wb.worksheets:
        for r in range(1, 26):
            a = str(ws.cell(row=r, column=1).value or "").strip().lower()
            b = str(ws.cell(row=r, column=2).value or "").strip().lower()
            d = str(ws.cell(row=r, column=4).value or "").strip().lower()
            if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
                return ws, r
    return wb.worksheets[0], 3


def _parse_estimate_items_for_amount(ws, header_row):
    """
    Mirror of your parse_estimate_items(), but only returns qty & rate for
    computing subtotal in Python.

    Format:
      A: Sl.No, B: Quantity, C: Unit, D: Item, E: Rate, F: Per, G: Unit, H: Amount
    """
    items = []
    max_row = min(ws.max_row, 5000)

    for r in range(header_row + 1, max_row + 1):
        desc_raw = ws.cell(row=r, column=4).value  # D
        desc = str(desc_raw or "").strip()
        rate_raw = ws.cell(row=r, column=5).value  # E
        amt_raw = ws.cell(row=r, column=8).value   # H

        desc_low = desc.lower()

        # STOP on ECV / Sub Total / Total
        if desc_low.startswith(("ecv", "sub total", "subtotal", "total")):
            break

        rate_str = "" if rate_raw is None else str(rate_raw).strip()
        is_rate_empty = (rate_str == "")

        # Heading row: description present, but rate empty OR description cell merged
        if desc and (is_rate_empty or _is_merged_cell_amt(ws, r, 4)):
            continue

        qty_raw = ws.cell(row=r, column=2).value  # B
        unit_raw = ws.cell(row=r, column=3).value  # C

        qty_str = "" if qty_raw is None else str(qty_raw).strip()
        unit_str = "" if unit_raw is None else str(unit_raw).strip()
        amt_str = "" if amt_raw is None else str(amt_raw).strip()

        all_blank = (
            desc == "" and
            is_rate_empty and
            qty_str == "" and
            unit_str == "" and
            amt_str == ""
        )
        if all_blank:
            continue

        if desc == "" and is_rate_empty:
            continue

        qty = _to_number_amt(qty_raw)
        rate = _to_number_amt(rate_raw)
        amt = _to_number_amt(amt_raw)

        # if only Amount given, but qty/rate missing, use that
        if amt != 0:
            if qty == 0 and rate != 0:
                qty = amt / rate
            elif rate == 0 and qty != 0:
                rate = amt / qty

        items.append({
            "qty": qty,
            "rate": rate,
        })

    return items


def _find_workslip_sheet_for_amount(wb):
    """
    Same idea as your find_workslip_sheet.
    """
    for ws in wb.worksheets:
        for r in range(1, 40):
            b = str(ws.cell(row=r, column=2).value or "").strip().lower()
            g = str(ws.cell(row=r, column=7).value or "").strip().lower()
            if "description" in b and ("qty" in g or "quantity" in g):
                return ws
    return wb.worksheets[0]


def _parse_workslip_items_for_amount(ws):
    """
    Mirror of parse_workslip_items(), but only returns qty & rate for subtotal.
    Automatically detects and uses the LAST workslip phase columns.
    Rate is always from Estimate Rate column for multi-phase workslips.
    """
    items = []
    max_row = min(ws.max_row, 5000)
    
    # Step 1: Find header row
    header_row = 8
    for r in range(1, 15):
        cell_val = str(ws.cell(row=r, column=1).value or "").strip().lower()
        if "sl" in cell_val:
            header_row = r
            break
    
    # Step 2: Scan header row to find all workslip/execution qty columns
    qty_columns = []
    amt_columns = []
    estimate_rate_col = 5  # Default - Rate (Estimate) column
    
    for c in range(1, 40):  # Scan more columns for multi-phase workslips
        header = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
        
        is_exec_col = ("execution" in header or "exec" in header or "workslip" in header)
        is_qty = ("qty" in header or "quantity" in header)
        is_amt = ("amount" in header or "amt" in header)
        is_estimate = ("est" in header or "estimate" in header)
        is_rate = ("rate" in header)
        
        # Find estimate rate column
        if is_estimate and is_rate:
            estimate_rate_col = c
        
        if is_estimate:
            continue
        
        # Skip More/Less/Remarks columns
        if "more" in header or "less" in header or "remark" in header:
            continue
            
        if is_exec_col and is_qty:
            qty_columns.append(c)
        elif is_exec_col and is_amt:
            amt_columns.append(c)
    
    # Step 3: Use the LAST workslip columns (latest phase)
    if qty_columns:
        qty_col = qty_columns[-1]  # Last qty column = latest workslip phase
        amt_col = amt_columns[-1] if amt_columns else qty_col + 1
        rate_col = estimate_rate_col  # Always use Estimate Rate
    else:
        # Fallback to old format
        qty_col = 7
        rate_col = 8
        amt_col = 9

    for r in range(1, max_row + 1):
        desc_raw = ws.cell(row=r, column=2).value  # B
        desc = str(desc_raw or "").strip()
        if desc == "":
            continue

        low = desc.lower()

        if low == "description of item":
            continue
        if low.startswith("sub total") or low.startswith("sub-total") or low.startswith("ecv"):
            break
        if "supplemental items" in low:
            continue

        if _is_merged_cell_amt(ws, r, 2):
            continue

        qty_raw = ws.cell(row=r, column=qty_col).value
        rate_raw = ws.cell(row=r, column=rate_col).value
        amt_raw = ws.cell(row=r, column=amt_col).value

        qty_exec = _to_number_amt(qty_raw)
        rate_exec = _to_number_amt(rate_raw)
        amt_exec = _to_number_amt(amt_raw)

        if amt_exec != 0:
            if qty_exec == 0 and rate_exec != 0:
                qty_exec = amt_exec / rate_exec
            elif rate_exec == 0 and qty_exec != 0:
                rate_exec = amt_exec / qty_exec

        # Pure heading: text but no qty, no rate, no amount
        if qty_exec == 0 and rate_exec == 0 and amt_exec == 0:
            continue

        items.append({
            "qty": qty_exec,
            "rate": rate_exec,
        })

    return items


def _read_tp_from_sheet_for_amount(ws):
    """
    Same T.P reader as your read_tp_from_sheet().
    Looks in B or D for 'T.P', starting with 'Add' or 'Deduct'.
    """
    tp_percent = None
    tp_type = None
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        for col in (2, 4):
            val = ws.cell(row=r, column=col).value
            if not val:
                continue
            s = str(val).strip()
            low = s.lower()
            if "t.p" in low:
                if low.startswith("add"):
                    tp_type = "Excess"
                elif low.startswith("deduct"):
                    tp_type = "Less"
                m = re.search(r"(\d+(\.\d+)?)", low)
                if m:
                    tp_percent = float(m.group(1))
                return tp_percent, tp_type
    return None, None


def _extract_total_amount_for_action(wb, action: str, request=None) -> float:
    """
    Compute the TOTAL amount for the bill that will be generated, using the
    SAME logic as your bill() function.

    - For estimate_first_part / estimate_first_final:
        * Treat uploaded file as Estimate
        * Find ALL estimate-like sheets (like bill() does)
        * For each sheet:
              subtotal_sheet = sum(Qty * Rate) over parsed items
              TP_sheet = subtotal_sheet * abs(tp_percent)/100
              total_sheet = subtotal_sheet +/- TP_sheet
        * Return sum of total_sheet over all such sheets.
          (This matches multi-sheet Bill_from_Estimate.xlsx.)

    - For workslip_first_part / workslip_first_final:
        * Treat uploaded file as WorkSlip
        * Find ALL workslip-like sheets (like bill() does)
        * For each sheet:
              subtotal_sheet = sum(Qty(Exec) * Rate(Exec)) over parsed items
              TP_sheet from sheet T.P row or session fallback
              total_sheet = subtotal_sheet +/- TP_sheet
        * Return sum of total_sheet over all such sheets.
          (Matches multi-sheet Bill_from_WorkSlip.xlsx.)

    - For all other actions (Nth & Part / Nth & Final):
        * We assume the uploaded Excel is already a Bill
        * Use _extract_total_amount_from_bill_wb(wb)
          which:
              - for 8-col bills reads Column H (Total row)
              - for 10-col Nth bills reads Column I (Total row)
    """
    action = (action or "").strip()

    # ---------- Small local helpers ----------

    def to_num(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    def is_merged(ws, row, col):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return True
        return False

    # --- ESTIMATE HELPERS (copied from your bill() logic, but only for amount) ---

    def looks_like_estimate_header(ws, r):
        a = str(ws.cell(row=r, column=1).value or "").strip().lower()
        b = str(ws.cell(row=r, column=2).value or "").strip().lower()
        c = str(ws.cell(row=r, column=3).value or "").strip().lower()
        d = str(ws.cell(row=r, column=4).value or "").strip().lower()
        e = str(ws.cell(row=r, column=5).value or "").strip().lower()

        has_sl = ("sl" in a) or ("s.no" in a) or ("serial" in a)
        has_qty = ("qty" in b or "quantity" in b or "qty" in c or "quantity" in c)
        has_desc = (
            "item" in c or "item" in d or "item" in e or
            "description" in c or "description" in d or "description" in e
        )
        return has_sl and has_qty and has_desc

    def find_all_estimate_sheets_amt(wb_inner):
        results = []
        for ws in wb_inner.worksheets:
            header_row = None
            max_scan = min(ws.max_row, 60)
            for r in range(1, max_scan + 1):
                if looks_like_estimate_header(ws, r):
                    header_row = r
                    break
            if header_row:
                results.append((ws, header_row))

        if not results:
            # fallback similar to find_estimate_sheet_and_header_row
            for ws in wb_inner.worksheets:
                for r in range(1, 26):
                    a = str(ws.cell(row=r, column=1).value or "").strip().lower()
                    b = str(ws.cell(row=r, column=2).value or "").strip().lower()
                    d = str(ws.cell(row=r, column=4).value or "").strip().lower()
                    if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
                        return [(ws, r)]
            # ultimate fallback
            return [(wb_inner.worksheets[0], 3)]
        return results

    def parse_estimate_items_amt(ws, header_row):
        items = []
        max_row = min(ws.max_row, 5000)

        for r in range(header_row + 1, max_row + 1):
            desc_raw = ws.cell(row=r, column=4).value  # D
            desc = str(desc_raw or "").strip()
            rate_raw = ws.cell(row=r, column=5).value  # E
            amt_raw = ws.cell(row=r, column=8).value   # H

            desc_low = desc.lower()

            # STOP on ECV / Sub Total / Total
            if desc_low.startswith(("ecv", "sub total", "subtotal", "total")):
                break

            rate_str = "" if rate_raw is None else str(rate_raw).strip()
            is_rate_empty = (rate_str == "")

            # Heading row: description present, but rate empty OR description cell merged
            if desc and (is_rate_empty or is_merged(ws, r, 4)):
                continue

            qty_raw = ws.cell(row=r, column=2).value  # B
            unit_raw = ws.cell(row=r, column=3).value  # C

            qty_str = "" if qty_raw is None else str(qty_raw).strip()
            unit_str = "" if unit_raw is None else str(unit_raw).strip()
            amt_str = "" if amt_raw is None else str(amt_raw).strip()

            all_blank = (
                desc == "" and
                is_rate_empty and
                qty_str == "" and
                unit_str == "" and
                amt_str == ""
            )
            if all_blank:
                continue

            if desc == "" and is_rate_empty:
                continue

            qty = to_num(qty_raw)
            rate = to_num(rate_raw)
            amt = to_num(amt_raw)

            # If only Amount given, infer qty or rate
            if amt != 0:
                if qty == 0 and rate != 0:
                    qty = amt / rate
                elif rate == 0 and qty != 0:
                    rate = amt / qty

            items.append({"qty": qty, "rate": rate})

        return items

    # --- WORKSLIP HELPERS (copied from your bill() logic, but only for amount) ---

    def find_all_workslip_sheets_amt(wb_inner):
        results = []
        for ws in wb_inner.worksheets:
            max_scan = min(ws.max_row, 60)
            for r in range(1, max_scan + 1):
                b = str(ws.cell(row=r, column=2).value or "").strip().lower()
                c = str(ws.cell(row=r, column=3).value or "").strip().lower()

                has_desc = (
                    "description of item" in b or
                    "description of item" in c or
                    ("description" in b and "item" in b) or
                    ("description" in c and "item" in c)
                )

                has_qty = False
                for col in range(5, 12):
                    t = str(ws.cell(row=r, column=col).value or "").strip().lower()
                    if "qty" in t or "quantity" in t:
                        has_qty = True
                        break

                if has_desc and has_qty:
                    results.append(ws)
                    break

        if not results:
            # fallback similar to your find_workslip_sheet
            for ws in wb_inner.worksheets:
                for r in range(1, 40):
                    b = str(ws.cell(row=r, column=2).value or "").strip().lower()
                    g = str(ws.cell(row=r, column=7).value or "").strip().lower()
                    if "description" in b and ("qty" in g or "quantity" in g):
                        return [ws]
            return [wb_inner.worksheets[0]]
        return results

    def parse_workslip_items_amt(ws):
        """
        Parse workslip items for amount calculation, detecting the last workslip phase columns.
        Rate is always from Estimate Rate column for multi-phase workslips.
        """
        items = []
        max_row = min(ws.max_row, 5000)
        
        # Step 1: Find header row
        header_row = 8
        for r in range(1, 15):
            cell_val = str(ws.cell(row=r, column=1).value or "").strip().lower()
            if "sl" in cell_val:
                header_row = r
                break
        
        # Step 2: Scan header row to find all workslip/execution qty columns
        qty_columns = []
        amt_columns = []
        estimate_rate_col = 5  # Default - Rate (Estimate) column
        
        for c in range(1, 40):  # Scan more columns for multi-phase workslips
            header = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
            
            is_exec_col = ("execution" in header or "exec" in header or "workslip" in header)
            is_qty = ("qty" in header or "quantity" in header)
            is_amt = ("amount" in header or "amt" in header)
            is_estimate = ("est" in header or "estimate" in header)
            is_rate = ("rate" in header)
            
            # Find estimate rate column
            if is_estimate and is_rate:
                estimate_rate_col = c
            
            if is_estimate:
                continue
            
            # Skip More/Less/Remarks columns
            if "more" in header or "less" in header or "remark" in header:
                continue
                
            if is_exec_col and is_qty:
                qty_columns.append(c)
            elif is_exec_col and is_amt:
                amt_columns.append(c)
        
        # Step 3: Use the LAST workslip columns (latest phase)
        if qty_columns:
            qty_col = qty_columns[-1]
            amt_col = amt_columns[-1] if amt_columns else qty_col + 1
            rate_col = estimate_rate_col  # Always use Estimate Rate
        else:
            qty_col = 7
            rate_col = 8
            amt_col = 9

        for r in range(1, max_row + 1):
            desc_raw = ws.cell(row=r, column=2).value  # B
            desc = str(desc_raw or "").strip()
            if desc == "":
                continue

            low = desc.lower()

            if low == "description of item":
                continue
            if low.startswith("sub total") or low.startswith("sub-total") or low.startswith("ecv"):
                break
            if "supplemental items" in low:
                continue

            if is_merged(ws, r, 2):
                continue

            qty_raw = ws.cell(row=r, column=qty_col).value
            rate_raw = ws.cell(row=r, column=rate_col).value
            amt_raw = ws.cell(row=r, column=amt_col).value

            qty_exec = to_num(qty_raw)
            rate_exec = to_num(rate_raw)
            amt_exec = to_num(amt_raw)

            if amt_exec != 0:
                if qty_exec == 0 and rate_exec != 0:
                    qty_exec = amt_exec / rate_exec
                elif rate_exec == 0 and qty_exec != 0:
                    rate_exec = amt_exec / qty_exec

            # Pure heading: text but no qty, no rate, no amount
            if qty_exec == 0 and rate_exec == 0 and amt_exec == 0:
                continue

            items.append({"qty": qty_exec, "rate": rate_exec})

        return items

    def read_tp_from_sheet_amt(ws):
        tp_percent = None
        tp_type = None
        max_row = ws.max_row
        for r in range(1, max_row + 1):
            for col in (2, 4):
                val = ws.cell(row=r, column=col).value
                if not val:
                    continue
                s = str(val).strip()
                low = s.lower()
                if "t.p" in low:
                    if low.startswith("add"):
                        tp_type = "Excess"
                    elif low.startswith("deduct"):
                        tp_type = "Less"
                    m = re.search(r"(\d+(\.\d+)?)", low)
                    if m:
                        tp_percent = float(m.group(1))
                    return tp_percent, tp_type
        return None, None

    # --------------- CASE 1: Estimate â†’ First / First & Final ---------------

    if action in ("estimate_first_part", "estimate_first_final"):
        sheets_info = find_all_estimate_sheets_amt(wb)
        if not sheets_info:
            return 0.0

        total_all_sheets = 0.0

        # T.P from session (same as bill())
        tp_percent_session = 0.0
        tp_type_session = "Excess"
        if request is not None:
            tp_percent_session = to_num(request.session.get("ws_tp_percent", 0.0))
            tp_type_session = (request.session.get("ws_tp_type") or "Excess")

        for (ws_est, header_row) in sheets_info:
            items = parse_estimate_items_amt(ws_est, header_row)
            if not items:
                continue

            subtotal = 0.0
            for it in items:
                subtotal += it["qty"] * it["rate"]

            tp_percent = tp_percent_session
            tp_type = tp_type_session

            tp_abs = subtotal * abs(tp_percent) / 100.0
            if tp_type == "Less":
                total_sheet = subtotal - tp_abs
            else:
                total_sheet = subtotal + tp_abs

            total_all_sheets += total_sheet

        return total_all_sheets

    # --------------- CASE 2: WorkSlip â†’ First / First & Final ---------------

    if action in ("workslip_first_part", "workslip_first_final"):
        ws_list = find_all_workslip_sheets_amt(wb)
        if not ws_list:
            return _extract_total_amount_from_bill_wb(wb)

        total_all_sheets = 0.0

        for ws_ws in ws_list:
            items = parse_workslip_items_amt(ws_ws)
            if not items:
                continue

            subtotal = 0.0
            for it in items:
                subtotal += it["qty"] * it["rate"]

            # T.P from sheet; if not found, use session  -  same as bill()
            tp_percent, tp_type = read_tp_from_sheet_amt(ws_ws)

            if request is not None:
                if tp_percent is None:
                    tp_percent = to_num(request.session.get("ws_tp_percent", 0.0))
                if not tp_type:
                    tp_type = (request.session.get("ws_tp_type") or "Excess")
            else:
                if tp_percent is None:
                    tp_percent = 0.0
                if not tp_type:
                    tp_type = "Excess"

            tp_abs = subtotal * abs(tp_percent) / 100.0
            if tp_type == "Less":
                total_sheet = subtotal - tp_abs
            else:
                total_sheet = subtotal + tp_abs

            total_all_sheets += total_sheet

        return total_all_sheets

    # --------------- CASE 3: Nth & Part / Nth & Final (N â‰¥ 2) ---------------

    # For these, the uploaded Excel is a Bill already.
    # _extract_total_amount_from_bill_wb:
    #  - detects 8-col vs 10-col
    #  - reads:
    #       H (col 8) of 'Total' row for 1st / First & Final
    #       I (col 9) of 'Total' row for Nth bills.
    return _extract_total_amount_from_bill_wb(wb)



def _build_mb_details_string(mb_measure_no, mb_measure_from, mb_measure_to,
                             mb_abs_no, mb_abs_from, mb_abs_to):
    return (
        f"MB.No. {mb_measure_no} P.No. {mb_measure_from} to {mb_measure_to} (Measurements) "
        f"& MB.No. {mb_abs_no} P.No. {mb_abs_from} to {mb_abs_to} (Abstract)"
    )


def _resolve_cc_header(action, nth_number_str=None):
    """
    Rebuild the CC header line based on 'action' (same actions used in bill view):

      - estimate_first_part      -> 'CC First & Part Bill'
      - estimate_first_final     -> 'CC First & Final Bill'
      - workslip_first_part      -> 'CC First & Part Bill'
      - workslip_first_final     -> 'CC First & Final Bill'
      - firstpart_nth_part, nth_nth_part   -> 'CC <Nth> & Part Bill'
      - firstpart_2nd_final, nth_nth_final -> 'CC <Nth> & Final Bill'
    """
    def ordinal_word(n):
        mapping = {
            1: "First",
            2: "Second",
            3: "Third",
            4: "Fourth",
            5: "Fifth",
            6: "Sixth",
            7: "Seventh",
            8: "Eighth",
            9: "Ninth",
            10: "Tenth",
        }
        if n in mapping:
            return mapping[n]
        if 10 < n % 100 < 14:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        return f"{n}{suffix}"

    base_action = action or ""

    if base_action in ("estimate_first_part", "workslip_first_part"):
        return "CC First & Part Bill"
    if base_action in ("estimate_first_final", "workslip_first_final"):
        return "CC First & Final Bill"

    # Nth-type actions
    try:
        n_val = int(nth_number_str or "2")
    except Exception:
        n_val = 2
    if n_val < 2:
        n_val = 2
    ord_word = ordinal_word(n_val)

    if base_action in ("firstpart_nth_part", "nth_nth_part"):
        return f"CC {ord_word} & Part Bill"
    if base_action in ("firstpart_2nd_final", "nth_nth_final"):
        return f"CC {ord_word} & Final Bill"

    # fallback
    return "CC Bill"


def _fill_excel_template(template_filename, context_dict):
    """
    Open Excel template from BILL_TEMPLATES_DIR, replace {{KEY}} placeholders
    with values from context_dict, return openpyxl Workbook.
    """
    template_path = os.path.join(BILL_TEMPLATES_DIR, template_filename)
    wb = load_workbook(template_path)

    # placeholders: {{NAME_OF_WORK}}, etc.
    placeholders = list(context_dict.keys())

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    text = cell.value
                    # replace {{KEY}} style placeholders
                    for key in placeholders:
                        placeholder = "{{" + key + "}}"
                        if placeholder in text:
                            text = text.replace(placeholder, str(context_dict[key]))
                    cell.value = text
    return wb


def _fill_docx_template(template_filename, context_dict):
    """
    Open Word template from BILL_TEMPLATES_DIR, replace {{KEY}} placeholders
    in paragraphs and tables, return python-docx Document.
    """
    template_path = os.path.join(BILL_TEMPLATES_DIR, template_filename)
    doc = Document(template_path)
    placeholders = list(context_dict.keys())

    # Helper to replace in a python-docx paragraph (handling runs)
    def replace_in_paragraph(paragraph):
        if not paragraph.text:
            return
        # Join runs to single string
        text = paragraph.text
        replaced = text
        for key in placeholders:
            ph = "{{" + key + "}}"
            if ph in replaced:
                replaced = replaced.replace(ph, str(context_dict[key]))
        if replaced != text:
            # Clear runs, set one run with replaced text (keeps style of first run)
            for _ in range(len(paragraph.runs)):
                paragraph.runs[0].text = ""
                paragraph.runs.pop(0)
            paragraph.add_run(replaced)

    for p in doc.paragraphs:
        replace_in_paragraph(p)

    for table in doc.tables:
        for row in table.rows:
            for cell in row.cells:
                for p in cell.paragraphs:
                    replace_in_paragraph(p)

    return doc

# -----------------------
# BILL FROM WORKSLIP
# -----------------------
# -----------------------
# BILL  -  from Estimate / from WorkSlip
# -----------------------
# imports for bill/templating consolidated at top of file

# BILL_TEMPLATES_DIR, _number_to_words_rupees, _extract_header_data_fuzzy_from_wb,
# _extract_total_amount_from_bill_wb, _build_mb_details_string,
# _resolve_cc_header, _fill_excel_template, _fill_docx_template
# are assumed to be defined ABOVE this function (as in your file).


