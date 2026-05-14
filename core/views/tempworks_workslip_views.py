"""
Temporary Works — Workslip module.

Multi-event Temporary Works produce a workslip with a different layout from
the regular workslip: each item shows a description header row, and below it
each event is listed as a roman-numeral sub-row with qty / rate / amount and
execution columns. Per-event AE rows appear when exec qty exceeds estimate.

Session keys (namespaced tw_ws_*):
  tw_ws_entries           - list copied from temp_entries (multi mode only)
  tw_ws_events_list       - global events registry
  tw_ws_exec_map          - {entry_id: {event_id: qty_exec}}
  tw_ws_work_name
  tw_ws_category          - 'electrical' or 'civil'
  tw_ws_selected_backend_id
  tw_ws_target_workslip   - 1, 2, ... (chain support, future)
  tw_ws_source_temp_id    - SavedWork.id of source tempwork
"""
import io
import json
import logging
import re
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from ..utils_excel import build_temp_day_rates, load_backend

logger = logging.getLogger(__name__)


def _to_roman_lower(n):
    try:
        n = int(n)
    except (TypeError, ValueError):
        return ""
    if n <= 0:
        return ""
    vals = [
        (1000, "m"), (900, "cm"), (500, "d"), (400, "cd"),
        (100, "c"), (90, "xc"), (50, "l"), (40, "xl"),
        (10, "x"), (9, "ix"), (5, "v"), (4, "iv"), (1, "i"),
    ]
    out = []
    for v, sym in vals:
        while n >= v:
            out.append(sym)
            n -= v
    return "".join(out)


def _norm_name(s):
    s = "" if s is None else str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _rate_for_event_days(day_rates_for_item, days):
    """Look up the rate that applies to `days` from a {day_no: rate} map."""
    if not day_rates_for_item:
        return 0.0
    try:
        days_int = int(days)
    except (TypeError, ValueError):
        return 0.0
    if days_int <= 0:
        return 0.0
    # Prefer exact match; fall back to nearest lower day count
    if days_int in day_rates_for_item:
        try:
            return float(day_rates_for_item[days_int] or 0)
        except (TypeError, ValueError):
            return 0.0
    candidates = sorted(d for d in day_rates_for_item.keys() if isinstance(d, int) and d <= days_int)
    if candidates:
        try:
            return float(day_rates_for_item[candidates[-1]] or 0)
        except (TypeError, ValueError):
            return 0.0
    return 0.0


def _coerce_float(v, default=0.0):
    try:
        if v in (None, ""):
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _build_view_rows(entries, events_list, day_rates_by_item, exec_map):
    """
    Build a flat list of view rows from tw_ws_entries for template rendering.
    Each row is a dict with kind = 'header' or 'event'.
    """
    rows = []
    sl = 1
    for entry in entries or []:
        if (entry or {}).get("mode") != "multi":
            continue
        item_name = entry.get("name") or ""
        entry_id = entry.get("id") or ""
        events = entry.get("events") or []
        valid_events = []
        for ev in events:
            ev_name = (ev.get("event_name") or "").strip()
            ev_days = int(_coerce_float(ev.get("days"), 0))
            ev_qty = _coerce_float(ev.get("qty"), 0.0)
            if ev_name and ev_days > 0 and ev_qty > 0:
                valid_events.append({
                    "event_id": ev.get("event_id") or "",
                    "event_name": ev_name,
                    "days": ev_days,
                    "qty_est": ev_qty,
                })
        if not valid_events:
            continue

        rows.append({
            "kind": "header",
            "sl": sl,
            "entry_id": entry_id,
            "item_name": item_name,
            "desc": item_name,
        })
        item_day_rates = day_rates_by_item.get(_norm_name(item_name)) or day_rates_by_item.get(item_name) or {}
        for i, ev in enumerate(valid_events, start=1):
            roman = _to_roman_lower(i)
            day_word = "day" if ev["days"] == 1 else "days"
            desc = f"{roman}. {ev['event_name']} for {ev['days']} {day_word}"
            rate = _rate_for_event_days(item_day_rates, ev["days"])
            exec_qty = _coerce_float(
                (exec_map.get(entry_id) or {}).get(ev["event_id"]), 0.0
            )
            rows.append({
                "kind": "event",
                "entry_id": entry_id,
                "event_id": ev["event_id"],
                "roman": roman,
                "desc": desc,
                "event_name": ev["event_name"],
                "days": ev["days"],
                "qty_est": ev["qty_est"],
                "rate": rate,
                "amt_est": round(ev["qty_est"] * rate, 2),
                "qty_exec": exec_qty,
                "amt_exec": round(exec_qty * rate, 2),
            })
        sl += 1
    return rows


@login_required(login_url="login")
def temp_workslip(request):
    """Main Temporary Works Workslip page (multi mode only)."""
    entries = request.session.get("tw_ws_entries") or []
    events_list = request.session.get("tw_ws_events_list") or []
    exec_map = request.session.get("tw_ws_exec_map") or {}
    work_name = request.session.get("tw_ws_work_name") or ""
    category = request.session.get("tw_ws_category") or "electrical"
    backend_id = request.session.get("tw_ws_selected_backend_id")

    if not entries:
        return redirect("dashboard")

    # Load backend for rate lookup
    try:
        items_list, _groups_map, _units_map, _ws_src, filepath = load_backend(
            f"temp_{category}", settings.BASE_DIR,
            backend_id=backend_id, module_code="temp_works",
        )
        day_rates = build_temp_day_rates(filepath, items_list) or {}
    except Exception:
        logger.exception("[TEMP_WS] load_backend failed (category=%s, backend_id=%s)", category, backend_id)
        day_rates = {}

    # build day_rates_by_item keyed both by normalized and original name
    day_rates_by_item = {}
    for k, v in (day_rates or {}).items():
        day_rates_by_item[k] = v
        day_rates_by_item[_norm_name(k)] = v

    if request.method == "POST":
        action = request.POST.get("action") or ""
        exec_map_str = request.POST.get("exec_map", "")
        new_exec_map = exec_map
        if exec_map_str:
            try:
                parsed = json.loads(exec_map_str)
                if isinstance(parsed, dict):
                    new_exec_map = {}
                    for ent_id, evs in parsed.items():
                        if not isinstance(evs, dict):
                            continue
                        new_exec_map[str(ent_id)] = {
                            str(ev_id): _coerce_float(q, 0.0)
                            for ev_id, q in evs.items()
                        }
            except Exception:
                logger.exception("[TEMP_WS] bad exec_map payload")

        wn_form = (request.POST.get("ws_work_name") or "").strip()
        if wn_form:
            work_name = wn_form
            request.session["tw_ws_work_name"] = work_name

        request.session["tw_ws_exec_map"] = new_exec_map
        request.session.modified = True
        exec_map = new_exec_map

        if action == "download_workslip":
            return _download_workslip_excel(
                entries=entries,
                events_list=events_list,
                exec_map=exec_map,
                day_rates_by_item=day_rates_by_item,
                work_name=work_name,
            )
        # action == "update_preview" -> fall through to render

    view_rows = _build_view_rows(entries, events_list, day_rates_by_item, exec_map)

    # Compute totals for preview
    total_est = sum(r["amt_est"] for r in view_rows if r["kind"] == "event")
    total_exec = sum(r["amt_exec"] for r in view_rows if r["kind"] == "event")

    return render(request, "core/temp_workslip.html", {
        "view_rows": view_rows,
        "work_name": work_name,
        "category": category,
        "total_est": round(total_est, 2),
        "total_exec": round(total_exec, 2),
    })


def _download_workslip_excel(entries, events_list, exec_map, day_rates_by_item, work_name):
    """Build the Temp Workslip Excel workbook and return as HttpResponse."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WorkSlip"

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFC8C8C8")

    headers = [
        "Sl.No", "Description of Item", "Unit",
        "Qty (Estimate)", "Rate", "Amount (Estimate)",
        "Qty (Execution)", "Amount (Execution)",
        "More", "Less", "Remarks",
    ]

    # Top title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="TEMPORARY WORKS — WORKSLIP")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Work name row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    wn_cell = ws.cell(row=2, column=1, value=f"Name of the work : {work_name}" if work_name else "Name of the work :")
    wn_cell.font = Font(bold=True)
    wn_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = 4
    for col_idx, txt in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=txt)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = header_fill
        c.border = border_all

    # Column widths
    widths = [6, 60, 10, 14, 12, 16, 14, 16, 10, 10, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    out_row = header_row + 1
    sl_counter = 1
    grand_est = 0.0
    grand_exec = 0.0
    grand_more = 0.0

    for entry in entries or []:
        if (entry or {}).get("mode") != "multi":
            continue
        item_name = entry.get("name") or ""
        entry_id = entry.get("id") or ""
        events = entry.get("events") or []
        valid_events = []
        for ev in events:
            ev_name = (ev.get("event_name") or "").strip()
            try:
                ev_days = int(ev.get("days") or 0)
            except (TypeError, ValueError):
                ev_days = 0
            try:
                ev_qty = float(ev.get("qty") or 0)
            except (TypeError, ValueError):
                ev_qty = 0.0
            if ev_name and ev_days > 0 and ev_qty > 0:
                valid_events.append({
                    "event_id": ev.get("event_id") or "",
                    "event_name": ev_name,
                    "days": ev_days,
                    "qty_est": ev_qty,
                })
        if not valid_events:
            continue

        # Header row (plain, not bold)
        ws.cell(row=out_row, column=1, value=sl_counter).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=out_row, column=2, value=item_name).alignment = Alignment(horizontal="justify", vertical="top", wrap_text=True)
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=out_row, column=c_idx).border = border_all
        out_row += 1

        item_day_rates = day_rates_by_item.get(_norm_name(item_name)) or day_rates_by_item.get(item_name) or {}

        for i, ev in enumerate(valid_events, start=1):
            roman = _to_roman_lower(i)
            day_word = "day" if ev["days"] == 1 else "days"
            desc = f"{roman}. {ev['event_name']} for {ev['days']} {day_word}"
            rate = _rate_for_event_days(item_day_rates, ev["days"])
            qty_est = float(ev["qty_est"])
            qty_exec = _coerce_float((exec_map.get(entry_id) or {}).get(ev["event_id"]), 0.0)

            excess = round(max(0.0, qty_exec - qty_est), 2)
            base_qty = round(min(qty_exec, qty_est), 2) if qty_est > 0 else round(qty_exec, 2)

            # Event base row
            ws.cell(row=out_row, column=1, value="")  # no sl on event rows
            ws.cell(row=out_row, column=2, value=desc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(row=out_row, column=3, value="")  # unit blank for tempworks events
            ws.cell(row=out_row, column=4, value=round(qty_est, 2))
            ws.cell(row=out_row, column=5, value=round(rate, 2))
            ws.cell(row=out_row, column=6, value=f"=D{out_row}*E{out_row}")
            ws.cell(row=out_row, column=7, value=base_qty)
            ws.cell(row=out_row, column=8, value=f"=G{out_row}*E{out_row}")
            # More/Less are filled on AE row if excess; here only the deficit Less applies
            if excess <= 0 and 0 < qty_exec < qty_est:
                ws.cell(row=out_row, column=9, value="")  # More
                ws.cell(row=out_row, column=10, value=f"=F{out_row}-H{out_row}")  # Less
                remark = "Less as per estimated"
            elif qty_exec == 0:
                remark = "Deleted"
            elif excess > 0:
                remark = ""  # remark on AE row
            else:
                remark = ""
            ws.cell(row=out_row, column=11, value=remark)
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=out_row, column=c_idx).border = border_all
            grand_est += round(qty_est * rate, 2)
            grand_exec += round(base_qty * rate, 2)
            out_row += 1

            if excess > 0:
                # AE row for this event
                ae_desc = f"{roman}.AE {ev['event_name']} for {ev['days']} {day_word}"
                ws.cell(row=out_row, column=1, value="")
                ws.cell(row=out_row, column=2, value=ae_desc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.cell(row=out_row, column=3, value="")
                ws.cell(row=out_row, column=4, value="")
                ws.cell(row=out_row, column=5, value=round(rate, 2))
                ws.cell(row=out_row, column=6, value="")
                ws.cell(row=out_row, column=7, value=excess)
                ws.cell(row=out_row, column=8, value=f"=G{out_row}*E{out_row}")
                ws.cell(row=out_row, column=9, value=f"=G{out_row}*E{out_row}")  # More
                ws.cell(row=out_row, column=10, value="")
                ws.cell(row=out_row, column=11, value="Excess as per estimated")
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=out_row, column=c_idx).border = border_all
                grand_exec += round(excess * rate, 2)
                grand_more += round(excess * rate, 2)
                out_row += 1

        sl_counter += 1

    # Total row
    tot_cell = ws.cell(row=out_row, column=2, value="Total")
    tot_cell.font = Font(bold=True)
    tot_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=out_row, column=6, value=round(grand_est, 2)).font = Font(bold=True)
    ws.cell(row=out_row, column=8, value=round(grand_exec, 2)).font = Font(bold=True)
    if grand_more:
        ws.cell(row=out_row, column=9, value=round(grand_more, 2)).font = Font(bold=True)
    for c_idx in range(1, len(headers) + 1):
        ws.cell(row=out_row, column=c_idx).border = border_all

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", work_name).strip("_") or "TempWorks"
    filename = f"{safe_name}_WorkSlip.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
