# core/views.py

import json
import os
import re
import logging
from copy import copy

import inflect
from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher
# (moved to top-level imports)

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()

# Define BILL_TEMPLATES_DIR for template operations
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")

_inflect_engine = inflect.engine()


def _apply_print_settings(wb, landscape=False):
    """
    Apply standard print settings to all sheets in a workbook:
      - A4 paper, fit all columns to 1 page width
      - Portrait (default) or Landscape
      - Times New Roman font on all populated cells
      - Repeat header row(s) at top of every printed page
    """
    for ws in wb.worksheets:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # unlimited pages vertically
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Auto-detect header row(s) for print titles
        # Look for rows containing "Sl" / "S.No" / "Sl.No" in column A (typical header)
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        header_row = 0
        for r in range(1, min(max_r + 1, 20)):  # scan first 20 rows
            val = ws.cell(row=r, column=1).value
            if val is not None:
                val_str = str(val).strip().lower().replace('.', '').replace(' ', '')
                if val_str in ('sl', 'sno', 'slno', 'srno', 'sino'):
                    header_row = r
                    break
        if header_row > 0:
            ws.print_title_rows = '1:{}'.format(header_row)

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    f = cell.font
                    cell.font = Font(
                        name='Times New Roman',
                        bold=f.bold,
                        italic=f.italic,
                        size=f.size,
                        color=f.color,
                        underline=f.underline,
                        strikethrough=f.strikethrough,
                    )


# ============================================================================
# PHASE 3B: HELPER FUNCTIONS FOR ORGANIZATION & ASYNC PROCESSING
# ============================================================================

def get_org_from_request(request):
    """
    Safely extract organization from request.
    
    Returns:
        Organization object (auto-creates if not available)
        
    For single-tenant/development mode, automatically creates
    a default organization and membership for the user.
    """
    if hasattr(request, 'organization') and request.organization:
        return request.organization
    
    # Auto-create organization for logged-in users (single-tenant mode)
    if request.user.is_authenticated:
        from core.models import Organization, Membership
        from django.utils.text import slugify
        
        # Try to find existing membership
        membership = Membership.objects.filter(user=request.user).select_related('organization').first()
        
        if membership:
            request.organization = membership.organization
            return membership.organization
        
        # Create default organization for user
        org_name = f"{request.user.username}'s Organization"
        org_slug = slugify(org_name)[:255]
        
        # Make slug unique if needed
        base_slug = org_slug
        counter = 1
        while Organization.objects.filter(slug=org_slug).exists():
            org_slug = f"{base_slug}-{counter}"
            counter += 1
        
        org, created = Organization.objects.get_or_create(
            name=org_name,
            defaults={
                'slug': org_slug,
                'owner': request.user,
                'is_active': True
            }
        )
        
        # Create membership
        Membership.objects.get_or_create(
            user=request.user,
            organization=org,
            defaults={'role': 'owner'}
        )
        
        request.organization = org
        return org
    
    # Not authenticated - raise error
    from django.http import Http404
    raise Http404("Please login to continue.")


def check_org_access(request, obj):
    """
    Verify that an object belongs to the user's organization.
    
    Args:
        request: Django request object (has organization from middleware)
        obj: Model instance to check (must have organization FK)
        
    Returns:
        True if object belongs to user's org
        
    Raises:
        Http404 if object doesn't belong to user's org
    """
    org = get_org_from_request(request)
    if hasattr(obj, 'organization') and obj.organization != org:
        from django.http import Http404
        raise Http404("You don't have permission to access this object.")
    return True


def enqueue_excel_task(job_id, task_name='excel_parse', **kwargs):
    """
    Enqueue an Excel processing task.
    Runs synchronously if Celery is not available.
    
    Args:
        job_id: ID of Job object to update
        task_name: 'excel_parse', 'generate_bill', 'generate_workslip'
        **kwargs: Additional arguments for the task (must include upload_id for excel_parse)
        
    Returns:
        Task result or mock object with id
    """
    from django.conf import settings
    
    # Get upload_id from kwargs or from job
    upload_id = kwargs.get('upload_id')
    if not upload_id:
        try:
            job = Job.objects.get(id=job_id)
            if job.upload:
                upload_id = job.upload.id
        except Job.DoesNotExist:
            pass
    
    # Try Celery first, fall back to sync if connection fails
    try:
        if task_name == 'generate_bill':
            task = generate_bill_pdf.delay(job_id, kwargs.get('project_id'))
        elif task_name == 'generate_workslip':
            task = generate_workslip_pdf.delay(job_id, kwargs.get('project_id'))
        elif task_name == 'generate_bill_document':
            task = generate_bill_document_task.delay(job_id)
        elif upload_id:
            task = process_excel_upload.delay(upload_id)
        else:
            raise ValueError("upload_id required for excel_parse task")
        return task
    except Exception as e:
        # Celery not available, run synchronously
        logger.warning(f"Celery not available ({e}), running task synchronously")
        
        # Create a mock task result
        class MockTask:
            def __init__(self):
                import uuid
                self.id = str(uuid.uuid4())
        
        # Run task synchronously
        if task_name == 'generate_bill':
            generate_bill_pdf(job_id, kwargs.get('project_id'))
        elif task_name == 'generate_workslip':
            generate_workslip_pdf(job_id, kwargs.get('project_id'))
        elif task_name == 'generate_bill_document':
            generate_bill_document_task(job_id)
        elif upload_id:
            process_excel_upload(upload_id)
        
        return MockTask()


def create_job_for_excel(request, upload=None, job_type='excel_parse', metadata=None):
    """
    Create a Job object and enqueue task for Excel processing.
    
    Args:
        request: Django request (has organization)
        upload: Upload object (optional, will be created if not provided)
        job_type: Type of job ('excel_parse', 'generate_bill', 'generate_workslip')
        metadata: Dict with additional metadata
        
    Returns:
        Tuple: (job, celery_task)
        
    Example:
        job, task = create_job_for_excel(request, upload, 'generate_bill', {'project_id': 123})
        return JsonResponse({'job_id': job.id, 'status_url': reverse('job_status', args=[job.id])})
    """
    org = get_org_from_request(request)
    
    # Create upload if not provided
    if not upload:
        upload = Upload.objects.create(
            filename=metadata.get('filename', 'export.xlsx') if metadata else 'export.xlsx',
            file_size=0,
            status='processing'
        )
    
    # Create job
    job = Job.objects.create(
        organization=org,
        user=request.user,
        upload=upload,
        job_type=job_type,
        status='queued',
        result={'metadata': metadata or {}}  # Store metadata in result field
    )
    
    # Enqueue task
    task_kwargs = {}
    if metadata:
        task_kwargs.update({k: v for k, v in metadata.items() if k not in ('filename', 'job_id')})
    
    task = enqueue_excel_task(job.id, job_type, **task_kwargs)
    
    # Store task ID in job
    job.celery_task_id = task.id
    job.save()
    
    return job, task


# ============================================================================
# END PHASE 3B HELPERS
# ============================================================================


def _number_to_words_rupees(n):
    """
    Convert number to words in Indian numbering system:
      12345.67 -> 'Twelve thousand three hundred and forty-five rupees only'
      1234567 -> 'Twelve lakh thirty-four thousand five hundred and sixty-seven rupees only'
    """
    try:
        integer_part = int(round(float(n)))
    except Exception:
        integer_part = 0

    if integer_part == 0:
        return "Zero rupees only"

    # Indian numbering system
    crores = integer_part // 10000000
    integer_part %= 10000000
    lakhs = integer_part // 100000
    integer_part %= 100000
    thousands = integer_part // 1000
    integer_part %= 1000
    hundreds = integer_part

    parts = []
    if crores > 0:
        parts.append(f"{_inflect_engine.number_to_words(crores)} crore" + ("s" if crores > 1 else ""))
    if lakhs > 0:
        parts.append(f"{_inflect_engine.number_to_words(lakhs)} lakh" + ("s" if lakhs > 1 else ""))
    if thousands > 0:
        parts.append(f"{_inflect_engine.number_to_words(thousands)} thousand")
    if hundreds > 0:
        parts.append(_inflect_engine.number_to_words(hundreds))

    words = " ".join(parts)
    words = words.replace("-", " ")
    return f"{words} rupees only".capitalize()


def _format_indian_number(num):
    """
    Format a number in Indian numbering system with commas.
    Example: 300000 -> "3,00,000" (lakhs), 10000000 -> "1,00,00,000" (crores)
    
    Indian system: 1,00,00,00,000 (crores, lakhs, thousands, hundreds)
    Pattern: last 3 digits, then groups of 2
    """
    try:
        # Handle string input
        if isinstance(num, str):
            num = float(num.replace(',', '').replace(' ', ''))
        
        num = float(num)
        
        # Handle negative numbers
        is_negative = num < 0
        num = abs(num)
        
        # Check if it has decimals
        if num == int(num):
            # No decimals - format as integer
            num_str = str(int(num))
        else:
            # Has decimals - format with 2 decimal places
            num_str = f"{num:.2f}"
            integer_part, decimal_part = num_str.split('.')
            # Format the integer part only
            num_str = integer_part
        
        # Apply Indian comma formatting
        if len(num_str) <= 3:
            formatted = num_str
        else:
            # Last 3 digits
            last_three = num_str[-3:]
            remaining = num_str[:-3]
            
            # Group remaining by 2
            groups = []
            while remaining:
                groups.append(remaining[-2:] if len(remaining) >= 2 else remaining)
                remaining = remaining[:-2]
            
            groups.reverse()
            formatted = ','.join(groups) + ',' + last_three
        
        # Add decimal part back if it exists
        if num != int(num):
            formatted += f".{f'{num:.2f}'.split('.')[1]}"
        
        result = f"-{formatted}" if is_negative else formatted
        return result
        
    except (ValueError, TypeError):
        return str(num)


def _get_current_financial_year():
    """
    Get the current financial year in format "2025-26".
    Financial year runs from April 1 to March 31.
    """
    today = timezone.now().date()
    
    if today.month >= 4:  # April onwards
        fy_start = today.year
        fy_end = (today.year + 1) % 100
    else:  # January to March
        fy_start = today.year - 1
        fy_end = today.year % 100
    
    return f"{fy_start}-{fy_end:02d}"


def _get_current_date_formatted():
    """
    Get current date in format "DD-MM-YYYY"
    """
    today = timezone.now().date()
    return today.strftime("%d-%m-%Y")


def _format_date_to_ddmmyyyy(date_str):
    """
    Convert date from yyyy-mm-dd (HTML date input format) to dd-mm-yyyy format.
    Returns the original string if parsing fails or if already in correct format.
    """
    if not date_str or not date_str.strip():
        return ""
    
    date_str = date_str.strip()
    
    # Check if already in dd-mm-yyyy format
    if len(date_str) == 10 and date_str[2] == '-' and date_str[5] == '-':
        return date_str
    
    # Try to parse yyyy-mm-dd format
    try:
        from datetime import datetime
        parsed = datetime.strptime(date_str, "%Y-%m-%d")
        return parsed.strftime("%d-%m-%Y")
    except (ValueError, TypeError):
        # Return as-is if parsing fails
        return date_str


def _get_letter_settings(user):
    """
    Get the letter settings for a user.
    Returns the LetterSettings object or None if not set.
    """
    try:
        return LetterSettings.objects.get(user=user)
    except LetterSettings.DoesNotExist:
        return None


