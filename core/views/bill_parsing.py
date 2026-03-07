# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import _format_indian_number

def _make_placeholder(name):
    def view(request, *args, **kwargs):
        return HttpResponse(f"Placeholder view: {name}")
    view.__name__ = name
    return view

def __getattr__(name: str):
    # Common view names referenced by estimate_site/urls.py â€” if any are
    # missing, return a lightweight placeholder so the server can start.
    if name.startswith("_"):
        raise AttributeError(name)
    if name not in _placeholder_views:
        _placeholder_views[name] = _make_placeholder(name)
    return _placeholder_views[name]


def _safe_float(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def _extract_header_data_fuzzy_from_wb(wb):
    """
    Scan all sheets, top ~40 rows, trying to find:
      - Name of work
      - Estimate amount / ECV (optional, not strictly needed here)
      - Admin sanction (ignored for docs)
      - Technical sanction (ignored for docs)
      - Agreement
      - Agency

    We are NOT strict about exact label text: things like
      'Name of work', 'Name of the work', 'Work name', etc
      'Agreement', 'Agt', 'Agrmt', ...
      'Agency', 'Contractor', ...
    are all accepted.

    We return values *after* the colon if present.

    Returns dict:
      {
        "name_of_work": "...",
        "agreement": "...",
        "agency": "...",
      }
    (others may be empty)
    """
    header = {
        "name_of_work": "",
        "estimate_amount": "",
        "admin_sanction": "",
        "tech_sanction": "",
        "agreement": "",
        "agency": "",
        "mb_details": "",
    }

    def clean_value_for_header(val):
        if not val:
            return ""
        s = str(val).strip()
        # if there's a colon, return RHS only
        if ":" in s:
            s = s.split(":", 1)[1].strip()
        return s

    for ws in wb.worksheets:
        max_row = min(ws.max_row, 40)
        max_col = min(ws.max_column, 20)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                raw = ws.cell(row=r, column=c).value
                if raw is None:
                    continue
                s_full = str(raw).strip()
                low = s_full.lower()
                # skip obviously numeric things
                if not low:
                    continue

                # tokens set
                tokens = set(low.replace(":", " ").replace(".", " ").split())

                # ---- Name of work ----
                if not header["name_of_work"]:
                    # look for 'name' + 'work' anywhere
                    if "name" in tokens and "work" in tokens:
                        header["name_of_work"] = clean_value_for_header(s_full)
                        continue

                # ---- Estimate Amount / ECV (optional) ----
                if not header["estimate_amount"]:
                    if "estimate" in tokens and ("amount" in tokens or "ecv" in tokens):
                        header["estimate_amount"] = clean_value_for_header(s_full)
                        continue

                # ---- Admin sanction (optional, not used in docs) ----
                if not header["admin_sanction"]:
                    if "admin" in low and "sanction" in low:
                        header["admin_sanction"] = clean_value_for_header(s_full)
                        continue

                # ---- Technical sanction (optional, not used in docs) ----
                if not header["tech_sanction"]:
                    if ("tech" in low or "technical" in low) and "sanction" in low:
                        header["tech_sanction"] = clean_value_for_header(s_full)
                        continue

                # ---- Agreement (Agt / Agrmt etc.) ----
                if not header["agreement"]:
                    if (
                        "agreement" in low
                        or "agrmt" in low
                        or "agt" in tokens
                        or "agt." in tokens
                    ):
                        header["agreement"] = clean_value_for_header(s_full)
                        continue

                # ---- Agency (Contractor / Firm etc.) ----
                if not header["agency"]:
                    if "agency" in tokens or "contractor" in tokens or "firm" in tokens:
                        header["agency"] = clean_value_for_header(s_full)
                        continue

                # ---- MB Details (M.B.No / MB No / Measurement Book) ----
                if not header["mb_details"]:
                    if _cell_has_mb_details(low, tokens):
                        header["mb_details"] = clean_value_for_header(s_full)
                        continue

    return header


def _cell_has_mb_details(low, tokens):
    """
    Check if a cell's text contains MB (Measurement Book) details.
    Matches patterns like:
      - M.B.No, M.B. No, MB No, MB.No, M.B.Nos
      - Measurement Book
      - mb no, mb details
    """
    # Direct pattern checks on lowercased text
    import re
    # M.B.No / M.B. No / M.B.Nos / MB.No / MB No
    if re.search(r'm\.?b\.?\s*no', low):
        return True
    # "measurement" + "book" anywhere
    if "measurement" in tokens and "book" in tokens:
        return True
    # "mb" token + "details" or "no" nearby
    if "mb" in tokens and ("details" in tokens or "no" in tokens or "nos" in tokens):
        return True
    return False


def _extract_header_data_from_sheet(ws):
    """
    Extract header data (name of work, agreement, agency, MB details, etc.) from a specific sheet.
    
    This is a per-sheet version of _extract_header_data_fuzzy_from_wb that reads
    data from the first ~40 rows of the given worksheet.
    
    Returns dict:
      {
        "name_of_work": "...",
        "estimate_amount": "...",
        "admin_sanction": "...",
        "tech_sanction": "...",
        "agreement": "...",
        "agency": "...",
        "mb_details": "...",
      }
    """
    header = {
        "name_of_work": "",
        "estimate_amount": "",
        "admin_sanction": "",
        "tech_sanction": "",
        "agreement": "",
        "agency": "",
        "mb_details": "",
    }

    def clean_value_for_header(val):
        if not val:
            return ""
        s = str(val).strip()
        # if there's a colon, return RHS only
        if ":" in s:
            s = s.split(":", 1)[1].strip()
        return s

    max_row = min(ws.max_row, 40)
    max_col = min(ws.max_column, 20)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            if raw is None:
                continue
            s_full = str(raw).strip()
            low = s_full.lower()
            # skip obviously numeric things
            if not low:
                continue

            # tokens set
            tokens = set(low.replace(":", " ").replace(".", " ").split())

            # ---- Name of work ----
            if not header["name_of_work"]:
                # look for 'name' + 'work' anywhere
                if "name" in tokens and "work" in tokens:
                    header["name_of_work"] = clean_value_for_header(s_full)
                    continue

            # ---- Estimate Amount / ECV (optional) ----
            if not header["estimate_amount"]:
                if "estimate" in tokens and ("amount" in tokens or "ecv" in tokens):
                    header["estimate_amount"] = clean_value_for_header(s_full)
                    continue

            # ---- Admin sanction (optional, not used in docs) ----
            if not header["admin_sanction"]:
                if "admin" in low and "sanction" in low:
                    header["admin_sanction"] = clean_value_for_header(s_full)
                    continue

            # ---- Technical sanction (optional, not used in docs) ----
            if not header["tech_sanction"]:
                if ("tech" in low or "technical" in low) and "sanction" in low:
                    header["tech_sanction"] = clean_value_for_header(s_full)
                    continue

            # ---- Agreement (Agt / Agrmt etc.) ----
            if not header["agreement"]:
                if (
                    "agreement" in low
                    or "agrmt" in low
                    or "agt" in tokens
                    or "agt." in tokens
                ):
                    header["agreement"] = clean_value_for_header(s_full)
                    continue

            # ---- Agency (Contractor / Firm etc.) ----
            if not header["agency"]:
                if "agency" in tokens or "contractor" in tokens or "firm" in tokens:
                    header["agency"] = clean_value_for_header(s_full)
                    continue

            # ---- MB Details (M.B.No / MB No / Measurement Book) ----
            if not header["mb_details"]:
                if _cell_has_mb_details(low, tokens):
                    header["mb_details"] = clean_value_for_header(s_full)
                    continue

    return header


def _detect_bill_format(ws):
    """
    Try to detect if this sheet looks like:
      - 'First Bill' 8-column format (S.No, Quantity, Unit, Item, Rate, Per, Unit, Amount)
      - 'Nth Bill' 10-column format with 'Quantity Till Date' headers

    Returns one of:
      "first_bill_8col", "nth_bill_10col", or None
    """
    max_scan_row = min(ws.max_row, 20)
    for r in range(1, max_scan_row + 1):
        a = str(ws.cell(row=r, column=1).value or "").strip().lower()
        b = str(ws.cell(row=r, column=2).value or "").strip().lower()
        c = str(ws.cell(row=r, column=3).value or "").strip().lower()
        d = str(ws.cell(row=r, column=4).value or "").strip().lower()
        # fifth column value not needed for detection here
        
        # First Bill style: "sl" in A, "quantity" in B, "item"/"description" in D
        if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
            # assume 8-col bill
            return "first_bill_8col"

        # Nth Bill style: "sl" in A and "quantity till date" in C
        if "sl" in a and "quantity till date" in c:
            return "nth_bill_10col"

    return None


def _extract_total_amount_from_bill_wb(wb):
    """
    Try to find the current bill amount (the 'Total' row) in a Bill workbook.

    - For 'First Bill' 8-col format -> read from column H in 'Total' row.
    - For 'Nth Bill' 10-col format -> read from column I (Since Last Amount) in 'Total' row.
    - Fallback: last numeric cell in the row containing 'Total'.

    Returns float.
    """
    if not wb.worksheets:
        return 0.0

    ws = wb.worksheets[0]  # assume bill is in first sheet
    return _extract_total_amount_from_single_sheet(ws)


def _extract_total_amount_from_single_sheet(ws):
    """
    Extract total amount from a single bill sheet (worksheet).
    
    - For 'First Bill' 8-col format -> read from column H in 'Total' row.
    - For 'Nth Bill' 10-col format -> read from column I (Since Last Amount) in 'Total' row.
    - Fallback: last numeric cell in the row containing 'Total'.

    Returns float.
    """
    if not ws:
        return 0.0
    
    bill_format = _detect_bill_format(ws)

    max_row = ws.max_row
    total_row = None

    for r in range(1, max_row + 1):
        found_total_label = False
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str):
                low = val.strip().lower()
                if low.startswith("total"):  # ignore 'sub total' rows by checking exact 'total'
                    # but allow 'total' and 'total amount' etc.
                    if not low.startswith("sub total") and not low.startswith("subtotal"):
                        found_total_label = True
                        break
        if found_total_label:
            total_row = r
            break

    if total_row is None:
        return 0.0

    # Format-specific read
    if bill_format == "first_bill_8col":
        # H = col 8
        amount_val = ws.cell(row=total_row, column=8).value
        if amount_val is not None and str(amount_val).strip() != "":
            return _safe_float(amount_val)

    if bill_format == "nth_bill_10col":
        # 'Since Last Amount' total is in column I (9)
        amount_val = ws.cell(row=total_row, column=9).value
        if amount_val is not None and str(amount_val).strip() != "":
            return _safe_float(amount_val)

    # Fallback: last numeric cell in the Total row
    last_numeric = 0.0
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=total_row, column=c).value
        if isinstance(val, (int, float)):
            last_numeric = float(val)
    return last_numeric


# Duplicate helper definitions removed here; canonical `_build_mb_details_string`
# and `_resolve_cc_header` are defined earlier in the file and will be used.




# -----------------------
# BILL FROM WORKSLIP
# -----------------------
# -----------------------
# BILL  -  from Estimate / from WorkSlip
# -----------------------
# imports consolidated at top of file


# ---------- Module-level helpers for `bill()` (extracted from nested defs) ----------
def to_number(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def is_merged_cell(ws, row, col):
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return True
    return False


def find_estimate_sheet_and_header_row(wb):
    for ws in wb.worksheets:
        for r in range(1, 26):
            a = str(ws.cell(row=r, column=1).value or "").strip().lower()
            b = str(ws.cell(row=r, column=2).value or "").strip().lower()
            d = str(ws.cell(row=r, column=4).value or "").strip().lower()
            if "sl" in a and "quantity" in b and ("item" in d or "description" in d):
                return ws, r
    return wb.worksheets[0], 3


def looks_like_header(ws, r):
    a = str(ws.cell(row=r, column=1).value or "").strip().lower()
    b = str(ws.cell(row=r, column=2).value or "").strip().lower()
    c = str(ws.cell(row=r, column=3).value or "").strip().lower()
    d = str(ws.cell(row=r, column=4).value or "").strip().lower()
    e = str(ws.cell(row=r, column=5).value or "").strip().lower()

    has_sl = ("sl" in a) or ("s.no" in a) or ("serial" in a)
    has_qty = ("qty" in b or "quantity" in b or "qty" in c or "quantity" in c)
    has_desc = (
        "item" in c or "item" in d or "item" in e or
        "description" in c or "description" in d or "description" in e
    )
    return has_sl and has_qty and has_desc


def find_all_estimate_sheets_and_header_rows(wb):
    results = []

    for ws in wb.worksheets:
        header_row = None
        max_scan = min(ws.max_row, 60)
        for r in range(1, max_scan + 1):
            if looks_like_header(ws, r):
                header_row = r
                break
        if header_row:
            results.append((ws, header_row))

    if not results:
        for ws in wb.worksheets:
            hr = None
            max_scan = min(ws.max_row or 0, 30)
            for r in range(1, max_scan + 1):
                if looks_like_header(ws, r):
                    hr = r
                    break
            if not hr:
                hr = 3
            results.append((ws, hr))

    return results


def parse_estimate_items(ws, header_row):
    items = []
    max_row = min(ws.max_row, 5000)

    for r in range(header_row + 1, max_row + 1):
        desc_raw = ws.cell(row=r, column=4).value
        desc = str(desc_raw or "").strip()
        rate_raw = ws.cell(row=r, column=5).value
        amt_raw = ws.cell(row=r, column=8).value

        desc_low = desc.lower()

        if desc_low.startswith(("ecv", "sub total", "subtotal", "total")):
            break

        rate_str = "" if rate_raw is None else str(rate_raw).strip()
        is_rate_empty = (rate_str == "")

        if desc and (is_rate_empty or is_merged_cell(ws, r, 4)):
            continue

        qty_raw = ws.cell(row=r, column=2).value
        unit_raw = ws.cell(row=r, column=3).value

        qty_str = "" if qty_raw is None else str(qty_raw).strip()
        unit_str = "" if unit_raw is None else str(unit_raw).strip()
        amt_str = "" if amt_raw is None else str(amt_raw).strip()

        all_blank = (
            desc == "" and
            is_rate_empty and
            qty_str == "" and
            unit_str == "" and
            amt_str == ""
        )

        if all_blank:
            continue

        if desc == "" and is_rate_empty:
            continue

        qty = to_number(qty_raw)
        unit = unit_str
        rate = to_number(rate_raw)
        amt = to_number(amt_raw)

        if amt != 0:
            if qty == 0 and rate != 0:
                qty = amt / rate
            elif rate == 0 and qty != 0:
                rate = amt / qty

        items.append({
            "qty": qty,
            "unit": unit,
            "desc": desc,
            "rate": rate,
        })

    return items


def find_workslip_sheet(wb):
    for ws in wb.worksheets:
        for r in range(1, 40):
            b = str(ws.cell(row=r, column=2).value or "").strip().lower()
            g = str(ws.cell(row=r, column=7).value or "").strip().lower()
            if "description" in b and ("qty" in g or "quantity" in g):
                return ws
    return wb.worksheets[0]


def find_all_workslip_sheets(wb):
    results = []
    for ws in wb.worksheets:
        max_scan = min(ws.max_row, 60)
        for r in range(1, max_scan + 1):
            b = str(ws.cell(row=r, column=2).value or "").strip().lower()
            c = str(ws.cell(row=r, column=3).value or "").strip().lower()

            has_desc = (
                "description of item" in b or
                "description of item" in c or
                ("description" in b and "item" in b) or
                ("description" in c and "item" in c)
            )

            has_qty = False
            for col in range(5, 12):
                t = str(ws.cell(row=r, column=col).value or "").strip().lower()
                if "qty" in t or "quantity" in t:
                    has_qty = True
                    break

            if has_desc and has_qty:
                results.append(ws)
                break

    if not results:
        results.append(find_workslip_sheet(wb))

    return results


def extract_tp_from_workslip(ws):
    """Extract T.P percentage and type (Less/Excess) from a workslip sheet."""
    max_scan = min(ws.max_row, 400)
    last_sub_total = None
    import re
    
    def has_tp_keyword(text):
        """Check if text contains T.P or tender premium keyword."""
        return "tp" in text or "tender premium" in text or "tenderpremium" in text

    for r in range(1, max_scan + 1):
        # Build row text for easy matching
        row_text_parts = []
        for col in range(1, 15):
            row_text_parts.append(str(ws.cell(row=r, column=col).value or ""))
        row_text = " ".join(row_text_parts).strip().lower()
        # Normalize "t.p" "t.p." "t. p" etc. to "tp" for matching
        row_text_norm = re.sub(r't\s*\.\s*p\.?', 'tp', row_text)

        # Track the latest Sub Total value (amount columns to the right)
        if "sub total" in row_text and "total" in row_text:
            # This may match both, but keep as potential
            pass
        if "sub total" in row_text and not row_text.startswith("total"):
            for amt_col in [8, 9, 10, 11, 12]:
                sub_val = to_number(ws.cell(row=r, column=amt_col).value)
                if sub_val is not None:
                    last_sub_total = sub_val
                    break

        # Direct percent extraction if present (check both original and normalized)
        if (has_tp_keyword(row_text) or has_tp_keyword(row_text_norm)) and ("%" in row_text or "percent" in row_text):
            percent_match = re.search(r"(\d+\.?\d*)\s*%", row_text)
            if percent_match:
                tp_percent = float(percent_match.group(1))
                # Check for Less/Deduct keywords vs Excess/Add
                is_less = "less" in row_text or "deduct" in row_text
                tp_type = "Less" if is_less else "Excess"
                print(f"DEBUG extract_tp_from_workslip: Found TP row with percent. row_text='{row_text[:100]}' percent={tp_percent} type={tp_type}")
                return tp_percent, tp_type

        # If no percent text, but TP row has an amount, derive percent from last_sub_total
        if has_tp_keyword(row_text) or has_tp_keyword(row_text_norm):
            for amt_col in [8, 9, 10, 11, 12]:
                tp_val = to_number(ws.cell(row=r, column=amt_col).value)
                if tp_val is None:
                    continue
                if last_sub_total and last_sub_total != 0:
                    tp_percent = abs(tp_val) * 100.0 / abs(last_sub_total)
                    # Check for Less/Deduct keywords vs Excess/Add; also use sign of amount
                    is_less = "less" in row_text or "deduct" in row_text or tp_val < 0
                    tp_type = "Less" if is_less else "Excess"
                    print(f"DEBUG extract_tp_from_workslip: Found TP row with amount. row_text='{row_text[:100]}' tp_val={tp_val} derived percent={tp_percent} type={tp_type}")
                    return tp_percent, tp_type

    # Default if not found
    print(f"DEBUG extract_tp_from_workslip: No TP found, using default 0.0 Excess")
    return 0.0, "Excess"


def parse_workslip_items(ws):
    """
    Parse workslip items, automatically detecting the last workslip phase columns.
    
    For multi-phase workslips (Workslip-1, Workslip-2, etc.), this function will
    find and use the LAST workslip's Qty/Amt columns for bill generation.
    
    The Rate is ALWAYS taken from the Estimate Rate column (column 5), as
    multi-phase workslips don't have separate Rate columns per phase.
    """
    items = []
    max_row = min(ws.max_row, 5000)
    
    # Step 1: Find header row and detect column structure
    header_row = 8  # Default
    for r in range(1, 15):
        cell_val = str(ws.cell(row=r, column=1).value or "").strip().lower()
        if "sl" in cell_val:
            header_row = r
            break
    
    # Step 2: Scan header row to find all workslip/execution qty columns
    # We need to find the LAST Qty column - this is the latest workslip phase
    qty_columns = []  # List of column indices for execution qty columns
    amt_columns = []  # List of column indices for execution amt columns
    estimate_rate_col = 5  # Default - Rate (Estimate) column
    
    for c in range(1, 40):  # Scan more columns for multi-phase workslips
        header = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
        
        # Check for workslip/execution columns
        is_exec_col = ("execution" in header or "exec" in header or "workslip" in header)
        is_qty = ("qty" in header or "quantity" in header)
        is_amt = ("amount" in header or "amt" in header)
        is_estimate = ("est" in header or "estimate" in header)
        is_rate = ("rate" in header)
        
        # Find estimate rate column
        if is_estimate and is_rate:
            estimate_rate_col = c
        
        # Skip estimate columns for qty/amt detection
        if is_estimate:
            continue
        
        # Skip More/Less/Remarks columns
        if "more" in header or "less" in header or "remark" in header:
            continue
            
        if is_exec_col and is_qty:
            qty_columns.append(c)
        elif is_exec_col and is_amt:
            amt_columns.append(c)
    
    # Step 3: Determine the columns to use
    # Use the LAST Qty column (latest phase) and its corresponding Amount column
    # Rate is always from Estimate Rate column
    if qty_columns:
        qty_col = qty_columns[-1]  # Last qty column = latest workslip phase
        # The Amount column should be immediately after Qty in multi-phase format
        amt_col = amt_columns[-1] if amt_columns else qty_col + 1
        rate_col = estimate_rate_col  # Always use Estimate Rate
    else:
        # Fallback to default columns (old single-phase format: col 7=Qty, 8=Rate, 9=Amt)
        qty_col = 7
        rate_col = 8  # Old format had Rate at col 8
        amt_col = 9
    
    print(f"DEBUG parse_workslip_items: Using columns - qty={qty_col}, rate={rate_col}, amt={amt_col} (detected from header row {header_row})")

    for r in range(1, max_row + 1):
        desc_raw = ws.cell(row=r, column=2).value
        desc = str(desc_raw or "").strip()
        if desc == "":
            continue

        low = desc.lower()

        if low == "description of item":
            continue
        if low.startswith("sub total") or low.startswith("sub-total") or low.startswith("ecv"):
            break
        if "supplemental items" in low:
            continue

        if is_merged_cell(ws, r, 2):
            continue

        is_ae = low.startswith("ae")

        qty_raw = ws.cell(row=r, column=qty_col).value
        rate_raw = ws.cell(row=r, column=rate_col).value
        amt_raw = ws.cell(row=r, column=amt_col).value

        qty_exec = to_number(qty_raw)
        rate_exec = to_number(rate_raw)
        amt_exec = to_number(amt_raw)

        if amt_exec != 0:
            if qty_exec == 0 and rate_exec != 0:
                qty_exec = amt_exec / rate_exec
            elif rate_exec == 0 and qty_exec != 0:
                rate_exec = amt_exec / qty_exec

        # Skip items with zero executed quantity (don't include in bill)
        if qty_exec == 0:
            continue

        unit = str(ws.cell(row=r, column=3).value or "").strip()

        items.append({
            "qty": qty_exec,
            "unit": unit,
            "desc": desc,
            "rate": rate_exec,
            "is_ae": is_ae,
        })

    return items


def find_nth_bill_sheet_and_header_row(wb):
    for ws in wb.worksheets:
        for r in range(1, 40):
            a = str(ws.cell(row=r, column=1).value or "").strip().lower()
            c = str(ws.cell(row=r, column=3).value or "").strip().lower()
            if "sl" in a and "quantity till date" in c:
                return ws, r
    return wb.worksheets[0], 10


def parse_first_bill_for_nth(ws, header_row):
    """Parse items from a First Bill (8-column format).
    
    header_row is the row with column headers (S.No, Quantity, Unit, Item, Rate, Per, Unit, Amount).
    We skip that header row and start parsing from header_row + 1.
    Only treat rows as items if they have a valid amount in column H.
    """
    items = []
    max_row = min(ws.max_row, 5000)

    for r in range(header_row + 1, max_row + 1):
        desc_raw = ws.cell(row=r, column=4).value
        desc = str(desc_raw or "").strip()
        rate_raw = ws.cell(row=r, column=5).value
        amt_cell_raw = ws.cell(row=r, column=8).value

        desc_low = desc.lower()

        if desc_low.startswith(("ecv", "sub total", "subtotal", "total")):
            break

        # Require amount to be non-empty and non-zero to be considered a valid item
        amt_val = to_number(amt_cell_raw)
        if amt_val == 0 or amt_val is None:
            # No amount = not a data item
            continue

        rate_str = "" if rate_raw is None else str(rate_raw).strip()
        is_rate_empty = (rate_str == "")

        if (
            desc
            and is_rate_empty
            and not desc_low.startswith(("ecv", "sub total", "subtotal", "total"))
        ) or (
            desc
            and is_merged_cell(ws, r, 4)
        ):
            continue

        prev_qty_raw = ws.cell(row=r, column=2).value
        unit_raw = ws.cell(row=r, column=3).value

        qty_str = "" if prev_qty_raw is None else str(prev_qty_raw).strip()
        unit_str = "" if unit_raw is None else str(unit_raw).strip()

        if desc == "" and is_rate_empty:
            continue

        unit = unit_str
        qty_val = to_number(prev_qty_raw)
        rate_val = to_number(rate_raw)
        prev_amount_val = amt_val

        if (amt_cell_raw is None or str(amt_cell_raw).strip() == "") and (qty_val != 0 or rate_val != 0):
            prev_amount_val = qty_val * rate_val

        items.append({
            "desc": desc,
            "unit": unit,
            "rate": rate_val,
            "prev_qty": qty_val,
            "prev_amount": prev_amount_val,
        })

    return items


def parse_nth_bill_for_next(ws, header_row):
    """Parse items from an Nth bill (11-column format with Unit column).
    
    Format: A: S.No, B: Item, C: Quantity, D: Unit, E: Rate, F: Total Value,
            G-H: Deduct Previous (Qty, Amt), I-J: Since Last (Qty, Amt), K: Remarks
    
    We look at the header row to determine if there's a Unit column.
    """
    items = []
    max_row = min(ws.max_row, 5000)
    start_row = header_row + 2
    
    # Check if this is an 11-column format with Unit column
    # by examining the header row for 'Unit' in column D or similar patterns
    has_unit_column = False
    for col in range(3, 6):  # Check columns C, D, E
        hdr_val = str(ws.cell(row=header_row, column=col).value or "").strip().lower()
        if hdr_val == "unit":
            has_unit_column = True
            break
    
    for r in range(start_row, max_row + 1):
        desc_raw = ws.cell(row=r, column=2).value
        desc = str(desc_raw or "").strip()
        if desc == "":
            continue

        low = desc.lower()
        if low.startswith("sub total") or low.startswith("subtotal"):
            break

        if has_unit_column:
            # 11-column format: A: S.No, B: Item, C: Qty, D: Unit, E: Rate, F: Total Value
            unit_raw = ws.cell(row=r, column=4).value
            rate_raw = ws.cell(row=r, column=5).value
            prev_qty_raw = ws.cell(row=r, column=3).value
            amt_cell_raw = ws.cell(row=r, column=6).value
        else:
            # 10-column format: A: S.No, B: Item, C: Qty, D: Rate, E: Total Value
            unit_raw = None
            rate_raw = ws.cell(row=r, column=4).value
            prev_qty_raw = ws.cell(row=r, column=3).value
            amt_cell_raw = ws.cell(row=r, column=5).value

        unit_str = str(unit_raw or "").strip() if unit_raw else ""
        qty_val = to_number(prev_qty_raw)
        rate_val = to_number(rate_raw)
        prev_amount_val = to_number(amt_cell_raw)

        if (amt_cell_raw is None or str(amt_cell_raw).strip() == "") and (qty_val != 0 or rate_val != 0):
            prev_amount_val = qty_val * rate_val

        items.append({
            "desc": desc,
            "unit": unit_str,
            "rate": rate_val,
            "prev_qty": qty_val,
            "prev_amount": prev_amount_val,
        })

    return items


def read_tp_from_sheet(ws):
    tp_percent = None
    tp_type = None
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        for col in (2, 4):
            val = ws.cell(row=r, column=col).value
            if not val:
                continue
            s = str(val).strip()
            low = s.lower()
            if "t.p" in low:
                if low.startswith("add"):
                    tp_type = "Excess"
                elif low.startswith("deduct"):
                    tp_type = "Less"
                m = re.search(r"(\d+(\.\d+)?)", low)
                if m:
                    tp_percent = float(m.group(1))
                return tp_percent, tp_type
    return None, None


def singular_unit(plural):
    p = str(plural or "").strip()
    if p.lower().endswith("s") and len(p) > 1:
        return p[:-1]
    return p


def ordinal_word(n):
    mapping = {
        1: "First",
        2: "Second",
        3: "Third",
        4: "Fourth",
        5: "Fifth",
        6: "Sixth",
        7: "Seventh",
        8: "Eighth",
        9: "Ninth",
        10: "Tenth",
    }
    if n in mapping:
        return mapping[n]
    if 10 < n % 100 < 14:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        if suffix is None:
            suffix = "th"
    return f"{n}{suffix}"


