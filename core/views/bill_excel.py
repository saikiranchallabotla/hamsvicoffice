# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import (_apply_print_settings, _format_indian_number,
    _number_to_words_rupees, _format_date_to_ddmmyyyy)
from .bill_parsing import (singular_unit, ordinal_word, to_number, is_merged_cell,
    read_tp_from_sheet, extract_tp_from_workslip)

def create_first_bill_sheet(
    wb_out,
    sheet_name,
    items,
    header_data,
    title_text,
    tp_percent,
    tp_type,
    mb_measure_no, mb_measure_p_from, mb_measure_p_to,
    mb_abs_no, mb_abs_p_from, mb_abs_p_to,
    doi, doc, domr, dobr
):
    if wb_out.worksheets and wb_out.worksheets[0].cell(row=1, column=1).value is None and len(wb_out.worksheets) == 1:
        ws_bill = wb_out.active
        ws_bill.title = sheet_name
    else:
        ws_bill = wb_out.create_sheet(title=sheet_name)

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFC8C8C8")
    subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")

    ws_bill.merge_cells("A1:H1")
    c = ws_bill["A1"]
    c.value = title_text
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")

    work_val = header_data.get("name_of_work", "").strip()
    est_val = header_data.get("estimate_amount", "").strip()
    adm_val = header_data.get("admin_sanction", "").strip()
    tech_val = header_data.get("tech_sanction", "").strip()
    agt_val = header_data.get("agreement", "").strip()
    agency_val = header_data.get("agency", "").strip()

    ws_bill.merge_cells("A2:H2")
    c2 = ws_bill["A2"]
    c2.value = f"Name of the work : {work_val}" if work_val else "Name of the work :"
    c2.font = Font(bold=True)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    ws_bill.merge_cells("A3:H3")
    c3 = ws_bill["A3"]
    c3.value = f"Estimate Amount : {est_val}" if est_val else "Estimate Amount :"
    c3.font = Font(bold=True)
    c3.alignment = Alignment(horizontal="left", vertical="center")

    ws_bill.merge_cells("A4:H4")
    c4 = ws_bill["A4"]
    c4.value = (
        f"Ref. to Administrative sanction : {adm_val}"
        if adm_val else "Ref. to Administrative sanction :"
    )
    c4.font = Font(bold=True)
    c4.alignment = Alignment(horizontal="left", vertical="center")

    ws_bill.merge_cells("A5:H5")
    c5 = ws_bill["A5"]
    c5.value = (
        f"Ref. to Technical sanction : {tech_val}"
        if tech_val else "Ref. to Technical sanction :"
    )
    c5.font = Font(bold=True)
    c5.alignment = Alignment(horizontal="left", vertical="center")

    ws_bill.merge_cells("A6:H6")
    c6 = ws_bill["A6"]
    c6.value = (
        f"Ref. to Agreement : {agt_val}"
        if agt_val else "Ref. to Agreement :"
    )
    c6.font = Font(bold=True)
    c6.alignment = Alignment(horizontal="left", vertical="center")

    ws_bill.merge_cells("A7:H7")
    c7 = ws_bill["A7"]
    c7.value = (
        f"Name of the Agency : {agency_val}"
        if agency_val else "Name of the Agency :"
    )
    c7.font = Font(bold=True)
    c7.alignment = Alignment(horizontal="left", vertical="center")

    ws_bill.merge_cells("A8:H8")
    c8 = ws_bill["A8"]
    c8.value = (
        f"M.B.No Details: MB.No. {mb_measure_no} P.No. {mb_measure_p_from} to {mb_measure_p_to} (Measurements)   "
        f"&   MB.No. {mb_abs_no} P.No. {mb_abs_p_from} to {mb_abs_p_to} (Abstract)"
    )
    c8.font = Font(bold=True)
    c8.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_bill.merge_cells("A9:H9")
    c9 = ws_bill["A9"]
    c9.value = f"DOI : {doi}    DOC : {doc}    DOMR : {domr}    DOBR : {dobr}"
    c9.font = Font(bold=True)
    c9.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(1, 10):
        for c_idx in range(1, 9):
            cell = ws_bill.cell(row=r, column=c_idx)
            cell.border = border_all

    header_row = 10
    headers = ["S.No", "Quantity", "Unit", "Item", "Rate", "Per", "Unit", "Amount"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws_bill.cell(row=header_row, column=col_idx, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        cell.fill = header_fill

    ws_bill.column_dimensions["A"].width = 6
    ws_bill.column_dimensions["B"].width = 10
    ws_bill.column_dimensions["C"].width = 10
    ws_bill.column_dimensions["D"].width = 45
    ws_bill.column_dimensions["E"].width = 10
    ws_bill.column_dimensions["F"].width = 6
    ws_bill.column_dimensions["G"].width = 10
    ws_bill.column_dimensions["H"].width = 15

    data_start = header_row + 1
    row_idx = data_start
    slno = 1

    for it in items:
        qty = round(float(it.get("qty", 0) or 0), 2)
        unit_pl = str(it.get("unit") or "").strip()
        desc = it.get("desc") or ""
        rate = round(float(it.get("rate", 0.0) or 0), 2)
        is_ae = bool(it.get("is_ae", False))

        if is_ae:
            ae_num = it.get("ae_number", "")
            ws_bill.cell(row=row_idx, column=1, value=None)
            desc = f"AE{ae_num}" if ae_num else "AE"
        else:
            ws_bill.cell(row=row_idx, column=1, value=slno)

        ws_bill.cell(row=row_idx, column=2, value=qty)
        ws_bill.cell(row=row_idx, column=3, value=unit_pl)
        ws_bill.cell(row=row_idx, column=4, value=desc)
        ws_bill.cell(row=row_idx, column=5, value=rate)
        ws_bill.cell(row=row_idx, column=6, value=1)
        ws_bill.cell(row=row_idx, column=7, value=singular_unit(unit_pl))
        ws_bill.cell(row=row_idx, column=8, value=f"=ROUND(B{row_idx}*E{row_idx},2)")

        fmt_money = '#,##0.00'
        for c_idx in range(1, 9):
            cell = ws_bill.cell(row=row_idx, column=c_idx)
            cell.border = border_all
            if c_idx == 4:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx in (2, 5, 8):
                cell.number_format = fmt_money

        if not is_ae:
            slno += 1

        row_idx += 1

    last_item_row = row_idx - 1

    sub_row = row_idx
    ws_bill.cell(row=sub_row, column=4, value="Sub Total Amount")
    ws_bill.cell(row=sub_row, column=8, value=f"=ROUND(SUM(H{data_start}:H{last_item_row}),2)")
    ws_bill.cell(row=sub_row, column=8).number_format = '#,##0.00'

    for c_idx in range(1, 9):
        cell = ws_bill.cell(row=sub_row, column=c_idx)
        cell.font = Font(bold=True)
        cell.border = border_all
        cell.fill = subtotal_fill
        if c_idx == 4:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    tp_row = sub_row + 1
    tp_percent = float(tp_percent or 0.0)
    tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

    # Show Add or Deduct explicitly based on tp_type
    label_prefix = "Deduct" if tp_type == "Less" else "Add"
    label_tp = f"{label_prefix} T.P @ {tp_percent} % {tp_type}"
    ws_bill.cell(row=tp_row, column=4, value=label_tp)
    ws_bill.cell(row=tp_row, column=8, value=f"=ROUND(H{sub_row}*{abs(tp_percent)}/100,2)")
    ws_bill.cell(row=tp_row, column=8).number_format = '#,##0.00'

    for c_idx in range(1, 9):
        cell = ws_bill.cell(row=tp_row, column=c_idx)
        cell.font = Font(bold=True)
        cell.border = border_all
        cell.fill = subtotal_fill
        if c_idx == 4:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row = tp_row + 1
    ws_bill.cell(row=total_row, column=4, value="Total")

    if tp_type == "Less":
        ws_bill.cell(row=total_row, column=8, value=f"=ROUND(H{sub_row}-H{tp_row},2)")
    else:
        ws_bill.cell(row=total_row, column=8, value=f"=ROUND(H{sub_row}+H{tp_row},2)")
    ws_bill.cell(row=total_row, column=8).number_format = '#,##0.00'

    for c_idx in range(1, 9):
        cell = ws_bill.cell(row=total_row, column=c_idx)
        cell.font = Font(bold=True)
        cell.border = border_all
        cell.fill = subtotal_fill
        if c_idx == 4:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(1, ws_bill.max_row + 1):
        ws_bill.row_dimensions[r].height = None


def build_first_bill_wb(items, header_data, title_text,
                        tp_percent, tp_type,
                        mb_measure_no, mb_measure_p_from, mb_measure_p_to,
                        mb_abs_no, mb_abs_p_from, mb_abs_p_to,
                        doi, doc, domr, dobr):
    wb_out = Workbook()
    create_first_bill_sheet(
        wb_out,
        sheet_name="Bill",
        items=items,
        header_data=header_data,
        title_text=title_text,
        tp_percent=tp_percent,
        tp_type=tp_type,
        mb_measure_no=mb_measure_no,
        mb_measure_p_from=mb_measure_p_from,
        mb_measure_p_to=mb_measure_p_to,
        mb_abs_no=mb_abs_no,
        mb_abs_p_from=mb_abs_p_from,
        mb_abs_p_to=mb_abs_p_to,
        doi=doi,
        doc=doc,
        domr=domr,
        dobr=dobr,
    )
    return wb_out


def build_nth_bill_wb(items, header_data, title_text,
                      tp_percent, tp_type,
                      mb_measure_no, mb_measure_p_from, mb_measure_p_to,
                      mb_abs_no, mb_abs_p_from, mb_abs_p_to,
                      doi, doc, domr, dobr):
    """
    Build a complete Nth bill workbook with 11-column format:
    Sl.No, Item, Quantity Till Date, Unit, Rate per Unit, Total Value Till Date,
    Deduct Previous (Qty, Amount), Since Last (Qty, Amount), Remarks
    """
    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "Bill"

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFC8C8C8")
    subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")

    ws.merge_cells("A1:K1")
    c1 = ws["A1"]
    c1.value = title_text
    c1.font = Font(bold=True, size=14)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    work_val = header_data.get("name_of_work", "").strip()
    est_val = header_data.get("estimate_amount", "").strip()
    adm_val = header_data.get("admin_sanction", "").strip()
    tech_val = header_data.get("tech_sanction", "").strip()
    agt_val = header_data.get("agreement", "").strip()
    agency_val = header_data.get("agency", "").strip()

    ws.merge_cells("A2:K2")
    c2 = ws["A2"]
    c2.value = f"Name of the work : {work_val}" if work_val else "Name of the work :"
    c2.font = Font(bold=True)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A3:K3")
    c3 = ws["A3"]
    c3.value = f"Estimate Amount : {est_val}" if est_val else "Estimate Amount :"
    c3.font = Font(bold=True)
    c3.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:K4")
    c4 = ws["A4"]
    c4.value = (
        f"Ref. to Administrative sanction : {adm_val}"
        if adm_val else "Ref. to Administrative sanction :"
    )
    c4.font = Font(bold=True)
    c4.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A5:K5")
    c5 = ws["A5"]
    c5.value = (
        f"Ref. to Technical sanction : {tech_val}"
        if tech_val else "Ref. to Technical sanction :"
    )
    c5.font = Font(bold=True)
    c5.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A6:K6")
    c6 = ws["A6"]
    c6.value = (
        f"Ref. to Agreement : {agt_val}"
        if agt_val else "Ref. to Agreement :"
    )
    c6.font = Font(bold=True)
    c6.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A7:K7")
    c7 = ws["A7"]
    c7.value = (
        f"Name of the Agency : {agency_val}"
        if agency_val else "Name of the Agency :"
    )
    c7.font = Font(bold=True)
    c7.alignment = Alignment(horizontal="left", vertical="center")

    # -------- ROW 8: Label + MB details (NTH BILL) IN ONE MERGED CELL --------
    ws.merge_cells("A8:K8")
    c8 = ws["A8"]
    c8.value = (
        f"M.B.No Details: MB.No. {mb_measure_no} P.No. {mb_measure_p_from} to {mb_measure_p_to} (Measurements)   "
        f"&   MB.No. {mb_abs_no} P.No. {mb_abs_p_from} to {mb_abs_p_to} (Abstract)"
    )
    c8.font = Font(bold=True)
    c8.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A9:K9")
    c9 = ws["A9"]
    c9.value = f"DOI : {doi}    DOC : {doc}    DOMR : {domr}    DOBR : {dobr}"
    c9.font = Font(bold=True)
    c9.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(1, 10):
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = border_all

    # Merge header cells: S.No, Item, Qty Till Date, Unit, Rate, Total Value, Remarks span 2 rows
    for col in [1, 2, 3, 4, 5, 6, 11]:
        ws.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)

    ws.merge_cells("G10:H10")  # Deduct Previous
    ws.merge_cells("I10:J10")  # Since Last

    ws.cell(row=10, column=1, value="S.No")
    ws.cell(row=10, column=2, value="Item")
    ws.cell(row=10, column=3, value="Quantity Till Date")
    ws.cell(row=10, column=4, value="Unit")
    ws.cell(row=10, column=5, value="Rate per Unit")
    ws.cell(row=10, column=6, value="Total Value till date")
    ws.cell(row=10, column=7, value="Deduct Previous Measurements")
    ws.cell(row=10, column=9, value="Since Last Measurements")
    ws.cell(row=10, column=11, value="Remarks")

    ws.cell(row=11, column=7, value="Quantity")
    ws.cell(row=11, column=8, value="Amount")
    ws.cell(row=11, column=9, value="Quantity")
    ws.cell(row=11, column=10, value="Amount")

    for r in (10, 11):
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 20

    data_start = 12
    r = data_start
    sl = 1

    for it in items:
        desc = it.get("desc") or ""
        unit = it.get("unit") or ""
        rate = round(float(it.get("rate", 0.0) or 0), 2)
        prev_qty = round(float(it.get("prev_qty", 0.0) or 0), 2)
        prev_amount = round(float(it.get("prev_amount", 0.0) or 0), 2)
        is_ae = bool(it.get("is_ae", False))

        if is_ae:
            ae_num = it.get("ae_number", "")
            ws.cell(row=r, column=1, value=None)
            desc = f"AE{ae_num}" if ae_num else "AE"
        else:
            ws.cell(row=r, column=1, value=sl)

        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=None)  # Quantity Till Date (to be filled)
        ws.cell(row=r, column=4, value=unit)
        ws.cell(row=r, column=5, value=rate)
        ws.cell(row=r, column=6, value=f"=ROUND(C{r}*E{r},2)")
        ws.cell(row=r, column=7, value=prev_qty)
        ws.cell(row=r, column=8, value=prev_amount)
        ws.cell(row=r, column=9, value=f"=ROUND(C{r}-G{r},2)")
        ws.cell(row=r, column=10, value=f"=ROUND(F{r}-H{r},2)")
        ws.cell(row=r, column=11, value="")

        fmt_money = '#,##0.00'
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = border_all
            if col == 2:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (3, 5, 6, 7, 8, 9, 10):
                cell.number_format = fmt_money

        r += 1
        if not is_ae:
            sl += 1

    last_item_row = r - 1

    sub_row = r
    ws.cell(row=sub_row, column=2, value="Sub Total")
    ws.cell(row=sub_row, column=6, value=f"=ROUND(SUM(F{data_start}:F{last_item_row}),2)")
    ws.cell(row=sub_row, column=8, value=f"=ROUND(SUM(H{data_start}:H{last_item_row}),2)")
    ws.cell(row=sub_row, column=10, value=f"=ROUND(SUM(J{data_start}:J{last_item_row}),2)")

    tp_row = sub_row + 1
    tp_percent = float(tp_percent or 0.0)
    tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

    # Show Add or Deduct explicitly based on tp_type
    label_prefix = "Deduct" if tp_type == "Less" else "Add"
    ws.cell(row=tp_row, column=2, value=f"{label_prefix} T.P @ {tp_percent} % {tp_type}")
    ws.cell(row=tp_row, column=6, value=f"=ROUND(F{sub_row}*{abs(tp_percent)}/100,2)")
    ws.cell(row=tp_row, column=8, value=f"=ROUND(H{sub_row}*{abs(tp_percent)}/100,2)")
    ws.cell(row=tp_row, column=10, value=f"=ROUND(J{sub_row}*{abs(tp_percent)}/100,2)")

    total_row = tp_row + 1
    ws.cell(row=total_row, column=2, value="Total")

    if tp_type == "Less":
        ws.cell(row=total_row, column=6, value=f"=ROUND(F{sub_row}-F{tp_row},2)")
        ws.cell(row=total_row, column=8, value=f"=ROUND(H{sub_row}-H{tp_row},2)")
        ws.cell(row=total_row, column=10, value=f"=ROUND(J{sub_row}-J{tp_row},2)")
    else:
        ws.cell(row=total_row, column=6, value=f"=ROUND(F{sub_row}+F{tp_row},2)")
        ws.cell(row=total_row, column=8, value=f"=ROUND(H{sub_row}+H{tp_row},2)")
        ws.cell(row=total_row, column=10, value=f"=ROUND(J{sub_row}+J{tp_row},2)")

    for rr in [sub_row, tp_row, total_row]:
        for col in range(1, 12):
            cell = ws.cell(row=rr, column=col)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = subtotal_fill
            if col == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (3, 5, 6, 7, 8, 9, 10):
                cell.number_format = '#,##0.00'

    for rr in range(1, ws.max_row + 1):
        ws.row_dimensions[rr].height = None

    return wb_out


def _populate_nth_bill_sheet(ws, items, header_data, title_text,
                             tp_percent, tp_type,
                             mb_measure_no, mb_measure_p_from, mb_measure_p_to,
                             mb_abs_no, mb_abs_p_from, mb_abs_p_to,
                             doi, doc, domr, dobr):
    """Populate an existing worksheet with Nth bill data.
    
    This is similar to build_nth_bill_wb but works on an existing sheet.
    Uses 11-column format with Unit column.
    """
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFC8C8C8")
    subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")

    ws.merge_cells("A1:K1")
    c1 = ws["A1"]
    c1.value = title_text
    c1.font = Font(bold=True, size=14)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    work_val = header_data.get("name_of_work", "").strip()
    est_val = header_data.get("estimate_amount", "").strip()
    adm_val = header_data.get("admin_sanction", "").strip()
    tech_val = header_data.get("tech_sanction", "").strip()
    agt_val = header_data.get("agreement", "").strip()
    agency_val = header_data.get("agency", "").strip()

    ws.merge_cells("A2:K2")
    c2 = ws["A2"]
    c2.value = f"Name of the work : {work_val}" if work_val else "Name of the work :"
    c2.font = Font(bold=True)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A3:K3")
    c3 = ws["A3"]
    c3.value = f"Estimate Amount : {est_val}" if est_val else "Estimate Amount :"
    c3.font = Font(bold=True)
    c3.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:K4")
    c4 = ws["A4"]
    c4.value = (
        f"Ref. to Administrative sanction : {adm_val}"
        if adm_val else "Ref. to Administrative sanction :"
    )
    c4.font = Font(bold=True)
    c4.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A5:K5")
    c5 = ws["A5"]
    c5.value = (
        f"Ref. to Technical sanction : {tech_val}"
        if tech_val else "Ref. to Technical sanction :"
    )
    c5.font = Font(bold=True)
    c5.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A6:K6")
    c6 = ws["A6"]
    c6.value = (
        f"Ref. to Agreement : {agt_val}"
        if agt_val else "Ref. to Agreement :"
    )
    c6.font = Font(bold=True)
    c6.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A7:K7")
    c7 = ws["A7"]
    c7.value = (
        f"Name of the Agency : {agency_val}"
        if agency_val else "Name of the Agency :"
    )
    c7.font = Font(bold=True)
    c7.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A8:K8")
    c8 = ws["A8"]
    c8.value = (
        f"M.B.No Details: MB.No. {mb_measure_no} P.No. {mb_measure_p_from} to {mb_measure_p_to} (Measurements)   "
        f"&   MB.No. {mb_abs_no} P.No. {mb_abs_p_from} to {mb_abs_p_to} (Abstract)"
    )
    c8.font = Font(bold=True)
    c8.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A9:K9")
    c9 = ws["A9"]
    c9.value = f"DOI : {doi}    DOC : {doc}    DOMR : {domr}    DOBR : {dobr}"
    c9.font = Font(bold=True)
    c9.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(1, 10):
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = border_all

    # Merge header cells: S.No, Item, Qty Till Date, Unit, Rate, Total Value, Remarks span 2 rows
    for col in [1, 2, 3, 4, 5, 6, 11]:
        ws.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)

    ws.merge_cells("G10:H10")  # Deduct Previous
    ws.merge_cells("I10:J10")  # Since Last

    ws.cell(row=10, column=1, value="S.No")
    ws.cell(row=10, column=2, value="Item")
    ws.cell(row=10, column=3, value="Quantity Till Date")
    ws.cell(row=10, column=4, value="Unit")
    ws.cell(row=10, column=5, value="Rate per Unit")
    ws.cell(row=10, column=6, value="Total Value till date")
    ws.cell(row=10, column=7, value="Deduct Previous Measurements")
    ws.cell(row=10, column=9, value="Since Last Measurements")
    ws.cell(row=10, column=11, value="Remarks")

    ws.cell(row=11, column=7, value="Quantity")
    ws.cell(row=11, column=8, value="Amount")
    ws.cell(row=11, column=9, value="Quantity")
    ws.cell(row=11, column=10, value="Amount")

    for r in (10, 11):
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 20

    data_start = 12
    r = data_start
    sl = 1

    for it in items:
        desc = it.get("desc") or ""
        unit = it.get("unit") or ""
        rate = round(float(it.get("rate", 0.0) or 0), 2)
        prev_qty = round(float(it.get("prev_qty", 0.0) or 0), 2)
        prev_amount = round(float(it.get("prev_amount", 0.0) or 0), 2)
        is_ae = bool(it.get("is_ae", False))

        if is_ae:
            ae_num = it.get("ae_number", "")
            ws.cell(row=r, column=1, value=None)
            desc = f"AE{ae_num}" if ae_num else "AE"
        else:
            ws.cell(row=r, column=1, value=sl)

        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=None)  # Quantity Till Date (to be filled)
        ws.cell(row=r, column=4, value=unit)
        ws.cell(row=r, column=5, value=rate)
        ws.cell(row=r, column=6, value=f"=ROUND(C{r}*E{r},2)")
        ws.cell(row=r, column=7, value=prev_qty)
        ws.cell(row=r, column=8, value=prev_amount)
        ws.cell(row=r, column=9, value=f"=ROUND(C{r}-G{r},2)")
        ws.cell(row=r, column=10, value=f"=ROUND(F{r}-H{r},2)")
        ws.cell(row=r, column=11, value="")

        fmt_money = '#,##0.00'
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = border_all
            if col == 2:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (3, 5, 6, 7, 8, 9, 10):
                cell.number_format = fmt_money

        r += 1
        if not is_ae:
            sl += 1

    last_item_row = r - 1

    sub_row = r
    ws.cell(row=sub_row, column=2, value="Sub Total")
    ws.cell(row=sub_row, column=6, value=f"=ROUND(SUM(F{data_start}:F{last_item_row}),2)")
    ws.cell(row=sub_row, column=8, value=f"=ROUND(SUM(H{data_start}:H{last_item_row}),2)")
    ws.cell(row=sub_row, column=10, value=f"=ROUND(SUM(J{data_start}:J{last_item_row}),2)")

    tp_row = sub_row + 1
    tp_percent = float(tp_percent or 0.0)
    tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

    label_prefix = "Deduct" if tp_type == "Less" else "Add"
    ws.cell(row=tp_row, column=2, value=f"{label_prefix} T.P @ {tp_percent} % {tp_type}")
    ws.cell(row=tp_row, column=6, value=f"=ROUND(F{sub_row}*{abs(tp_percent)}/100,2)")
    ws.cell(row=tp_row, column=8, value=f"=ROUND(H{sub_row}*{abs(tp_percent)}/100,2)")
    ws.cell(row=tp_row, column=10, value=f"=ROUND(J{sub_row}*{abs(tp_percent)}/100,2)")

    total_row = tp_row + 1
    ws.cell(row=total_row, column=2, value="Total")

    if tp_type == "Less":
        ws.cell(row=total_row, column=6, value=f"=ROUND(F{sub_row}-F{tp_row},2)")
        ws.cell(row=total_row, column=8, value=f"=ROUND(H{sub_row}-H{tp_row},2)")
        ws.cell(row=total_row, column=10, value=f"=ROUND(J{sub_row}-J{tp_row},2)")
    else:
        ws.cell(row=total_row, column=6, value=f"=ROUND(F{sub_row}+F{tp_row},2)")
        ws.cell(row=total_row, column=8, value=f"=ROUND(H{sub_row}+H{tp_row},2)")
        ws.cell(row=total_row, column=10, value=f"=ROUND(J{sub_row}+J{tp_row},2)")

    for rr in [sub_row, tp_row, total_row]:
        for col in range(1, 12):
            cell = ws.cell(row=rr, column=col)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = subtotal_fill
            if col == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (3, 5, 6, 7, 8, 9, 10):
                cell.number_format = '#,##0.00'

    for rr in range(1, ws.max_row + 1):
        ws.row_dimensions[rr].height = None


