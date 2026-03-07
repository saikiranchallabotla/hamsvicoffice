# Auto-generated from core/views.py split
import json
import os
import re
import logging
from copy import copy

import inflect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from django.urls import reverse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string

import io
from io import BytesIO
from difflib import SequenceMatcher

from ..models import Project, SelfFormattedTemplate, Estimate, Organization, Membership, Upload, Job, OutputFile, LetterSettings
from ..decorators import org_required, role_required

logger = logging.getLogger(__name__)
from ..tasks import process_excel_upload, generate_bill_pdf, generate_workslip_pdf, generate_bill_document_task
from ..utils_excel import load_backend, copy_block_with_styles_and_formulas, build_temp_day_rates

p_engine = inflect.engine()
BILL_TEMPLATES_DIR = os.path.join(settings.BASE_DIR, "core", "templates", "core", "bill_templates")
_inflect_engine = inflect.engine()

from .utils import _apply_print_settings
from .bill_parsing import (parse_estimate_items, parse_workslip_items,
    find_estimate_sheet_and_header_row, find_workslip_sheet, find_all_workslip_sheets,
    read_tp_from_sheet, extract_tp_from_workslip, to_number, is_merged_cell,
    ordinal_word, singular_unit, _extract_header_data_from_sheet)
from .bill_excel import build_first_bill_wb, build_nth_bill_wb

def build_estimate_wb(ws_src, blocks):
    """Minimal helper to build an estimate workbook from source and blocks.

    This is a conservative stub that returns an empty workbook with a single
    'Estimate' sheet so callers expecting a Workbook can proceed. Replace
    with real implementation when available.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    return wb
    print('DEBUG: defined build_estimate_wb')

    def parse_estimate_items(ws, header_row):
        """
        Estimate / Bill-like format:
          A: Sl.No, B: Quantity, C: Unit, D: Item, E: Rate, F: Per, G: Unit, H: Amount

        Rules:
          - HEADING ROW (skip, but do NOT stop):
              * D has text AND (E is empty OR D is merged)
          - Repeat this skipping for all such heading rows
            until we hit:
              * 'ECV', 'Sub Total', 'Subtotal', 'Total' â†’ STOP.
          - Completely blank rows are just skipped (never used as stop condition).
        """
        items = []
        max_row = min(ws.max_row, 5000)

        for r in range(header_row + 1, max_row + 1):
            desc_raw = ws.cell(row=r, column=4).value  # D
            desc = str(desc_raw or "").strip()
            rate_raw = ws.cell(row=r, column=5).value  # E
            amt_raw = ws.cell(row=r, column=8).value   # H

            desc_low = desc.lower()

            # ---- HARD STOP on totals / ECV rows ----
            if desc_low.startswith(("ecv", "sub total", "subtotal", "total")):
                break

            rate_str = "" if rate_raw is None else str(rate_raw).strip()
            is_rate_empty = (rate_str == "")

            # ---------- HEADING ROW ----------
            if desc and (is_rate_empty or is_merged_cell(ws, r, 4)):
                # pure heading â†’ skip, but DO NOT stop
                continue

            qty_raw = ws.cell(row=r, column=2).value  # B
            unit_raw = ws.cell(row=r, column=3).value  # C

            qty_str = "" if qty_raw is None else str(qty_raw).strip()
            unit_str = "" if unit_raw is None else str(unit_raw).strip()
            amt_str = "" if amt_raw is None else str(amt_raw).strip()

            all_blank = (
                desc == "" and
                is_rate_empty and
                qty_str == "" and
                unit_str == "" and
                amt_str == ""
            )

            # Completely blank row â†’ just skip
            if all_blank:
                continue

            if desc == "" and is_rate_empty:
                # still junk, skip
                continue

            # ---------- REAL ITEM ROW ----------
            qty = to_number(qty_raw)
            unit = unit_str
            rate = to_number(rate_raw)
            amt = to_number(amt_raw)

            if amt != 0:
                if qty == 0 and rate != 0:
                    qty = amt / rate
                elif rate == 0 and qty != 0:
                    rate = amt / qty

            items.append({
                "qty": qty,
                "unit": unit,
                "desc": desc,
                "rate": rate,
            })

        return items

    def find_workslip_sheet(wb):
        """
        Single-sheet heuristic for backward compatibility.
        """
        for ws in wb.worksheets:
            for r in range(1, 40):
                b = str(ws.cell(row=r, column=2).value or "").strip().lower()
                g = str(ws.cell(row=r, column=7).value or "").strip().lower()
                if "description" in b and ("qty" in g or "quantity" in g):
                    return ws
        return wb.worksheets[0]

    def find_all_workslip_sheets(wb):
        """
        NEW: multi-sheet heuristic for WorkSlips.
        """
        results = []
        for ws in wb.worksheets:
            max_scan = min(ws.max_row, 60)
            for r in range(1, max_scan + 1):
                b = str(ws.cell(row=r, column=2).value or "").strip().lower()
                c = str(ws.cell(row=r, column=3).value or "").strip().lower()

                has_desc = (
                    "description of item" in b or
                    "description of item" in c or
                    ("description" in b and "item" in b) or
                    ("description" in c and "item" in c)
                )

                has_qty = False
                for col in range(5, 12):
                    t = str(ws.cell(row=r, column=col).value or "").strip().lower()
                    if "qty" in t or "quantity" in t:
                        has_qty = True
                        break

                if has_desc and has_qty:
                    results.append(ws)
                    break

        if not results:
            results.append(find_workslip_sheet(wb))

        return results

    def parse_workslip_items(ws):
        """
        Parse workslip items, automatically detecting the last workslip phase columns.
        
        For multi-phase workslips (Workslip-1, Workslip-2, etc.), this function will
        find and use the LAST workslip's Qty/Amt columns for bill generation.
        Rate is always from Estimate Rate column for multi-phase workslips.
        """
        items = []
        max_row = min(ws.max_row, 5000)
        
        # Step 1: Find header row and detect column structure
        header_row = 8  # Default
        for r in range(1, 15):
            cell_val = str(ws.cell(row=r, column=1).value or "").strip().lower()
            if "sl" in cell_val:
                header_row = r
                break
        
        # Step 2: Scan header row to find all workslip/execution qty columns
        qty_columns = []
        amt_columns = []
        estimate_rate_col = 5  # Default - Rate (Estimate) column
        
        for c in range(1, 40):  # Scan more columns for multi-phase workslips
            header = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
            
            # Check for workslip/execution qty columns
            is_exec_col = ("execution" in header or "exec" in header or "workslip" in header)
            is_qty = ("qty" in header or "quantity" in header)
            is_amt = ("amount" in header or "amt" in header)
            is_estimate = ("est" in header or "estimate" in header)
            is_rate = ("rate" in header)
            
            # Find estimate rate column
            if is_estimate and is_rate:
                estimate_rate_col = c
            
            # Skip estimate columns - we only want execution/workslip columns
            if is_estimate:
                continue
            
            # Skip More/Less/Remarks columns
            if "more" in header or "less" in header or "remark" in header:
                continue
                
            if is_exec_col and is_qty:
                qty_columns.append(c)
            elif is_exec_col and is_amt:
                amt_columns.append(c)
        
        # Step 3: Determine the columns to use
        # Use the LAST Qty column (latest phase) and its corresponding Amount column
        if qty_columns:
            qty_col = qty_columns[-1]  # Last qty column = latest workslip phase
            amt_col = amt_columns[-1] if amt_columns else qty_col + 1
            rate_col = estimate_rate_col  # Always use Estimate Rate
        else:
            # Fallback to old format
            qty_col = 7
            rate_col = 8
            amt_col = 9
        
        print(f"DEBUG parse_workslip_items (local): Using columns - qty={qty_col}, rate={rate_col}, amt={amt_col}")

        for r in range(1, max_row + 1):
            desc_raw = ws.cell(row=r, column=2).value  # B
            desc = str(desc_raw or "").strip()
            if desc == "":
                continue

            low = desc.lower()

            if low == "description of item":
                continue
            if low.startswith("sub total") or low.startswith("sub-total") or low.startswith("ecv"):
                break
            if "supplemental items" in low:
                continue

            # Heading row in merged cell -> skip
            if is_merged_cell(ws, r, 2):
                continue

            is_ae = low.startswith("ae")

            qty_raw = ws.cell(row=r, column=qty_col).value
            rate_raw = ws.cell(row=r, column=rate_col).value
            amt_raw = ws.cell(row=r, column=amt_col).value

            qty_exec = to_number(qty_raw)
            rate_exec = to_number(rate_raw)
            amt_exec = to_number(amt_raw)

            if amt_exec != 0:
                if qty_exec == 0 and rate_exec != 0:
                    qty_exec = amt_exec / rate_exec
                elif rate_exec == 0 and qty_exec != 0:
                    rate_exec = amt_exec / qty_exec

            # Skip items with zero executed quantity (don't include in bill)
            if qty_exec == 0:
                continue

            unit = str(ws.cell(row=r, column=3).value or "").strip()  # C

            items.append({
                "qty": qty_exec,
                "unit": unit,
                "desc": desc,
                "rate": rate_exec,
                "is_ae": is_ae,
            })

        return items

    def find_nth_bill_sheet_and_header_row(wb):
        for ws in wb.worksheets:
            for r in range(1, 40):
                a = str(ws.cell(row=r, column=1).value or "").strip().lower()
                c = str(ws.cell(row=r, column=3).value or "").strip().lower()
                if "sl" in a and "quantity till date" in c:
                    return ws, r
        return wb.worksheets[0], 10

    def parse_first_bill_for_nth(ws, header_row):
        """
        First & Part / First & Final Bill (8-column format):
          A: Sl.No, B: Quantity, C: Unit, D: Item, E: Rate, F: Per, G: Unit, H: Amount

        Prev qty  = Column B
        Prev amt  = Column H; if H has no usable value (formula without cache),
                    fallback to B * E.

        Heading rows:
          - Description cell (D) is merged, OR
          - D has text but E is empty (and not 'Total' / 'Sub Total' / 'ECV').

        Stop only on totals rows, never on blanks.
        """
        items = []
        max_row = min(ws.max_row, 5000)

        for r in range(header_row + 1, max_row + 1):
            desc_raw = ws.cell(row=r, column=4).value  # D
            desc = str(desc_raw or "").strip()
            rate_raw = ws.cell(row=r, column=5).value  # E

            desc_low = desc.lower()

            if desc_low.startswith(("ecv", "sub total", "subtotal", "total")):
                break

            rate_str = "" if rate_raw is None else str(rate_raw).strip()
            is_rate_empty = (rate_str == "")

            # Heading row â†’ skip only
            if (
                desc
                and is_rate_empty
                and not desc_low.startswith(("ecv", "sub total", "subtotal", "total"))
            ) or (
                desc
                and is_merged_cell(ws, r, 4)
            ):
                continue

            prev_qty_raw = ws.cell(row=r, column=2).value  # B
            amt_cell_raw = ws.cell(row=r, column=8).value  # H
            unit_raw = ws.cell(row=r, column=3).value      # C

            qty_str = "" if prev_qty_raw is None else str(prev_qty_raw).strip()
            unit_str = "" if unit_raw is None else str(unit_raw).strip()
            amt_str = "" if amt_cell_raw is None else str(amt_cell_raw).strip()

            all_blank = (
                desc == "" and is_rate_empty and qty_str == "" and unit_str == "" and amt_str == ""
            )
            if all_blank:
                continue

            if desc == "" and is_rate_empty:
                continue

            unit = unit_str
            qty_val = to_number(prev_qty_raw)
            rate_val = to_number(rate_raw)
            prev_amount_val = to_number(amt_cell_raw)

            if (amt_cell_raw is None or str(amt_cell_raw).strip() == "") and (qty_val != 0 or rate_val != 0):
                prev_amount_val = qty_val * rate_val

            items.append({
                "desc": desc,
                "unit": unit,
                "rate": rate_val,
                "prev_qty": qty_val,
                "prev_amount": prev_amount_val,
            })

        return items

    def parse_nth_bill_for_next(ws, header_row):
        """
        Nth bill format (supports both 10-column and 11-column with Unit):

        10-column: A: S.No, B: Item, C: Qty, D: Rate, E: Total Value
        11-column: A: S.No, B: Item, C: Qty, D: Unit, E: Rate, F: Total Value
        
        Prev qty  = column C (Quantity Till Date)
        Prev amt  = column E or F (Total Value till date)
        """
        items = []
        max_row = min(ws.max_row, 5000)
        start_row = header_row + 2  # skip row 10 & 11 headings
        
        # Check if this is an 11-column format with Unit column
        has_unit_column = False
        for col in range(3, 6):  # Check columns C, D, E
            hdr_val = str(ws.cell(row=header_row, column=col).value or "").strip().lower()
            if hdr_val == "unit":
                has_unit_column = True
                break

        for r in range(start_row, max_row + 1):
            desc_raw = ws.cell(row=r, column=2).value  # B
            desc = str(desc_raw or "").strip()
            if desc == "":
                continue

            low = desc.lower()
            if low.startswith("sub total") or low.startswith("subtotal"):
                break

            if has_unit_column:
                # 11-column format: A: S.No, B: Item, C: Qty, D: Unit, E: Rate, F: Total Value
                unit_raw = ws.cell(row=r, column=4).value  # D
                rate_raw = ws.cell(row=r, column=5).value  # E
                prev_qty_raw = ws.cell(row=r, column=3).value  # C
                amt_cell_raw = ws.cell(row=r, column=6).value  # F
            else:
                # 10-column format: A: S.No, B: Item, C: Qty, D: Rate, E: Total Value
                unit_raw = None
                rate_raw = ws.cell(row=r, column=4).value  # D
                prev_qty_raw = ws.cell(row=r, column=3).value  # C
                amt_cell_raw = ws.cell(row=r, column=5).value  # E

            unit_str = str(unit_raw or "").strip() if unit_raw else ""
            qty_val = to_number(prev_qty_raw)
            rate_val = to_number(rate_raw)
            prev_amount_val = to_number(amt_cell_raw)

            if (amt_cell_raw is None or str(amt_cell_raw).strip() == "") and (qty_val != 0 or rate_val != 0):
                prev_amount_val = qty_val * rate_val

            items.append({
                "desc": desc,
                "unit": unit_str,
                "rate": rate_val,
                "prev_qty": qty_val,
                "prev_amount": prev_amount_val,
            })

        return items

    def read_tp_from_sheet(ws):
        """
        Generic TP reader: look for a cell (B or D) containing 'T.P'
        starting with 'Add' or 'Deduct'.
        """
        tp_percent = None
        tp_type = None
        max_row = ws.max_row
        for r in range(1, max_row + 1):
            for col in (2, 4):
                val = ws.cell(row=r, column=col).value
                if not val:
                    continue
                s = str(val).strip()
                low = s.lower()
                if "t.p" in low:
                    if low.startswith("add"):
                        tp_type = "Excess"
                    elif low.startswith("deduct"):
                        tp_type = "Less"
                    m = re.search(r"(\d+(\.\d+)?)", low)
                    if m:
                        tp_percent = float(m.group(1))
                    return tp_percent, tp_type
        return None, None

    def singular_unit(plural):
        p = str(plural or "").strip()
        if p.lower().endswith("s") and len(p) > 1:
            return p[:-1]
        return p

    def ordinal_word(n):
        mapping = {
            1: "First",
            2: "Second",
            3: "Third",
            4: "Fourth",
            5: "Fifth",
            6: "Sixth",
            7: "Seventh",
            8: "Eighth",
            9: "Ninth",
            10: "Tenth",
        }
        if n in mapping:
            return mapping[n]
        if 10 < n % 100 < 14:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
            if suffix is None:
                suffix = "th"
        return f"{n}{suffix}"

    def create_first_bill_sheet(
        wb_out,
        sheet_name,
        items,
        header_data,
        title_text,
        tp_percent,
        tp_type,
        mb_measure_no, mb_measure_p_from, mb_measure_p_to,
        mb_abs_no, mb_abs_p_from, mb_abs_p_to,
        doi, doc, domr, dobr
    ):
        """
        Create ONE 8-column bill sheet in an existing workbook.
        Same layout as your original First Bill.
        """
        # first sheet: reuse active; others: create new
        if wb_out.worksheets and wb_out.worksheets[0].cell(row=1, column=1).value is None and len(wb_out.worksheets) == 1:
            ws_bill = wb_out.active
            ws_bill.title = sheet_name
        else:
            ws_bill = wb_out.create_sheet(title=sheet_name)

        thin = Side(border_style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="FFC8C8C8")
        subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")

        ws_bill.merge_cells("A1:H1")
        c = ws_bill["A1"]
        c.value = title_text
        c.font = Font(bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")

        work_val = header_data.get("name_of_work", "").strip()
        est_val = header_data.get("estimate_amount", "").strip()
        adm_val = header_data.get("admin_sanction", "").strip()
        tech_val = header_data.get("tech_sanction", "").strip()
        agt_val = header_data.get("agreement", "").strip()
        agency_val = header_data.get("agency", "").strip()

        ws_bill.merge_cells("A2:H2")
        c2 = ws_bill["A2"]
        c2.value = f"Name of the work : {work_val}" if work_val else "Name of the work :"
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="left", vertical="center")

        ws_bill.merge_cells("A3:H3")
        c3 = ws_bill["A3"]
        c3.value = f"Estimate Amount : {est_val}" if est_val else "Estimate Amount :"
        c3.font = Font(bold=True)
        c3.alignment = Alignment(horizontal="left", vertical="center")

        ws_bill.merge_cells("A4:H4")
        c4 = ws_bill["A4"]
        c4.value = (
            f"Ref. to Administrative sanction : {adm_val}"
            if adm_val else "Ref. to Administrative sanction :"
        )
        c4.font = Font(bold=True)
        c4.alignment = Alignment(horizontal="left", vertical="center")

        ws_bill.merge_cells("A5:H5")
        c5 = ws_bill["A5"]
        c5.value = (
            f"Ref. to Technical sanction : {tech_val}"
            if tech_val else "Ref. to Technical sanction :"
        )
        c5.font = Font(bold=True)
        c5.alignment = Alignment(horizontal="left", vertical="center")

        ws_bill.merge_cells("A6:H6")
        c6 = ws_bill["A6"]
        c6.value = (
            f"Ref. to Agreement : {agt_val}"
            if agt_val else "Ref. to Agreement :"
        )
        c6.font = Font(bold=True)
        c6.alignment = Alignment(horizontal="left", vertical="center")

        ws_bill.merge_cells("A7:H7")
        c7 = ws_bill["A7"]
        c7.value = (
            f"Name of the Agency : {agency_val}"
            if agency_val else "Name of the Agency :"
        )
        c7.font = Font(bold=True)
        c7.alignment = Alignment(horizontal="left", vertical="center")

        # -------- ROW 8: Label + MB details (FIRST BILL) IN ONE MERGED CELL --------
        ws_bill.merge_cells("A8:H8")
        c8 = ws_bill["A8"]
        c8.value = (
            f"M.B.No Details: MB.No. {mb_measure_no} P.No. {mb_measure_p_from} to {mb_measure_p_to} (Measurements)   "
            f"&   MB.No. {mb_abs_no} P.No. {mb_abs_p_from} to {mb_abs_p_to} (Abstract)"
        )
        c8.font = Font(bold=True)
        c8.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws_bill.merge_cells("A9:H9")
        c9 = ws_bill["A9"]
        c9.value = f"DOI : {doi}    DOC : {doc}    DOMR : {domr}    DOBR : {dobr}"
        c9.font = Font(bold=True)
        c9.alignment = Alignment(horizontal="left", vertical="center")

        for r in range(1, 10):
            for c_idx in range(1, 9):
                cell = ws_bill.cell(row=r, column=c_idx)
                cell.border = border_all

        header_row = 10
        headers = ["S.No", "Quantity", "Unit", "Item", "Rate", "Per", "Unit", "Amount"]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws_bill.cell(row=header_row, column=col_idx, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            cell.fill = header_fill

        ws_bill.column_dimensions["A"].width = 6
        ws_bill.column_dimensions["B"].width = 10
        ws_bill.column_dimensions["C"].width = 10
        ws_bill.column_dimensions["D"].width = 45
        ws_bill.column_dimensions["E"].width = 10
        ws_bill.column_dimensions["F"].width = 6
        ws_bill.column_dimensions["G"].width = 10
        ws_bill.column_dimensions["H"].width = 15

        data_start = header_row + 1
        row_idx = data_start
        slno = 1

        for it in items:
            qty = it.get("qty", 0)
            unit_pl = str(it.get("unit") or "").strip()
            desc = it.get("desc") or ""
            rate = it.get("rate", 0.0)
            is_ae = bool(it.get("is_ae", False))

            if is_ae:
                ws_bill.cell(row=row_idx, column=1, value=None)
            else:
                ws_bill.cell(row=row_idx, column=1, value=slno)

            ws_bill.cell(row=row_idx, column=2, value=qty)
            ws_bill.cell(row=row_idx, column=3, value=unit_pl)
            ws_bill.cell(row=row_idx, column=4, value=desc)
            ws_bill.cell(row=row_idx, column=5, value=rate)
            ws_bill.cell(row=row_idx, column=6, value=1)
            ws_bill.cell(row=row_idx, column=7, value=singular_unit(unit_pl))
            ws_bill.cell(row=row_idx, column=8, value=f"=B{row_idx}*E{row_idx}")

            for c_idx in range(1, 9):
                cell = ws_bill.cell(row=row_idx, column=c_idx)
                cell.border = border_all
                if c_idx == 4:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            if not is_ae:
                slno += 1

            row_idx += 1

        last_item_row = row_idx - 1

        sub_row = row_idx
        ws_bill.cell(row=sub_row, column=4, value="Sub Total Amount")
        ws_bill.cell(row=sub_row, column=8, value=f"=SUM(H{data_start}:H{last_item_row})")

        for c_idx in range(1, 9):
            cell = ws_bill.cell(row=sub_row, column=c_idx)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = subtotal_fill
            if c_idx == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        tp_row = sub_row + 1
        tp_percent = float(tp_percent or 0.0)
        tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

        label_tp = f"Add/ Deduct T.P @ {tp_percent} % {tp_type}"
        ws_bill.cell(row=tp_row, column=4, value=label_tp)
        ws_bill.cell(row=tp_row, column=8, value=f"=H{sub_row}*{abs(tp_percent)}/100")

        for c_idx in range(1, 9):
            cell = ws_bill.cell(row=tp_row, column=c_idx)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = subtotal_fill
            if c_idx == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        total_row = tp_row + 1
        ws_bill.cell(row=total_row, column=4, value="Total")

        if tp_type == "Less":
            ws_bill.cell(row=total_row, column=8, value=f"=H{sub_row}-H{tp_row}")
        else:
            ws_bill.cell(row=total_row, column=8, value=f"=H{sub_row}+H{tp_row}")

        for c_idx in range(1, 9):
            cell = ws_bill.cell(row=total_row, column=c_idx)
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.fill = subtotal_fill
            if c_idx == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # let Excel auto-fit default heights
        for r in range(1, ws_bill.max_row + 1):
            ws_bill.row_dimensions[r].height = None

    def build_first_bill_wb(items, header_data, title_text,
                            tp_percent, tp_type,
                            mb_measure_no, mb_measure_p_from, mb_measure_p_to,
                            mb_abs_no, mb_abs_p_from, mb_abs_p_to,
                            doi, doc, domr, dobr):
        """
        Backward-compatible wrapper: build a single-sheet bill workbook.
        """
        wb_out = Workbook()
        create_first_bill_sheet(
            wb_out,
            sheet_name="Bill",
            items=items,
            header_data=header_data,
            title_text=title_text,
            tp_percent=tp_percent,
            tp_type=tp_type,
            mb_measure_no=mb_measure_no,
            mb_measure_p_from=mb_measure_p_from,
            mb_measure_p_to=mb_measure_p_to,
            mb_abs_no=mb_abs_no,
            mb_abs_p_from=mb_abs_p_from,
            mb_abs_p_to=mb_abs_p_to,
            doi=doi,
            doc=doc,
            domr=domr,
            dobr=dobr,
        )
        return wb_out

    def build_nth_bill_wb(items, header_data, title_text,
                          tp_percent, tp_type,
                          mb_measure_no, mb_measure_p_from, mb_measure_p_to,
                          mb_abs_no, mb_abs_p_from, mb_abs_p_to,
                          doi, doc, domr, dobr):
        """
        Same as your earlier Nth bill format, single sheet only.
        Uses 11-column format with Unit column.
        """
        wb_out = Workbook()
        ws = wb_out.active
        ws.title = "Bill"

        thin = Side(border_style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="FFC8C8C8")
        subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")

        ws.merge_cells("A1:K1")
        c1 = ws["A1"]
        c1.value = title_text
        c1.font = Font(bold=True, size=14)
        c1.alignment = Alignment(horizontal="center", vertical="center")

        work_val = header_data.get("name_of_work", "").strip()
        est_val = header_data.get("estimate_amount", "").strip()
        adm_val = header_data.get("admin_sanction", "").strip()
        tech_val = header_data.get("tech_sanction", "").strip()
        agt_val = header_data.get("agreement", "").strip()
        agency_val = header_data.get("agency", "").strip()

        ws.merge_cells("A2:K2")
        c2 = ws["A2"]
        c2.value = f"Name of the work : {work_val}" if work_val else "Name of the work :"
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A3:K3")
        c3 = ws["A3"]
        c3.value = f"Estimate Amount : {est_val}" if est_val else "Estimate Amount :"
        c3.font = Font(bold=True)
        c3.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A4:K4")
        c4 = ws["A4"]
        c4.value = (
            f"Ref. to Administrative sanction : {adm_val}"
            if adm_val else "Ref. to Administrative sanction :"
        )
        c4.font = Font(bold=True)
        c4.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A5:K5")
        c5 = ws["A5"]
        c5.value = (
            f"Ref. to Technical sanction : {tech_val}"
            if tech_val else "Ref. to Technical sanction :"
        )
        c5.font = Font(bold=True)
        c5.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A6:K6")
        c6 = ws["A6"]
        c6.value = (
            f"Ref. to Agreement : {agt_val}"
            if agt_val else "Ref. to Agreement :"
        )
        c6.font = Font(bold=True)
        c6.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A7:K7")
        c7 = ws["A7"]
        c7.value = (
            f"Name of the Agency : {agency_val}"
            if agency_val else "Name of the Agency :"
        )
        c7.font = Font(bold=True)
        c7.alignment = Alignment(horizontal="left", vertical="center")

        # -------- ROW 8: Label + MB details (NTH BILL) IN ONE MERGED CELL --------
        ws.merge_cells("A8:K8")
        c8 = ws["A8"]
        c8.value = (
            f"M.B.No Details: MB.No. {mb_measure_no} P.No. {mb_measure_p_from} to {mb_measure_p_to} (Measurements)   "
            f"&   MB.No. {mb_abs_no} P.No. {mb_abs_p_from} to {mb_abs_p_to} (Abstract)"
        )
        c8.font = Font(bold=True)
        c8.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.merge_cells("A9:K9")
        c9 = ws["A9"]
        c9.value = f"DOI : {doi}    DOC : {doc}    DOMR : {domr}    DOBR : {dobr}"
        c9.font = Font(bold=True)
        c9.alignment = Alignment(horizontal="left", vertical="center")

        for r in range(1, 10):
            for col in range(1, 12):
                cell = ws.cell(row=r, column=col)
                cell.border = border_all

        # Merge header cells: S.No, Item, Qty Till Date, Unit, Rate, Total Value, Remarks span 2 rows
        for col in [1, 2, 3, 4, 5, 6, 11]:
            ws.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)

        ws.merge_cells("G10:H10")  # Deduct Previous
        ws.merge_cells("I10:J10")  # Since Last

        ws.cell(row=10, column=1, value="S.No")
        ws.cell(row=10, column=2, value="Item")
        ws.cell(row=10, column=3, value="Quantity Till Date")
        ws.cell(row=10, column=4, value="Unit")
        ws.cell(row=10, column=5, value="Rate per Unit")
        ws.cell(row=10, column=6, value="Total Value till date")
        ws.cell(row=10, column=7, value="Deduct Previous Measurements")
        ws.cell(row=10, column=9, value="Since Last Measurements")
        ws.cell(row=10, column=11, value="Remarks")

        ws.cell(row=11, column=7, value="Quantity")
        ws.cell(row=11, column=8, value="Amount")
        ws.cell(row=11, column=9, value="Quantity")
        ws.cell(row=11, column=10, value="Amount")

        for r in (10, 11):
            for col in range(1, 12):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(bold=True)
                cell.border = border_all
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 16
        ws.column_dimensions["I"].width = 14
        ws.column_dimensions["J"].width = 16
        ws.column_dimensions["K"].width = 20

        data_start = 12
        r = data_start
        sl = 1

        for it in items:
            desc = it.get("desc") or ""
            unit = it.get("unit") or ""
            rate = it.get("rate", 0.0)
            prev_qty = it.get("prev_qty", 0.0)
            prev_amount = it.get("prev_amount", 0.0)

            ws.cell(row=r, column=1, value=sl)
            ws.cell(row=r, column=2, value=desc)
            ws.cell(row=r, column=3, value=None)  # Quantity Till Date (to be filled)
            ws.cell(row=r, column=4, value=unit)
            ws.cell(row=r, column=5, value=rate)
            ws.cell(row=r, column=6, value=f"=C{r}*E{r}")
            ws.cell(row=r, column=7, value=prev_qty)
            ws.cell(row=r, column=8, value=prev_amount)
            ws.cell(row=r, column=9, value=f"=C{r}-G{r}")
            ws.cell(row=r, column=10, value=f"=F{r}-H{r}")
            ws.cell(row=r, column=11, value="")

            for col in range(1, 12):
                cell = ws.cell(row=r, column=col)
                cell.border = border_all
                if col == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            r += 1
            sl += 1

        last_item_row = r - 1

        sub_row = r
        ws.cell(row=sub_row, column=2, value="Sub Total")
        ws.cell(row=sub_row, column=6, value=f"=SUM(F{data_start}:F{last_item_row})")
        ws.cell(row=sub_row, column=8, value=f"=SUM(H{data_start}:H{last_item_row})")
        ws.cell(row=sub_row, column=10, value=f"=SUM(J{data_start}:J{last_item_row})")

        tp_row = sub_row + 1
        tp_percent = float(tp_percent or 0.0)
        tp_type = tp_type if tp_type in ("Less", "Excess") else "Excess"

        label_prefix = "Deduct" if tp_type == "Less" else "Add"
        ws.cell(row=tp_row, column=2, value=f"{label_prefix} T.P @ {tp_percent} % {tp_type}")
        ws.cell(row=tp_row, column=6, value=f"=F{sub_row}*{abs(tp_percent)}/100")
        ws.cell(row=tp_row, column=8, value=f"=H{sub_row}*{abs(tp_percent)}/100")
        ws.cell(row=tp_row, column=10, value=f"=J{sub_row}*{abs(tp_percent)}/100")

        total_row = tp_row + 1
        ws.cell(row=total_row, column=2, value="Total")

        if tp_type == "Less":
            ws.cell(row=total_row, column=6, value=f"=F{sub_row}-F{tp_row}")
            ws.cell(row=total_row, column=8, value=f"=H{sub_row}-H{tp_row}")
            ws.cell(row=total_row, column=10, value=f"=J{sub_row}-J{tp_row}")
        else:
            ws.cell(row=total_row, column=6, value=f"=F{sub_row}+F{tp_row}")
            ws.cell(row=total_row, column=8, value=f"=H{sub_row}+H{tp_row}")
            ws.cell(row=total_row, column=10, value=f"=J{sub_row}+J{tp_row}")

        for rr in [sub_row, tp_row, total_row]:
            for col in range(1, 12):
                cell = ws.cell(row=rr, column=col)
                cell.font = Font(bold=True)
                cell.border = border_all
                cell.fill = subtotal_fill
                if col == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for rr in range(1, ws.max_row + 1):
            ws.row_dimensions[rr].height = None

        return wb_out

    # ---------- main bill() logic ----------
    print('DEBUG: entering main bill logic')

    if request.method == "GET":
        print("DEBUG: bill handling GET -> rendering bill.html")
        resp = render(request, "core/bill.html")
        print("DEBUG: render returned type", type(resp))
        return resp

    action = request.POST.get("action", "").strip()
    uploaded = request.FILES.get("bill_file")

    if not uploaded:
        return render(request, "core/bill.html", {
            "error": "Please upload an Excel file.",
        })

    mb_measure_no = request.POST.get("mb_measure_no", "").strip()
    mb_measure_p_from = request.POST.get("mb_measure_p_from", "").strip()
    mb_measure_p_to = request.POST.get("mb_measure_p_to", "").strip()
    mb_abs_no = request.POST.get("mb_abstract_no", "").strip()
    mb_abs_p_from = request.POST.get("mb_abstract_p_from", "").strip()
    mb_abs_p_to = request.POST.get("mb_abstract_p_to", "").strip()
    doi = request.POST.get("doi", "").strip()
    doc = request.POST.get("doc", "").strip()
    domr = request.POST.get("domr", "").strip()
    dobr = request.POST.get("dobr", "").strip()

    tp_percent_session = to_number(request.session.get("ws_tp_percent", 0.0))
    tp_type_session = request.session.get("ws_tp_type", "Excess")

    try:
        wb_in = load_workbook(uploaded, data_only=True)
        print('DEBUG: loaded workbook, sheets=', [s.title for s in wb_in.worksheets])
    except Exception as e:
        return render(request, "core/bill.html", {
            "error": f"Error reading uploaded Excel: {e}",
        })

    print('DEBUG: action=', action)
    # SECTION 1  -  Bill from Estimate (multi-sheet)
    if action in ("estimate_first_part", "estimate_first_final"):
        sheet_info_list = find_all_estimate_sheets_and_header_rows(wb_in)

        # Build one output workbook with one Bill sheet per estimate sheet
        wb_out = Workbook()

        any_items = False
        for idx, (ws_est, header_row) in enumerate(sheet_info_list, start=1):
            header_data = extract_header_data(ws_est)
            items = parse_estimate_items(ws_est, header_row)
            if not items:
                continue

            any_items = True
            title_text = "CC First & Part Bill" if action == "estimate_first_part" else "CC First & Final Bill"
            tp_percent = tp_percent_session
            tp_type = tp_type_session

            if len(sheet_info_list) == 1:
                sheet_name = "Bill"
            else:
                # Name by sheet: 'Bill_<SheetName>' (trim to Excel limit)
                base = ws_est.title.strip() or f"Sheet{idx}"
                sheet_name = f"Bill_{base}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]

            create_first_bill_sheet(
                wb_out,
                sheet_name=sheet_name,
                items=items,
                header_data=header_data,
                title_text=title_text,
                tp_percent=tp_percent,
                tp_type=tp_type,
                mb_measure_no=mb_measure_no,
                mb_measure_p_from=mb_measure_p_from,
                mb_measure_p_to=mb_measure_p_to,
                mb_abs_no=mb_abs_no,
                mb_abs_p_from=mb_abs_p_from,
                mb_abs_p_to=mb_abs_p_to,
                doi=doi,
                doc=doc,
                domr=domr,
                dobr=dobr,
            )

        if not any_items:
            return render(request, "core/bill.html", {
                "error": "Could not detect any items in the uploaded Estimate (all sheets).",
            })

        _apply_print_settings(wb_out)
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="Bill_from_Estimate.xlsx"'
        wb_out.save(resp)
        return resp

    # SECTION 2  -  Bill from WorkSlip (multi-sheet)
    if action in ("workslip_first_part", "workslip_first_final"):
        ws_list = find_all_workslip_sheets(wb_in)

        wb_out = Workbook()
        any_items = False

        for idx, ws_ws in enumerate(ws_list, start=1):
            header_data = extract_header_data(ws_ws)
            items = parse_workslip_items(ws_ws)
            if not items:
                continue

            any_items = True
            title_text = "CC First & Part Bill" if action == "workslip_first_part" else "CC First & Final Bill"

            tp_percent, tp_type = read_tp_from_sheet(ws_ws)
            if tp_percent is None:
                tp_percent = tp_percent_session
            if tp_type is None:
                tp_type = tp_type_session

            if len(ws_list) == 1:
                sheet_name = "Bill"
            else:
                base = ws_ws.title.strip() or f"WS{idx}"
                sheet_name = f"Bill_{base}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]

            create_first_bill_sheet(
                wb_out,
                sheet_name=sheet_name,
                items=items,
                header_data=header_data,
                title_text=title_text,
                tp_percent=tp_percent,
                tp_type=tp_type,
                mb_measure_no=mb_measure_no,
                mb_measure_p_from=mb_measure_p_from,
                mb_measure_p_to=mb_measure_p_to,
                mb_abs_no=mb_abs_no,
                mb_abs_p_from=mb_abs_p_from,
                mb_abs_p_to=mb_abs_p_to,
                doi=doi,
                doc=doc,
                domr=domr,
                dobr=dobr,
            )

        if not any_items:
            return render(request, "core/bill.html", {
                "error": "Could not detect any executed items in the uploaded WorkSlip (all sheets).",
            })

        _apply_print_settings(wb_out)
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="Bill_from_WorkSlip.xlsx"'
        wb_out.save(resp)
        return resp

    # SECTION 3  -  Nth & Part / 2nd & Final from First & Part Bill
    if action in ("firstpart_nth_part", "firstpart_2nd_final"):
        ws_first, header_row = find_estimate_sheet_and_header_row(wb_in)
        header_data = extract_header_data(ws_first)
        items = parse_first_bill_for_nth(ws_first, header_row)

        if not items:
            return render(request, "core/bill.html", {
                "error": "Could not detect any items in the uploaded First & Part / First & Final Bill.",
            })

        if action == "firstpart_nth_part":
            nth_str = request.POST.get("nth_number", "").strip()
            try:
                nth_val = int(nth_str)
            except Exception:
                nth_val = 2
            if nth_val < 2:
                nth_val = 2
            ord_word = ordinal_word(nth_val)
            title_text = f"CC {ord_word} & Part Bill"
        else:
            ord_word = ordinal_word(2)
            title_text = f"CC {ord_word} & Final Bill"

        tp_percent, tp_type = read_tp_from_sheet(ws_first)
        if tp_percent is None:
            tp_percent = tp_percent_session
        if tp_type is None:
            tp_type = tp_type_session

        wb_out = build_nth_bill_wb(
            items=items,
            header_data=header_data,
            title_text=title_text,
            tp_percent=tp_percent,
            tp_type=tp_type,
            mb_measure_no=mb_measure_no,
            mb_measure_p_from=mb_measure_p_from,
            mb_measure_p_to=mb_measure_p_to,
            mb_abs_no=mb_abs_no,
            mb_abs_p_from=mb_abs_p_from,
            mb_abs_p_to=mb_abs_p_to,
            doi=doi,
            doc=doc,
            domr=domr,
            dobr=dobr,
        )

        _apply_print_settings(wb_out)
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "Nth_Bill_from_FirstPart.xlsx" if action == "firstpart_nth_part" else "Second_Final_from_FirstPart.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb_out.save(resp)
        return resp

    # SECTION 4  -  Nth & Part / Nth & Final from Nth & Part Bill
    if action in ("nth_nth_part", "nth_nth_final"):
        ws_nth, header_row = find_nth_bill_sheet_and_header_row(wb_in)
        header_data = extract_header_data(ws_nth)
        items = parse_nth_bill_for_next(ws_nth, header_row)

        if not items:
            return render(request, "core/bill.html", {
                "error": "Could not detect any items in the uploaded Nth & Part Bill.",
            })

        nth_str = request.POST.get("nth_number", "").strip()
        try:
            nth_val = int(nth_str)
        except Exception:
            nth_val = 2
        if nth_val < 2:
            nth_val = 2
        ord_word = ordinal_word(nth_val)

        if action == "nth_nth_part":
            title_text = f"CC {ord_word} & Part Bill"
        else:
            title_text = f"CC {ord_word} & Final Bill"

        tp_percent, tp_type = read_tp_from_sheet(ws_nth)
        if tp_percent is None:
            tp_percent = tp_percent_session
        if tp_type is None:
            tp_type = tp_type_session

        wb_out = build_nth_bill_wb(
            items=items,
            header_data=header_data,
            title_text=title_text,
            tp_percent=tp_percent,
            tp_type=tp_type,
            mb_measure_no=mb_measure_no,
            mb_measure_p_from=mb_measure_p_from,
            mb_measure_p_to=mb_measure_p_to,
            mb_abs_no=mb_abs_no,
            mb_abs_p_from=mb_abs_p_from,
            mb_abs_p_to=mb_abs_p_to,
            doi=doi,
            doc=doc,
            domr=domr,
            dobr=dobr,
        )

        _apply_print_settings(wb_out)
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "Nth_Bill_from_NthPart.xlsx" if action == "nth_nth_part" else "Nth_Final_from_NthPart.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb_out.save(resp)
        return resp

    return render(request, "core/bill.html", {
        "error": "This action is not implemented or is invalid.",
    })





# ---------- Helpers only for amount calculation for Covering/Movement ----------

