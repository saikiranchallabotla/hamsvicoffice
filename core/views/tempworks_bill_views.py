"""
Temporary Works — Bill module.

Two-sheet workbook:
  Sheet 1 (Bill Summary) — one row per item, totals across events
  Sheet 2 (Bill Detail)  — per-event rows mirroring the Workslip layout

Session keys (tw_bill_*):
  tw_bill_entries
  tw_bill_events_list
  tw_bill_exec_map           - {entry_id: {event_id: qty_bill}}
  tw_bill_prev_exec          - cumulative qty from previous bills (carry-forward)
  tw_bill_work_name
  tw_bill_category
  tw_bill_selected_backend_id
  tw_bill_number
  tw_bill_source_workslip_id
"""
import io
import json
import logging
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from ..utils_excel import build_temp_day_rates, load_backend
from .tempworks_workslip_views import (
    _build_desc_map, _build_unit_map, _coerce_float, _norm_name, _rate_for_event_days, _to_roman_lower,
)

logger = logging.getLogger(__name__)


def _build_bill_view_rows(entries, exec_map, prev_exec, day_rates_by_item):
    rows = []
    sl = 1
    for entry in entries or []:
        if (entry or {}).get("mode") != "multi":
            continue
        item_name = entry.get("name") or ""
        entry_id = entry.get("id") or ""
        events = entry.get("events") or []
        valid_events = []
        for ev in events:
            ev_name = (ev.get("event_name") or "").strip()
            ev_days = int(_coerce_float(ev.get("days"), 0))
            ev_qty = _coerce_float(ev.get("qty"), 0.0)
            ev_id = ev.get("event_id") or ""
            bill_q = _coerce_float((exec_map.get(entry_id) or {}).get(ev_id), 0.0)
            prev_q = _coerce_float((prev_exec.get(entry_id) or {}).get(ev_id), 0.0)
            if not ev_name or ev_days <= 0:
                continue
            # Include event if it has an estimate OR any billed activity
            if ev_qty <= 0 and bill_q <= 0 and prev_q <= 0:
                continue
            valid_events.append({
                "event_id": ev_id,
                "event_name": ev_name,
                "days": ev_days,
                "qty_est": ev_qty,
            })
        if not valid_events:
            continue

        rows.append({
            "kind": "header",
            "sl": sl,
            "entry_id": entry_id,
            "item_name": item_name,
            "desc": item_name,
        })
        item_day_rates = day_rates_by_item.get(_norm_name(item_name)) or day_rates_by_item.get(item_name) or {}
        for i, ev in enumerate(valid_events, start=1):
            roman = _to_roman_lower(i)
            day_word = "day" if ev["days"] == 1 else "days"
            desc = f"{roman}. {ev['event_name']} for {ev['days']} {day_word}"
            rate = _rate_for_event_days(item_day_rates, ev["days"])
            bill_qty = _coerce_float((exec_map.get(entry_id) or {}).get(ev["event_id"]), 0.0)
            prev_qty = _coerce_float((prev_exec.get(entry_id) or {}).get(ev["event_id"]), 0.0)
            net_qty = round(bill_qty - prev_qty, 2)
            rows.append({
                "kind": "event",
                "entry_id": entry_id,
                "event_id": ev["event_id"],
                "roman": roman,
                "desc": desc,
                "event_name": ev["event_name"],
                "days": ev["days"],
                "qty_est": ev["qty_est"],
                "rate": rate,
                "amt_est": round(ev["qty_est"] * rate, 2),
                "qty_bill": bill_qty,
                "qty_prev": prev_qty,
                "qty_net": net_qty,
                "amt_net": round(net_qty * rate, 2),
            })
        sl += 1
    return rows


@login_required(login_url="login")
def temp_bill(request):
    entries = request.session.get("tw_bill_entries") or []
    exec_map = request.session.get("tw_bill_exec_map") or {}
    prev_exec = request.session.get("tw_bill_prev_exec") or {}
    work_name = request.session.get("tw_bill_work_name") or ""
    category = request.session.get("tw_bill_category") or "electrical"
    backend_id = request.session.get("tw_bill_selected_backend_id")
    bill_number = request.session.get("tw_bill_number") or 1

    if not entries:
        return redirect("dashboard")

    items_list = []
    filepath = None
    try:
        items_list, _gm, _um, _ws_src, filepath = load_backend(
            f"temp_{category}", settings.BASE_DIR,
            backend_id=backend_id, module_code="temp_works",
        )
        day_rates = build_temp_day_rates(filepath, items_list) or {}
    except Exception:
        logger.exception("[TEMP_BILL] load_backend failed (category=%s)", category)
        day_rates = {}
        _gm, _um = {}, {}

    unit_by_item = _build_unit_map(_um or {}, _gm or {})

    day_rates_by_item = {}
    for k, v in (day_rates or {}).items():
        day_rates_by_item[k] = v
        day_rates_by_item[_norm_name(k)] = v

    desc_by_item = _build_desc_map(items_list, filepath)

    if request.method == "POST":
        action = request.POST.get("action") or ""
        exec_map_str = request.POST.get("exec_map", "")
        new_exec_map = exec_map
        if exec_map_str:
            try:
                parsed = json.loads(exec_map_str)
                if isinstance(parsed, dict):
                    new_exec_map = {}
                    for ent_id, evs in parsed.items():
                        if not isinstance(evs, dict):
                            continue
                        new_exec_map[str(ent_id)] = {
                            str(ev_id): _coerce_float(q, 0.0)
                            for ev_id, q in evs.items()
                        }
            except Exception:
                logger.exception("[TEMP_BILL] bad exec_map payload")

        wn_form = (request.POST.get("ws_work_name") or "").strip()
        if wn_form:
            work_name = wn_form
            request.session["tw_bill_work_name"] = work_name

        request.session["tw_bill_exec_map"] = new_exec_map
        request.session.modified = True
        exec_map = new_exec_map

        if action == "download_bill":
            return _download_bill_excel(
                entries=entries, exec_map=exec_map, prev_exec=prev_exec,
                day_rates_by_item=day_rates_by_item, desc_by_item=desc_by_item,
                unit_by_item=unit_by_item,
                work_name=work_name, bill_number=bill_number,
            )

    view_rows = _build_bill_view_rows(entries, exec_map, prev_exec, day_rates_by_item)
    total_net = sum(r["amt_net"] for r in view_rows if r["kind"] == "event")

    return render(request, "core/temp_bill.html", {
        "view_rows": view_rows,
        "work_name": work_name,
        "category": category,
        "bill_number": bill_number,
        "total_net": round(total_net, 2),
        "has_previous": any(r.get("qty_prev") for r in view_rows if r["kind"] == "event"),
    })


def _aggregate_per_item(entries, exec_map, prev_exec, day_rates_by_item):
    """Return a list of per-item totals for the Bill Summary sheet."""
    summary = []
    for entry in entries or []:
        if (entry or {}).get("mode") != "multi":
            continue
        item_name = entry.get("name") or ""
        entry_id = entry.get("id") or ""
        events = entry.get("events") or []
        item_day_rates = day_rates_by_item.get(_norm_name(item_name)) or day_rates_by_item.get(item_name) or {}

        qty_est_total = 0.0
        qty_bill_total = 0.0
        qty_prev_total = 0.0
        amt_net_total = 0.0
        # Simple rate for summary: weighted by event days. We use per-event rate
        # to compute amount; report a notional "rate" = total_amt/total_qty when usable.
        for ev in events:
            ev_name = (ev.get("event_name") or "").strip()
            ev_days = int(_coerce_float(ev.get("days"), 0))
            ev_qty_est = _coerce_float(ev.get("qty"), 0.0)
            if not (ev_name and ev_days > 0):
                continue
            ev_id = ev.get("event_id") or ""
            rate = _rate_for_event_days(item_day_rates, ev_days)
            qty_bill = _coerce_float((exec_map.get(entry_id) or {}).get(ev_id), 0.0)
            qty_prev = _coerce_float((prev_exec.get(entry_id) or {}).get(ev_id), 0.0)
            # Summary: only count events that have actually been billed
            if qty_bill <= 0 and qty_prev <= 0:
                continue
            qty_net = qty_bill - qty_prev

            qty_est_total += ev_qty_est
            qty_bill_total += qty_bill
            qty_prev_total += qty_prev
            amt_net_total += qty_net * rate

        if qty_est_total <= 0 and qty_bill_total <= 0:
            continue
        summary.append({
            "item_name": item_name,
            "qty_est": round(qty_est_total, 2),
            "qty_prev": round(qty_prev_total, 2),
            "qty_bill": round(qty_bill_total, 2),
            "qty_net": round(qty_bill_total - qty_prev_total, 2),
            "amt_net": round(amt_net_total, 2),
        })
    return summary


def _download_bill_excel(entries, exec_map, prev_exec, day_rates_by_item, desc_by_item, unit_by_item, work_name, bill_number):
    """Build the Temp Bill Excel matching the normal-bill format.

    B1 (first bill): 8-col layout — S.No | Quantity | Unit | Item | Rate | Per | Unit | Amount
    B2+ (nth bills): 11-col layout — S.No | Item | Quantity Till Date | Unit | Rate per Unit
                     | Total Value till date | Deduct Previous (Qty, Amt) | Since Last (Qty, Amt) | Remarks

    Under each item the events appear as roman-numeral sub-rows (no S.No) with
    qty/rate/amount. Events with zero current+previous qty are skipped. AE{n}
    rows are appended per event when bill_qty exceeds the estimate.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill"

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFC8C8C8")
    subtotal_fill = PatternFill("solid", fgColor="FFE6E6E6")

    is_first = int(bill_number or 1) == 1
    n_cols = 8 if is_first else 11
    title_text = (
        f"FIRST AND PART BILL — TEMP WORKS BILL {bill_number}"
        if is_first
        else f"{bill_number}TH AND PART BILL — TEMP WORKS"
    )

    # Title + Name of the work
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1, value=title_text)
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    wn = ws.cell(row=2, column=1, value=f"Name of the work : {work_name}" if work_name else "Name of the work :")
    wn.font = Font(bold=True)
    wn.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(1, 3):
        for c_idx in range(1, n_cols + 1):
            ws.cell(row=r, column=c_idx).border = border_all

    # Column headers
    if is_first:
        header_row = 3
        headers = ["S.No", "Quantity", "Unit", "Item", "Rate", "Per", "Unit", "Amount"]
        for col_idx, text in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col_idx, value=text)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = header_fill
            c.border = border_all
        widths = {"A": 6, "B": 12, "C": 10, "D": 50, "E": 12, "F": 6, "G": 10, "H": 16}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        data_start = header_row + 1
    else:
        # Merged 2-row header for nth bill
        for col in [1, 2, 3, 4, 5, 6, 11]:
            ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        ws.merge_cells("G3:H3")
        ws.merge_cells("I3:J3")
        ws.cell(row=3, column=1, value="S.No")
        ws.cell(row=3, column=2, value="Item")
        ws.cell(row=3, column=3, value="Quantity Till Date")
        ws.cell(row=3, column=4, value="Unit")
        ws.cell(row=3, column=5, value="Rate per Unit")
        ws.cell(row=3, column=6, value="Total Value till date")
        ws.cell(row=3, column=7, value="Deduct Previous Measurements")
        ws.cell(row=3, column=9, value="Since Last Measurements")
        ws.cell(row=3, column=11, value="Remarks")
        ws.cell(row=4, column=7, value="Quantity")
        ws.cell(row=4, column=8, value="Amount")
        ws.cell(row=4, column=9, value="Quantity")
        ws.cell(row=4, column=10, value="Amount")
        for r in (3, 4):
            for col in range(1, 12):
                c = ws.cell(row=r, column=col)
                c.font = Font(bold=True)
                c.border = border_all
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = {"A": 6, "B": 50, "C": 14, "D": 8, "E": 12, "F": 16,
                  "G": 12, "H": 14, "I": 12, "J": 14, "K": 22}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        data_start = 5

    fmt_money = "#,##0.00"
    fmt_qty = "#,##0.##"

    def write_event_row(r, desc, qty, rate, qty_prev, remark, is_ae=False, sl_value=None, unit=""):
        """Write a single data row in the active layout."""
        amt = round(qty * rate, 2)
        if is_first:
            ws.cell(row=r, column=1, value=sl_value if sl_value is not None else None)
            ws.cell(row=r, column=2, value=round(qty, 2))
            ws.cell(row=r, column=3, value=unit or "")
            ws.cell(row=r, column=4, value=desc)
            ws.cell(row=r, column=5, value=round(rate, 2))
            ws.cell(row=r, column=6, value=1)
            ws.cell(row=r, column=7, value=unit or "")
            ws.cell(row=r, column=8, value=f"=ROUND(B{r}*E{r},2)")
            desc_col = 4
            money_cols = (5, 8)
            qty_cols = (2,)
        else:
            ws.cell(row=r, column=1, value=sl_value if sl_value is not None else None)
            ws.cell(row=r, column=2, value=desc)
            ws.cell(row=r, column=3, value=round(qty, 2))
            ws.cell(row=r, column=4, value=unit or "")
            ws.cell(row=r, column=5, value=round(rate, 2))
            ws.cell(row=r, column=6, value=f"=ROUND(C{r}*E{r},2)")
            ws.cell(row=r, column=7, value=round(qty_prev, 2))
            ws.cell(row=r, column=8, value=round(qty_prev * rate, 2))
            ws.cell(row=r, column=9, value=f"=C{r}-G{r}")
            ws.cell(row=r, column=10, value=f"=ROUND(F{r}-H{r},2)")
            ws.cell(row=r, column=11, value=remark or "")
            desc_col = 2
            money_cols = (5, 6, 8, 10)
            qty_cols = (3, 7, 9)
        for c_idx in range(1, n_cols + 1):
            c = ws.cell(row=r, column=c_idx)
            c.border = border_all
            if c_idx == desc_col:
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx in money_cols:
                c.number_format = fmt_money
            elif c_idx in qty_cols:
                c.number_format = fmt_qty
        return amt

    def write_item_header(r, sl, header_desc):
        """Header row carrying the item description (no qty/amount of its own)."""
        if is_first:
            ws.cell(row=r, column=1, value=sl)
            ws.cell(row=r, column=4, value=header_desc)
            desc_col = 4
        else:
            ws.cell(row=r, column=1, value=sl)
            ws.cell(row=r, column=2, value=header_desc)
            desc_col = 2
        for c_idx in range(1, n_cols + 1):
            c = ws.cell(row=r, column=c_idx)
            c.border = border_all
            if c_idx == desc_col:
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = data_start
    sl_counter = 1
    item_first_data_row = None

    for entry in entries or []:
        if (entry or {}).get("mode") != "multi":
            continue
        item_name = entry.get("name") or ""
        entry_id = entry.get("id") or ""
        events = entry.get("events") or []

        valid_events = []
        for ev in events:
            ev_name = (ev.get("event_name") or "").strip()
            try:
                ev_days = int(ev.get("days") or 0)
            except (TypeError, ValueError):
                ev_days = 0
            try:
                ev_qty = float(ev.get("qty") or 0)
            except (TypeError, ValueError):
                ev_qty = 0.0
            ev_id = ev.get("event_id") or ""
            qb = _coerce_float((exec_map.get(entry_id) or {}).get(ev_id), 0.0)
            qp = _coerce_float((prev_exec.get(entry_id) or {}).get(ev_id), 0.0)
            if not ev_name or ev_days <= 0:
                continue
            if qb <= 0 and qp <= 0:
                continue
            valid_events.append({
                "event_id": ev_id,
                "event_name": ev_name,
                "days": ev_days,
                "qty_est": ev_qty,
            })
        if not valid_events:
            continue

        # Item description header row
        header_desc = (desc_by_item or {}).get(item_name) or (desc_by_item or {}).get(_norm_name(item_name)) or item_name
        write_item_header(row_idx, sl_counter, header_desc)
        if item_first_data_row is None:
            item_first_data_row = row_idx
        row_idx += 1

        item_day_rates = day_rates_by_item.get(_norm_name(item_name)) or day_rates_by_item.get(item_name) or {}
        item_unit = (unit_by_item or {}).get(item_name) or (unit_by_item or {}).get(_norm_name(item_name)) or ""
        ae_counter = 0

        for i, ev in enumerate(valid_events, start=1):
            roman = _to_roman_lower(i)
            day_word = "day" if ev["days"] == 1 else "days"
            ev_desc = f"{roman}. {ev['event_name']} for {ev['days']} {day_word}"
            rate = _rate_for_event_days(item_day_rates, ev["days"])
            qty_est = float(ev["qty_est"])
            qty_bill = _coerce_float((exec_map.get(entry_id) or {}).get(ev["event_id"]), 0.0)
            qty_prev = _coerce_float((prev_exec.get(entry_id) or {}).get(ev["event_id"]), 0.0)
            excess = round(max(0.0, qty_bill - qty_est), 2)
            base_qty = qty_bill if excess <= 0 else qty_est

            remark = ""
            if excess > 0:
                remark = ""
            elif qty_bill == 0 and qty_prev > 0:
                remark = "Deleted"
            elif 0 < qty_bill < qty_est:
                remark = "Less as per estimated"

            write_event_row(row_idx, ev_desc, base_qty, rate, min(qty_prev, base_qty), remark, unit=item_unit)
            row_idx += 1

            if excess > 0:
                ae_counter += 1
                ae_prev = max(0.0, qty_prev - qty_est)
                write_event_row(
                    row_idx,
                    f"AE{ae_counter}",
                    excess,
                    rate,
                    ae_prev,
                    "Excess as per estimated",
                    is_ae=True,
                    unit=item_unit,
                )
                row_idx += 1

        sl_counter += 1

    last_data_row = row_idx - 1

    # ---- Totals ----
    if item_first_data_row is None:
        # Nothing was written (no billed events). Add a placeholder note.
        ws.cell(row=row_idx, column=1, value="No billed events.")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", work_name).strip("_") or "TempWorks"
        filename = f"{safe_name}_Bill_{bill_number}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    sub_row = row_idx
    if is_first:
        ws.cell(row=sub_row, column=4, value="Sub Total Amount")
        ws.cell(row=sub_row, column=8, value=f"=ROUND(SUM(H{item_first_data_row}:H{last_data_row}),2)")
        total_cols = (8,)
        desc_col = 4
    else:
        ws.cell(row=sub_row, column=2, value="Sub Total")
        ws.cell(row=sub_row, column=6, value=f"=ROUND(SUM(F{item_first_data_row}:F{last_data_row}),2)")
        ws.cell(row=sub_row, column=8, value=f"=ROUND(SUM(H{item_first_data_row}:H{last_data_row}),2)")
        ws.cell(row=sub_row, column=10, value=f"=ROUND(SUM(J{item_first_data_row}:J{last_data_row}),2)")
        total_cols = (6, 8, 10)
        desc_col = 2

    for c_idx in range(1, n_cols + 1):
        c = ws.cell(row=sub_row, column=c_idx)
        c.font = Font(bold=True)
        c.border = border_all
        c.fill = subtotal_fill
        if c_idx == desc_col:
            c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")
        if c_idx in total_cols:
            c.number_format = fmt_money

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", work_name).strip("_") or "TempWorks"
    filename = f"{safe_name}_Bill_{bill_number}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
